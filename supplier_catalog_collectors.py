"""Supplier-first catalog collectors backed by proven read-only endpoints.

Manager contributes authenticated clients and parsers only.  No collector
loads Manager's active/tracked-products catalog.
"""

from __future__ import annotations

import asyncio
import csv
import html as html_module
import hashlib
import json
import os
import re
import sys
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from urllib.parse import parse_qs, urlencode, urlparse, urlunparse

from abw_discovery import normalize_abw_candidates, valid_abw_identifier
from qogita_discovery import normalize_qogita_candidates
from qudo_discovery import normalize_qudo_candidates, valid_qudo_identifier
from supplier_catalog import (
    SupplierCatalogGeneration,
    canonical_gtin14,
    candidates_to_cache_records,
    supplier_product_cache_key,
)
from umma_discovery import normalize_umma_candidates


MANAGER_ROOT = Path(__file__).resolve().parent.parent / "Glow-Up-Manager"
QOGITA_COVERAGE = {
    "type": "partial_catalog",
    "description": (
        "Union dei cataloghi dei seller alias Qogita configurati; non equivale "
        "al catalogo buyer completo (la ricerca globale sincrona è limitata a 10.000 record)"
    ),
    "complete": False,
}
QOGITA_EXPORT_COLUMNS = (
    "GTIN", "Name", "Category", "Brand", "€ Lowest Price inc. shipping",
    "Unit", "Lowest Priced Offer Inventory", "Is a pre-order?",
    "Estimated Delivery Time (weeks)", "Number of Offers",
    "Total Inventory of All Offers", "Product Link",
)
QOGITA_EXPORT_COVERAGE = {
    "type": "filtered_catalog",
    "description": (
        "Export asincrono buyer Qogita; la copertura prodotto dipende dai filtri "
        "persistiti con la richiesta e resta distinta dall'enrichment offerte/tier"
    ),
    "complete": False,
}
UMMA_COVERAGE = {
    "type": "partial_catalog",
    "description": (
        "Enumerazione globale UMMA /search/product senza keyword; il conteggio "
        "restituito non coincide ancora con righe e ID unici enumerati"
    ),
    "complete": False,
}
ABW_COVERAGE = {
    "type": "full_relevant_catalog",
    "description": (
        "Export ufficiale ABW Beauty per tutti i brand; enumera le righe catalogo "
        "ma non contiene tutte le fasce Standard necessarie alla copertura scenari"
    ),
    "complete": True,
}
QUDO_COVERAGE = {
    "type": "full_relevant_catalog",
    "description": (
        "Indice pubblico WooCommerce filtrato alle variation Seller=QUDO; una scansione "
        "completa identifier-only ha dimostrato il traversal, ma non tutte le offerte "
        "espongono un GTIN valido e il full scenario enrichment resta separato"
    ),
    "complete": True,
}


class SupplierCollectorError(RuntimeError):
    """Collector failure with persistence-safe, non-secret diagnostics."""

    def __init__(self, message, *, code="collector_failed", diagnostics=None):
        super().__init__(message)
        self.code = code
        self.diagnostics = dict(diagnostics or {})


def _qudo_collector_error(exc, *, phase, client, **context):
    diagnostics = {
        "phase": phase,
        "request_count": int(getattr(client, "request_count", 0) or 0),
        "remote_status": getattr(exc, "remote_status", None),
        **{key: value for key, value in context.items() if value is not None},
    }
    return SupplierCollectorError(
        f"Qudo collector failed during {phase}",
        code=getattr(exc, "code", "qudo_collection_failed"),
        diagnostics=diagnostics,
    )


def _load_manager_environment(manager_root: Path = MANAGER_ROOT) -> None:
    source = manager_root / "src"
    if not source.is_dir():
        raise RuntimeError("Manager source is unavailable")
    if str(source) not in sys.path:
        sys.path.insert(0, str(source))
    env_path = manager_root / ".env"
    if not env_path.is_file():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        clean = line.strip()
        if not clean or clean.startswith("#") or "=" not in clean:
            continue
        key, value = clean.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def _iso_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _apply_run_id(candidates, run_id, observed_at):
    for candidate in candidates:
        for scenario in candidate.get("scenarios") or []:
            scenario["snapshot_id"] = run_id or "dry-run"
            scenario["snapshot_at"] = observed_at
            scenario["freshness_status"] = "fresh"
    return candidates


def umma_enumeration_proof(total_count, row_ids):
    clean = [str(value) for value in row_ids if str(value or "")]
    unique = set(clean)
    numeric = [int(value) for value in clean if value.isdigit()]
    return {
        "source_count": int(total_count or 0),
        "enumerated_count": len(clean),
        "unique_count": len(unique),
        "duplicate_count": len(clean) - len(unique),
        "enumeration_gap": max(0, int(total_count or 0) - len(unique)),
        "stable_order": "orderById=DESC",
        "monotonic_desc": len(numeric) == len(clean) and all(
            left > right for left, right in zip(numeric, numeric[1:])
        ),
    }


def compare_umma_enumerations(first_ids, second_ids):
    first, second = set(map(str, first_ids)), set(map(str, second_ids))
    return {
        "same_set": first == second,
        "only_first": len(first - second),
        "only_second": len(second - first),
        "union_count": len(first | second),
    }


def _catalog_product(
    supplier, supplier_product_id, *, canonical_ean=None, raw_identifiers=(),
    identifier_type=None, supplier_option_id=None, supplier_sku=None,
    brand="", title="", metadata=None,
):
    return {
        "canonical_product_key": supplier_product_cache_key(
            supplier, supplier_product_id,
            supplier_option_id=supplier_option_id, supplier_sku=supplier_sku,
            fallback_identifier=canonical_ean,
        ),
        "canonical_ean": canonical_ean,
        "canonical_gtin": canonical_gtin14(canonical_ean),
        "identifier_type": identifier_type,
        "raw_identifiers": list(raw_identifiers),
        "supplier_product_id": str(supplier_product_id or ""),
        "supplier_option_id": str(supplier_option_id or "") or None,
        "supplier_sku": str(supplier_sku or "") or None,
        "brand": str(brand or ""), "title": str(title or ""),
        "size_value": None, "size_unit": None, "pack_count": None,
        "metadata": dict(metadata or {}),
    }


def _qogita_variant_fid_from_url(value):
    path = urlparse(str(value or "").strip()).path
    match = re.search(r"/products/([^/]+)", path, re.I)
    return match.group(1) if match else None


def _normalized_decimal_text(value):
    try:
        decimal = Decimal(str(value or "").strip().replace(",", "."))
    except InvalidOperation as exc:
        raise ValueError(f"Invalid Qogita catalog decimal: {value!r}") from exc
    return format(decimal.normalize(), "f")


def qogita_catalog_fingerprint(row):
    """Fingerprint catalog signals; it is not proof that offer tiers are unchanged."""
    fields = {
        "gtin": str(row.get("GTIN") or "").strip(),
        "name": str(row.get("Name") or "").strip(),
        "category": str(row.get("Category") or "").strip(),
        "brand": str(row.get("Brand") or "").strip(),
        "lowest_price": _normalized_decimal_text(row.get("€ Lowest Price inc. shipping")),
        "unit": int(str(row.get("Unit") or "0").strip()),
        "lowest_offer_inventory": int(str(row.get("Lowest Priced Offer Inventory") or "0").strip()),
        "preorder": str(row.get("Is a pre-order?") or "").strip().casefold(),
        "delivery_weeks": str(row.get("Estimated Delivery Time (weeks)") or "").strip(),
        "number_of_offers": int(str(row.get("Number of Offers") or "0").strip()),
        "total_inventory": int(str(row.get("Total Inventory of All Offers") or "0").strip()),
        "product_link": str(row.get("Product Link") or "").strip(),
    }
    return hashlib.sha256(
        json.dumps(fields, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()


class QogitaCatalogExportReader:
    """Stream official Qogita CSV exports without loading the catalog in memory."""

    coverage = QOGITA_EXPORT_COVERAGE

    def __init__(self, source_file):
        self.source_file = Path(source_file).expanduser().resolve()
        if not self.source_file.is_file():
            raise FileNotFoundError(self.source_file)

    def metadata(self):
        metadata = {}
        with self.source_file.open("r", encoding="utf-8-sig", newline="") as handle:
            for values in csv.reader(handle):
                if "GTIN" in values and "Name" in values:
                    metadata["columns"] = [value.strip() for value in values]
                    break
                if len(values) >= 2 and values[0].strip():
                    metadata[values[0].strip()] = values[1].strip()
        metadata["source_file_name"] = self.source_file.name
        metadata["source_type"] = "official_qogita_async_catalog_export"
        return metadata

    def rows(self, *, limit=None):
        if limit is not None and int(limit) <= 0:
            raise ValueError("limit must be a positive integer")
        with self.source_file.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            header = None
            for values in reader:
                if "GTIN" in values and "Name" in values:
                    header = [value.strip() for value in values]
                    break
            if header is None or any(column not in header for column in QOGITA_EXPORT_COLUMNS):
                raise ValueError("Qogita catalog header is missing required columns")
            emitted = 0
            for index, values in enumerate(reader, start=1):
                if not values or not any(str(value).strip() for value in values):
                    continue
                padded = values + [""] * max(0, len(header) - len(values))
                yield index, dict(zip(header, padded))
                emitted += 1
                if limit is not None and emitted >= int(limit):
                    break

    @staticmethod
    def product_from_row(raw, *, row_number=None):
        gtin = str(raw.get("GTIN") or "").strip()
        canonical = canonical_gtin14(gtin)
        if not canonical:
            raise ValueError(
                f"Invalid Qogita GTIN at catalog row {row_number or '?'}: {gtin!r}"
            )
        product_url = str(raw.get("Product Link") or "").strip() or None
        variant_fid = _qogita_variant_fid_from_url(product_url)
        fingerprint = qogita_catalog_fingerprint(raw)
        source_row = str(raw.get("Source Row") or "").strip() or None
        return {
                    "canonical_product_key": supplier_product_cache_key(
                        "qogita", gtin, fallback_identifier=gtin,
                    ),
                    "canonical_ean": gtin if len(gtin) == 13 else None,
                    "canonical_gtin": canonical,
                    "identifier_type": "EAN" if len(gtin) == 13 else (
                        "UPC" if len(gtin) == 12 else "GTIN"
                    ),
                    "raw_identifiers": [{"value": gtin, "type": "GTIN_EXPORT"}],
                    "supplier_product_id": gtin,
                    "supplier_option_id": None,
                    "supplier_sku": None,
                    "brand": str(raw.get("Brand") or "").strip(),
                    "title": str(raw.get("Name") or "").strip(),
                    "size_value": None,
                    "size_unit": None,
                    "pack_count": None,
                    "catalog_fingerprint": fingerprint,
                    "variant_fid": variant_fid,
                    "product_url": product_url,
                    "enrichment_status": (
                        "unresolved_variant" if not variant_fid else "enrichment_pending"
                    ),
                    "offer_tier_observed_at": None,
                    "metadata": {
                        "category": str(raw.get("Category") or "").strip(),
                        "lowest_price_inc_shipping": _normalized_decimal_text(
                            raw.get("€ Lowest Price inc. shipping")
                        ),
                        "unit": int(str(raw.get("Unit") or "0").strip()),
                        "lowest_priced_offer_inventory": int(str(
                            raw.get("Lowest Priced Offer Inventory") or "0"
                        ).strip()),
                        "is_preorder": str(raw.get("Is a pre-order?") or "").strip().casefold() == "yes",
                        "estimated_delivery_weeks": (
                            int(str(raw.get("Estimated Delivery Time (weeks)")).strip())
                            if str(raw.get("Estimated Delivery Time (weeks)") or "").strip()
                            else None
                        ),
                        "number_of_offers": int(str(raw.get("Number of Offers") or "0").strip()),
                        "total_inventory_all_offers": int(str(
                            raw.get("Total Inventory of All Offers") or "0"
                        ).strip()),
                        "source_row": source_row,
                        "catalog_fingerprint": fingerprint,
                        "variant_fid": variant_fid,
                        "product_url": product_url,
                        "commercial_values_are_indicative": True,
                    },
        }

    def products(self, *, limit=None, skip_invalid=False):
        for row_number, raw in self.rows(limit=limit):
            try:
                yield self.product_from_row(raw, row_number=row_number)
            except (TypeError, ValueError):
                if not skip_invalid:
                    raise


def _generation(coverage, candidates, *, pages, requests, retries=0,
                rate_limits=0, server_errors=0, diagnostics=None, complete=None,
                catalog_products=(), scenario_enrichment_status=None,
                scenario_enrichment_count=None):
    products, scenarios = candidates_to_cache_records(candidates)
    # Acquisition records retain supplier-owned raw identifiers and richer source
    # metadata; scenario-derived records are only a fallback for adapters without
    # a separate catalog row.
    merged_products = {row["canonical_product_key"]: row for row in products}
    merged_products.update({
        row["canonical_product_key"]: dict(row) for row in catalog_products
    })
    diagnostics = diagnostics or {}
    return SupplierCatalogGeneration(
        supplier=str((candidates[0].get("scenarios") or [{}])[0].get("supplier") if candidates else diagnostics.get("supplier")),
        coverage_type=coverage["type"],
        coverage_description=coverage["description"],
        coverage_complete=coverage["complete"] if complete is None else bool(complete),
        products=tuple(merged_products.values()), scenarios=scenarios, page_count=pages,
        request_count=requests, retry_count=retries,
        rate_limit_count=rate_limits, server_error_count=server_errors,
        source_type=diagnostics.get("source_type"),
        source_count=diagnostics.get("source_count"),
        enumerated_count=diagnostics.get("enumerated_count"),
        unique_count=diagnostics.get("unique_count"),
        completeness_status=str(diagnostics.get("completeness_status") or "partial_catalog"),
        completeness_reason=str(diagnostics.get("completeness_reason") or coverage["description"]),
        export_generated_at=diagnostics.get("export_generated_at"),
        upstream_catalog_version=diagnostics.get("upstream_catalog_version"),
        product_catalog_coverage_type=str(
            diagnostics.get("product_catalog_coverage_type") or coverage["type"]
        ),
        product_catalog_coverage_complete=(
            coverage["complete"] if complete is None else bool(complete)
        ),
        scenario_enrichment_status=(
            scenario_enrichment_status
            or ("partial" if scenarios else "none")
        ),
        scenario_enrichment_count=(
            len(scenarios) if scenario_enrichment_count is None
            else int(scenario_enrichment_count)
        ),
        scenario_enrichment_observed_at=diagnostics.get("export_generated_at"),
        diagnostics=diagnostics,
    )


class QogitaCatalogCollector:
    coverage = QOGITA_COVERAGE

    def __init__(self, seller_aliases, *, manager_root=MANAGER_ROOT, pacing_seconds=0.4):
        self.seller_aliases = tuple(sorted({str(v).strip() for v in seller_aliases if str(v).strip()}))
        self.manager_root = Path(manager_root)
        self.pacing_seconds = pacing_seconds
        if not self.seller_aliases:
            raise ValueError("At least one Qogita seller alias is required")

    def __call__(self, *, run_id, limit=None, dry_run=False):
        return asyncio.run(self._collect(run_id=run_id, limit=limit))

    async def _collect(self, *, run_id, limit):
        _load_manager_environment(self.manager_root)
        from purchase_prices.qogita_seller_catalog import parse_seller_catalog_page
        from qogita.client import QogitaClient
        from qogita.service import load_config

        client = QogitaClient(load_config())
        rows = []
        catalog_products = []
        pages = 0
        observed = _iso_now()
        try:
            await client.login()
            for alias in self.seller_aliases:
                page = 1
                expected_total = None
                collected = 0
                while True:
                    payload = await client.get_seller_catalog_page(alias, page=page, size=100)
                    pages += 1
                    products, total, has_next = parse_seller_catalog_page(payload, seller_alias=alias)
                    if expected_total is None:
                        expected_total = total
                    elif total != expected_total:
                        raise RuntimeError("Qogita count changed during pagination")
                    for product in products:
                        collected += 1
                        catalog_products.append(_catalog_product(
                            "qogita", product.variant_fid,
                            canonical_ean=product.gtin,
                            raw_identifiers=({"value": product.gtin, "type": "EAN"},),
                            identifier_type="EAN", supplier_option_id=product.offer_qid,
                            brand=product.brand, title=product.name,
                            metadata={
                                "seller_alias": alias, "category": product.category_name,
                                "image_url": product.image_url, "product_url": product.product_url,
                                "inventory": product.inventory, "selling_unit": product.selling_unit,
                            },
                        ))
                        for tier in product.tiers:
                            rows.append({
                                "run_id": run_id or "dry-run", "seller_alias": alias,
                                "gtin": product.gtin, "variant_fid": product.variant_fid,
                                "offer_qid": product.offer_qid, "product_name": product.name,
                                "brand": product.brand, "category_name": product.category_name,
                                "image_url": product.image_url, "inventory": product.inventory,
                                "selling_unit": product.selling_unit, "product_url": product.product_url,
                                "observed_at": observed, "tier_mov": str(tier.mov),
                                "currency": tier.currency, "tier_price": str(tier.price),
                                "is_active": tier.is_active,
                            })
                        if limit and len({row["gtin"] for row in rows}) >= limit:
                            break
                    if limit and len({row["gtin"] for row in rows}) >= limit:
                        break
                    if collected >= total:
                        break
                    if not has_next or not products:
                        raise RuntimeError("Qogita pagination ended before total count")
                    page += 1
                    await asyncio.sleep(self.pacing_seconds)
                if limit and len({row["gtin"] for row in rows}) >= limit:
                    break
        finally:
            await client.close()
        candidates, normalizer_diagnostics = normalize_qogita_candidates(
            rows, now=datetime.now(timezone.utc)
        )
        _apply_run_id(candidates, run_id, observed)
        diagnostics = {
            "supplier": "qogita", "seller_aliases": list(self.seller_aliases),
            "source_type": "configured_seller_catalogs",
            "source_count": len(catalog_products),
            "enumerated_count": len(catalog_products),
            "unique_count": len({row["canonical_product_key"] for row in catalog_products}),
            "completeness_status": "full_relevant_catalog" if limit is None else "partial_catalog",
            "completeness_reason": (
                "Official full-account export was identified, but this collector still "
                "enumerates only configured seller aliases"
            ),
            "normalizer": normalizer_diagnostics,
        }
        return _generation(
            self.coverage, candidates, pages=pages, requests=client.request_count,
            diagnostics=diagnostics, complete=self.coverage["complete"] and limit is None,
            catalog_products=catalog_products,
        )


class UmmaCatalogCollector:
    coverage = UMMA_COVERAGE
    PAGE_SIZE = 100

    def __init__(self, *, manager_root=MANAGER_ROOT, pacing_seconds=0.2):
        self.manager_root = Path(manager_root)
        self.pacing_seconds = pacing_seconds

    def __call__(self, *, run_id, limit=None, dry_run=False):
        return asyncio.run(self._collect(run_id=run_id, limit=limit))

    async def _collect(self, *, run_id, limit):
        _load_manager_environment(self.manager_root)
        from purchase_prices.fx import EcbFxClient, FxError
        from purchase_prices.umma import (
            ACCOUNT_MINIMUM_CURRENCY, ACCOUNT_MINIMUM_ORDER, PRICE_BASIS,
            PRICE_BASIS_SOURCE, PRICING_SCOPE, SOURCE, UmmaClient, UmmaError,
            normalize_offer, parse_product_modes,
        )

        client, fx_client = UmmaClient(), EcbFxClient()
        pages = 0
        observed = _iso_now()
        details = []
        catalog_products = []
        try:
            skip, total, product_ids, product_id_set, search_product_ids = 0, None, [], set(), []
            while total is None or skip < total:
                payload = await client._get("/search/product", params={
                    "skip": str(skip), "take": str(self.PAGE_SIZE),
                    "orderById": "DESC",
                })
                pages += 1
                current_total = int(payload.get("totalCount") or 0)
                if total is None:
                    total = current_total
                elif current_total != total:
                    raise RuntimeError("UMMA totalCount changed during pagination")
                items = payload.get("items") or []
                for item in items:
                    product_id = str(item.get("id") or "") if isinstance(item, dict) else ""
                    if product_id:
                        search_product_ids.append(product_id)
                    if product_id and product_id not in product_id_set:
                        product_id_set.add(product_id)
                        product_ids.append(product_id)
                        if limit and len(product_ids) >= limit:
                            break
                if limit and len(product_ids) >= limit:
                    break
                skip += self.PAGE_SIZE
                if skip < total:
                    await asyncio.sleep(self.pacing_seconds)
            try:
                fx = await fx_client.latest_usd_to_eur()
            except FxError:
                fx = None
            rows = []
            rejected = []
            for index, product_id in enumerate(product_ids):
                if index:
                    await asyncio.sleep(self.pacing_seconds)
                detail = await client._get(f"/product/{product_id}")
                details.append(detail)
                option_barcodes = []
                for mapper in detail.get("mapperSaleProducts") or []:
                    option = mapper.get("productOption") if isinstance(mapper, dict) else None
                    barcode = str((option or {}).get("barcode") or "").strip()
                    if barcode and barcode not in option_barcodes:
                        option_barcodes.append(barcode)
                    if not isinstance(option, dict):
                        continue
                    canonical = None
                    identifier_type = None
                    if re.fullmatch(r"\d{13}ED", barcode):
                        from umma_discovery import valid_ean13
                        canonical = barcode[:13] if valid_ean13(barcode[:13]) else None
                        identifier_type = "EAN" if canonical else None
                    elif re.fullmatch(r"\d{8}|\d{12}|\d{13}|\d{14}", barcode):
                        canonical = barcode
                        identifier_type = {8: "GTIN-8", 12: "UPC", 13: "EAN", 14: "GTIN"}[len(barcode)]
                    catalog_products.append(_catalog_product(
                        "umma", product_id, canonical_ean=canonical,
                        raw_identifiers=({"value": barcode, "type": "UMMA_BARCODE"},) if barcode else (),
                        identifier_type=identifier_type,
                        supplier_option_id=option.get("id"),
                        supplier_sku=option.get("sku") or option.get("erpSku"),
                        brand=detail.get("brandName") or detail.get("brand"),
                        title=option.get("englishName") or detail.get("englishName"),
                        metadata={"mapper_sale_product_id": mapper.get("id")},
                    ))
                for barcode in option_barcodes:
                    try:
                        offers = parse_product_modes(detail, expected_gtin=barcode)
                    except UmmaError as exc:
                        if exc.code in {"not_found", "no_offers"}:
                            rejected.append({"product_id": product_id, "barcode": barcode, "reason": exc.code})
                            continue
                        raise
                    for normalized in (normalize_offer(offer, fx) for offer in offers):
                        offer = normalized.source
                        rows.append({
                            "run_id": run_id or "dry-run", "gtin": offer.gtin,
                            "supplier_product_id": offer.supplier_product_id,
                            "mapper_sale_product_id": offer.mapper_sale_product_id,
                            "product_option_id": offer.product_option_id,
                            "supplier_sku": offer.supplier_sku, "product_name": offer.product_name,
                            "sales_mode": offer.sales_mode, "observed_at": observed,
                            "original_unit_price": str(offer.original_unit_price),
                            "original_currency": offer.original_currency,
                            "price_basis": PRICE_BASIS, "price_basis_source": PRICE_BASIS_SOURCE,
                            "fx_usd_to_eur_rate": str(fx.usd_to_eur) if fx else None,
                            "fx_reference_rate": str(fx.eur_to_usd_reference) if fx else None,
                            "fx_rate_date": fx.rate_date.isoformat() if fx else None,
                            "fx_source": fx.source if fx else None, "fx_stale": bool(fx.stale) if fx else False,
                            "net_unit_price_eur": str(normalized.net_unit_price_eur) if normalized.net_unit_price_eur is not None else None,
                            "vat_rate_percent": "22", "vat_amount_eur": str(normalized.vat_amount_eur) if normalized.vat_amount_eur is not None else None,
                            "gross_unit_price_eur": str(normalized.gross_unit_price_eur) if normalized.gross_unit_price_eur is not None else None,
                            "available_quantity": offer.available_quantity,
                            "availability_status": offer.availability_status,
                            "minimum_product_quantity": offer.minimum_product_quantity,
                            "selling_unit": offer.selling_unit, "maximum_quantity": offer.maximum_quantity,
                            "lead_time": offer.lead_time, "minimum_order_value": str(ACCOUNT_MINIMUM_ORDER),
                            "minimum_order_currency": ACCOUNT_MINIMUM_CURRENCY,
                            "pricing_scope": PRICING_SCOPE, "source": SOURCE,
                        })
        finally:
            await client.close()
            await fx_client.close()
        candidates, normalizer_diagnostics = normalize_umma_candidates(
            rows, now=datetime.now(timezone.utc)
        )
        _apply_run_id(candidates, run_id, observed)
        proof = umma_enumeration_proof(total, search_product_ids)
        diagnostics = {
            "supplier": "umma", "search_total_count": total,
            "search_rows_received": len(search_product_ids),
            "unique_product_ids": len(product_ids),
            **proof,
            "enumeration_order": "orderById=DESC",
            "source_type": "global_search_index",
            "source_count": total,
            "enumerated_count": len(search_product_ids),
            "unique_count": len(product_ids),
            "completeness_status": "partial_catalog",
            "completeness_reason": (
                "Stable ID ordering is proven, but upstream totalCount exceeds the "
                "terminating enumerated set"
            ),
            "product_details_loaded": len(details), "rejected_options": rejected,
            "normalizer": normalizer_diagnostics,
        }
        return _generation(
            self.coverage, candidates, pages=pages,
            requests=client.request_count + int(getattr(fx_client, "request_count", 0) or 0),
            diagnostics=diagnostics, complete=False,
            catalog_products=catalog_products,
        )


class AbwCatalogCollector:
    coverage = ABW_COVERAGE
    FIRST_PAGE = "/en/beauty/list.html/bcc.15022_bpt.46"

    def __init__(self, *, manager_root=MANAGER_ROOT, pacing_seconds=0.25,
                 source_file=None):
        self.manager_root = Path(manager_root)
        self.pacing_seconds = pacing_seconds
        self.source_file = Path(source_file).expanduser().resolve() if source_file else None

    def __call__(self, *, run_id, limit=None, dry_run=False):
        if self.source_file is not None:
            return self._collect_export(run_id=run_id, limit=limit)
        if limit is None:
            raise RuntimeError(
                "ABW full acquisition is not enabled: integrate and validate the official All Products catalog download first"
            )
        if int(limit) > 36:
            raise ValueError("ABW audit samples are limited to the first 36 All Beauty records")
        return asyncio.run(self._collect(run_id=run_id, limit=limit))

    @staticmethod
    def _decimal(value):
        if value is None or isinstance(value, bool):
            return None
        try:
            parsed = Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError):
            return None
        return parsed if parsed.is_finite() and parsed > 0 else None

    @staticmethod
    def _export_timestamp(value):
        if isinstance(value, datetime):
            parsed = value
        else:
            parsed = datetime.strptime(str(value or "").strip(), "%d-%b-%Y")
        return parsed.replace(tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")

    def _collect_export(self, *, run_id, limit):
        """Parse an official ABW Beauty XLSX without contacting ABW or Manager."""
        if not self.source_file.is_file():
            raise FileNotFoundError(self.source_file)
        from openpyxl import load_workbook

        workbook = load_workbook(self.source_file, read_only=True, data_only=True)
        try:
            if "ABW Beauty Product Catalog" not in workbook.sheetnames:
                raise RuntimeError("Unsupported ABW export sheet")
            sheet = workbook["ABW Beauty Product Catalog"]
            # Apache POI emits this export with a stale A1:A1 worksheet dimension.
            # Reset it so openpyxl streams the complete XML rather than row 1 only.
            if sheet.max_row == 1 and sheet.max_column == 1:
                sheet.reset_dimensions()
            export_generated_at = self._export_timestamp(sheet.cell(1, 2).value)
            headers = [
                str(cell.value or "").strip()
                for cell in next(sheet.iter_rows(min_row=2, max_row=2, max_col=24))
            ]
            expected = [
                "ABW \nCatalog No.", "UPC", "Brand", "Product Name", "Option Name",
                "Wholesale Unit \nWeight / pc (g) \n#", "ABW Selling \nPrice (EUR)",
                "Availability", "Box Qty 1", "Box Selling \nPrice (EUR)",
                "Price / piece \n(EUR)", "Box Qty 1 \nAvailability", "Box Qty 2",
                "Box Selling \nPrice (EUR)", "Price / piece \n(EUR)",
                "Box Qty 2 \nAvailability", "Box Qty 3", "Box Selling \nPrice (EUR)",
                "Price / piece \n(EUR)", "Box Qty 3 \nAvailability", "Box Qty 4",
                "Box Selling \nPrice (EUR)", "Price / piece \n(EUR)",
                "Box Qty 4 \nAvailability",
            ]
            if headers != expected:
                raise RuntimeError("Unsupported ABW export headers")

            rows = []
            catalog_products = []
            invalid_identifiers = []
            missing_box_totals = []
            source_count = 0
            for values in sheet.iter_rows(min_row=3, max_col=24, values_only=True):
                catalog_no = str(values[0] or "").strip()
                if not catalog_no:
                    continue
                if catalog_no.startswith(("* ", "# ")):
                    continue
                source_count += 1
                if limit and source_count > int(limit):
                    break
                raw_identifier = str(values[1] or "").strip()
                composite = re.fullmatch(r"(\d{13})\s*x\s*(\d+)", raw_identifier, re.I)
                composite_quantity = int(composite.group(2)) if composite else None
                composite_ean = composite.group(1) if composite else None
                canonical = (
                    raw_identifier if valid_abw_identifier(raw_identifier)
                    else composite_ean if composite_ean and valid_abw_identifier(composite_ean)
                    else None
                )
                brand = str(values[2] or "").strip()
                product_name = str(values[3] or "").strip()
                option_name = str(values[4] or "").strip()
                title = " · ".join(value for value in (product_name, option_name) if value)
                availability = str(values[7] or "").strip()
                metadata = {
                    "catalog_no": catalog_no,
                    "option_name": option_name,
                    "wholesale_unit_weight_g": values[5],
                    "availability_text": availability,
                    "source_file_name": self.source_file.name,
                    "source_scope": "ABW Beauty Product Catalog",
                }
                catalog_products.append(_catalog_product(
                    "abw", catalog_no, canonical_ean=canonical,
                    raw_identifiers=({
                        "value": raw_identifier,
                        "type": "EAN_X_QUANTITY" if composite_quantity else "UPC_EXPORT",
                    },)
                    if raw_identifier else (),
                    identifier_type={8: "GTIN-8", 12: "UPC", 13: "EAN", 14: "GTIN"}.get(
                        len(canonical or "")
                    ),
                    supplier_option_id=catalog_no,
                    brand=brand, title=title, metadata=metadata,
                ))
                if not canonical:
                    invalid_identifiers.append({"catalog_no": catalog_no, "raw_identifier": raw_identifier})
                    for slot in range(4):
                        offset = 8 + slot * 4
                        quantity = self._decimal(values[offset])
                        total = self._decimal(values[offset + 1])
                        shown_unit = self._decimal(values[offset + 2])
                        if quantity is not None and total is None:
                            missing_box_totals.append({
                                "catalog_no": catalog_no, "slot": slot + 1,
                                "quantity": str(quantity),
                                "shown_unit_price": str(shown_unit) if shown_unit else None,
                            })
                    continue
                base_price = self._decimal(values[6])
                if composite_quantity and base_price is not None:
                    rows.append({
                        "run_id": run_id or "dry-run", "seller_sku": catalog_no,
                        "gtin": canonical, "supplier_product_id": catalog_no,
                        "option_product_id": catalog_no, "product_name": title,
                        "brand": brand, "mode": "bulk_box",
                        "condition_key": f"export_bulk_product_{composite_quantity}",
                        "condition_label": f"Box {composite_quantity}",
                        "tier_min_quantity": composite_quantity, "tier_max_quantity": None,
                        "pack_size": composite_quantity, "pack_price": str(base_price),
                        "net_unit_price_eur": None,
                        "displayed_unit_price": str(self._decimal(values[10]) or ""),
                        "currency": "EUR",
                        "price_source": "official_abw_catalog_export_bulk_product_total",
                        "price_basis": "net_pack_total_divided_by_quantity",
                        "available_quantity": None, "availability_status": "available_to_order",
                        "stock_text": availability, "lead_time": availability,
                        "warehouse": "unspecified", "discount_label": "",
                        "product_url": "", "minimum_order_value": "250",
                        "minimum_order_currency": "USD", "observed_at": export_generated_at,
                        "source": "official_abw_catalog_export",
                    })
                    continue
                if base_price is not None:
                    rows.append({
                        "run_id": run_id or "dry-run", "seller_sku": catalog_no,
                        "gtin": canonical, "supplier_product_id": catalog_no,
                        "option_product_id": catalog_no, "product_name": title,
                        "brand": brand, "mode": "standard",
                        "condition_key": "export_standard",
                        "condition_label": "Standard (prezzo catalogo)",
                        "tier_min_quantity": 1, "tier_max_quantity": None,
                        "pack_size": None, "pack_price": None,
                        "net_unit_price_eur": str(base_price), "currency": "EUR",
                        "price_source": "official_abw_catalog_export",
                        "price_basis": "net_unit_price", "available_quantity": None,
                        "availability_status": "available_to_order",
                        "stock_text": availability, "lead_time": availability,
                        "warehouse": "unspecified", "discount_label": "",
                        "product_url": "", "minimum_order_value": "250",
                        "minimum_order_currency": "USD", "observed_at": export_generated_at,
                        "source": "official_abw_catalog_export",
                    })
                for slot in range(4):
                    offset = 8 + slot * 4
                    quantity = self._decimal(values[offset])
                    total = self._decimal(values[offset + 1])
                    shown_unit = self._decimal(values[offset + 2])
                    box_availability = str(values[offset + 3] or "").strip()
                    if quantity is None:
                        continue
                    if total is None:
                        missing_box_totals.append({
                            "catalog_no": catalog_no, "slot": slot + 1,
                            "quantity": str(quantity),
                            "shown_unit_price": str(shown_unit) if shown_unit else None,
                        })
                        continue
                    rows.append({
                        "run_id": run_id or "dry-run", "seller_sku": catalog_no,
                        "gtin": canonical, "supplier_product_id": catalog_no,
                        "option_product_id": catalog_no, "product_name": title,
                        "brand": brand, "mode": "bulk_box",
                        "condition_key": f"export_box_{slot + 1}_{int(quantity)}",
                        "condition_label": f"Box {int(quantity)}", "tier_min_quantity": int(quantity),
                        "tier_max_quantity": None, "pack_size": int(quantity),
                        "pack_price": str(total), "net_unit_price_eur": str(shown_unit) if shown_unit else None,
                        "displayed_unit_price": str(shown_unit) if shown_unit else None,
                        "currency": "EUR", "price_source": "official_abw_catalog_export_box_total",
                        "price_basis": "net_pack_total_divided_by_quantity",
                        "available_quantity": None, "availability_status": "available_to_order",
                        "stock_text": box_availability, "lead_time": box_availability,
                        "warehouse": "unspecified", "discount_label": "",
                        "product_url": "", "minimum_order_value": "250",
                        "minimum_order_currency": "USD", "observed_at": export_generated_at,
                        "source": "official_abw_catalog_export",
                    })
        finally:
            workbook.close()

        candidates, normalizer_diagnostics = normalize_abw_candidates(
            rows, now=datetime.fromisoformat(export_generated_at.replace("Z", "+00:00")),
        )
        _apply_run_id(candidates, run_id, export_generated_at)
        diagnostics = {
            "supplier": "abw", "source_type": "official_abw_beauty_xlsx",
            "source_count": source_count, "enumerated_count": source_count,
            "unique_count": len({row["canonical_product_key"] for row in catalog_products}),
            "completeness_status": "full_relevant_catalog" if limit is None else "partial_catalog",
            "completeness_reason": (
                "Product/option rows and authoritative box totals are enumerated, but the export "
                "does not expose Standard quantity ranges, warehouse, numeric stock, or separate product/option IDs"
            ),
            "export_generated_at": export_generated_at,
            "upstream_catalog_version": f"abw-beauty-{export_generated_at[:10]}",
            "source_file_name": self.source_file.name,
            "invalid_or_missing_identifiers": invalid_identifiers,
            "bulk_boxes_missing_authoritative_total": missing_box_totals,
            "normalizer": normalizer_diagnostics,
        }
        return _generation(
            self.coverage, candidates, pages=1, requests=0,
            diagnostics=diagnostics, complete=limit is None,
            scenario_enrichment_status="partial" if rows else "none",
            catalog_products=catalog_products,
        )

    async def _collect(self, *, run_id, limit):
        _load_manager_environment(self.manager_root)
        from purchase_prices.abw import (
            BASE_URL, MINIMUM_ORDER_CURRENCY, MINIMUM_ORDER_VALUE, PRICE_BASIS,
            PRICE_SOURCE, SOURCE, AbwClient, AbwError, _browse_result,
            _exact_gtin, _page_config, parse_abw_product,
        )
        client = AbwClient()
        observed = _iso_now()
        pages = 0
        rows = []
        catalog_products = []
        skipped = []
        try:
            await client.login()
            first_response = await client._request("GET", self.FIRST_PAGE)
            first = _browse_result(first_response.text)
            total_pages = int(first.get("totalPage") or 0)
            first_url = str(first.get("firstPageUrl") or first_response.url)
            catalog_rows = []
            for page_number in range(1, total_pages + 1):
                if page_number == 1:
                    browse = first
                else:
                    parsed = urlparse(first_url)
                    query = parse_qs(parsed.query)
                    query["pn"] = [str(page_number)]
                    page_url = urlunparse(parsed._replace(query=urlencode(query, doseq=True)))
                    response = await client._request("GET", page_url)
                    config = _page_config(response.text)
                    client._set_security(config.get("security"))
                    browse = _browse_result(response.text)
                pages += 1
                if int(browse.get("currentPageNumber") or 0) != page_number:
                    raise RuntimeError("ABW page correlation failed")
                for entry in browse.get("products") or []:
                    product = entry.get("product") if isinstance(entry, dict) else None
                    product_id = str((product or {}).get("productId") or "")
                    product_url = str(entry.get("url") or "") if isinstance(entry, dict) else ""
                    if product_id and product_id not in {row[0] for row in catalog_rows}:
                        catalog_rows.append((product_id, product_url))
                        if limit and len(catalog_rows) >= limit:
                            break
                if limit and len(catalog_rows) >= limit:
                    break
                await asyncio.sleep(self.pacing_seconds)
            for index, (product_id, product_url) in enumerate(catalog_rows):
                if index:
                    await asyncio.sleep(self.pacing_seconds)
                response = await client._request(
                    "GET", "/rest/product/v1/full-detail-product",
                    params={"productId": product_id},
                    headers={"Referer": product_url or BASE_URL},
                )
                payload = response.json()
                gtin = _exact_gtin(payload)
                product_payload = payload.get("product") if isinstance(payload, dict) else {}
                catalog_products.append(_catalog_product(
                    "abw", product_id, canonical_ean=gtin,
                    raw_identifiers=({"value": gtin, "type": "EAN"},) if gtin else (),
                    identifier_type="EAN" if gtin and len(gtin) == 13 else None,
                    supplier_option_id=(payload.get("selectedProductOption") or {}).get("productId") if isinstance(payload, dict) else None,
                    brand=(product_payload or {}).get("brandName"),
                    title=(product_payload or {}).get("name"),
                    metadata={"product_url": product_url, "category": "Beauty"},
                ))
                if not gtin:
                    skipped.append({"product_id": product_id, "reason": "identifier_missing"})
                    continue
                try:
                    options = parse_abw_product(payload, expected_gtin=gtin, product_url=product_url)
                except AbwError as exc:
                    if exc.code == "unavailable":
                        skipped.append({"product_id": product_id, "gtin": gtin, "reason": "unavailable"})
                        continue
                    raise
                for option in options:
                    rows.append({
                        "run_id": run_id or "dry-run", "seller_sku": option.product_name,
                        "gtin": option.gtin, "supplier_product_id": option.supplier_product_id,
                        "option_product_id": option.option_product_id,
                        "product_name": option.product_name, "brand": option.brand,
                        "mode": option.mode, "condition_key": option.condition_key,
                        "condition_label": option.condition_label,
                        "tier_min_quantity": option.tier_min_quantity,
                        "tier_max_quantity": option.tier_max_quantity,
                        "pack_size": option.pack_size,
                        "pack_price": str(option.pack_price) if option.pack_price is not None else None,
                        "net_unit_price_eur": str(option.net_unit_price),
                        "currency": option.currency, "price_source": PRICE_SOURCE,
                        "price_basis": PRICE_BASIS, "vat_rate": "0.22",
                        "vat_amount": str(option.vat_amount),
                        "gross_unit_price": str(option.gross_unit_price),
                        "available_quantity": option.available_quantity,
                        "availability_status": option.availability_status,
                        "stock_text": option.stock_text, "lead_time": option.lead_time,
                        "warehouse": option.warehouse, "discount_label": option.discount_label,
                        "product_url": option.product_url,
                        "minimum_order_value": str(MINIMUM_ORDER_VALUE),
                        "minimum_order_currency": MINIMUM_ORDER_CURRENCY,
                        "observed_at": observed, "source": SOURCE,
                    })
        finally:
            await client.close()
        candidates, normalizer_diagnostics = normalize_abw_candidates(
            rows, now=datetime.now(timezone.utc)
        )
        _apply_run_id(candidates, run_id, observed)
        diagnostics = {
            "supplier": "abw", "category": "All Beauty",
            "catalog_total_count": int(first.get("totalCount") or 0),
            "source_type": "all_beauty_sample",
            "source_count": int(first.get("totalCount") or 0),
            "enumerated_count": len(catalog_rows),
            "unique_count": len({row["canonical_product_key"] for row in catalog_products}),
            "completeness_status": "full_relevant_catalog" if limit is None else "partial_catalog",
            "completeness_reason": (
                "Official All Products export can be requested, but the emailed file "
                "has not yet been downloaded and parsed"
            ),
            "product_details_loaded": len(catalog_rows), "skipped": skipped,
            "normalizer": normalizer_diagnostics,
        }
        return _generation(
            self.coverage, candidates, pages=pages, requests=client.request_count,
            retries=client.retry_count, rate_limits=client.rate_limit_count,
            diagnostics=diagnostics, complete=False,
            catalog_products=catalog_products,
        )


def _json_ld_gtins(page_html):
    gtins = []
    for block in re.findall(
        r"<script[^>]+type=[\"']application/ld\+json[\"'][^>]*>(.*?)</script>",
        page_html, re.I | re.S,
    ):
        try:
            payload = json.loads(html_module.unescape(block))
        except json.JSONDecodeError:
            continue
        stack = [payload]
        while stack:
            value = stack.pop()
            if isinstance(value, dict):
                for key in ("gtin", "gtin8", "gtin12", "gtin13", "gtin14", "ean", "upc"):
                    raw = str(value.get(key) or "").strip()
                    if raw and raw not in gtins:
                        gtins.append(raw)
                stack.extend(value.values())
            elif isinstance(value, list):
                stack.extend(value)
    return gtins


def _qudo_page_gtins(page_html):
    """Read the product-owned EAN marker, falling back to schema.org data."""
    return _qudo_page_identifier_evidence(page_html)["selected_gtins"]


def _qudo_page_identifier_evidence(page_html):
    """Return Qudo identifiers together with their public-page provenance."""
    marker_gtins = []
    for raw in re.findall(
        r'<h2[^>]*class=["\'][^"\']*\bqudo-ean\b[^"\']*["\'][^>]*>\s*EAN:\s*(\d{8,14})\s*</h2>',
        page_html, re.I,
    ):
        if raw not in marker_gtins:
            marker_gtins.append(raw)
    json_ld_gtins = _json_ld_gtins(page_html)
    valid_marker = [value for value in marker_gtins if valid_qudo_identifier(value)]
    valid_json_ld = [value for value in json_ld_gtins if valid_qudo_identifier(value)]
    selected = valid_marker or valid_json_ld or marker_gtins or json_ld_gtins
    return {
        "marker_gtins": marker_gtins,
        "json_ld_gtins": json_ld_gtins,
        "selected_gtins": selected,
        "identifier_source": "html_marker" if valid_marker else (
            "json_ld" if valid_json_ld else "html_marker_invalid" if marker_gtins else (
                "json_ld_invalid" if json_ld_gtins else None
            )
        ),
    }


def _qudo_variation_metadata(payload):
    """Allowlisted commercial diagnostics for identifier-unresolved Qudo rows."""
    if not isinstance(payload, dict):
        return {}
    prices = payload.get("prices") if isinstance(payload.get("prices"), dict) else {}
    limits = payload.get("add_to_cart") if isinstance(payload.get("add_to_cart"), dict) else {}
    return {
        "variation_price_minor": prices.get("price"),
        "currency_code": prices.get("currency_code"),
        "currency_minor_unit": prices.get("currency_minor_unit"),
        "stock_quantity": limits.get("maximum"),
        "minimum_product_quantity": limits.get("minimum"),
        "selling_unit": limits.get("multiple_of"),
        "is_in_stock": payload.get("is_in_stock"),
        "supplier_sku": payload.get("sku"),
    }


class QudoCatalogCollector:
    coverage = QUDO_COVERAGE
    PAGE_SIZE = 100

    def __init__(self, *, manager_root=MANAGER_ROOT, pacing_seconds=0.2):
        self.manager_root = Path(manager_root)
        self.pacing_seconds = pacing_seconds

    def __call__(self, *, run_id, limit=None, dry_run=False):
        return asyncio.run(self._collect(run_id=run_id, limit=limit))

    async def _collect(self, *, run_id, limit):
        _load_manager_environment(self.manager_root)
        from purchase_prices.qudo import (
            PRICE_BASIS, PRICING_SCOPE, SOURCE, QudoClient, parse_product_page,
            parse_store_offer, select_qudo_variation,
        )
        client = QudoClient()
        observed = _iso_now()
        pages = 0
        parent_rows = []
        catalog_products = []
        skipped = []
        valid_gtin_products = 0
        try:
            page_number, total_pages, global_total = 1, None, None
            while total_pages is None or page_number <= total_pages:
                try:
                    response = await client._get(
                        "/wp-json/wc/store/v1/products",
                        params={"per_page": str(self.PAGE_SIZE), "page": str(page_number)},
                    )
                except Exception as exc:
                    raise _qudo_collector_error(
                        exc, phase="catalog_index", client=client,
                        page_number=page_number, page_size=self.PAGE_SIZE,
                    ) from exc
                pages += 1
                payload = response.json()
                current_total = int(response.headers.get("X-WP-Total") or 0)
                current_pages = int(response.headers.get("X-WP-TotalPages") or 0)
                if total_pages is None:
                    total_pages, global_total = current_pages, current_total
                elif (current_pages, current_total) != (total_pages, global_total):
                    raise RuntimeError("Qudo catalog count changed during pagination")
                for row in payload:
                    try:
                        select_qudo_variation(row)
                    except Exception:
                        continue
                    parent_rows.append(row)
                    if limit and len(parent_rows) >= limit:
                        break
                if limit and len(parent_rows) >= limit:
                    break
                page_number += 1
                if page_number <= total_pages:
                    await asyncio.sleep(self.pacing_seconds)
            rows = []
            consecutive_failures = 0
            for index, parent in enumerate(parent_rows):
                if index:
                    await asyncio.sleep(self.pacing_seconds)
                permalink = str(parent.get("permalink") or "")
                variation_id = select_qudo_variation(parent)
                brand = ""
                for attribute in parent.get("attributes") or []:
                    if str(attribute.get("name") or "").casefold() == "brand":
                        brand = str(((attribute.get("terms") or [{}])[0]).get("name") or "")
                try:
                    page_response = await client._get(permalink)
                    identifier_evidence = _qudo_page_identifier_evidence(page_response.text)
                    gtins = identifier_evidence["selected_gtins"]
                except Exception as exc:
                    catalog_products.append(_catalog_product(
                        "qudo", parent.get("id"), raw_identifiers=(),
                        supplier_option_id=variation_id, supplier_sku=parent.get("sku"),
                        brand=brand, title=parent.get("name"),
                        metadata={
                            "product_url": permalink,
                            "is_purchasable": parent.get("is_purchasable"),
                            "identifier_status": "request_failed",
                            "error_code": getattr(exc, "code", type(exc).__name__),
                        },
                    ))
                    skipped.append({
                        "product_id": str(parent.get("id") or ""),
                        "reason": "identifier_request_failed",
                    })
                    consecutive_failures += 1
                    if consecutive_failures >= 20:
                        raise _qudo_collector_error(
                            exc, phase="product_identifier", client=client,
                            product_id=str(parent.get("id") or ""),
                            variation_id=str(variation_id or ""),
                            consecutive_failures=consecutive_failures,
                        ) from exc
                    continue
                if len(gtins) != 1:
                    variation_metadata = {}
                    try:
                        variation_response = await client._get(
                            f"/wp-json/wc/store/v1/products/{variation_id}"
                        )
                        variation_metadata = _qudo_variation_metadata(variation_response.json())
                    except Exception as exc:
                        variation_metadata = {
                            "variation_error_code": getattr(exc, "code", type(exc).__name__)
                        }
                    catalog_products.append(_catalog_product(
                        "qudo", parent.get("id"), raw_identifiers=tuple(
                            {"value": value, "type": "HTML_MARKER"}
                            for value in identifier_evidence["marker_gtins"]
                        ) + tuple(
                            {"value": value, "type": "JSON_LD"}
                            for value in identifier_evidence["json_ld_gtins"]
                        ), supplier_option_id=variation_id, supplier_sku=parent.get("sku"),
                        brand=brand, title=parent.get("name"),
                        metadata={
                            "product_url": permalink,
                            "is_purchasable": parent.get("is_purchasable"),
                            "identifier_status": "missing" if not gtins else "ambiguous",
                            "identifier_source": identifier_evidence["identifier_source"],
                            **variation_metadata,
                        },
                    ))
                    skipped.append({
                        "product_id": str(parent.get("id") or ""),
                        "reason": "identifier_missing" if not gtins else "identifier_ambiguous",
                    })
                    consecutive_failures = 0
                    continue
                catalog_products.append(_catalog_product(
                    "qudo", parent.get("id"), canonical_ean=gtins[0],
                    raw_identifiers=tuple(
                        {"value": value, "type": "HTML_MARKER"}
                        for value in identifier_evidence["marker_gtins"]
                    ) + tuple(
                        {"value": value, "type": "JSON_LD"}
                        for value in identifier_evidence["json_ld_gtins"]
                        if value not in identifier_evidence["marker_gtins"]
                    ),
                    identifier_type="EAN" if len(gtins[0]) == 13 else None,
                    supplier_option_id=variation_id, supplier_sku=parent.get("sku"),
                    brand=brand, title=parent.get("name"),
                    metadata={
                        "product_url": permalink,
                        "is_purchasable": parent.get("is_purchasable"),
                        "identifier_source": identifier_evidence["identifier_source"],
                        "identifier_marker_gtins": identifier_evidence["marker_gtins"],
                        "identifier_json_ld_gtins": identifier_evidence["json_ld_gtins"],
                    },
                ))
                valid_gtin_products += 1
                try:
                    page = parse_product_page(
                        page_response.text, product_url=permalink, expected_gtin=gtins[0]
                    )
                    variation_response = await client._get(
                        f"/wp-json/wc/store/v1/products/{variation_id}"
                    )
                    offer = parse_store_offer(page, parent, variation_response.json())
                except Exception as exc:
                    skipped.append({
                        "product_id": str(parent.get("id") or ""),
                        "gtin": gtins[0],
                        "reason": "scenario_enrichment_failed",
                        "error_code": getattr(exc, "code", type(exc).__name__),
                    })
                    consecutive_failures += 1
                    if consecutive_failures >= 20:
                        raise _qudo_collector_error(
                            exc, phase="scenario_enrichment", client=client,
                            product_id=str(parent.get("id") or ""),
                            variation_id=str(variation_id or ""),
                            consecutive_failures=consecutive_failures,
                        ) from exc
                    continue
                rows.append({
                    "run_id": run_id or "dry-run", "seller_sku": offer.supplier_sku,
                    "gtin": offer.gtin, "supplier": "QUDO",
                    "supplier_product_id": offer.supplier_product_id,
                    "supplier_offer_id": offer.supplier_offer_id,
                    "supplier_sku": offer.supplier_sku, "product_name": offer.product_name,
                    "brand": brand, "observed_at": observed, "currency": offer.currency,
                    "unit_price": str(offer.net_unit_price), "price_basis": PRICE_BASIS,
                    "pricing_scope": PRICING_SCOPE,
                    "available_quantity": offer.available_quantity,
                    "availability_status": offer.availability_status,
                    "minimum_product_quantity": offer.minimum_product_quantity,
                    "selling_unit": offer.selling_unit,
                    "minimum_order_value": str(offer.minimum_order_value),
                    "minimum_order_currency": offer.minimum_order_currency,
                    "product_url": offer.product_url, "source": SOURCE,
                    "identifier_source": identifier_evidence["identifier_source"],
                })
                consecutive_failures = 0
        finally:
            await client.close()
        candidates, normalizer_diagnostics = normalize_qudo_candidates(
            rows, now=datetime.now(timezone.utc)
        )
        _apply_run_id(candidates, run_id, observed)
        diagnostics = {
            "supplier": "qudo", "global_catalog_total": global_total,
            "qudo_offer_products": len(parent_rows),
            "source_type": "woocommerce_store_api_plus_product_html",
            "source_count": global_total,
            "enumerated_count": len(parent_rows),
            "unique_count": len({str(row.get("id") or "") for row in parent_rows}),
            "completeness_status": "full_relevant_catalog" if limit is None else "partial_catalog",
            "completeness_reason": (
                "All QUDO seller variations in the stable public index were traversed"
                if limit is None else
                "Sample run; a sample can never establish or publish complete coverage"
            ) + "; products without a usable GTIN remain diagnostic and scenario coverage is partial",
            "products_without_usable_identifier": skipped,
            "valid_gtin_products": valid_gtin_products,
            "canonical_gtin_products": sum(
                1 for product in catalog_products if product.get("canonical_gtin")
            ),
            "scenario_products": len(rows),
            "scenario_success_ratio": (
                len(rows) / len(parent_rows) if parent_rows else 0
            ),
            "normalized_scenario_products": len(candidates),
            "normalizer": normalizer_diagnostics,
        }
        return _generation(
            self.coverage, candidates, pages=pages, requests=client.request_count,
            diagnostics=diagnostics, complete=limit is None,
            scenario_enrichment_status="partial" if rows else "none",
            catalog_products=catalog_products,
        )


def build_collector(supplier, *, qogita_seller_aliases=(), manager_root=MANAGER_ROOT,
                    source_file=None):
    supplier = str(supplier or "").casefold()
    if supplier == "qogita":
        return QogitaCatalogCollector(qogita_seller_aliases, manager_root=manager_root)
    if supplier == "umma":
        return UmmaCatalogCollector(manager_root=manager_root)
    if supplier == "abw":
        return AbwCatalogCollector(manager_root=manager_root, source_file=source_file)
    if supplier == "qudo":
        return QudoCatalogCollector(manager_root=manager_root)
    raise ValueError(f"Unsupported supplier: {supplier}")
