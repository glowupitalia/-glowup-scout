import logging
import math
import os
import re
import tempfile
import time
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP, localcontext

import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles import Protection
from openpyxl.utils import get_column_letter


logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

ELIGIBLE_STATUS = "TROVATO CON OFFERTE"
CENT = Decimal("0.01")
PERCENT = Decimal("0.01")
FBA_VAT_RATE = Decimal("0.22")
FBA_VAT_MULTIPLIER = Decimal("1") + FBA_VAT_RATE
DEFAULT_REFERRAL_RATE = Decimal("0.19")
TARGET_MARGINS = (15, 20, 25)
FEE_BATCH_SIZE = 20
FEE_BATCH_INTERVAL_SECONDS = 2.0

RESULT_COLUMNS = [
    "EAN",
    "Brand",
    "Titolo",
    "Costo",
    "BSR Beauty",
    "Prezzo riferimento",
    "Venditori FBA",
    "Venditori totali",
    "Margine attuale %",
    "Prezzo 15%",
    "Prezzo 20%",
    "Prezzo 25%",
    "Score",
    "Opportunità",
    "Link Offerte Amazon",
]

TECHNICAL_COLUMNS = [
    "Row ID",
    "ASIN",
    "FBA fee netta",
    "FBA fee IVA inclusa",
    "Referral Fee",
    "Referral rate",
    "Referral source",
    "Price source",
    "Economics status",
]

RESULT_ROW_ID_COLUMN = len(RESULT_COLUMNS) + 1


class ProductFeeParseError(ValueError):
    pass


def _sanitize_log_text(value, limit=500):
    """Keep Amazon diagnostics useful without logging authentication material."""
    text = str(value or "").replace("\n", " ").replace("\r", " ")
    patterns = (
        r"(?i)bearer\s+[^\s,;]+",
        r"(?i)(access[_ -]?token|refresh[_ -]?token|client[_ -]?secret|"
        r"authorization)\s*[:=]\s*[^\s,;]+",
    )
    for pattern in patterns:
        text = re.sub(pattern, "[REDACTED]", text)
    return text[:limit]


def _fee_response_diagnostics(entry):
    wrapper = entry if isinstance(entry, dict) else {}
    has_result = isinstance(wrapper.get("FeesEstimateResult"), dict)
    result = wrapper.get("FeesEstimateResult") if has_result else wrapper
    result = result if isinstance(result, dict) else {}
    identifier = result.get("FeesEstimateIdentifier") or {}
    estimate = result.get("FeesEstimate")
    estimate_dict = estimate if isinstance(estimate, dict) else {}
    details = estimate_dict.get("FeeDetailList")
    details_list = details if isinstance(details, list) else []
    fee_types = [
        str(detail.get("FeeType"))
        for detail in details_list
        if isinstance(detail, dict) and detail.get("FeeType")
    ]
    fba = _find_fee(details_list, "FBAFees")
    included = (fba or {}).get("IncludedFeeDetailList")
    included_list = included if isinstance(included, list) else []
    pick_and_pack = _find_fee(included_list, "FBAPickAndPack")
    error = result.get("Error") or {}
    return {
        "result": result,
        "identifier": identifier if isinstance(identifier, dict) else {},
        "status": result.get("Status"),
        "error_code": _sanitize_log_text(error.get("Code")),
        "error_message": _sanitize_log_text(error.get("Message")),
        "has_result": has_result,
        "has_estimate": isinstance(estimate, dict),
        "has_details": isinstance(details, list),
        "has_fba": fba is not None,
        "has_included": isinstance(included, list),
        "has_pick_and_pack": pick_and_pack is not None,
        "has_referral": _find_fee(details_list, "ReferralFee") is not None,
        "fee_types": fee_types,
        "keys": sorted(result.keys()),
        "estimate_keys": sorted(estimate_dict.keys()),
    }


def _log_fee_item(candidate, entry, correlation_method, fba_status):
    diagnostics = _fee_response_diagnostics(entry)
    logger.info(
        "PRODUCT FEES ITEM | asin=%s correlation_id=%s correlated=%s "
        "correlation_method=%s status=%s error_code=%s error_message=%s "
        "has_FeesEstimateResult=%s has_FeesEstimate=%s "
        "has_FeeDetailList=%s has_FBAFees=%s "
        "has_IncludedFeeDetailList=%s has_FBAPickAndPack=%s "
        "has_ReferralFee=%s fee_types=%s fba_status=%s",
        candidate["asin"],
        candidate["identifier"],
        entry is not None,
        correlation_method,
        diagnostics["status"] or "<missing>",
        diagnostics["error_code"] or "<none>",
        diagnostics["error_message"] or "<none>",
        diagnostics["has_result"],
        diagnostics["has_estimate"],
        diagnostics["has_details"],
        diagnostics["has_fba"],
        diagnostics["has_included"],
        diagnostics["has_pick_and_pack"],
        diagnostics["has_referral"],
        diagnostics["fee_types"],
        fba_status,
    )


def to_int(value):
    try:
        if value in ["", None, "None"]:
            return None
        parsed = int(float(str(value).replace(",", ".")))
        return parsed if parsed >= 0 else None
    except Exception:
        return None


def to_decimal(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        cleaned = cleaned.replace("€", "").replace("EUR", "").strip()
        if "," in cleaned and "." in cleaned:
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", ".")
        value = cleaned
    try:
        parsed = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None
    return parsed if parsed.is_finite() else None


def money_value(value):
    parsed = to_decimal(value)
    if parsed is None:
        return None
    return parsed.quantize(CENT, rounding=ROUND_HALF_UP)


def select_reference_price(pricing):
    """Select the positive landed price used by fees and economics."""
    pricing = pricing or {}
    candidates = (
        ("buy_box", pricing.get("Buy Box Amount")),
        ("buy_box", pricing.get("Buy Box")),
        ("min_fba", pricing.get("Prezzo minimo FBA Amount")),
        ("min_fba", pricing.get("Prezzo minimo FBA")),
        ("min_fbm", pricing.get("Prezzo minimo FBM Amount")),
        ("min_fbm", pricing.get("Prezzo minimo FBM")),
    )
    for source, raw_price in candidates:
        price = to_decimal(raw_price)
        if price is not None and price > 0:
            return price, source
    return None, "missing_price"


def bsr_points(bsr_beauty):
    bsr = to_int(bsr_beauty)
    if bsr is None:
        return 0
    if bsr <= 1000:
        return 50
    if bsr <= 5000:
        return 45
    if bsr <= 10000:
        return 40
    if bsr <= 25000:
        return 30
    if bsr <= 50000:
        return 15
    return 5


def fba_seller_points(venditori_fba, venditori_totali):
    total = to_int(venditori_totali)
    fba = to_int(venditori_fba)
    if total in (None, 0) or fba is None:
        return 0
    if fba == 0:
        return 20
    if fba <= 2:
        return 18
    if fba <= 4:
        return 14
    if fba <= 6:
        return 10
    if fba <= 10:
        return 5
    return 0


def total_seller_points(venditori_totali):
    total = to_int(venditori_totali)
    if total in (None, 0):
        return 0
    if total <= 3:
        return 10
    if total <= 6:
        return 7
    if total <= 10:
        return 4
    return 0


def margin_points(margin_percent):
    margin = to_decimal(margin_percent)
    if margin is None or margin < Decimal("10"):
        return 0
    if margin < Decimal("15"):
        return 4
    if margin < Decimal("20"):
        return 14
    if margin < Decimal("25"):
        return 18
    return 20


def opportunity_from_score(score):
    score = to_int(score)
    if score is None:
        return "🔴 Debole"
    if score >= 85:
        return "🟢 Eccellente"
    if score >= 70:
        return "🟢 Ottima"
    if score >= 55:
        return "🟡 Interessante"
    if score >= 40:
        return "🟠 Da valutare"
    return "🔴 Debole"


def opportunity_score(
    bsr_beauty,
    venditori_fba,
    venditori_totali,
    margin_percent,
):
    score = (
        bsr_points(bsr_beauty)
        + fba_seller_points(venditori_fba, venditori_totali)
        + total_seller_points(venditori_totali)
        + margin_points(margin_percent)
    )
    return score, opportunity_from_score(score)


def _find_fee(details, fee_type):
    return next(
        (
            detail
            for detail in details or []
            if isinstance(detail, dict) and detail.get("FeeType") == fee_type
        ),
        None,
    )


def _parse_amount(money):
    if not isinstance(money, dict):
        return None
    amount = to_decimal(money.get("Amount"))
    if amount is None or amount < 0:
        return None
    currency = str(money.get("CurrencyCode") or "").upper()
    if currency and currency != "EUR":
        raise ProductFeeParseError(
            f"unexpected fee currency: {currency}"
        )
    return amount


def _fba_amounts(detail):
    final_fee = _parse_amount((detail or {}).get("FinalFee"))
    if final_fee is None:
        raise ProductFeeParseError("FBA FinalFee.Amount missing")

    fee_amount = _parse_amount((detail or {}).get("FeeAmount"))
    promotion = _parse_amount((detail or {}).get("FeePromotion")) or Decimal("0")
    tax = _parse_amount((detail or {}).get("TaxAmount"))
    base_net = fee_amount - promotion if fee_amount is not None else final_fee
    if base_net < 0:
        raise ProductFeeParseError("invalid promoted FBA fee")

    tax_handling = "manager_vat_22"
    if tax is None or tax == 0:
        net = final_fee
        gross = final_fee * FBA_VAT_MULTIPLIER
        tax_amount = gross - net
    elif abs(final_fee - (base_net + tax)) <= CENT:
        net = base_net
        gross = final_fee
        tax_amount = tax
        tax_handling = "amazon_tax_included_in_final"
    elif abs(final_fee - base_net) <= CENT:
        net = final_fee
        gross = final_fee + tax
        tax_amount = tax
        tax_handling = "amazon_tax_separate"
    else:
        raise ProductFeeParseError("ambiguous Amazon FBA tax structure")

    return {
        "fba_fee_net": net,
        "fba_tax": tax_amount,
        "fba_fee_gross": gross,
        "tax_handling": tax_handling,
    }


def parse_product_fee_result(entry, reference_price=None):
    result = (entry or {}).get("FeesEstimateResult") or entry or {}
    identifier = result.get("FeesEstimateIdentifier") or {}
    status = result.get("Status")
    if status != "Success":
        error = result.get("Error") or {}
        message = error.get("Message") or (
            f"Product Fees status: {status or 'missing'}"
        )
        raise ProductFeeParseError(message)

    estimate = result.get("FeesEstimate") or {}
    fee_details = estimate.get("FeeDetailList") or []
    fba_fee = _find_fee(fee_details, "FBAFees")
    pick_and_pack = _find_fee(
        (fba_fee or {}).get("IncludedFeeDetailList") or [],
        "FBAPickAndPack",
    )

    selected_fee = None
    source = None
    if fba_fee and _parse_amount(fba_fee.get("FinalFee")) is not None:
        selected_fee = fba_fee
        source = "FBAFees"
    elif pick_and_pack and _parse_amount(pick_and_pack.get("FinalFee")) is not None:
        selected_fee = pick_and_pack
        source = "FBAPickAndPack"
    if selected_fee is None:
        raise ProductFeeParseError("FBAFees and FBAPickAndPack missing")

    parsed = _fba_amounts(selected_fee)
    referral = _find_fee(fee_details, "ReferralFee")
    referral_amount = _parse_amount((referral or {}).get("FinalFee"))
    price = to_decimal(reference_price)
    if price is None:
        price = _parse_amount(
            (identifier.get("PriceToEstimateFees") or {}).get("ListingPrice")
        )
    referral_rate = None
    if referral_amount is not None and price is not None and price > 0:
        referral_rate = referral_amount / price

    parsed.update({
        "asin": str(identifier.get("IdValue") or "").strip(),
        "identifier": str(
            identifier.get("SellerInputIdentifier") or ""
        ).strip(),
        "source": source,
        "referral_fee": referral_amount,
        "referral_rate": referral_rate,
        "currency": "EUR",
    })
    return parsed


def calculate_economics(price, product_cost, fee_estimate):
    selling_price = to_decimal(price)
    cost = to_decimal(product_cost)
    if selling_price is None or selling_price <= 0:
        return {"status": "missing_buy_box"}
    if cost is None or cost < 0:
        return {"status": "missing_product_cost"}
    if not fee_estimate:
        return {"status": "missing_fba_fee"}

    fba_gross = to_decimal(fee_estimate.get("fba_fee_gross"))
    if fba_gross is None or fba_gross < 0:
        return {"status": "missing_fba_fee"}

    referral_fee = to_decimal(fee_estimate.get("referral_fee"))
    referral_rate = to_decimal(fee_estimate.get("referral_rate"))
    if referral_fee is None or referral_fee <= 0:
        referral_rate = DEFAULT_REFERRAL_RATE
        referral_fee = selling_price * referral_rate
        referral_source = "fallback_19_percent"
    else:
        referral_source = "amazon_referral_fee"
        if referral_rate is None or referral_rate < 0:
            referral_rate = referral_fee / selling_price

    if referral_rate < 0 or referral_rate >= 1:
        return {"status": "invalid_referral_rate"}

    with localcontext() as context:
        context.prec = 40
        profit = selling_price - cost - referral_fee - fba_gross
        margin_percent = profit / selling_price * Decimal("100")
        targets = {}
        for target in TARGET_MARGINS:
            target_rate = Decimal(target) / Decimal("100")
            denominator = Decimal("1") - referral_rate - target_rate
            targets[target] = (
                None
                if denominator <= 0
                else (cost + fba_gross) / denominator
            )

    return {
        "status": "ready",
        "referral_source": referral_source,
        "referral_fee": referral_fee.quantize(CENT, rounding=ROUND_HALF_UP),
        "referral_rate": referral_rate,
        "fba_fee_net": to_decimal(fee_estimate.get("fba_fee_net")),
        "fba_fee_gross": fba_gross.quantize(CENT, rounding=ROUND_HALF_UP),
        "profit": profit.quantize(CENT, rounding=ROUND_HALF_UP),
        "margin_percent": margin_percent.quantize(PERCENT, rounding=ROUND_HALF_UP),
        "target_prices": {
            target: (
                value.quantize(CENT, rounding=ROUND_HALF_UP)
                if value is not None
                else None
            )
            for target, value in targets.items()
        },
    }


def estimate_fba_fallback(catalog):
    """Reserved for verified weight/dimension or internal fallback rules."""
    has_logistics = any(
        catalog.get(key)
        for key in (
            "Peso prodotto",
            "Peso package",
            "Dimensioni prodotto",
            "Dimensioni package",
        )
    )
    return {
        "estimate": None,
        "status": (
            "logistics_present_no_verified_tariff"
            if has_logistics
            else "insufficient_logistics_data"
        ),
    }


def _chunks(items, size=FEE_BATCH_SIZE):
    for start in range(0, len(items), size):
        yield items[start:start + size]


def _progress(callback, value):
    if callback is None:
        return None
    try:
        callback(min(1.0, max(0.0, value)))
        return callback
    except Exception:
        logger.exception(
            "PROGRESS UPDATE FAILED; continuing without UI progress"
        )
        return None


def _base_result(ean, cost, input_order):
    return {
        "EAN": ean,
        "Brand": "",
        "Titolo": "",
        "Costo": cost,
        "BSR Beauty": None,
        "Prezzo riferimento": None,
        "Venditori FBA": None,
        "Venditori totali": None,
        "Margine attuale %": None,
        "Prezzo 15%": None,
        "Prezzo 20%": None,
        "Prezzo 25%": None,
        "Score": 0,
        "Opportunità": opportunity_from_score(0),
        "Link Offerte Amazon": "",
        "ASIN": "",
        "Stato": "TROVATO",
        "Errore": "",
        "FBA Fee Status": "not_requested",
        "Economics Status": "not_calculated",
        "_Price source": "missing_price",
        "_Input order": input_order,
    }


def analyze_products(
    df_input,
    costo_col,
    token,
    search_catalog,
    search_pricing,
    safe_call,
    search_fees_batch=None,
    progress_callback=None,
    throttle_seconds=0.7,
    fee_batch_interval_seconds=FEE_BATCH_INTERVAL_SECONDS,
    sleep_func=time.sleep,
    source_file=None,
):
    """Analyze catalog/pricing rows, then enrich eligible rows with batch fees."""
    results = []
    total_products = len(df_input)
    active_progress_callback = progress_callback

    logger.info(
        "PROCESSING PRODUCTS | products=%s file=%s",
        total_products,
        source_file or "<unknown>",
    )

    for position, (_, input_row) in enumerate(df_input.iterrows(), start=1):
        ean_value = str(input_row["EAN"]).strip()
        costo_value = input_row[costo_col] if costo_col else ""
        result = _base_result(ean_value, costo_value, position - 1)

        try:
            catalog = safe_call(search_catalog, ean_value, token)
            if catalog:
                pricing = safe_call(search_pricing, catalog["ASIN"], token)
                reference_price, price_source = select_reference_price(pricing)
                result.update({
                    "ASIN": catalog.get("ASIN", ""),
                    "Titolo": catalog.get("Titolo", ""),
                    "Brand": catalog.get("Brand", ""),
                    "BSR Beauty": catalog.get("BSR Beauty"),
                    "Prezzo riferimento": (
                        float(reference_price)
                        if reference_price is not None else None
                    ),
                    "_Price source": price_source,
                    "Venditori totali": pricing.get("Venditori totali"),
                    "Venditori FBA": pricing.get("Venditori FBA"),
                    "Link Offerte Amazon": (
                        "https://www.amazon.it/gp/offer-listing/"
                        f"{catalog.get('ASIN', '')}"
                    ),
                    "Peso prodotto": catalog.get("Peso prodotto"),
                    "Peso package": catalog.get("Peso package"),
                    "Dimensioni prodotto": catalog.get("Dimensioni prodotto"),
                    "Dimensioni package": catalog.get("Dimensioni package"),
                    "Product Type": catalog.get("Product Type"),
                    "_Catalog attributes": catalog.get("_Catalog attributes"),
                    "_Catalog dimensions": catalog.get("_Catalog dimensions"),
                })
            else:
                result["FBA Fee Status"] = "catalog_not_found"
        except Exception as exc:
            logger.exception(
                "PRODUCT PROCESSING FAILED | phase=PROCESSING PRODUCTS "
                "file=%s ean=%s",
                source_file or "<unknown>",
                ean_value,
            )
            result["Stato"] = "ERRORE API / LIMITE AMAZON"
            result["Errore"] = str(exc)
            result["FBA Fee Status"] = "catalog_or_pricing_error"

        results.append(result)
        if active_progress_callback is not None and total_products:
            active_progress_callback = _progress(
                active_progress_callback,
                position / total_products * 0.8,
            )
        if throttle_seconds:
            sleep_func(throttle_seconds)

    fee_candidates = []
    for row_index, result in enumerate(results):
        asin = str(result.get("ASIN") or "").strip()
        price = to_decimal(result.get("Prezzo riferimento"))
        if asin and price is not None and price > 0:
            identifier = f"glowup-scout|{result['_Input order']}|{asin}"
            fee_candidates.append({
                "row_index": row_index,
                "asin": asin,
                "price": float(price),
                "identifier": identifier,
            })
        elif asin:
            result["FBA Fee Status"] = "missing_price"

    if search_fees_batch is None:
        for candidate in fee_candidates:
            results[candidate["row_index"]]["FBA Fee Status"] = (
                "fee_estimator_unavailable"
            )
    else:
        batches = list(_chunks(fee_candidates))
        for batch_number, batch in enumerate(batches, start=1):
            try:
                entries = safe_call(search_fees_batch, batch, token) or []
                logger.info(
                    "PRODUCT FEES BATCH | file=%s batch=%s "
                    "candidates_sent=%s results_received=%s",
                    source_file or "<unknown>",
                    batch_number,
                    len(batch),
                    len(entries),
                )
                entries_by_identifier = {}
                entries_by_asin = {}
                for entry in entries:
                    raw_result = (
                        (entry or {}).get("FeesEstimateResult") or entry or {}
                    )
                    identifier = raw_result.get("FeesEstimateIdentifier") or {}
                    request_identifier = str(
                        identifier.get("SellerInputIdentifier") or ""
                    ).strip()
                    asin = str(identifier.get("IdValue") or "").strip()
                    if request_identifier:
                        entries_by_identifier[request_identifier] = entry
                    if asin:
                        entries_by_asin.setdefault(asin, entry)

                for candidate in batch:
                    result = results[candidate["row_index"]]
                    entry = entries_by_identifier.get(candidate["identifier"])
                    correlation_method = "SellerInputIdentifier"
                    if entry is None:
                        entry = entries_by_asin.get(candidate["asin"])
                        correlation_method = "ASIN"
                    if entry is None:
                        correlation_method = "none"
                        result["FBA Fee Status"] = "fee_result_missing"
                        result["_Fee correlation"] = correlation_method
                        _log_fee_item(
                            candidate, entry, correlation_method,
                            result["FBA Fee Status"],
                        )
                        continue
                    result["_Fee correlation"] = correlation_method
                    try:
                        estimate = parse_product_fee_result(
                            entry,
                            reference_price=candidate["price"],
                        )
                    except ProductFeeParseError as exc:
                        result["FBA Fee Status"] = "amazon_fee_invalid"
                        result["Errore FBA"] = str(exc)
                        diagnostics = _fee_response_diagnostics(entry)
                        logger.info(
                            "PRODUCT FEES PARSE REJECTED | asin=%s "
                            "correlation_id=%s error=%s result_keys=%s "
                            "estimate_keys=%s",
                            candidate["asin"],
                            candidate["identifier"],
                            _sanitize_log_text(exc),
                            diagnostics["keys"],
                            diagnostics["estimate_keys"],
                        )
                        _log_fee_item(
                            candidate, entry, correlation_method,
                            result["FBA Fee Status"],
                        )
                        continue
                    result["_Fee estimate"] = estimate
                    result["FBA Fee Status"] = estimate["source"]
                    _log_fee_item(
                        candidate, entry, correlation_method,
                        result["FBA Fee Status"],
                    )
            except Exception as exc:
                logger.exception(
                    "FEE BATCH FAILED | file=%s batch=%s size=%s",
                    source_file or "<unknown>",
                    batch_number,
                    len(batch),
                )
                for candidate in batch:
                    result = results[candidate["row_index"]]
                    result["FBA Fee Status"] = "fee_batch_error"
                    result["Errore FBA"] = str(exc)

            if active_progress_callback is not None and batches:
                active_progress_callback = _progress(
                    active_progress_callback,
                    0.8 + batch_number / len(batches) * 0.2,
                )
            if batch_number < len(batches) and fee_batch_interval_seconds:
                sleep_func(fee_batch_interval_seconds)

    for result in results:
        fee_estimate = result.get("_Fee estimate")
        if not fee_estimate and result.get("ASIN"):
            fallback = estimate_fba_fallback(result)
            result["FBA Fallback Status"] = fallback["status"]
            fee_estimate = fallback["estimate"]
        economics = calculate_economics(
            result.get("Prezzo riferimento"),
            result.get("Costo"),
            fee_estimate,
        )
        result["Economics Status"] = economics["status"]
        if economics["status"] == "ready":
            result["Margine attuale %"] = float(economics["margin_percent"])
            for target in TARGET_MARGINS:
                value = economics["target_prices"][target]
                result[f"Prezzo {target}%"] = (
                    float(value) if value is not None else None
                )
            result["_Economics"] = economics

        if result.get("ASIN"):
            logger.info(
                "PRODUCT FEES ECONOMICS | asin=%s correlation_id=%s "
                "correlation_method=%s fba_status=%s economics_status=%s",
                result.get("ASIN"),
                (
                    f"glowup-scout|{result['_Input order']}|"
                    f"{result.get('ASIN')}"
                ),
                result.get("_Fee correlation", "none"),
                result.get("FBA Fee Status"),
                result.get("Economics Status"),
            )

        score, opportunity = opportunity_score(
            result.get("BSR Beauty"),
            result.get("Venditori FBA"),
            result.get("Venditori totali"),
            result.get("Margine attuale %"),
        )
        result["Score"] = score
        result["Opportunità"] = opportunity

    return finalize_results(results)


def finalize_results(results):
    """Apply status normalization and deterministic business ordering."""
    df_results = pd.DataFrame(results)
    if df_results.empty:
        return df_results

    if "Venditori totali" in df_results.columns:
        for idx in df_results.index:
            asin = str(df_results.at[idx, "ASIN"]) if "ASIN" in df_results else ""
            sellers = df_results.at[idx, "Venditori totali"]
            if asin in ["", "None", "nan"]:
                if df_results.at[idx, "Stato"] == "TROVATO":
                    df_results.at[idx, "Stato"] = "NON TROVATO SU AMAZON"
            elif pd.isna(sellers) or sellers == 0:
                df_results.at[idx, "Stato"] = "TROVATO SENZA OFFERTE"
            else:
                df_results.at[idx, "Stato"] = ELIGIBLE_STATUS

    df_results["Score"] = pd.to_numeric(
        df_results.get("Score"), errors="coerce"
    ).fillna(0)
    df_results["BSR Beauty"] = pd.to_numeric(
        df_results.get("BSR Beauty"), errors="coerce"
    )
    if "_Input order" not in df_results:
        df_results["_Input order"] = range(len(df_results))
    return df_results.sort_values(
        by=["Score", "BSR Beauty", "_Input order"],
        ascending=[False, True, True],
        na_position="last",
        kind="mergesort",
    ).reset_index(drop=True)


def summarize_results(df_results):
    """Return lightweight counters without serializing result rows to the UI."""
    total = len(df_results)
    eligible = 0
    if "Stato" in df_results.columns:
        eligible = int((df_results["Stato"] == ELIGIBLE_STATUS).sum())
    return {
        "total": total,
        "eligible": eligible,
        "not_eligible": total - eligible,
    }


def _excel_value(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, float) and not math.isfinite(value):
        return None
    return value


def _excel_row_id(source_row, export_position):
    return _excel_value(source_row.get("_Input order", export_position))


def _technical_excel_row(source_row, export_position):
    economics = source_row.get("_Economics")
    economics = economics if isinstance(economics, dict) else {}
    return [
        _excel_row_id(source_row, export_position),
        _excel_value(source_row.get("ASIN")),
        _excel_value(economics.get("fba_fee_net")),
        _excel_value(economics.get("fba_fee_gross")),
        _excel_value(economics.get("referral_fee")),
        _excel_value(economics.get("referral_rate")),
        _excel_value(economics.get("referral_source")),
        _excel_value(source_row.get("_Price source")),
        _excel_value(source_row.get("Economics Status")),
    ]


def _technical_lookup_excel_formula(row, technical_column, last_data_row):
    return (
        f"INDEX('Dati'!${technical_column}$2:${technical_column}"
        f"${last_data_row},MATCH($P{row},'Dati'!$A$2:$A"
        f"${last_data_row},0))"
    )


def _margin_excel_formula(row, last_data_row):
    fba_gross = _technical_lookup_excel_formula(row, "D", last_data_row)
    referral_rate = _technical_lookup_excel_formula(row, "F", last_data_row)
    return (
        f'=IFERROR(IF(AND(ISNUMBER(F{row}),F{row}>0,ISNUMBER(D{row}),'
        f"ISNUMBER({fba_gross}),ISNUMBER({referral_rate})),"
        f"ROUND((F{row}-D{row}-(F{row}*{referral_rate})-"
        f'{fba_gross})/F{row},4),""),"")'
    )


def _target_price_excel_formula(row, target_rate, last_data_row):
    fba_gross = _technical_lookup_excel_formula(row, "D", last_data_row)
    referral_rate = _technical_lookup_excel_formula(row, "F", last_data_row)
    return (
        f'=IFERROR(IF(AND(ISNUMBER(D{row}),ISNUMBER({fba_gross}),'
        f"ISNUMBER({referral_rate}),"
        f"(1-{referral_rate}-{target_rate})>0),"
        f"ROUND((D{row}+{fba_gross})/"
        f'(1-{referral_rate}-{target_rate}),2),""),"")'
    )


def _score_excel_formula(row):
    bsr = (
        f"IF(AND(ISNUMBER(E{row}),E{row}>=0),"
        f"IF(E{row}<=1000,50,IF(E{row}<=5000,45,"
        f"IF(E{row}<=10000,40,IF(E{row}<=25000,30,"
        f"IF(E{row}<=50000,15,5))))),0)"
    )
    fba = (
        f"IF(AND(ISNUMBER(H{row}),H{row}>0,ISNUMBER(G{row}),G{row}>=0),"
        f"IF(G{row}=0,20,IF(G{row}<=2,18,IF(G{row}<=4,14,"
        f"IF(G{row}<=6,10,IF(G{row}<=10,5,0))))),0)"
    )
    total = (
        f"IF(AND(ISNUMBER(H{row}),H{row}>0),"
        f"IF(H{row}<=3,10,IF(H{row}<=6,7,IF(H{row}<=10,4,0))),0)"
    )
    margin = (
        f"IF(ISNUMBER(I{row}),IF(I{row}<10%,0,IF(I{row}<15%,4,"
        f"IF(I{row}<20%,14,IF(I{row}<25%,18,20)))),0)"
    )
    return f"={bsr}+{fba}+{total}+{margin}"


def _opportunity_excel_formula(row):
    return (
        f'=IF(M{row}>=85,"🟢 Eccellente",IF(M{row}>=70,"🟢 Ottima",'
        f'IF(M{row}>=55,"🟡 Interessante",IF(M{row}>=40,'
        f'"🟠 Da valutare","🔴 Debole"))))'
    )


def write_results_excel(df_results, output_file):
    """Write the protected, formula-driven 15-column workbook atomically."""
    output_path = os.path.abspath(output_file)
    output_dir = os.path.dirname(output_path)
    os.makedirs(output_dir, exist_ok=True)

    temporary_file = tempfile.NamedTemporaryFile(
        prefix=".glowup_scout_output_",
        suffix=".xlsx",
        dir=output_dir,
        delete=False,
    )
    temporary_path = temporary_file.name
    temporary_file.close()
    export = df_results.reindex(columns=RESULT_COLUMNS).copy()
    export = export.map(_excel_value)

    logger.info("WRITING EXCEL | file=%s rows=%s", output_path, len(export))
    try:
        with pd.ExcelWriter(temporary_path, engine="openpyxl") as writer:
            export.to_excel(writer, index=False, sheet_name="Risultati")
            workbook = writer.book
            ws = writer.sheets["Risultati"]
            data_ws = workbook.create_sheet("Dati")
            data_ws.append(TECHNICAL_COLUMNS)
            for export_position, (_, source_row) in enumerate(
                df_results.iterrows(), start=1
            ):
                data_ws.append(
                    _technical_excel_row(source_row, export_position)
                )
            data_ws.sheet_state = "hidden"

            ws.cell(row=1, column=RESULT_ROW_ID_COLUMN).value = "Row ID"
            for export_position, (_, source_row) in enumerate(
                df_results.iterrows(), start=1
            ):
                ws.cell(
                    row=export_position + 1,
                    column=RESULT_ROW_ID_COLUMN,
                ).value = _excel_row_id(source_row, export_position)
            row_id_letter = get_column_letter(RESULT_ROW_ID_COLUMN)
            ws.column_dimensions[row_id_letter].hidden = True

            workbook.calculation.calcMode = "auto"
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True
            workbook.calculation.calcOnSave = True
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{row_id_letter}{ws.max_row}"
            ws.print_area = f"A1:O{ws.max_row}"
            header_fill = PatternFill("solid", fgColor="1F4E78")
            for cell in ws[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            headers = {
                ws.cell(row=1, column=col).value: col
                for col in range(1, ws.max_column + 1)
            }
            link_col = headers["Link Offerte Amazon"]
            money_columns = [
                headers[name]
                for name in (
                    "Costo", "Prezzo riferimento", "Prezzo 15%",
                    "Prezzo 20%", "Prezzo 25%",
                )
            ]
            margin_col = headers["Margine attuale %"]
            cost_col = headers["Costo"]
            target_columns = {
                15: headers["Prezzo 15%"],
                20: headers["Prezzo 20%"],
                25: headers["Prezzo 25%"],
            }
            score_col = headers["Score"]
            opportunity_col = headers["Opportunità"]
            editable_fill = PatternFill("solid", fgColor="EAF2FE")
            ws.cell(row=1, column=cost_col).comment = Comment(
                "Modifica questi valori per simulare margine, prezzi target, "
                "Score e Opportunità.",
                "GlowUp Scout",
            )
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=link_col)
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"
                for column in money_columns:
                    ws.cell(row=row, column=column).number_format = '€ #,##0.00'
                cost_cell = ws.cell(row=row, column=cost_col)
                cost_cell.protection = Protection(locked=False)
                cost_cell.fill = editable_fill
                ws.cell(row=row, column=margin_col).value = (
                    _margin_excel_formula(row, data_ws.max_row)
                )
                ws.cell(row=row, column=margin_col).number_format = "0.00%"
                for target, column in target_columns.items():
                    ws.cell(row=row, column=column).value = (
                        _target_price_excel_formula(
                            row, target / 100, data_ws.max_row
                        )
                    )
                ws.cell(row=row, column=score_col).value = (
                    _score_excel_formula(row)
                )
                ws.cell(row=row, column=opportunity_col).value = (
                    _opportunity_excel_formula(row)
                )

            ws.protection.sheet = True
            ws.protection.autoFilter = False
            ws.protection.sort = False
            ws.protection.selectLockedCells = True
            ws.protection.selectUnlockedCells = False

            widths = {
                "EAN": 16,
                "Brand": 20,
                "Titolo": 55,
                "Costo": 12,
                "BSR Beauty": 14,
                "Prezzo riferimento": 18,
                "Venditori FBA": 15,
                "Venditori totali": 17,
                "Margine attuale %": 18,
                "Prezzo 15%": 13,
                "Prezzo 20%": 13,
                "Prezzo 25%": 13,
                "Score": 10,
                "Opportunità": 20,
                "Link Offerte Amazon": 45,
            }
            for name, width in widths.items():
                ws.column_dimensions[
                    get_column_letter(headers[name])
                ].width = width

        os.replace(temporary_path, output_path)
    except Exception:
        logger.exception(
            "EXCEL GENERATION FAILED | phase=WRITING EXCEL file=%s",
            output_path,
        )
        try:
            os.unlink(temporary_path)
        except FileNotFoundError:
            pass
        raise

    logger.info("EXCEL GENERATED | file=%s", output_path)
    return output_path
