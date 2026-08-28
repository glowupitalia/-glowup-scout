import json
import os
import subprocess
import tempfile
import unittest
import sqlite3
import requests
from copy import deepcopy
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook
from streamlit.testing.v1 import AppTest

from batch_analysis import opportunity_score
from discovery import (
    DISCOVERY_SCHEMA_VERSION,
    DiscoveryCheckpointStore,
    _fee_element_retry_delay,
    _fees_batch_with_retry,
    classify_product_fee_entry,
    default_filters,
    discovery_checkpoint_compatibility,
    discovery_funnel_view,
    normalize_discovery_state,
    recalculate_diagnostic_funnel,
    run_discovery,
)
from discovery_amazon import (
    RefreshingTokenProvider,
    _request_with_retry,
    beauty_rank,
    build_item_offers_batch_requests,
    catalog_identifier_batches,
    classify_catalog_identifier,
    correlate_catalog_items,
    normalize_commercial_identifier,
    parse_item_offers_batch,
    search_catalog_by_gtins_batch,
)
from discovery_excel import (
    OPPORTUNITY_COLUMNS,
    SCENARIO_COLUMNS,
    write_discovery_excel,
)
from purchase_scenarios import (
    assign_scenario_roles,
    canonicalize_target_prices,
    recommended_scenario,
    target_price,
)
from product_fees import ProductFeeBatchResults
from qogita_discovery import _sqlite_rows, normalize_qogita_candidates, valid_gtin
from qogita_refresh import (
    QogitaRefreshConfigurationError,
    build_manager_subprocess_environment,
    inspect_qogita_cache,
    refresh_qogita_seller_catalogs,
    snapshots_advanced,
)


class FakeResponse:
    def __init__(self, payload=None, status=200, headers=None):
        self.payload = payload or {}
        self.status_code = status
        self.headers = headers or {}

    def json(self):
        return self.payload

    def raise_for_status(self):
        if self.status_code >= 400:
            import requests
            response = requests.Response()
            response.status_code = self.status_code
            raise requests.HTTPError(response=response)


def qogita_row(
    gtin="8801234567890", *, price="10", mov="100", stock=5,
    active=True, selling_unit=1, seller_alias="seller",
    observed_at="2026-08-20T10:00:00Z",
):
    return {
        "gtin": gtin,
        "variant_fid": "fid",
        "offer_qid": f"offer-{gtin}",
        "product_name": "Crema",
        "brand": "Brand",
        "category_name": "Beauty",
        "image_url": "",
        "inventory": stock,
        "selling_unit": selling_unit,
        "product_url": "https://qogita.example/product",
        "observed_at": observed_at,
        "seller_alias": seller_alias,
        "tier_mov": mov,
        "currency": "EUR",
        "tier_price": price,
        "is_active": active,
    }


def test_ean(index):
    body = f"5901234{int(index):05d}"
    total = sum(
        int(value) * (1 if position % 2 == 0 else 3)
        for position, value in enumerate(body)
    )
    return body + str((-total) % 10)


def catalog_mapping(gtins, _job_id):
    return {
        gtin: {
            "status": "resolved",
            "asin": f"B{index:09d}",
            "amazon_title": f"Amazon {gtin}",
            "amazon_brand": "Brand",
            "bsr_beauty": 5000,
            "beauty_status": "display_group_beauty",
            "product_type": "BEAUTY",
        }
        for index, gtin in enumerate(gtins, start=1)
    }


def unique_catalog_mapping(gtins, _job_id):
    return {
        gtin: {
            "status": "resolved",
            "asin": f"B{int(gtin[7:12]) + 1:09d}",
            "amazon_title": f"Amazon {gtin}", "amazon_brand": "Brand",
            "bsr_beauty": 5000, "beauty_status": "display_group_beauty",
            "product_type": "BEAUTY",
        }
        for gtin in gtins
    }


def pricing_mapping(asins, _job_id):
    return {
        asin: {
            "status": "success",
            "Venditori FBA": 2,
            "Venditori totali": 4,
            "Seller count source": "summary_number_of_offers",
            "reference_price": 30,
            "price_source": "buy_box",
        }
        for asin in asins
    }


def fee_result(asin, identifier, *, referral=4.5):
    details = [{
        "FeeType": "FBAFees",
        "FeeAmount": {"Amount": 4, "CurrencyCode": "EUR"},
        "FeePromotion": {"Amount": 0, "CurrencyCode": "EUR"},
        "FinalFee": {"Amount": 4, "CurrencyCode": "EUR"},
    }]
    if referral is not None:
        details.insert(0, {
            "FeeType": "ReferralFee",
            "FeeAmount": {"Amount": referral, "CurrencyCode": "EUR"},
            "FeePromotion": {"Amount": 0, "CurrencyCode": "EUR"},
            "FinalFee": {"Amount": referral, "CurrencyCode": "EUR"},
        })
    return {"FeesEstimateResult": {
        "Status": "Success",
        "FeesEstimateIdentifier": {
            "IdValue": asin,
            "SellerInputIdentifier": identifier,
            "PriceToEstimateFees": {"ListingPrice": {"Amount": 30, "CurrencyCode": "EUR"}},
        },
        "FeesEstimate": {"FeeDetailList": details},
    }}


def fee_batch(requests_, _token, *, referral=4.5):
    return [fee_result(row["asin"], row["identifier"], referral=referral) for row in requests_]


def fee_error_result(
    asin,
    identifier,
    *,
    status="ServerError",
    code="InternalError",
    error_type="Receiver",
    message="There is an internal service failure.",
):
    return {"FeesEstimateResult": {
        "Status": status,
        "FeesEstimateIdentifier": {
            "IdValue": asin,
            "SellerInputIdentifier": identifier,
            "PriceToEstimateFees": {
                "ListingPrice": {"Amount": 30, "CurrencyCode": "EUR"}
            },
        },
        "Error": {"Code": code, "Type": error_type, "Message": message},
    }}


class QogitaDiscoveryTests(unittest.TestCase):
    def test_loads_latest_qogita_catalog_generation_read_only(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Path(directory) / "manager.db"
            connection = sqlite3.connect(database)
            connection.executescript("""
                CREATE TABLE qogita_seller_catalog_runs (
                    run_id TEXT, seller_alias TEXT, observed_at TEXT, status TEXT
                );
                CREATE TABLE qogita_seller_product_snapshots (
                    run_id TEXT, gtin TEXT, variant_fid TEXT, offer_qid TEXT,
                    product_name TEXT, brand TEXT, category_name TEXT,
                    image_url TEXT, inventory INTEGER, selling_unit INTEGER,
                    product_url TEXT, observed_at TEXT
                );
                CREATE TABLE qogita_seller_tier_snapshots (
                    run_id TEXT, offer_qid TEXT, tier_mov TEXT, currency TEXT,
                    tier_price TEXT, is_active BOOLEAN, observed_at TEXT
                );
                INSERT INTO qogita_seller_catalog_runs VALUES
                    ('old','seller','2026-01-01T00:00:00Z','success'),
                    ('new','seller','2026-08-20T00:00:00Z','success');
                INSERT INTO qogita_seller_product_snapshots VALUES
                    ('old','11111111','f1','o1','Old','B','Beauty','',1,1,'','2026-01-01T00:00:00Z'),
                    ('new','22222222','f2','o2','New','B','Beauty','',4,1,'','2026-08-20T00:00:00Z');
                INSERT INTO qogita_seller_tier_snapshots VALUES
                    ('old','o1','100','EUR','9',1,'2026-01-01T00:00:00Z'),
                    ('new','o2','100','EUR','10',1,'2026-08-20T00:00:00Z');
            """)
            connection.commit()
            connection.close()
            rows = _sqlite_rows(database)
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["gtin"], "22222222")

    def test_gtin_validation_deduplication_vat_and_least_mov_tier(self):
        rows = [
            qogita_row(price="8", mov="500"),
            qogita_row(price="10", mov="100"),
            qogita_row(gtin="invalid"),
            qogita_row(gtin="12345678", stock=0),
        ]
        candidates, diagnostics = normalize_qogita_candidates(
            rows, minimum_stock=1,
            now=datetime(2026, 8, 20, 12, 0, tzinfo=timezone.utc),
        )
        self.assertTrue(valid_gtin("8801234567890"))
        self.assertFalse(valid_gtin("invalid"))
        self.assertEqual(len(candidates), 1)
        candidate = candidates[0]
        self.assertEqual(candidate["mov"], Decimal("100"))
        self.assertEqual(candidate["cost_net"], Decimal("10.00"))
        self.assertEqual(candidate["cost_vat"], Decimal("2.20"))
        self.assertEqual(candidate["cost_gross"], Decimal("12.20"))
        self.assertEqual(len(candidate["scenarios"]), 2)
        self.assertEqual(diagnostics["invalid_gtin"], 1)
        self.assertEqual(diagnostics["below_stock"], 1)

    def test_manager_compatible_inactive_flag_is_diagnostic_but_non_eur_is_rejected(self):
        inactive = qogita_row(active=False)
        non_eur = qogita_row(gtin="12345678")
        non_eur["currency"] = "USD"
        candidates, diagnostics = normalize_qogita_candidates(
            [inactive, non_eur],
            now=datetime(2026, 8, 20, 12, 0, tzinfo=timezone.utc),
        )
        self.assertEqual(len(candidates), 1)
        self.assertFalse(candidates[0]["tier_is_active"])
        self.assertEqual(diagnostics["missing_price"], 1)

    def test_stock_must_cover_selling_unit(self):
        candidates, diagnostics = normalize_qogita_candidates([
            qogita_row(stock=5, selling_unit=6, active=False),
        ])
        self.assertEqual(candidates, [])
        self.assertEqual(diagnostics["below_selling_unit"], 1)


class QogitaRefreshTests(unittest.TestCase):
    NOW = datetime(2026, 8, 20, 12, 0, tzinfo=timezone.utc)

    def manager_tree(self, root):
        (root / ".venv/bin").mkdir(parents=True)
        (root / ".venv/bin/python").touch()
        (root / "src").mkdir()
        (root / "scripts").mkdir()
        (root / "scripts/sync_qogita_seller_catalog.py").touch()

    def test_cache_ttl_is_strictly_less_than_twenty_four_hours(self):
        fresh = inspect_qogita_cache([
            qogita_row(observed_at="2026-08-19T12:00:01Z"),
        ], now=self.NOW)
        stale = inspect_qogita_cache([
            qogita_row(observed_at="2026-08-19T12:00:00Z"),
        ], now=self.NOW)
        self.assertTrue(fresh["fresh"])
        self.assertFalse(stale["fresh"])
        self.assertEqual(stale["stale_aliases"], ["seller"])

    def test_snapshots_must_advance_for_every_alias(self):
        before = {"A": "2026-08-18T10:00:00Z", "B": "2026-08-18T10:00:00Z"}
        self.assertTrue(snapshots_advanced(before, {
            "A": "2026-08-20T10:00:00Z", "B": "2026-08-20T11:00:00Z",
        }, ["A", "B"]))
        self.assertFalse(snapshots_advanced(before, {
            "A": "2026-08-20T10:00:00Z", "B": "2026-08-18T10:00:00Z",
        }, ["A", "B"]))

    def test_pythonpath_absent_is_set_to_manager_src_without_mutating_source(self):
        original = {"PATH": "/usr/bin"}
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            self.manager_tree(root)
            environment = build_manager_subprocess_environment(
                root, environ=original,
            )
        self.assertEqual(environment["PYTHONPATH"], str(root.resolve() / "src"))
        self.assertEqual(environment["PATH"], "/usr/bin")
        self.assertEqual(original, {"PATH": "/usr/bin"})

    def test_existing_pythonpath_is_preserved_after_manager_src_using_pathsep(self):
        original = {"PYTHONPATH": "/existing/one" + os.pathsep + "/existing/two"}
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            self.manager_tree(root)
            environment = build_manager_subprocess_environment(
                root, environ=original,
            )
        self.assertEqual(
            environment["PYTHONPATH"].split(os.pathsep),
            [str(root.resolve() / "src"), "/existing/one", "/existing/two"],
        )
        self.assertEqual(
            original["PYTHONPATH"],
            "/existing/one" + os.pathsep + "/existing/two",
        )

    def test_missing_manager_src_is_explicit(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            (root / ".venv/bin").mkdir(parents=True)
            (root / ".venv/bin/python").touch()
            (root / "scripts").mkdir()
            (root / "scripts/sync_qogita_seller_catalog.py").touch()
            with self.assertRaises(QogitaRefreshConfigurationError) as raised:
                build_manager_subprocess_environment(root, environ={})
        self.assertEqual(raised.exception.code, "manager_src_missing")

    def test_missing_manager_script_is_explicit_and_subprocess_is_not_started(self):
        calls = []
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            (root / ".venv/bin").mkdir(parents=True)
            (root / ".venv/bin/python").touch()
            (root / "src").mkdir()
            result = refresh_qogita_seller_catalogs(
                ["SELLER"], manager_root=root,
                runner=lambda *_args, **_kwargs: calls.append(True),
            )
        self.assertEqual(result["error_code"], "manager_script_missing")
        self.assertEqual(calls, [])

    def test_manager_runner_delegates_auth_token_and_all_aliases_without_credentials(self):
        calls = []

        def runner(command, **kwargs):
            calls.append((command, kwargs))
            alias = command[-1]
            return SimpleNamespace(
                returncode=0,
                stdout='{"status":"success","seller_alias":"%s"}\n' % alias,
                stderr="",
            )

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            self.manager_tree(root)
            result = refresh_qogita_seller_catalogs(
                ["SELLER-B", "SELLER-A"], manager_root=root, runner=runner,
            )
        self.assertEqual(result["status"], "success")
        self.assertEqual(result["updated_aliases"], ["SELLER-A", "SELLER-B"])
        self.assertEqual([row[0][-1] for row in calls], ["SELLER-A", "SELLER-B"])
        self.assertTrue(all(
            kwargs["env"]["PYTHONPATH"].split(os.pathsep)[0]
            == str(root.resolve() / "src")
            for _, kwargs in calls
        ))
        self.assertTrue(all(
            row[0][1].endswith("sync_qogita_seller_catalog.py")
            for row in calls
        ))

    def test_manager_lock_conflict_is_a_safe_failure(self):
        def runner(_command, **_kwargs):
            return SimpleNamespace(
                returncode=0,
                stdout='{"status":"skipped","reason":"already_running"}\n',
                stderr="",
            )

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            self.manager_tree(root)
            result = refresh_qogita_seller_catalogs(
                ["SELLER"], manager_root=root, runner=runner,
            )
        self.assertEqual(result["status"], "failed")
        self.assertEqual(result["error_code"], "already_running")

    def test_manager_database_import_succeeds_offline_with_scout_environment(self):
        manager_root = Path(__file__).resolve().parents[2] / "Glow-Up-Manager"
        environment = build_manager_subprocess_environment(
            manager_root, environ={"PATH": os.environ.get("PATH", "")},
        )
        completed = subprocess.run(
            [str(manager_root / ".venv/bin/python"), "-c", "import database"],
            cwd=manager_root, env=environment, capture_output=True, text=True,
            check=False, timeout=10,
        )
        self.assertEqual(completed.returncode, 0, completed.stderr)


class AmazonDiscoveryTests(unittest.TestCase):
    def test_ean_13_classifies_as_ean(self):
        self.assertEqual(classify_catalog_identifier("8809080826140"), "EAN")

    def test_upc_12_classifies_as_upc(self):
        self.assertEqual(classify_catalog_identifier("012345678905"), "UPC")

    def test_gtin_14_classifies_as_gtin(self):
        self.assertEqual(classify_catalog_identifier("01234567890128"), "GTIN")

    def test_non_numeric_identifier_is_invalid(self):
        self.assertIsNone(classify_catalog_identifier("88090808A6140"))

    def test_unsupported_identifier_length_is_invalid(self):
        self.assertIsNone(classify_catalog_identifier("12345678"))
        self.assertIsNone(classify_catalog_identifier("123456789012345"))

    def test_catalog_mixed_batch_is_separated_by_identifier_type(self):
        calls = []

        def request(method, url, **kwargs):
            calls.append((method, url, kwargs))
            return FakeResponse({"items": []})

        provider = RefreshingTokenProvider(lambda: "token")
        result = search_catalog_by_gtins_batch(
            [
                "8809080826140", "012345678905", "01234567890128",
                "8809532220700", "invalid",
            ],
            provider,
            marketplace_id="APJ6JRA9NG5V4", request_func=request,
            sleep_func=lambda _: None,
        )
        self.assertEqual(result, [])
        self.assertEqual(result.invalid_identifiers, ("invalid",))
        params = [call[2]["params"] for call in calls]
        self.assertEqual(
            [(row["identifiersType"], row["identifiers"]) for row in params],
            [
                ("EAN", "8809080826140,8809532220700"),
                ("UPC", "012345678905"),
                ("GTIN", "01234567890128"),
            ],
        )
        self.assertTrue(all(
            row["includedData"] == (
                "summaries,identifiers,salesRanks,productTypes,images,"
                "relationships,attributes,classifications,dimensions"
            )
            for row in params
        ))

    def test_catalog_batch_has_at_most_twenty_identifiers(self):
        twenty = [f"8809080826{index:03d}" for index in range(20)]
        batches, invalid = catalog_identifier_batches(twenty)
        self.assertEqual(invalid, [])
        self.assertEqual(len(batches), 1)
        self.assertEqual(len(batches[0][1]), 20)
        with self.assertRaises(ValueError):
            catalog_identifier_batches(twenty + ["8809080826999"])

    def test_catalog_partial_ean_batch_marks_unreturned_input_not_found(self):
        eans = ["8809080826140", "8809532220700", "8809080820438"]

        def item(asin, ean):
            return {"asin": asin, "identifiers": [{"identifiers": [
                {"identifierType": "EAN", "identifier": ean}
            ]}]}

        def request(_method, _url, **_kwargs):
            return FakeResponse({"items": [
                item("B07PBXTKPY", eans[0]),
                item("B07QM9YV2S", eans[1]),
            ]})

        provider = RefreshingTokenProvider(lambda: "token")
        items = search_catalog_by_gtins_batch(
            eans, provider, marketplace_id="APJ6JRA9NG5V4",
            request_func=request, sleep_func=lambda _: None,
        )
        mapping = correlate_catalog_items(eans, items)
        self.assertEqual(mapping[eans[0]]["asin"], "B07PBXTKPY")
        self.assertEqual(mapping[eans[1]]["asin"], "B07QM9YV2S")
        self.assertEqual(mapping[eans[2]]["status"], "not_found")
        self.assertEqual(mapping[eans[2]]["identifier_type"], "EAN")

    def test_catalog_paginates_twenty_identifiers_until_all_are_received(self):
        eans = [f"8800000000{index:03d}" for index in range(20)]
        calls = []

        def item(asin, ean):
            return {"asin": asin, "identifiers": [{"identifiers": [
                {"identifierType": "EAN", "identifier": ean}
            ]}]}

        def request(_method, _url, **kwargs):
            calls.append(dict(kwargs["params"]))
            start = 10 if kwargs["params"].get("pageToken") else 0
            payload = {
                "items": [item(f"B{index:09d}", eans[index]) for index in range(start, start + 10)],
                "numberOfResults": 20,
            }
            if start == 0:
                payload["pagination"] = {"nextToken": "page-2"}
            return FakeResponse(payload)

        result = search_catalog_by_gtins_batch(
            eans, RefreshingTokenProvider(lambda: "token"), marketplace_id="IT",
            request_func=request, sleep_func=lambda _: None,
        )
        mapping = correlate_catalog_items(eans, result)
        self.assertEqual(len(calls), 2)
        self.assertEqual(calls[0]["pageSize"], 20)
        self.assertNotIn("pageToken", calls[0])
        self.assertEqual(calls[1]["pageToken"], "page-2")
        self.assertTrue(all(row["status"] == "resolved" for row in mapping.values()))
        self.assertEqual(result.batch_diagnostics[0], {
            "identifier_type": "EAN", "page_count": 2,
            "number_of_results": 20, "had_next_token": True,
            "items_received": 20, "input_identifier_count": 20,
            "complete": True, "error": None,
        })

    def test_catalog_supports_three_pages_and_deduplicates_asins(self):
        ean = "8809532221523"
        payloads = [
            {"items": [{"asin": "B000000001", "identifiers": [{"identifiers": [
                {"identifierType": "EAN", "identifier": ean}
            ]}]}], "pagination": {"nextToken": "two"}},
            {"items": [{"asin": "B000000001", "identifiers": [{"identifiers": [
                {"identifierType": "EAN", "identifier": ean}
            ]}]}], "pagination": {"nextToken": "three"}},
            {"items": [{"asin": "B000000002", "identifiers": [{"identifiers": [
                {"identifierType": "EAN", "identifier": ean}
            ]}]}]},
        ]

        result = search_catalog_by_gtins_batch(
            [ean], RefreshingTokenProvider(lambda: "token"), marketplace_id="IT",
            request_func=lambda *_args, **_kwargs: FakeResponse(payloads.pop(0)),
            sleep_func=lambda _: None,
        )
        mapping = correlate_catalog_items([ean], result)
        self.assertEqual(len(result), 2)
        self.assertEqual(mapping[ean]["status"], "ambiguous")
        self.assertEqual(
            {row["asin"] for row in mapping[ean]["listings"]},
            {"B000000001", "B000000002"},
        )
        self.assertEqual(result.batch_diagnostics[0]["page_count"], 3)

    def test_catalog_retries_a_transient_second_page(self):
        ean = "8809532221523"
        calls = []

        def request(_method, _url, **kwargs):
            calls.append(kwargs["params"].get("pageToken"))
            if len(calls) == 1:
                return FakeResponse({"items": [], "pagination": {"nextToken": "two"}})
            if len(calls) == 2:
                return FakeResponse(status=503)
            return FakeResponse({"items": [{"asin": "B000000001", "identifiers": [{
                "identifiers": [{"identifierType": "EAN", "identifier": ean}]
            }]}]})

        result = search_catalog_by_gtins_batch(
            [ean], RefreshingTokenProvider(lambda: "token"), marketplace_id="IT",
            request_func=request, sleep_func=lambda _: None,
        )
        self.assertEqual(calls, [None, "two", "two"])
        self.assertEqual(correlate_catalog_items([ean], result)[ean]["status"], "resolved")

    def test_failed_later_page_is_incomplete_not_not_found(self):
        eans = ["8809532221523", "8809532220748"]
        calls = []

        def request(_method, _url, **kwargs):
            calls.append(kwargs["params"].get("pageToken"))
            if calls[0] is None and len(calls) == 1:
                return FakeResponse({"items": [], "pagination": {"nextToken": "two"}})
            return FakeResponse(status=503)

        result = search_catalog_by_gtins_batch(
            eans, RefreshingTokenProvider(lambda: "token"), marketplace_id="IT",
            request_func=request, sleep_func=lambda _: None,
        )
        mapping = correlate_catalog_items(eans, result)
        self.assertEqual(len(calls), 5)
        self.assertTrue(all(row["status"] == "catalog_incomplete" for row in mapping.values()))
        self.assertFalse(result.batch_diagnostics[0]["complete"])
        self.assertEqual(result.batch_diagnostics[0]["page_count"], 1)

    def test_lwa_connection_error_is_retried_inside_request_boundary(self):
        import requests

        token_attempts = []
        sleeps = []

        def token():
            token_attempts.append(len(token_attempts) + 1)
            if len(token_attempts) < 3:
                raise requests.ConnectionError("diagnostic detail must not be logged")
            return "token"

        response = _request_with_retry(
            "GET", "https://example.invalid",
            token_provider=RefreshingTokenProvider(token),
            request_func=lambda *_args, **_kwargs: FakeResponse({}),
            sleep_func=sleeps.append,
            random_func=lambda: 0,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(token_attempts, [1, 2, 3])
        self.assertEqual(sleeps, [2, 4])

    def test_catalog_connection_outage_opens_circuit_for_remaining_types(self):
        import requests

        calls = []

        def request(*_args, **_kwargs):
            calls.append(True)
            raise requests.ConnectionError("offline")

        ean = "8809532221523"
        upc = "012345678905"
        result = search_catalog_by_gtins_batch(
            [ean, upc], RefreshingTokenProvider(lambda: "token"),
            marketplace_id="IT", request_func=request,
            sleep_func=lambda _: None,
        )
        mapping = correlate_catalog_items([ean, upc], result)
        self.assertEqual(len(calls), 4)
        self.assertEqual(mapping[ean]["status"], "catalog_incomplete")
        self.assertEqual(mapping[upc]["status"], "catalog_incomplete")
        self.assertEqual(result.batch_diagnostics[1]["error"], "circuit_open")

    def test_no_next_token_is_one_page_and_not_found_is_final(self):
        ean = "8809532221523"
        result = search_catalog_by_gtins_batch(
            [ean], RefreshingTokenProvider(lambda: "token"), marketplace_id="IT",
            request_func=lambda *_args, **_kwargs: FakeResponse({"items": []}),
            sleep_func=lambda _: None,
        )
        mapping = correlate_catalog_items([ean], result)
        self.assertEqual(mapping[ean]["status"], "not_found")
        self.assertEqual(mapping[ean]["diagnostics"]["page_count"], 1)
        self.assertFalse(mapping[ean]["diagnostics"]["had_next_token"])

    def test_gs1_identifiers_share_a_canonical_gtin14(self):
        ean = normalize_commercial_identifier("8809532221523", "EAN")
        gtin = normalize_commercial_identifier("08809532221523", "GTIN")
        upc = normalize_commercial_identifier("012345678905", "UPC")
        upc_gtin = normalize_commercial_identifier("00012345678905", "GTIN")
        self.assertEqual(ean["canonical_gtin14"], "08809532221523")
        self.assertEqual(ean["canonical_gtin14"], gtin["canonical_gtin14"])
        self.assertEqual(upc["canonical_gtin14"], upc_gtin["canonical_gtin14"])
        invalid = normalize_commercial_identifier("8809532221524", "EAN")
        self.assertIsNone(invalid["canonical_gtin14"])
        self.assertEqual(invalid["raw_identifier"], "8809532221524")

    def test_zero_padded_gtin_correlates_and_preserves_raw_identifier(self):
        ean = "8809532221523"
        raw_gtin = "08809532221523"
        item = {"asin": "B09ZB7GDJL", "identifiers": [{"identifiers": [
            {"identifierType": "GTIN", "identifier": raw_gtin}
        ]}]}
        mapping = correlate_catalog_items([ean], [item])[ean]
        self.assertEqual(mapping["status"], "resolved")
        identifier = mapping["listings"][0]["diagnostics"]["commercial_identifiers"][0]
        self.assertEqual(identifier["raw_identifier"], raw_gtin)
        self.assertEqual(identifier["raw_type"], "GTIN")
        self.assertEqual(identifier["canonical_gtin14"], raw_gtin)

    def test_invalid_identifier_is_not_sent_and_is_diagnostic(self):
        calls = []
        provider = RefreshingTokenProvider(lambda: "token")
        items = search_catalog_by_gtins_batch(
            ["not-numeric"], provider, marketplace_id="IT",
            request_func=lambda *args, **kwargs: calls.append((args, kwargs)),
        )
        self.assertEqual(calls, [])
        self.assertEqual(items.invalid_identifiers, ("not-numeric",))
        mapping = correlate_catalog_items(["not-numeric"], items)
        self.assertEqual(mapping["not-numeric"], {
            "status": "invalid_identifier", "identifier_type": None,
        })

    def test_beauty_rank_requires_explicit_beauty_display_group(self):
        item = {"salesRanks": [{"displayGroupRanks": [
            {"websiteDisplayGroup": "beauty_display_on_website", "title": "Bellezza", "rank": 4321}
        ]}]}
        self.assertEqual(beauty_rank(item), (4321, "display_group_beauty"))
        item["salesRanks"][0]["displayGroupRanks"][0]["websiteDisplayGroup"] = "grocery"
        item["salesRanks"][0]["displayGroupRanks"][0]["title"] = "Alimentari"
        self.assertEqual(beauty_rank(item), (None, "beauty_rank_unverified"))

    def test_catalog_correlation_handles_missing_ambiguous_and_duplicate_asin(self):
        def item(asin, gtin):
            return {"asin": asin, "identifiers": [{"identifiers": [
                {"identifierType": "EAN", "identifier": gtin}
            ]}]}
        identifiers = ["1111111111111", "2222222222222", "3333333333333"]
        mapping = correlate_catalog_items(
            identifiers,
            [
                item("B000000001", identifiers[0]),
                item("B000000002", identifiers[1]),
                item("B000000003", identifiers[1]),
            ],
        )
        self.assertEqual(mapping[identifiers[0]]["status"], "resolved")
        self.assertEqual(mapping[identifiers[0]]["identifier_type"], "EAN")
        self.assertEqual(mapping[identifiers[1]]["status"], "ambiguous")
        self.assertEqual(mapping[identifiers[2]]["status"], "not_found")

    def test_pricing_batch_size_counts_and_price_hierarchy(self):
        self.assertEqual(len(build_item_offers_batch_requests(["B000000001"] * 20, "IT")["requests"]), 20)
        with self.assertRaises(ValueError):
            build_item_offers_batch_requests(["B000000001"] * 21, "IT")
        entries = [{
            "request": {"uri": "/products/pricing/v0/items/B000000001/offers"},
            "status": {"statusCode": 200},
            "body": {"payload": {
                "ASIN": "B000000001",
                "Summary": {
                    "NumberOfOffers": [
                        {"condition": "new", "fulfillmentChannel": "Amazon", "OfferCount": 3},
                        {"condition": "new", "fulfillmentChannel": "Merchant", "OfferCount": 4},
                    ],
                    "BuyBoxPrices": [],
                    "LowestPrices": [
                        {"fulfillmentChannel": "Amazon", "LandedPrice": {"Amount": 20, "CurrencyCode": "EUR"}},
                        {"fulfillmentChannel": "Merchant", "LandedPrice": {"Amount": 18, "CurrencyCode": "EUR"}},
                    ],
                },
                "Offers": [],
            }},
        }]
        parsed = parse_item_offers_batch(entries)["B000000001"]
        self.assertEqual(parsed["Venditori FBA"], 3)
        self.assertEqual(parsed["Venditori totali"], 7)
        self.assertEqual(parsed["Seller count source"], "summary_number_of_offers")
        self.assertEqual(parsed["reference_price"], Decimal("20"))
        self.assertEqual(parsed["price_source"], "min_fba")

        entries[0]["body"]["payload"]["Summary"]["BuyBoxPrices"] = [{
            "LandedPrice": {"Amount": 25, "CurrencyCode": "EUR"}
        }]
        parsed = parse_item_offers_batch(entries)["B000000001"]
        self.assertEqual(parsed["reference_price"], Decimal("25"))
        self.assertEqual(parsed["price_source"], "buy_box")

    def test_retry_429_and_token_renewal_after_401(self):
        tokens = iter(["first", "second"])
        provider = RefreshingTokenProvider(lambda: next(tokens))
        statuses = iter([401, 429, 200])
        seen = []

        def request(_method, _url, **kwargs):
            seen.append(kwargs["headers"]["x-amz-access-token"])
            return FakeResponse({}, next(statuses))

        response = _request_with_retry(
            "GET", "https://example.invalid", token_provider=provider,
            request_func=request, sleep_func=lambda _: None,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(seen, ["first", "second", "second"])

    def test_product_fees_retries_429_and_renews_401_token(self):
        import requests
        fetched = []
        provider = RefreshingTokenProvider(
            lambda: fetched.append(len(fetched) + 1) or f"token-{len(fetched)}"
        )
        statuses = iter([401, 429, 200])

        def fees(_rows, _token):
            status = next(statuses)
            if status != 200:
                response = requests.Response()
                response.status_code = status
                raise requests.HTTPError(response=response)
            return ["ok"]

        result = _fees_batch_with_retry(
            fees, [{}], provider, job_id="job", sleep_func=lambda _: None
        )
        self.assertEqual(result, ["ok"])
        self.assertEqual(fetched, [1, 2])

    def test_http_200_product_fee_success_is_classified_success(self):
        entry = fee_result("B000000001", "fee|1")
        self.assertEqual(
            classify_product_fee_entry(entry)["classification"], "success"
        )

    def test_server_internal_receiver_error_is_retryable(self):
        diagnostics = classify_product_fee_entry(
            fee_error_result("B000000001", "fee|1")
        )
        self.assertEqual(diagnostics["classification"], "transient")
        self.assertEqual(diagnostics["status"], "ServerError")
        self.assertEqual(diagnostics["error_code"], "InternalError")
        self.assertEqual(diagnostics["error_type"], "Receiver")

    def test_client_fee_error_is_not_retryable(self):
        diagnostics = classify_product_fee_entry(fee_error_result(
            "B000000001", "fee|1", status="ClientError",
            code="InvalidInput", error_type="Sender",
            message="Invalid marketplace input",
        ))
        self.assertEqual(diagnostics["classification"], "permanent")

    def test_element_retry_delay_respects_backoff_retry_after_and_rate(self):
        entries = ProductFeeBatchResults(
            [], retry_after="3", rate_limit="0.5"
        )
        first = _fee_element_retry_delay(entries, 1)
        second = _fee_element_retry_delay(entries, 2)
        self.assertGreaterEqual(first, 3)
        self.assertLessEqual(first, 3.3)
        self.assertGreaterEqual(second, 4)
        self.assertLessEqual(second, 4.4)


class DiscoveryPipelineTests(unittest.TestCase):
    def run_pipeline(self, directory, **overrides):
        store = DiscoveryCheckpointStore(directory)
        arguments = {
            "checkpoint_store": store,
            "catalog_batch": catalog_mapping,
            "pricing_batch": pricing_mapping,
            "fees_batch": fee_batch,
            "token_provider": RefreshingTokenProvider(lambda: "token"),
            "qogita_loader": lambda: [qogita_row()],
            "qogita_refresher": lambda _aliases: (_ for _ in ()).throw(
                AssertionError("fresh cache must not refresh")
            ),
            "now_provider": lambda: datetime(
                2026, 8, 20, 12, 0, tzinfo=timezone.utc
            ),
            "sleep_func": lambda _: None,
            "pricing_batch_interval": 0,
            "fee_batch_interval": 0,
        }
        arguments.update(overrides)
        return run_discovery(default_filters(), **arguments)

    def test_fresh_cache_skips_refresh_and_persists_preflight(self):
        with tempfile.TemporaryDirectory() as directory:
            state = self.run_pipeline(directory)
        self.assertEqual(state["qogita_refresh_status"], "cache_fresh")
        self.assertEqual(
            state["qogita_snapshot_before"],
            {"seller": "2026-08-20T10:00:00Z"},
        )
        self.assertEqual(
            state["qogita_snapshot_after"],
            state["qogita_snapshot_before"],
        )
        self.assertEqual(state["qogita_seller_aliases_updated"], [])
        self.assertIsNone(state["qogita_refresh_started_at"])
        self.assertIsNone(state["qogita_refresh_completed_at"])

    def test_stale_cache_refreshes_all_aliases_and_reloads_generation(self):
        refreshed = {"value": False}
        refresh_calls = []

        def rows():
            observed = (
                "2026-08-20T11:00:00Z" if refreshed["value"]
                else "2026-08-18T10:00:00Z"
            )
            return [
                qogita_row(
                    "8801234567890", seller_alias="SELLER-A",
                    observed_at=observed, active=False,
                ),
                qogita_row(
                    "8801234567891", seller_alias="SELLER-B",
                    observed_at=observed, active=False,
                ),
            ]

        def refresh(aliases):
            refresh_calls.append(list(aliases))
            refreshed["value"] = True
            return {
                "status": "success", "updated_aliases": list(aliases),
                "duration_seconds": 1.25,
            }

        with tempfile.TemporaryDirectory() as directory:
            state = self.run_pipeline(
                directory, qogita_loader=rows, qogita_refresher=refresh,
            )
        self.assertEqual(refresh_calls, [["SELLER-A", "SELLER-B"]])
        self.assertEqual(state["status"], "completed")
        self.assertEqual(state["qogita_refresh_status"], "refreshed")
        self.assertEqual(
            state["qogita_seller_aliases_updated"],
            ["SELLER-A", "SELLER-B"],
        )
        self.assertEqual(
            state["qogita_snapshot_after"], {
                "SELLER-A": "2026-08-20T11:00:00Z",
                "SELLER-B": "2026-08-20T11:00:00Z",
            },
        )
        self.assertEqual(state["qogita_refresh_duration_seconds"], 1.25)
        self.assertEqual(len(state["results"]), 2)
        self.assertTrue(all(
            row["tier_is_active"] is False for row in state["results"]
        ))

    def test_refresh_failure_stops_before_every_amazon_phase(self):
        amazon_calls = []

        def amazon_forbidden(*_args):
            amazon_calls.append(True)
            raise AssertionError("Amazon must not run")

        with tempfile.TemporaryDirectory() as directory:
            store = DiscoveryCheckpointStore(directory)
            state = self.run_pipeline(
                directory,
                checkpoint_store=store,
                qogita_loader=lambda: [qogita_row(
                    observed_at="2026-08-18T10:00:00Z"
                )],
                qogita_refresher=lambda _aliases: {
                    "status": "failed", "updated_aliases": [],
                    "error_code": "authentication_failed",
                    "duration_seconds": 0.5,
                },
                catalog_batch=amazon_forbidden,
                pricing_batch=amazon_forbidden,
                fees_batch=amazon_forbidden,
            )
            persisted = store.load(state["job_id"])
            latest_incomplete = store.latest_incomplete()
        self.assertEqual(amazon_calls, [])
        self.assertEqual(state["status"], "qogita_refresh_failed")
        self.assertEqual(state["phase"], "qogita_refresh_failed")
        self.assertEqual(state["qogita_refresh_status"], "refresh_failed")
        self.assertEqual(state["qogita_refresh_error"], "authentication_failed")
        self.assertIsNone(state["qogita_snapshot_after"])
        self.assertEqual(persisted["qogita_refresh_status"], "refresh_failed")
        self.assertEqual(latest_incomplete["job_id"], state["job_id"])

    def test_success_without_a_new_fresh_generation_stops_before_amazon(self):
        amazon_calls = []

        def amazon_forbidden(*_args):
            amazon_calls.append(True)
            raise AssertionError("Amazon must not run")

        with tempfile.TemporaryDirectory() as directory:
            state = self.run_pipeline(
                directory,
                qogita_loader=lambda: [qogita_row(
                    observed_at="2026-08-18T10:00:00Z"
                )],
                qogita_refresher=lambda aliases: {
                    "status": "success", "updated_aliases": list(aliases),
                },
                catalog_batch=amazon_forbidden,
                pricing_batch=amazon_forbidden,
                fees_batch=amazon_forbidden,
            )
        self.assertEqual(amazon_calls, [])
        self.assertEqual(state["status"], "qogita_refresh_failed")
        self.assertEqual(
            state["qogita_refresh_error"], "refreshed_cache_not_fresh"
        )

    def test_resume_after_refresh_failure_retries_refresh(self):
        refreshed = {"value": False}
        refresh_calls = []

        def rows():
            return [qogita_row(observed_at=(
                "2026-08-20T11:00:00Z" if refreshed["value"]
                else "2026-08-18T10:00:00Z"
            ))]

        with tempfile.TemporaryDirectory() as directory:
            store = DiscoveryCheckpointStore(directory)
            failed = self.run_pipeline(
                directory, checkpoint_store=store, qogita_loader=rows,
                qogita_refresher=lambda aliases: {
                    "status": "failed", "updated_aliases": [],
                    "error_code": "temporary",
                },
            )

            def succeeds(aliases):
                refresh_calls.append(list(aliases))
                refreshed["value"] = True
                return {
                    "status": "success", "updated_aliases": list(aliases),
                }

            resumed = run_discovery(
                failed["filters"], checkpoint_store=store,
                catalog_batch=catalog_mapping, pricing_batch=pricing_mapping,
                fees_batch=fee_batch,
                token_provider=RefreshingTokenProvider(lambda: "token"),
                qogita_loader=rows, qogita_refresher=succeeds,
                job_id=failed["job_id"], sleep_func=lambda _: None,
                now_provider=lambda: datetime(
                    2026, 8, 20, 12, 0, tzinfo=timezone.utc
                ),
                pricing_batch_interval=0, fee_batch_interval=0,
            )
        self.assertEqual(refresh_calls, [["seller"]])
        self.assertEqual(resumed["status"], "completed")
        self.assertEqual(resumed["qogita_refresh_status"], "refreshed")

    def test_resume_after_published_refresh_does_not_refresh_again(self):
        with tempfile.TemporaryDirectory() as directory:
            store = DiscoveryCheckpointStore(directory)
            state = store.create(default_filters())
            state.update({
                "phase": "qogita_refreshing",
                "qogita_refresh_status": "refreshing",
                "qogita_snapshot_before": {
                    "seller": "2026-08-18T10:00:00Z"
                },
                "qogita_seller_aliases": ["seller"],
            })
            store.save(state)
            resumed = run_discovery(
                state["filters"], checkpoint_store=store,
                catalog_batch=catalog_mapping, pricing_batch=pricing_mapping,
                fees_batch=fee_batch,
                token_provider=RefreshingTokenProvider(lambda: "token"),
                qogita_loader=lambda: [qogita_row(
                    observed_at="2026-08-20T11:00:00Z"
                )],
                qogita_refresher=lambda _aliases: (_ for _ in ()).throw(
                    AssertionError("published refresh must not repeat")
                ),
                job_id=state["job_id"], sleep_func=lambda _: None,
                now_provider=lambda: datetime(
                    2026, 8, 20, 12, 0, tzinfo=timezone.utc
                ),
                pricing_batch_interval=0, fee_batch_interval=0,
            )
        self.assertEqual(resumed["status"], "completed")
        self.assertEqual(resumed["qogita_refresh_status"], "refreshed")

    def test_pipeline_filters_economics_score_ranking_and_funnel(self):
        rows = [qogita_row(), qogita_row(gtin="12345678", price="11")]
        with tempfile.TemporaryDirectory() as directory:
            state = self.run_pipeline(directory, qogita_loader=lambda: rows)
            self.assertEqual(state["status"], "completed")
            self.assertEqual(len(state["results"]), 2)
            self.assertGreaterEqual(state["results"][0]["margin_percent"], 15)
            expected = opportunity_score(5000, 2, 4, state["results"][0]["margin_percent"])
            self.assertEqual((state["results"][0]["score"], state["results"][0]["opportunity"]), expected)
            self.assertEqual(state["funnel"]["qogita_initial"], 2)
            self.assertEqual(state["funnel"]["competition_passed"], 2)
            self.assertEqual(state["funnel"]["fee_valid"], 2)
            self.assertEqual(state["funnel"]["margin_passed"], 2)
            self.assertEqual(state["funnel"]["margin_below_threshold"], 0)
            self.assertEqual(state["funnel"]["final_opportunities"], 2)
            self.assertTrue(all(
                row["evaluation_status"] == "margin_passed"
                for row in state["results"]
            ))

    def test_below_margin_candidate_keeps_complete_evaluation_in_checkpoint(self):
        with tempfile.TemporaryDirectory() as directory:
            store = DiscoveryCheckpointStore(directory)
            state = self.run_pipeline(
                directory,
                checkpoint_store=store,
                qogita_loader=lambda: [qogita_row(price="18")],
            )
            self.assertEqual(state["status"], "completed")
            self.assertEqual(state["results"], [])

            candidate = state["candidates"][0]
            self.assertEqual(candidate["economics"]["status"], "ready")
            self.assertLess(candidate["margin_percent"], 15)
            expected = opportunity_score(
                5000, 2, 4, candidate["margin_percent"]
            )
            self.assertEqual(
                (candidate["score"], candidate["opportunity"]), expected
            )
            self.assertTrue(all(
                target_price(candidate["economics"], target) is not None
                for target in (15, 20, 25)
            ))
            self.assertEqual(
                candidate["evaluation_status"], "margin_below_threshold"
            )
            self.assertEqual(
                candidate["exclusion_reason"], "margin_below_threshold"
            )
            self.assertEqual(state["funnel"]["margin_passed"], 0)
            self.assertEqual(state["funnel"]["margin_below_threshold"], 1)
            self.assertEqual(state["funnel"]["final_opportunities"], 0)

            persisted = store.load(state["job_id"])["candidates"][0]
            self.assertEqual(persisted["economics"]["status"], "ready")
            self.assertIn("margin_percent", persisted)
            self.assertIn("score", persisted)
            self.assertIn("opportunity", persisted)
            self.assertEqual(
                persisted["exclusion_reason"], "margin_below_threshold"
            )

    def test_resume_does_not_repeat_fees_for_below_margin_candidate(self):
        fee_calls = []

        def tracked_fees(requests_, token):
            fee_calls.append([row["asin"] for row in requests_])
            return fee_batch(requests_, token)

        with tempfile.TemporaryDirectory() as directory:
            store = DiscoveryCheckpointStore(directory)
            state = self.run_pipeline(
                directory,
                checkpoint_store=store,
                qogita_loader=lambda: [qogita_row(price="18")],
                fees_batch=tracked_fees,
            )
            self.assertEqual(len(fee_calls), 1)

            resumed = run_discovery(
                state["filters"], checkpoint_store=store,
                catalog_batch=lambda *_: (_ for _ in ()).throw(
                    AssertionError("catalog must not repeat")
                ),
                pricing_batch=lambda *_: (_ for _ in ()).throw(
                    AssertionError("pricing must not repeat")
                ),
                fees_batch=lambda *_: (_ for _ in ()).throw(
                    AssertionError("fees must not repeat")
                ),
                token_provider=RefreshingTokenProvider(lambda: "token"),
                qogita_loader=lambda: (_ for _ in ()).throw(
                    AssertionError("qogita must not reload")
                ),
                job_id=state["job_id"], sleep_func=lambda _: None,
                pricing_batch_interval=0, fee_batch_interval=0,
            )
            self.assertEqual(len(fee_calls), 1)
            self.assertEqual(resumed["status"], "completed")
            self.assertEqual(resumed["results"], [])
            self.assertEqual(
                resumed["candidates"][0]["exclusion_reason"],
                "margin_below_threshold",
            )


class MultiScenarioDiscoveryTests(unittest.TestCase):
    NOW = datetime(2026, 8, 20, 12, 0, tzinfo=timezone.utc)

    def run_pipeline(self, directory, **overrides):
        store = DiscoveryCheckpointStore(directory)
        arguments = {
            "checkpoint_store": store,
            "catalog_batch": catalog_mapping,
            "pricing_batch": pricing_mapping,
            "fees_batch": fee_batch,
            "token_provider": RefreshingTokenProvider(lambda: "token"),
            "qogita_loader": lambda: [qogita_row()],
            "qogita_refresher": lambda _aliases: (_ for _ in ()).throw(
                AssertionError("fresh cache must not refresh")
            ),
            "now_provider": lambda: self.NOW,
            "sleep_func": lambda _: None,
            "pricing_batch_interval": 0,
            "fee_batch_interval": 0,
        }
        arguments.update(overrides)
        return run_discovery(default_filters(), **arguments)

    def real_tier_rows(self):
        return [
            qogita_row(
                gtin="8809532220748", mov=mov, price=price, stock=4980,
                selling_unit=60, seller_alias="3VKO8Q",
                observed_at="2026-08-20T10:00:00Z",
            )
            for mov, price in (
                (500, "6.94"), (1500, "6.83"), (5000, "6.80"),
                (10000, "6.73"), (15000, "6.73"),
            )
        ]

    def run_rows(self, rows, **overrides):
        calls = {"catalog": [], "pricing": [], "fees": []}

        def catalog(values, job_id):
            calls["catalog"].append(list(values))
            return catalog_mapping(values, job_id)

        def pricing(values, job_id):
            calls["pricing"].append(list(values))
            return pricing_mapping(values, job_id)

        def fees(requests_, token):
            calls["fees"].append([row["asin"] for row in requests_])
            return fee_batch(requests_, token)

        with tempfile.TemporaryDirectory() as directory:
            arguments = dict(
                checkpoint_store=DiscoveryCheckpointStore(directory),
                catalog_batch=catalog, pricing_batch=pricing, fees_batch=fees,
                token_provider=RefreshingTokenProvider(lambda: "token"),
                qogita_loader=lambda: rows,
                qogita_refresher=lambda _: (_ for _ in ()).throw(
                    AssertionError("fresh fixture must not refresh")
                ),
                now_provider=lambda: self.NOW, sleep_func=lambda _: None,
                pricing_batch_interval=0, fee_batch_interval=0,
            )
            arguments.update(overrides)
            state = run_discovery(default_filters(), **arguments)
            persisted = arguments["checkpoint_store"].load(state["job_id"])
        return state, persisted, calls

    def test_real_qogita_product_preserves_five_stable_distinct_tiers(self):
        products, diagnostics = normalize_qogita_candidates(
            self.real_tier_rows(), now=self.NOW
        )
        self.assertEqual(len(products), 1)
        scenarios = products[0]["scenarios"]
        self.assertEqual(len(scenarios), 5)
        self.assertEqual(diagnostics["qogita_scenarios"], 5)
        self.assertNotEqual(scenarios[3]["scenario_id"], scenarios[4]["scenario_id"])
        self.assertEqual(
            scenarios[3]["cost_gross_unit_eur"],
            scenarios[4]["cost_gross_unit_eur"],
        )
        changed = self.real_tier_rows()
        changed[0]["tier_price"] = "5.55"
        changed[0]["observed_at"] = "2026-08-20T11:00:00Z"
        changed_id = normalize_qogita_candidates(changed, now=self.NOW)[0][0]["scenarios"][0]["scenario_id"]
        self.assertEqual(changed_id, scenarios[0]["scenario_id"])
        changed[0]["tier_mov"] = "500.00"
        formatted_id = normalize_qogita_candidates(changed, now=self.NOW)[0][0]["scenarios"][0]["scenario_id"]
        self.assertEqual(formatted_id, scenarios[0]["scenario_id"])

    def test_amazon_runs_once_then_economics_fans_out_to_all_tiers(self):
        state, persisted, calls = self.run_rows(self.real_tier_rows())
        self.assertEqual(calls["catalog"], [["8809532220748"]])
        self.assertEqual(len(calls["pricing"]), 1)
        self.assertEqual(len(calls["pricing"][0]), 1)
        self.assertEqual(len(calls["fees"]), 1)
        self.assertEqual(len(calls["fees"][0]), 1)
        self.assertEqual(len(state["amazon_observations"]), 1)
        self.assertEqual(len(state["results"]), 1)
        scenarios = state["results"][0]["scenarios"]
        self.assertEqual(len(scenarios), 5)
        self.assertEqual(len({row["margin_percent"] for row in scenarios}), 4)
        self.assertEqual(len({target_price(row["economics"], 15) for row in scenarios}), 4)
        self.assertEqual(len(persisted["candidates"][0]["scenarios"]), 5)
        self.assertEqual(len(persisted["amazon_observations"]), 1)
        self.assertEqual(state["funnel"]["qogita_products"], 1)
        self.assertEqual(state["funnel"]["qogita_scenarios"], 5)
        self.assertEqual(state["funnel"]["scenarios_evaluated"], 5)

    def test_new_checkpoint_has_valid_multiscenario_schema_and_recommendation(self):
        state, persisted, _ = self.run_rows(self.real_tier_rows())
        self.assertEqual(
            persisted["discovery_schema_version"], DISCOVERY_SCHEMA_VERSION
        )
        self.assertEqual(
            discovery_checkpoint_compatibility(persisted)["status"],
            "compatible",
        )
        product = persisted["results"][0]
        scenario = recommended_scenario(product)
        self.assertIsNotNone(scenario)
        self.assertIn("cost_gross_unit_eur", scenario)
        self.assertGreater(Decimal(str(scenario["cost_gross_unit_eur"])), 0)
        self.assertGreater(persisted["funnel"]["qogita_products"], 0)
        self.assertGreater(persisted["funnel"]["qogita_scenarios"], 0)

    def test_legacy_single_cost_checkpoint_is_not_synthesized_as_scenario(self):
        legacy = {
            "job_id": "legacy", "status": "completed", "phase": "completed",
            "funnel": {"qogita_initial": 80, "amazon_found": 40},
            "candidates": [{
                "gtin": "8809532221349", "cost_gross": "8.83",
                "cost_net": "7.24", "mov": "500", "asin": "B08W1ZHTL3",
            }],
            "results": [{
                "gtin": "8809532221349", "cost_gross": "8.83",
                "cost_net": "7.24", "mov": "500", "asin": "B08W1ZHTL3",
            }],
            "amazon_observations": [],
        }
        normalized = normalize_discovery_state(legacy)
        self.assertEqual(normalized["checkpoint_compatibility"], "legacy_incompatible")
        self.assertEqual(
            normalized["checkpoint_compatibility_reason"],
            "missing_purchase_scenarios",
        )
        self.assertNotIn("scenarios", normalized["results"][0])

    def test_legacy_checkpoint_stops_before_any_api_or_supplier_call(self):
        with tempfile.TemporaryDirectory() as directory:
            store = DiscoveryCheckpointStore(directory)
            legacy = {
                "job_id": "legacy", "status": "completed", "phase": "completed",
                "filters": default_filters(), "created_at": "2026-08-20T10:00:00Z",
                "updated_at": "2026-08-20T10:00:00Z", "funnel": {},
                "candidates": [{"gtin": "8809532221349", "cost_gross": "8.83"}],
                "results": [], "amazon_observations": [], "errors": [],
            }
            store.save(legacy)
            forbidden = lambda *_args, **_kwargs: (_ for _ in ()).throw(
                AssertionError("legacy rendering/resume must not call external stages")
            )
            result = run_discovery(
                default_filters(), checkpoint_store=store,
                catalog_batch=forbidden, pricing_batch=forbidden,
                fees_batch=forbidden,
                token_provider=RefreshingTokenProvider(lambda: "token"),
                qogita_loader=forbidden, job_id="legacy", sleep_func=lambda _: None,
            )
        self.assertEqual(result["status"], "legacy_incompatible")
        self.assertEqual(result["phase"], "legacy_incompatible")

    def test_multiscenario_funnel_view_uses_canonical_populated_keys(self):
        state, _, _ = self.run_rows(self.real_tier_rows())
        view = discovery_funnel_view(state)
        self.assertEqual(view["products"]["qogita_products"], 1)
        self.assertEqual(view["products"]["amazon_found"], 1)
        self.assertEqual(view["products"]["beauty_valid"], 1)
        self.assertEqual(view["products"]["bsr_passed"], 1)
        self.assertEqual(view["scenarios"]["qogita_scenarios"], 5)
        self.assertEqual(view["scenarios"]["scenarios_evaluated"], 5)

    def test_multiscenario_checkpoint_round_trip_preserves_recommendation(self):
        state, _, _ = self.run_rows(self.real_tier_rows())
        state = normalize_discovery_state(json.loads(json.dumps(state, default=str)))
        self.assertNotEqual(
            state["checkpoint_compatibility"], "legacy_incompatible"
        )
        product = state["results"][0]
        self.assertEqual(len(product["scenarios"]), 5)
        scenario = recommended_scenario(product)
        self.assertEqual(Decimal(str(scenario["account_mov"])), Decimal("10000"))
        self.assertGreater(Decimal(str(scenario["cost_gross_unit_eur"])), 0)

    def test_two_eans_resolving_to_one_asin_share_pricing_and_fees(self):
        rows = [
            qogita_row(gtin="8801234567890", observed_at="2026-08-20T10:00:00Z"),
            qogita_row(gtin="8801234567891", observed_at="2026-08-20T10:00:00Z"),
        ]
        catalog_calls = []
        pricing_calls = []
        fee_calls = []

        def catalog(values, _job_id):
            catalog_calls.append(list(values))
            return {value: {
                "status": "resolved", "identifier_type": "EAN",
                "asin": "B000SHARED", "amazon_title": "Shared",
                "amazon_brand": "Brand", "bsr_beauty": 5000,
                "beauty_status": "display_group_beauty", "product_type": "BEAUTY",
            } for value in values}

        def pricing(values, _job_id):
            pricing_calls.append(list(values))
            return pricing_mapping(values, _job_id)

        def fees(requests_, token):
            fee_calls.append([row["asin"] for row in requests_])
            return fee_batch(requests_, token)

        with tempfile.TemporaryDirectory() as directory:
            state = run_discovery(
                default_filters(), checkpoint_store=DiscoveryCheckpointStore(directory),
                catalog_batch=catalog, pricing_batch=pricing, fees_batch=fees,
                token_provider=RefreshingTokenProvider(lambda: "token"),
                qogita_loader=lambda: rows,
                qogita_refresher=lambda _: (_ for _ in ()).throw(AssertionError()),
                now_provider=lambda: self.NOW, sleep_func=lambda _: None,
                pricing_batch_interval=0, fee_batch_interval=0,
            )
        self.assertEqual(catalog_calls, [["8801234567890", "8801234567891"]])
        self.assertEqual(pricing_calls, [["B000SHARED"]])
        self.assertEqual(fee_calls, [["B000SHARED"]])
        self.assertEqual(len(state["amazon_observations"]), 1)
        self.assertEqual(len(state["results"]), 2)
        self.assertEqual(state["funnel"]["fee_valid"], 2)

    def test_completed_multiscenario_resume_reuses_shared_amazon_observation(self):
        with tempfile.TemporaryDirectory() as directory:
            store = DiscoveryCheckpointStore(directory)
            state = run_discovery(
                default_filters(), checkpoint_store=store,
                catalog_batch=catalog_mapping, pricing_batch=pricing_mapping,
                fees_batch=fee_batch,
                token_provider=RefreshingTokenProvider(lambda: "token"),
                qogita_loader=self.real_tier_rows,
                qogita_refresher=lambda _: (_ for _ in ()).throw(AssertionError()),
                now_provider=lambda: self.NOW, sleep_func=lambda _: None,
                pricing_batch_interval=0, fee_batch_interval=0,
            )
            resumed = run_discovery(
                state["filters"], checkpoint_store=store,
                catalog_batch=lambda *_: (_ for _ in ()).throw(
                    AssertionError("Catalog must not repeat")
                ),
                pricing_batch=lambda *_: (_ for _ in ()).throw(
                    AssertionError("Pricing must not repeat")
                ),
                fees_batch=lambda *_: (_ for _ in ()).throw(
                    AssertionError("Fees must not repeat")
                ),
                token_provider=RefreshingTokenProvider(lambda: "token"),
                qogita_loader=lambda: (_ for _ in ()).throw(
                    AssertionError("Qogita must not reload")
                ),
                job_id=state["job_id"], sleep_func=lambda _: None,
                pricing_batch_interval=0, fee_batch_interval=0,
            )
        self.assertEqual(resumed["status"], "completed")
        self.assertEqual(len(resumed["results"][0]["scenarios"]), 5)
        self.assertEqual(len(resumed["amazon_observations"]), 1)

    def test_scenario_roles_are_deterministic(self):
        scenarios = [
            {"scenario_id": "mov500", "account_mov": 500, "cost_gross_unit_eur": 12,
             "economics_status": "ready", "margin_percent": 14, "score": 60},
            {"scenario_id": "mov1500", "account_mov": 1500, "cost_gross_unit_eur": 11,
             "economics_status": "ready", "margin_percent": 15.2, "score": 74},
            {"scenario_id": "mov5000", "account_mov": 5000, "cost_gross_unit_eur": 9,
             "economics_status": "ready", "margin_percent": 25, "score": 80},
        ]
        roles = assign_scenario_roles(scenarios, 15)
        self.assertEqual(roles["scenario_base"], "mov500")
        self.assertEqual(roles["scenario_minimo_redditizio"], "mov1500")
        self.assertEqual(roles["scenario_migliore"], "mov5000")
        self.assertEqual(roles["scenario_raccomandato"], "mov5000")
        self.assertIn("Base", scenarios[0]["roles"])

    def test_base_below_margin_but_later_tier_passes_product(self):
        rows = [
            qogita_row(price="18", mov="500", observed_at="2026-08-20T10:00:00Z"),
            qogita_row(price="8", mov="1500", observed_at="2026-08-20T10:00:00Z"),
        ]
        state, _, _ = self.run_rows(rows)
        self.assertEqual(len(state["results"]), 1)
        product = state["results"][0]
        by_mov = {int(row["account_mov"]): row for row in product["scenarios"]}
        self.assertLess(by_mov[500]["margin_percent"], 15)
        self.assertGreaterEqual(by_mov[1500]["margin_percent"], 15)
        self.assertNotEqual(by_mov[500]["score"], by_mov[1500]["score"])
        self.assertEqual(
            product["scenario_roles"]["scenario_minimo_redditizio"],
            by_mov[1500]["scenario_id"],
        )
        self.assertEqual(state["funnel"]["scenarios_margin_passed"], 1)
        self.assertEqual(state["funnel"]["scenarios_margin_below_threshold"], 1)

    def test_all_scenarios_below_margin_are_persisted_but_not_final(self):
        rows = [
            qogita_row(price="18", mov="500", observed_at="2026-08-20T10:00:00Z"),
            qogita_row(price="17", mov="1500", observed_at="2026-08-20T10:00:00Z"),
        ]
        state, persisted, calls = self.run_rows(rows)
        self.assertEqual(state["results"], [])
        product = state["candidates"][0]
        self.assertEqual(product["evaluation_status"], "margin_below_threshold")
        self.assertEqual(product["exclusion_reason"], "margin_below_threshold")
        self.assertTrue(all("score" in row and "opportunity" in row for row in product["scenarios"]))
        self.assertTrue(all(
            set((row.get("economics") or {}).get("target_prices") or {}) == {"15", "20", "25"}
            for row in product["scenarios"]
        ))
        self.assertEqual(len(persisted["candidates"][0]["scenarios"]), 2)
        self.assertEqual(len(calls["fees"]), 1)

    def test_target_prices_are_canonical_at_runtime_and_after_json_checkpoint(self):
        state, persisted, _ = self.run_rows(self.real_tier_rows())
        runtime = state["results"][0]["scenarios"][0]["economics"]
        reloaded = persisted["results"][0]["scenarios"][0]["economics"]
        self.assertEqual(set(runtime["target_prices"]), {"15", "20", "25"})
        self.assertEqual(set(reloaded["target_prices"]), {"15", "20", "25"})
        self.assertEqual(target_price(runtime, 15), target_price(reloaded, "15"))
        legacy = {"target_prices": {15: 1, 20: 2, 25: 3}}
        canonicalize_target_prices(legacy)
        self.assertEqual(legacy["target_prices"], {"15": 1, "20": 2, "25": 3})

    def test_amazon_observation_separates_fba_and_referral_sources(self):
        state, _, _ = self.run_rows(self.real_tier_rows())
        observation = state["amazon_observations"][0]
        self.assertEqual(observation["fba_source"], "FBAFees")
        self.assertEqual(observation["referral_source"], "amazon_referral_fee")

        fallback_state, _, _ = self.run_rows(
            self.real_tier_rows(),
            fees_batch=lambda rows, token: fee_batch(rows, token, referral=None),
        )
        fallback = fallback_state["amazon_observations"][0]
        self.assertEqual(fallback["fba_source"], "FBAFees")
        self.assertEqual(fallback["referral_source"], "fallback_19_percent")
        self.assertEqual(Decimal(str(fallback["referral_rate"])), Decimal("0.19"))

    def test_legacy_observation_source_is_migrated_without_api_calls(self):
        with tempfile.TemporaryDirectory() as directory:
            store = DiscoveryCheckpointStore(directory)
            state = run_discovery(
                default_filters(), checkpoint_store=store,
                catalog_batch=catalog_mapping, pricing_batch=pricing_mapping,
                fees_batch=fee_batch,
                token_provider=RefreshingTokenProvider(lambda: "token"),
                qogita_loader=self.real_tier_rows,
                qogita_refresher=lambda _: (_ for _ in ()).throw(AssertionError()),
                now_provider=lambda: self.NOW, sleep_func=lambda _: None,
                pricing_batch_interval=0, fee_batch_interval=0,
            )
            observation = state["amazon_observations"][0]
            observation.pop("fba_source", None)
            observation["referral_source"] = "FBAFees"
            store.save(state)
            resumed = run_discovery(
                state["filters"], checkpoint_store=store,
                catalog_batch=lambda *_: (_ for _ in ()).throw(AssertionError()),
                pricing_batch=lambda *_: (_ for _ in ()).throw(AssertionError()),
                fees_batch=lambda *_: (_ for _ in ()).throw(AssertionError()),
                token_provider=RefreshingTokenProvider(lambda: "token"),
                qogita_loader=lambda: (_ for _ in ()).throw(AssertionError()),
                job_id=state["job_id"], sleep_func=lambda _: None,
                pricing_batch_interval=0, fee_batch_interval=0,
            )
        migrated = resumed["amazon_observations"][0]
        self.assertEqual(migrated["fba_source"], "FBAFees")
        self.assertEqual(migrated["referral_source"], "amazon_referral_fee")

    def test_competition_exclusions_are_persisted_without_fees_and_not_resumed(self):
        cases = (
            (5, 7, ["fba_sellers_above_threshold"]),
            (2, 9, ["total_sellers_above_threshold"]),
            (8, 15, ["fba_sellers_above_threshold", "total_sellers_above_threshold"]),
        )
        for fba, total, reasons in cases:
            with self.subTest(fba=fba, total=total), tempfile.TemporaryDirectory() as directory:
                fee_calls = []
                store = DiscoveryCheckpointStore(directory)

                def pricing(asins, _job):
                    return {asin: {
                        "status": "success", "Venditori FBA": fba,
                        "Venditori totali": total,
                        "Seller count source": "summary_number_of_offers",
                        "reference_price": 30, "price_source": "buy_box",
                    } for asin in asins}

                state = run_discovery(
                    default_filters(), checkpoint_store=store,
                    catalog_batch=catalog_mapping, pricing_batch=pricing,
                    fees_batch=lambda rows, token: fee_calls.append(rows),
                    token_provider=RefreshingTokenProvider(lambda: "token"),
                    qogita_loader=self.real_tier_rows,
                    qogita_refresher=lambda _: (_ for _ in ()).throw(AssertionError()),
                    now_provider=lambda: self.NOW, sleep_func=lambda _: None,
                    pricing_batch_interval=0, fee_batch_interval=0,
                )
                product = state["candidates"][0]
                self.assertEqual(state["results"], [])
                self.assertEqual(fee_calls, [])
                self.assertEqual(product["evaluation_status"], "competition_filtered")
                self.assertEqual(product["exclusion_reasons"], reasons)
                self.assertEqual(len(product["scenarios"]), 5)
                self.assertTrue(all("economics" not in row for row in product["scenarios"]))
                self.assertEqual(state["amazon_observations"], [])
                self.assertEqual(state["funnel"]["competition_filtered_products"], 1)
                self.assertEqual(
                    state["funnel"]["fba_threshold_excluded"],
                    int("fba_sellers_above_threshold" in reasons),
                )
                self.assertEqual(
                    state["funnel"]["total_sellers_threshold_excluded"],
                    int("total_sellers_above_threshold" in reasons),
                )
                resumed = run_discovery(
                    state["filters"], checkpoint_store=store,
                    catalog_batch=lambda *_: (_ for _ in ()).throw(AssertionError()),
                    pricing_batch=lambda *_: (_ for _ in ()).throw(AssertionError()),
                    fees_batch=lambda *_: (_ for _ in ()).throw(AssertionError()),
                    token_provider=RefreshingTokenProvider(lambda: "token"),
                    qogita_loader=lambda: (_ for _ in ()).throw(AssertionError()),
                    job_id=state["job_id"], sleep_func=lambda _: None,
                    pricing_batch_interval=0, fee_batch_interval=0,
                )
                self.assertEqual(
                    resumed["candidates"][0]["evaluation_status"],
                    "competition_filtered",
                )

    def test_pipeline_tracks_catalog_invalid_identifier_in_funnel(self):
        rows = [qogita_row(), qogita_row(gtin="12345678")]

        def catalog(gtins, _job_id):
            return {
                gtin: (
                    {"status": "invalid_identifier", "identifier_type": None}
                    if gtin == "12345678"
                    else catalog_mapping([gtin], _job_id)[gtin]
                )
                for gtin in gtins
            }

        with tempfile.TemporaryDirectory() as directory:
            state = self.run_pipeline(
                directory, qogita_loader=lambda: rows, catalog_batch=catalog,
            )
            self.assertEqual(state["status"], "completed")
            self.assertEqual(state["funnel"]["catalog_invalid_identifier"], 1)
            self.assertEqual(state["funnel"]["amazon_found"], 1)

    def test_catalog_incomplete_is_persisted_without_false_not_found(self):
        diagnostics = {
            "page_count": 1, "number_of_results": 20,
            "had_next_token": True, "items_received": 10,
            "input_identifier_count": 20, "complete": False,
            "error": "AmazonBatchError",
        }

        def catalog(gtins, _job_id):
            return {
                gtin: {
                    "status": "catalog_incomplete", "identifier_type": "EAN",
                    "listings": [], "diagnostics": diagnostics,
                }
                for gtin in gtins
            }

        with tempfile.TemporaryDirectory() as directory:
            store = DiscoveryCheckpointStore(directory)
            state = self.run_pipeline(
                directory, checkpoint_store=store, catalog_batch=catalog,
            )
            persisted = store.load(state["job_id"])
        candidate = persisted["candidates"][0]
        self.assertEqual(candidate["catalog_status"], "catalog_incomplete")
        self.assertNotEqual(candidate["catalog_status"], "not_found")
        self.assertEqual(candidate["catalog_diagnostics"]["page_count"], 1)
        self.assertFalse(candidate["catalog_diagnostics"]["complete"])
        self.assertEqual(state["funnel"]["amazon_found"], 0)

    def test_bsr_and_competition_filters_run_before_fees(self):
        fees_seen = []

        def catalog(gtins, _job):
            mapping = catalog_mapping(gtins, _job)
            mapping[gtins[0]]["bsr_beauty"] = 50000
            return mapping

        def fees(requests_, token):
            fees_seen.extend(requests_)
            return fee_batch(requests_, token)

        with tempfile.TemporaryDirectory() as directory:
            state = self.run_pipeline(directory, catalog_batch=catalog, fees_batch=fees)
            self.assertEqual(state["funnel"]["bsr_in_range"], 0)
            self.assertEqual(fees_seen, [])

    def test_fba_and_total_seller_limits_are_both_enforced(self):
        def pricing(asins, _job):
            return {
                asin: {
                    "status": "success", "Venditori FBA": 5,
                    "Venditori totali": 7, "reference_price": 30,
                    "price_source": "buy_box",
                }
                for asin in asins
            }
        with tempfile.TemporaryDirectory() as directory:
            state = self.run_pipeline(directory, pricing_batch=pricing)
            self.assertEqual(state["funnel"]["competition_passed"], 0)

        def pricing_total(asins, _job):
            return {
                asin: {
                    "status": "success", "Venditori FBA": 2,
                    "Venditori totali": 9, "reference_price": 30,
                    "price_source": "buy_box",
                }
                for asin in asins
            }
        with tempfile.TemporaryDirectory() as directory:
            state = self.run_pipeline(directory, pricing_batch=pricing_total)
            self.assertEqual(state["funnel"]["competition_passed"], 0)

    def test_referral_fallback_and_margin_filter(self):
        with tempfile.TemporaryDirectory() as directory:
            state = self.run_pipeline(
                directory,
                fees_batch=lambda rows, token: fee_batch(rows, token, referral=None),
            )
            self.assertEqual(state["results"][0]["economics"]["referral_source"], "fallback_19_percent")

    def test_partial_fee_batch_retries_only_transient_element(self):
        rows = [
            qogita_row(gtin="8809080826140"),
            qogita_row(gtin="8809532220700"),
        ]
        calls = []

        def fees(requests_, _token):
            calls.append([row["identifier"] for row in requests_])
            if len(calls) == 1:
                return [
                    fee_result(requests_[0]["asin"], requests_[0]["identifier"]),
                    fee_error_result(requests_[1]["asin"], requests_[1]["identifier"]),
                ]
            return [
                fee_result(requests_[0]["asin"], requests_[0]["identifier"])
            ]

        with tempfile.TemporaryDirectory() as directory:
            state = self.run_pipeline(
                directory, qogita_loader=lambda: rows, fees_batch=fees,
                catalog_batch=unique_catalog_mapping,
            )
            self.assertEqual(state["status"], "completed")
            self.assertEqual([len(call) for call in calls], [2, 1])
            self.assertNotIn(calls[0][0], calls[1])
            self.assertEqual(state["funnel"]["fee_valid"], 2)
            self.assertEqual(state["funnel"]["fee_pending"], 0)
            self.assertEqual(state["funnel"]["fee_invalid"], 0)
            self.assertEqual(len(state["results"]), 2)

    def test_all_transient_batch_isolates_six_listings_and_preserves_successes(self):
        rows = [
            qogita_row(gtin=gtin)
            for gtin in (
                "8809532220953", "8809532221349", "8809532221356",
                "8809532221783", "8809532221790", "8809532221943",
            )
        ]
        calls = []

        def fees(requests_, _token):
            calls.append([row["asin"] for row in requests_])
            if len(calls) == 1:
                return [
                    fee_error_result(row["asin"], row["identifier"])
                    for row in requests_
                ]
            self.assertEqual(len(requests_), 1)
            row = requests_[0]
            return [fee_result(row["asin"], row["identifier"])]

        with tempfile.TemporaryDirectory() as directory:
            state = self.run_pipeline(
                directory, qogita_loader=lambda: rows, fees_batch=fees,
                catalog_batch=unique_catalog_mapping,
            )

        self.assertEqual([len(call) for call in calls], [6, 1, 1, 1, 1, 1, 1])
        self.assertEqual(len({call[0] for call in calls[1:]}), 6)
        self.assertEqual(state["status"], "completed")
        self.assertEqual(state["funnel"]["competition_passed_listings"], 6)
        self.assertEqual(state["funnel"]["fee_valid_listings"], 6)
        self.assertEqual(state["funnel"]["fee_pending"], 0)
        self.assertEqual(len(state["amazon_observations"]), 6)
        self.assertEqual(len(state["opportunity_combinations"]), 6)
        self.assertTrue(all(
            row["fee_status"] == "valid" and row["fee_attempts"] == 2
            for row in state["amazon_observations"]
        ))

    def test_persistent_internal_error_becomes_unavailable_without_stopping_job(self):
        calls = []
        sleeps = []

        def fees(requests_, _token):
            calls.append([row["identifier"] for row in requests_])
            return ProductFeeBatchResults([
                fee_error_result(row["asin"], row["identifier"])
                for row in requests_
            ], retry_after="0", rate_limit="100")

        with tempfile.TemporaryDirectory() as directory:
            store = DiscoveryCheckpointStore(directory)
            state = self.run_pipeline(
                directory, checkpoint_store=store, fees_batch=fees,
                sleep_func=sleeps.append,
            )
            self.assertEqual([len(call) for call in calls], [1, 1, 1])
            self.assertEqual(len(sleeps), 2)
            self.assertTrue(2 <= sleeps[0] <= 2.2)
            self.assertTrue(4 <= sleeps[1] <= 4.4)
            self.assertEqual(state["status"], "completed")
            self.assertEqual(state["phase"], "completed")
            self.assertEqual(state["funnel"]["fee_valid"], 0)
            self.assertEqual(state["funnel"]["fee_pending"], 0)
            self.assertEqual(state["funnel"]["fee_unavailable"], 1)
            self.assertEqual(state["funnel"]["fee_invalid"], 0)
            self.assertEqual(state["candidates"][0]["fee_status"], "unavailable")
            self.assertEqual(
                state["candidates"][0]["fee_unavailable_reason"],
                "amazon_internal_error",
            )
            self.assertEqual(state["candidates"][0]["fee_attempts"], 3)
            self.assertIsNone(store.latest_incomplete())

    def test_ten_fee_targets_complete_with_one_unavailable(self):
        rows = [qogita_row(gtin=test_ean(index)) for index in range(10)]
        failed_asin = "B000000001"

        def fees(requests_, _token):
            return [
                fee_error_result(row["asin"], row["identifier"])
                if row["asin"] == failed_asin
                else fee_result(row["asin"], row["identifier"])
                for row in requests_
            ]

        with tempfile.TemporaryDirectory() as directory:
            state = self.run_pipeline(
                directory, qogita_loader=lambda: rows, fees_batch=fees,
                catalog_batch=unique_catalog_mapping,
            )
        self.assertEqual(state["status"], "completed")
        self.assertEqual(state["fee_target_count"], 10)
        self.assertEqual(state["fee_valid_count"], 9)
        self.assertEqual(state["fee_unavailable_count"], 1)
        self.assertTrue(state["fee_coverage_partial"])
        self.assertEqual(len(state["results"]), 9)
        unavailable = next(
            row for row in state["amazon_observations"]
            if row["fee_status"] == "unavailable"
        )
        self.assertNotIn("fee_estimate", unavailable)

    def test_hundred_fee_targets_complete_with_five_unavailable(self):
        rows = [qogita_row(gtin=test_ean(index)) for index in range(100)]
        failed_asins = {f"B{index:09d}" for index in range(1, 6)}

        def fees(requests_, _token):
            return [
                fee_error_result(row["asin"], row["identifier"])
                if row["asin"] in failed_asins
                else fee_result(row["asin"], row["identifier"])
                for row in requests_
            ]

        with tempfile.TemporaryDirectory() as directory:
            state = self.run_pipeline(
                directory, qogita_loader=lambda: rows, fees_batch=fees,
                catalog_batch=unique_catalog_mapping,
            )
        self.assertEqual(state["status"], "completed")
        self.assertEqual(state["fee_target_count"], 100)
        self.assertEqual(state["fee_valid_count"], 95)
        self.assertEqual(state["fee_unavailable_count"], 5)
        self.assertEqual(len(state["results"]), 95)

    def test_request_level_fee_outage_pauses_whole_job(self):
        def outage(*_):
            raise requests.ConnectionError("offline diagnostic detail")

        with tempfile.TemporaryDirectory() as directory:
            state = self.run_pipeline(directory, fees_batch=outage)
            self.assertEqual(state["status"], "waiting_retry")
            self.assertEqual(state["phase"], "fees_pending")
            self.assertTrue(state["resumable"])
            self.assertEqual(state["fee_outage_reason"], "ConnectionError")
            resumed = run_discovery(
                state["filters"], checkpoint_store=DiscoveryCheckpointStore(directory),
                catalog_batch=lambda *_: self.fail("Catalog must not repeat"),
                pricing_batch=lambda *_: self.fail("Pricing must not repeat"),
                fees_batch=fee_batch,
                token_provider=RefreshingTokenProvider(lambda: "token"),
                qogita_loader=lambda: self.fail("Supplier must not reload"),
                job_id=state["job_id"], sleep_func=lambda *_: None,
                pricing_batch_interval=0, fee_batch_interval=0,
            )
        self.assertEqual(resumed["status"], "completed")
        self.assertEqual(resumed["fee_valid_count"], 1)

    def test_high_element_failure_rate_opens_systemic_circuit(self):
        rows = [qogita_row(gtin=test_ean(index)) for index in range(10)]

        def fees(requests_, _token):
            return [
                fee_error_result(row["asin"], row["identifier"])
                for row in requests_
            ]

        with tempfile.TemporaryDirectory() as directory:
            state = self.run_pipeline(
                directory, qogita_loader=lambda: rows,
                catalog_batch=unique_catalog_mapping, fees_batch=fees,
            )
        self.assertEqual(state["status"], "waiting_retry")
        self.assertEqual(state["phase"], "fees_pending")
        self.assertEqual(state["fee_outage_reason"], "ElementFailureRate")
        self.assertTrue(all(
            row["fee_status"] == "retryable_error"
            for row in state["amazon_observations"]
        ))

    def test_resume_retries_only_pending_fees_without_catalog_or_pricing(self):
        def persistent(requests_, _token):
            return [
                fee_error_result(row["asin"], row["identifier"])
                for row in requests_
            ]

        with tempfile.TemporaryDirectory() as directory:
            store = DiscoveryCheckpointStore(directory)
            pending = self.run_pipeline(
                directory, checkpoint_store=store, fees_batch=persistent,
            )
            self.assertEqual(pending["candidates"][0]["fee_status"], "unavailable")
            fee_calls = []

            def succeeds(requests_, _token):
                fee_calls.append([row["asin"] for row in requests_])
                return [
                    fee_result(row["asin"], row["identifier"])
                    for row in requests_
                ]

            resumed = run_discovery(
                pending["filters"], checkpoint_store=store,
                catalog_batch=lambda *_: (_ for _ in ()).throw(
                    AssertionError("catalog must not repeat")
                ),
                pricing_batch=lambda *_: (_ for _ in ()).throw(
                    AssertionError("pricing must not repeat")
                ),
                fees_batch=succeeds,
                token_provider=RefreshingTokenProvider(lambda: "token"),
                qogita_loader=lambda: (_ for _ in ()).throw(
                    AssertionError("qogita must not reload")
                ),
                job_id=pending["job_id"], sleep_func=lambda _: None,
                pricing_batch_interval=0, fee_batch_interval=0,
            )
            self.assertEqual(fee_calls, [[pending["candidates"][0]["asin"]]])
            self.assertEqual(resumed["status"], "completed")
            self.assertEqual(resumed["phase"], "completed")
            self.assertEqual(resumed["funnel"]["fee_valid"], 1)
            self.assertEqual(resumed["funnel"]["fee_pending"], 0)
            self.assertEqual(len(resumed["results"]), 1)

    def test_legacy_completed_internal_error_checkpoint_can_resume_at_fees(self):
        with tempfile.TemporaryDirectory() as directory:
            store = DiscoveryCheckpointStore(directory)
            state = self.run_pipeline(directory)
            row = state["candidates"][0]
            row.pop("fee_estimate", None)
            row.pop("economics", None)
            row["fee_status"] = "invalid"
            row["fee_error"] = "There is an internal service failure."
            for key in ("fee_error_status", "fee_error_code", "fee_error_type"):
                row.pop(key, None)
            state["status"] = "completed"
            state["phase"] = "completed"
            state["results"] = []
            store.save(state)

            resumed = run_discovery(
                state["filters"], checkpoint_store=store,
                catalog_batch=lambda *_: (_ for _ in ()).throw(
                    AssertionError("catalog must not repeat")
                ),
                pricing_batch=lambda *_: (_ for _ in ()).throw(
                    AssertionError("pricing must not repeat")
                ),
                fees_batch=fee_batch,
                token_provider=RefreshingTokenProvider(lambda: "token"),
                qogita_loader=lambda: (_ for _ in ()).throw(
                    AssertionError("qogita must not reload")
                ),
                job_id=state["job_id"], sleep_func=lambda _: None,
                pricing_batch_interval=0, fee_batch_interval=0,
            )
            self.assertEqual(resumed["status"], "completed")
            self.assertEqual(resumed["funnel"]["fee_valid"], 1)
            self.assertEqual(len(resumed["results"]), 1)

    def test_client_fee_error_is_invalid_without_server_side_retry(self):
        calls = []

        def fees(requests_, _token):
            calls.append(len(requests_))
            return [fee_error_result(
                row["asin"], row["identifier"], status="ClientError",
                code="InvalidInput", error_type="Sender",
                message="Invalid marketplace input",
            ) for row in requests_]

        with tempfile.TemporaryDirectory() as directory:
            state = self.run_pipeline(directory, fees_batch=fees)
            self.assertEqual(calls, [1])
            self.assertEqual(state["status"], "completed")
            self.assertEqual(state["funnel"]["fee_pending"], 0)
            self.assertEqual(state["funnel"]["fee_invalid"], 1)
            self.assertEqual(state["candidates"][0]["fee_status"], "invalid")

    def test_checkpoint_atomic_and_resume_skips_completed_catalog(self):
        calls = {"catalog": 0, "pricing": 0}

        def catalog(gtins, job):
            calls["catalog"] += 1
            return catalog_mapping(gtins, job)

        def fail_pricing(_asins, _job):
            calls["pricing"] += 1
            raise RuntimeError("temporary")

        with tempfile.TemporaryDirectory() as directory:
            store = DiscoveryCheckpointStore(directory)
            with self.assertRaises(RuntimeError):
                self.run_pipeline(directory, catalog_batch=catalog, pricing_batch=fail_pricing)
            state = store.latest_incomplete()
            self.assertEqual(state["phase"], "bsr_filtered")
            self.assertEqual(list(Path(directory).glob(".*.json")), [])
            resumed = run_discovery(
                state["filters"], checkpoint_store=store,
                catalog_batch=catalog, pricing_batch=pricing_mapping,
                fees_batch=fee_batch,
                token_provider=RefreshingTokenProvider(lambda: "token"),
                qogita_loader=lambda: (_ for _ in ()).throw(AssertionError("must not reload")),
                job_id=state["job_id"], sleep_func=lambda _: None,
                pricing_batch_interval=0, fee_batch_interval=0,
            )
            self.assertEqual(resumed["status"], "completed")
            self.assertEqual(calls["catalog"], 1)


class DiscoveryDiagnosticFunnelTests(unittest.TestCase):
    @staticmethod
    def listing(status="competition_passed", *, reason=None, compatible=True):
        row = {
            "asin": "B000000001",
            "compatibility_status": "compatible" if compatible else "incompatible",
            "beauty_status": "display_group_beauty",
            "bsr_beauty": 5000,
            "evaluation_status": status,
        }
        if reason:
            row["exclusion_reason"] = reason
            row["exclusion_reasons"] = [reason]
        return row

    @staticmethod
    def product(key, listings, catalog_status="resolved"):
        return {
            "product_key": key,
            "catalog_status": catalog_status,
            "amazon_listings": listings,
        }

    def funnel(self, *products):
        state = {"candidates": list(products), "funnel": {}}
        return recalculate_diagnostic_funnel(state)

    def test_one_product_one_listing_counts_one_found_product(self):
        funnel = self.funnel(self.product("one", [self.listing()]))
        self.assertEqual(funnel["amazon_found"], 1)
        self.assertEqual(funnel["amazon_listings_found"], 1)

    def test_one_product_two_listings_does_not_double_count_found_product(self):
        funnel = self.funnel(self.product("one", [self.listing(), self.listing()]))
        self.assertEqual(funnel["amazon_found"], 1)
        self.assertEqual(funnel["amazon_listings_found"], 2)

    def test_ambiguous_product_with_persisted_listings_is_found(self):
        product = self.product("ambiguous", [self.listing(), self.listing()], "ambiguous")
        self.assertEqual(self.funnel(product)["amazon_found"], 1)

    def test_catalog_not_found_is_not_amazon_found(self):
        product = self.product("missing", [], "not_found")
        self.assertEqual(self.funnel(product)["amazon_found"], 0)

    def test_incompatible_listing_still_counts_as_found_product(self):
        product = self.product(
            "incompatible", [self.listing("catalog_incompatible", compatible=False)]
        )
        funnel = self.funnel(product)
        self.assertEqual(funnel["amazon_found"], 1)
        self.assertEqual(funnel["compatible_listings"], 0)

    def test_bsr_filtered_product_is_not_competition_filtered(self):
        product = self.product("bsr", [self.listing("bsr_filtered")])
        self.assertEqual(self.funnel(product)["competition_filtered_products"], 0)

    def test_non_beauty_product_is_not_competition_filtered(self):
        listing = self.listing("beauty_filtered")
        listing["beauty_status"] = "beauty_rank_unverified"
        product = self.product("beauty", [listing])
        self.assertEqual(self.funnel(product)["competition_filtered_products"], 0)

    def test_competition_passed_listing_does_not_filter_product(self):
        product = self.product("passed", [self.listing("competition_passed")])
        funnel = self.funnel(product)
        self.assertEqual(funnel["competition_passed_listings"], 1)
        self.assertEqual(funnel["competition_filtered_products"], 0)

    def test_fba_filtered_listing_counts_listing_and_product(self):
        product = self.product("fba", [self.listing(
            "competition_filtered", reason="fba_sellers_above_threshold",
        )])
        funnel = self.funnel(product)
        self.assertEqual(funnel["competition_filtered_products"], 1)
        self.assertEqual(funnel["fba_threshold_excluded"], 1)

    def test_one_listing_passes_one_fails_product_is_not_filtered(self):
        product = self.product("mixed", [
            self.listing("competition_passed"),
            self.listing(
                "competition_filtered", reason="fba_sellers_above_threshold",
            ),
        ])
        funnel = self.funnel(product)
        self.assertEqual(funnel["competition_filtered_products"], 0)
        self.assertEqual(funnel["fba_threshold_excluded"], 1)

    def test_two_listings_both_fail_count_one_filtered_product(self):
        product = self.product("failed", [
            self.listing(
                "competition_filtered", reason="fba_sellers_above_threshold",
            ),
            self.listing(
                "competition_filtered", reason="total_sellers_above_threshold",
            ),
        ])
        funnel = self.funnel(product)
        self.assertEqual(funnel["competition_filtered_products"], 1)
        self.assertEqual(funnel["fba_threshold_excluded"], 1)
        self.assertEqual(funnel["total_sellers_threshold_excluded"], 1)

    def test_listing_can_increment_both_competition_exclusion_counters(self):
        listing = self.listing("competition_filtered")
        listing["exclusion_reasons"] = [
            "fba_sellers_above_threshold", "total_sellers_above_threshold",
        ]
        listing["exclusion_reason"] = ",".join(listing["exclusion_reasons"])
        funnel = self.funnel(self.product("both", [listing]))
        self.assertEqual(funnel["fba_threshold_excluded"], 1)
        self.assertEqual(funnel["total_sellers_threshold_excluded"], 1)

    def test_large_multilisting_funnel_fixture_has_product_semantics(self):
        products = []
        for index in range(49):
            status = "competition_passed" if index < 17 else "bsr_filtered"
            listings = [self.listing(status)]
            if index == 0:
                listings.append(self.listing("bsr_filtered"))
            products.append(self.product(f"product-{index}", listings))
        state = {"candidates": products, "funnel": {"final_products": 14}}
        funnel = recalculate_diagnostic_funnel(state)
        self.assertEqual(funnel["amazon_found"], 49)
        self.assertEqual(funnel["amazon_listings_found"], 50)
        self.assertEqual(funnel["competition_passed_listings"], 17)
        self.assertEqual(funnel["competition_filtered_products"], 0)
        self.assertEqual(funnel["final_products"], 14)


class DiscoveryExcelTests(unittest.TestCase):
    def _result(self):
        rows = [
            qogita_row(price=price, mov=mov, observed_at="2026-08-20T10:00:00Z")
            for mov, price in ((500, "6.94"), (1500, "6.83"), (5000, "6.80"),
                               (10000, "6.73"), (15000, "6.73"))
        ]
        product = normalize_qogita_candidates(
            rows, now=datetime(2026, 8, 20, 12, 0, tzinfo=timezone.utc)
        )[0][0]
        for index, scenario in enumerate(product["scenarios"]):
            scenario.update({
                "economics_status": "ready", "margin_percent": 20 + index,
                "score": 80 + index, "opportunity": "🟢 Ottima",
                "roles": ["Raccomandato"] if index == 4 else [],
            })
        product.update({
            "scenario_roles": {
                "scenario_base": product["scenarios"][0]["scenario_id"],
                "scenario_minimo_redditizio": product["scenarios"][0]["scenario_id"],
                "scenario_migliore": product["scenarios"][-1]["scenario_id"],
                "scenario_raccomandato": product["scenarios"][-1]["scenario_id"],
            },
            "asin": "B000000001",
            "amazon_offers_url": "https://www.amazon.it/gp/offer-listing/B000000001",
            "amazon_observation": {
                "asin": "B000000001", "amazon_brand": "Brand", "amazon_title": "Crema",
                "bsr_beauty": 5000, "reference_price": Decimal("30"),
                "fba_sellers": 2, "total_sellers": 4,
                "price_source": "buy_box", "seller_count_source": "summary_number_of_offers",
                "observed_at": "2026-08-20T12:00:00Z",
                "fee_estimate": {
                    "fba_fee_net": Decimal("4"), "fba_fee_gross": Decimal("4.88"),
                    "referral_fee": Decimal("4.5"), "referral_rate": Decimal("0.15"),
                    "source": "amazon_referral_fee",
                },
            },
        })
        return product

    def test_export_multiscenario_formulas_hidden_ids_and_sort_safe_lookup(self):
        result = self._result()
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "discovery.xlsx"
            write_discovery_excel([result], path)
            workbook = load_workbook(path, data_only=False)
            self.assertEqual(workbook.sheetnames, [
                "Opportunità", "Tutti i risultati", "Listing Amazon",
                "Scenari", "Dati", "Parametri run",
            ])
            opportunity = workbook["Opportunità"]
            scenarios = workbook["Scenari"]
            self.assertEqual(
                [opportunity.cell(1, col).value for col in range(1, len(OPPORTUNITY_COLUMNS) + 1)],
                OPPORTUNITY_COLUMNS,
            )
            self.assertEqual(
                [scenarios.cell(1, col).value for col in range(1, len(SCENARIO_COLUMNS) + 1)],
                SCENARIO_COLUMNS,
            )
            self.assertEqual(scenarios.max_row, 6)
            for column in ("W", "X", "Y", "Z"):
                self.assertTrue(opportunity.column_dimensions[column].hidden)
            for column in ("AA", "AB", "AC", "AD"):
                self.assertTrue(scenarios.column_dimensions[column].hidden)
            self.assertTrue(scenarios["AA2"].protection.locked)
            self.assertFalse(scenarios["I2"].protection.locked)
            self.assertIn("MATCH($AC2,'Dati'!$A$2:$A$2,0)", scenarios["N2"].value)
            self.assertIn("O2", scenarios["N2"].value)
            self.assertIn("I2", scenarios["O2"].value)
            for column in ("N", "O", "P", "Q", "R", "S", "T"):
                self.assertTrue(str(scenarios[f"{column}2"].value).startswith("="))
            self.assertEqual(workbook["Dati"].sheet_state, "hidden")
            self.assertEqual(opportunity["V2"].hyperlink.target, result["amazon_offers_url"])
            self.assertTrue(opportunity.protection.sheet)
            self.assertTrue(scenarios.protection.sheet)
            self.assertFalse(opportunity.protection.autoFilter)
            self.assertFalse(opportunity.protection.sort)
            self.assertFalse(opportunity.protection.formatColumns)

    def test_opportunities_are_sorted_and_currency_inputs_remain_numeric(self):
        template = self._result()
        specifications = [
            ("8800000000001", 72, 30, 5),
            ("8800000000002", 100, 20, 20),
            ("8800000000003", 85, 50, 10),
            ("8800000000004", 100, 30, 4),
            ("8800000000005", 100, 30, 9),
            ("8800000000006", 43, 99, 50),
        ]
        products = []
        for index, (ean, score, margin, profit) in enumerate(specifications):
            product = deepcopy(template)
            product["canonical_ean"] = product["gtin"] = ean
            product["product_key"] = f"product-{index}"
            observation = product["amazon_observation"]
            observation["observation_id"] = f"observation-{index}"
            observation["reference_price"] = "30.1234"
            scenario = product["scenarios"][-1]
            scenario["cost_gross_unit_eur"] = "9.028125535928"
            scenario.update({"score": score, "margin_percent": margin, "profit": profit})
            products.append(product)

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "sorted.xlsx"
            write_discovery_excel(products, path)
            workbook = load_workbook(path, data_only=False)
            sheet = workbook["Opportunità"]
            self.assertEqual(
                [sheet[f"A{row}"].value for row in range(2, 8)],
                ["8800000000005", "8800000000004", "8800000000002",
                 "8800000000003", "8800000000001", "8800000000006"],
            )
            self.assertEqual(sheet["G2"].data_type, "n")
            self.assertAlmostEqual(sheet["G2"].value, 9.028125535928)
            self.assertEqual(sheet["G2"].number_format, '€ #,##0.00')
            self.assertFalse(sheet["G2"].protection.locked)
            self.assertEqual(sheet["J2"].data_type, "n")
            self.assertTrue(sheet["N2"].value.startswith("=IFERROR"))
            self.assertIn("G2", sheet["N2"].value)
            self.assertIn("N2", sheet["M2"].value)
            self.assertTrue(sheet["N2"].protection.locked)
            self.assertEqual(sheet["R2"].number_format, "0")
            self.assertEqual(sheet.auto_filter.ref, "A1:V7")
            self.assertFalse(sheet.protection.autoFilter)
            self.assertFalse(sheet.protection.sort)
            self.assertFalse(sheet.protection.formatColumns)
            for row in range(2, sheet.max_row + 1):
                for column in range(1, len(OPPORTUNITY_COLUMNS) + 1):
                    self.assertEqual(
                        sheet.cell(row, column).protection.locked,
                        column != 7,
                    )
                    value = sheet.cell(row, column).value
                    if isinstance(value, str) and value.startswith("="):
                        self.assertNotIn("#REF!", value)
            workbook.close()

    def test_complete_audit_export_preserves_every_filter_outcome(self):
        filters = {
            "bsr_min": 0, "bsr_max": 20000,
            "max_fba_sellers": 4, "max_total_sellers": 8,
            "minimum_margin": 15,
        }
        products = []
        observations = []

        def add_product(index, kind, *, margin=None, fba=2, total=4, bsr=5000):
            ean = f"880000000{index:04d}"
            product_key = f"product-{index}"
            asin = f"B{index:09d}"
            scenario = {
                "scenario_id": f"scenario-{index}", "product_key": product_key,
                "supplier": "qogita", "supplier_alias": "seller",
                "scenario_label": "MOV € 500", "scenario_type": "qogita_mov",
                "account_mov": Decimal("500"), "cost_gross_unit_eur": Decimal("10"),
                "stock": 10,
            }
            listing = {
                "listing_id": f"listing-{index}", "amazon_observation_id": f"obs-{index}",
                "asin": asin, "title": f"Amazon {kind}", "brand": "Brand",
                "compatibility_status": "compatible", "compatibility_reason": ["ean_match"],
                "beauty_status": "display_group_beauty",
                "display_group": "beauty_display_on_website", "bsr_beauty": bsr,
                "catalog_status": "resolved", "reference_price": Decimal("30"),
                "price_source": "buy_box", "min_fba_price": Decimal("30"),
                "min_fbm_price": Decimal("31"), "fba_sellers": fba,
                "total_sellers": total, "pricing_status": "success",
                "competition_status": "passed", "evaluation_status": "competition_passed",
            }
            if kind == "competition_fba":
                listing.update({
                    "evaluation_status": "competition_filtered",
                    "competition_status": "filtered", "fba_sellers": 5,
                    "exclusion_reason": "fba_sellers_above_threshold",
                    "exclusion_reasons": ["fba_sellers_above_threshold"],
                })
            elif kind == "competition_total":
                listing.update({
                    "evaluation_status": "competition_filtered",
                    "competition_status": "filtered", "total_sellers": 9,
                    "exclusion_reason": "total_sellers_above_threshold",
                    "exclusion_reasons": ["total_sellers_above_threshold"],
                })
            elif kind == "bsr":
                listing.update({"evaluation_status": "bsr_filtered", "bsr_beauty": 20001,
                                "exclusion_reason": "bsr_out_of_range"})
            elif kind == "not_beauty":
                listing.update({"evaluation_status": "beauty_filtered",
                                "beauty_status": "beauty_rank_unverified",
                                "display_group": "health_and_beauty_display_on_website",
                                "exclusion_reason": "not_beauty_display_group"})
            elif kind == "incompatible":
                listing.update({"evaluation_status": "catalog_incompatible",
                                "compatibility_status": "incompatible",
                                "compatibility_reason": ["volume_mismatch"]})
            elif kind == "fee_pending":
                listing.update({"fee_status": "fee_pending", "fee_attempts": 3,
                                "fee_error": "InternalError"})
            elif kind == "fee_invalid":
                listing.update({"fee_status": "invalid", "fee_attempts": 1,
                                "fee_error": "InvalidInput"})
            elif kind == "economics":
                listing["fee_status"] = "valid"

            observation = {
                "observation_id": f"obs-{index}", "asin": asin,
                "canonical_ean": ean, "amazon_brand": "Brand",
                "amazon_title": f"Amazon {kind}", "bsr_beauty": listing["bsr_beauty"],
                "reference_price": Decimal("30"), "price_source": "buy_box",
                "min_fba_price": Decimal("30"), "min_fbm_price": Decimal("31"),
                "fba_sellers": listing["fba_sellers"], "total_sellers": listing["total_sellers"],
                "fee_status": listing.get("fee_status"), "fee_attempts": listing.get("fee_attempts"),
                "diagnostics": {"product_keys": [product_key]},
            }
            combinations = []
            if margin is not None:
                observation.update({
                    "fee_status": "valid", "fba_fee_net": Decimal("4"),
                    "fba_fee_gross": Decimal("4.88"), "referral_fee": Decimal("4.5"),
                    "referral_rate": Decimal("0.15"), "referral_source": "amazon_referral_fee",
                    "fee_estimate": {"fba_fee_net": Decimal("4"),
                                     "fba_fee_gross": Decimal("4.88"),
                                     "referral_fee": Decimal("4.5"),
                                     "referral_rate": Decimal("0.15")},
                })
                listing["fee_status"] = "valid"
                status = "margin_passed" if Decimal(str(margin)) >= 15 else "margin_below_threshold"
                combinations.append({
                    "combination_id": f"combination-{index}",
                    "scenario_id": scenario["scenario_id"], "asin": asin,
                    "amazon_observation_id": observation["observation_id"],
                    "cost_gross_unit_eur": Decimal("10"),
                    "price_reference": Decimal("30"), "profit": Decimal("4.497"),
                    "margin_percent": Decimal(str(margin)), "score": 60,
                    "opportunity": "🟡 Interessante", "evaluation_status": status,
                    "diagnostics": {"economics_status": "ready"},
                })
            product = {
                "product_key": product_key, "gtin": ean, "canonical_ean": ean,
                "brand": "Brand", "title": f"Supplier {kind}",
                "catalog_status": "resolved", "scenarios": [scenario],
                "amazon_listings": [listing], "amazon_observations": [observation],
                "opportunity_combinations": combinations,
                "combination_roles": {
                    "recommended_combination": combinations[0]["combination_id"]
                    if combinations else None,
                },
            }
            products.append(product)
            observations.append(observation)
            return product

        opportunity = add_product(1, "opportunity", margin=20)
        # Cross-supplier and multi-listing remain a single product.
        opportunity["scenarios"].append({
            **opportunity["scenarios"][0], "scenario_id": "scenario-1-umma",
            "supplier": "umma", "scenario_label": "U-Quick",
        })
        opportunity["amazon_listings"].append({
            **opportunity["amazon_listings"][0], "listing_id": "listing-1-alt",
            "asin": "B000000099", "amazon_observation_id": None,
            "evaluation_status": "bsr_filtered", "bsr_beauty": 50000,
            "exclusion_reason": "bsr_out_of_range", "fee_status": None,
        })
        add_product(2, "margin", margin=Decimal("14.99"))
        add_product(3, "competition_fba")
        add_product(4, "competition_total")
        add_product(5, "bsr")
        add_product(6, "not_beauty")
        add_product(7, "incompatible")
        add_product(8, "fee_pending")
        add_product(9, "fee_invalid")
        add_product(10, "economics")
        products.append({
            "product_key": "product-11", "gtin": "8800000000011",
            "brand": "Brand", "title": "Supplier not found",
            "catalog_status": "not_found", "scenarios": [{
                "scenario_id": "scenario-11", "supplier": "qudo",
                "scenario_label": "Qudo", "cost_gross_unit_eur": Decimal("9"),
            }], "amazon_listings": [], "opportunity_combinations": [],
        })
        state = {
            "job_id": "audit-job", "status": "completed", "phase": "completed",
            "created_at": "2026-08-24T10:00:00Z", "completed_at": "2026-08-24T10:05:00Z",
            "selected_suppliers": ["qogita", "umma", "abw", "qudo"],
            "filters": filters, "candidates": products, "results": [opportunity],
            "amazon_observations": observations,
        }

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "audit.xlsx"
            write_discovery_excel(state, path)
            workbook = load_workbook(path, data_only=False)

        self.assertEqual(workbook["Opportunità"].max_row, 2)
        all_results = workbook["Tutti i risultati"]
        listings = workbook["Listing Amazon"]
        scenarios = workbook["Scenari"]
        self.assertEqual(all_results.max_row, 13)
        self.assertEqual(listings.max_row, 12)
        self.assertEqual(scenarios.max_row, 3)
        status_values = {cell.value for cell in all_results["T"] if cell.row > 1}
        self.assertTrue({
            "Opportunità", "Sotto soglia margine", "Escluso per concorrenza",
            "Fuori range BSR", "Non Beauty", "Listing incompatibile",
            "Fee in attesa", "Fee non valida", "Non trovato su Amazon",
            "Economia non disponibile",
        }.issubset(status_values))
        reasons = [cell.value for cell in all_results["U"] if cell.row > 1]
        self.assertIn("Margine 14.99% < soglia 15.00%", reasons)
        self.assertIn("Venditori FBA 5 > limite 4", reasons)
        self.assertIn("Venditori totali 9 > limite 8", reasons)
        self.assertIn("BSR 20.001 > massimo 20.000", reasons)
        suppliers = [cell.value for cell in all_results["D"] if cell.row > 1]
        self.assertIn("QOGITA · UMMA", suppliers)
        self.assertEqual(workbook["Dati"].sheet_state, "hidden")
        self.assertEqual(workbook["Parametri run"]["B2"].value, "audit-job")
        metadata = {
            row[0].value: row[1].value
            for row in workbook["Parametri run"].iter_rows(min_row=2)
        }
        self.assertEqual(metadata["Prodotti trovati Amazon"], 10)
        self.assertEqual(metadata["Pagine Amazon trovate"], 11)
        self.assertEqual(metadata["Prodotti esclusi per concorrenza"], 2)
        self.assertGreaterEqual(
            metadata["Pagine Amazon con concorrenza valida"], 1,
        )

    def test_zero_opportunities_still_exports_diagnostic_workbook(self):
        product = self._result()
        product["evaluation_status"] = "margin_below_threshold"
        observation = product["amazon_observation"]
        product["amazon_observations"] = [observation]
        product["amazon_listings"] = [{
            "asin": observation["asin"], "title": observation["amazon_title"],
            "brand": observation["amazon_brand"], "compatibility_status": "compatible",
            "beauty_status": "display_group_beauty", "bsr_beauty": 5000,
            "catalog_status": "resolved", "evaluation_status": "competition_passed",
            "amazon_observation_id": product["product_key"], "fee_status": "valid",
        }]
        state = {
            "job_id": "zero-job", "status": "completed", "phase": "completed",
            "filters": {"minimum_margin": 99}, "candidates": [product], "results": [],
        }
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "zero.xlsx"
            write_discovery_excel(state, path)
            workbook = load_workbook(path, data_only=False)
        self.assertEqual(workbook["Opportunità"].max_row, 1)
        self.assertGreater(workbook["Tutti i risultati"].max_row, 1)
        self.assertGreater(workbook["Listing Amazon"].max_row, 1)
        self.assertGreater(workbook["Scenari"].max_row, 1)

    def test_unavailable_fee_is_preserved_and_excluded_from_opportunities(self):
        product = self._result()
        product["product_key"] = "fee-unavailable-product"
        product["opportunity_combinations"] = []
        product["evaluation_status"] = "economics_unavailable"
        product["exclusion_reason"] = "amazon_fee_unavailable"
        product["amazon_listings"] = [{
            "asin": "B000000001", "title": "Crema", "brand": "Brand",
            "amazon_observation_id": "fee-unavailable-observation",
            "compatibility_status": "compatible",
            "beauty_status": "display_group_beauty", "bsr_beauty": 5000,
            "catalog_status": "resolved", "pricing_status": "success",
            "competition_status": "passed", "evaluation_status": "economics_unavailable",
            "reference_price": Decimal("30"), "fba_sellers": 2,
            "total_sellers": 4, "fee_status": "unavailable",
            "fee_attempts": 3, "fee_unavailable_reason": "amazon_internal_error",
            "exclusion_reason": "amazon_fee_unavailable",
        }]
        state = {
            "job_id": "fee-partial", "status": "completed", "phase": "completed",
            "filters": {"minimum_margin": 15}, "candidates": [product], "results": [],
            "fee_target_count": 1, "fee_valid_count": 0,
            "fee_unavailable_count": 1, "fee_invalid_count": 0,
            "fee_coverage_partial": True,
        }
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "partial-fees.xlsx"
            write_discovery_excel(state, path)
            workbook = load_workbook(path, data_only=False)
        listing_headers = {
            cell.value: cell.column for cell in workbook["Listing Amazon"][1]
        }
        self.assertEqual(
            workbook["Listing Amazon"].cell(2, listing_headers["Fee status"]).value,
            "unavailable",
        )
        self.assertEqual(workbook["Opportunità"].max_row, 1)
        metadata = {
            row[0].value: row[1].value
            for row in workbook["Parametri run"].iter_rows(min_row=2)
        }
        self.assertEqual(metadata["Fee Amazon non disponibili"], 1)
        self.assertTrue(metadata["Copertura Fee parziale"])

    def test_large_multiscenario_export(self):
        template = self._result()
        products = []
        for index in range(1340):
            import copy
            product = copy.deepcopy(template)
            product["gtin"] = f"{8800000000000 + index}"
            product["product_key"] = f"product-{index}"
            product["amazon_observation"]["asin"] = f"B{index:09d}"
            for scenario_index, scenario in enumerate(product["scenarios"]):
                scenario["scenario_id"] = f"scenario-{index}-{scenario_index}"
                scenario["product_key"] = product["product_key"]
            product["scenario_roles"]["scenario_raccomandato"] = product["scenarios"][-1]["scenario_id"]
            products.append(product)
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "large.xlsx"
            write_discovery_excel(products, path)
            workbook = load_workbook(path, read_only=False, data_only=False)
            self.assertEqual(workbook["Scenari"].max_row, 6701)


class DiscoveryUITests(unittest.TestCase):
    @staticmethod
    def _load_result(app, state, output=b"workbook"):
        app.session_state["ui_state"] = "discovery_result"
        app.session_state["discovery_status"] = "completed"
        app.session_state["discovery_result"] = {
            "state": state, "output_bytes": output,
        }
        return app.run()

    def test_legacy_result_shows_clean_message_without_rendering_product_card(self):
        app = AppTest.from_file("app_glowup.py", default_timeout=20).run()
        legacy_state = {
            "job_id": "legacy-ui", "status": "completed", "phase": "completed",
            "funnel": {
                "qogita_initial": 80, "amazon_found": 40,
                "beauty_valid_bsr": 39, "bsr_in_range": 7,
                "competition_passed": 5, "fee_valid": 5,
                "final_opportunities": 4,
            },
            "candidates": [], "amazon_observations": [],
            "results": [{
                "gtin": "8809532221349", "brand": "HARUHARU",
                "asin": "B08W1ZHTL3", "cost_gross": "8.83", "mov": "500",
            }],
        }
        self._load_result(app, legacy_state)
        self.assertFalse(app.exception)
        self.assertTrue(any(
            "versione precedente della Discovery" in element.value
            for element in app.markdown
        ))
        self.assertEqual(
            len([button for button in app.button if button.label == "Confronta scenari"]),
            0,
        )
        self.assertEqual(len(app.metric), 0)

    def test_partial_fee_coverage_shows_warning(self):
        app = AppTest.from_file("app_glowup.py", default_timeout=20).run()
        state = {
            "discovery_schema_version": DISCOVERY_SCHEMA_VERSION,
            "job_id": "partial-fees", "status": "completed", "phase": "completed",
            "candidates": [], "results": [], "amazon_observations": [],
            "fee_target_count": 843, "fee_valid_count": 842,
            "fee_unavailable_count": 1, "fee_coverage_partial": True,
            "funnel": {
                "qogita_products": 0, "qogita_scenarios": 0,
                "fee_target_count": 843, "fee_valid_count": 842,
                "fee_unavailable_count": 1,
            },
        }
        self._load_result(app, state)
        self.assertFalse(app.exception)
        self.assertTrue(any(
            "842/843 Fee Amazon disponibili" in warning.value
            for warning in app.warning
        ))

    def test_one_product_card_and_local_scenario_detail_toggle(self):
        app = AppTest.from_file("app_glowup.py", default_timeout=20).run()
        scenario = {
            "scenario_id": "scenario-1", "product_key": "product-1",
            "canonical_ean": "8809532220748", "supplier": "qogita",
            "scenario_type": "qogita_mov", "scenario_label": "MOV € 500,00",
            "cost_gross_unit_eur": 8.47, "cost_net_unit_eur": 6.94,
            "account_mov": 500, "margin_percent": 25.0, "score": 80,
            "opportunity": "🟢 Ottima", "roles": ["Base", "Raccomandato"],
            "stock": 100,
            "economics": {"target_prices": {"15": 20, "20": 22, "25": 24}},
        }
        product = {
            "product_key": "product-1", "gtin": "8809532220748",
            "brand": "Brand", "title": "Title", "asin": "B000000001",
            "amazon_offers_url": "https://www.amazon.it/gp/offer-listing/B000000001",
            "scenarios": [scenario],
            "scenario_roles": {"scenario_raccomandato": "scenario-1"},
            "amazon_observation": {
                "amazon_brand": "Brand", "amazon_title": "Title",
                "bsr_beauty": 5000, "reference_price": 30,
                "fba_sellers": 2, "total_sellers": 4,
            },
        }
        loaded_result = json.loads(json.dumps({
            "state": {
                "discovery_schema_version": DISCOVERY_SCHEMA_VERSION,
                "job_id": "test", "results": [product],
                "candidates": [product], "amazon_observations": [],
                "qogita_refresh_status": "cache_fresh",
                "qogita_snapshot_after": {},
                "funnel": {
                    "qogita_products": 1, "qogita_scenarios": 1,
                    "amazon_found": 1, "beauty_valid": 1, "bsr_passed": 1,
                    "competition_passed": 1, "fee_valid": 1,
                    "final_opportunities": 1, "scenarios_evaluated": 1,
                    "scenarios_margin_passed": 1,
                    "scenarios_margin_below_threshold": 0,
                },
            },
            "output_bytes": b"workbook",
        }, default=lambda value: value.decode() if isinstance(value, bytes) else value))
        loaded_result["output_bytes"] = b"workbook"
        self._load_result(app, loaded_result["state"], loaded_result["output_bytes"])
        self.assertFalse(app.exception)
        toggles = [button for button in app.button if button.label == "Confronta scenari"]
        self.assertEqual(len(toggles), 1)
        toggles[0].click()
        app.run()
        self.assertFalse(app.exception)
        self.assertEqual(len(app.dataframe), 1)
        self.assertEqual(
            len([button for button in app.button if button.label == "Confronta scenari"]),
            1,
        )


if __name__ == "__main__":
    unittest.main()
