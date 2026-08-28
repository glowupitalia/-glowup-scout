import tempfile
import unittest
from decimal import Decimal
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from streamlit.testing.v1 import AppTest

from discovery import (
    DiscoveryCheckpointStore,
    _select_recommended_combination,
    default_filters,
    normalize_discovery_state,
    run_discovery,
    validate_filters,
)
from discovery_amazon import beauty_rank, correlate_catalog_items
from discovery_excel import write_discovery_excel
from purchase_scenarios import recommended_combination
from tests.testsuite_discovery import fee_error_result, fee_result, qogita_row
from tests.testsuite_umma_discovery import (
    NOW,
    real_5820_three_mode_rows,
)
from umma_discovery import normalize_umma_candidates


EAN = "8809640735820"


def catalog_item(
    asin, *, pack=1, bsr=5000,
    display_group="beauty_display_on_website", volume=30,
    volume_unit="milliliters",
):
    return {
        "asin": asin,
        "summaries": [{
            "marketplaceId": "IT", "brand": "ANUA",
            "manufacturer": "ANUA INC.",
            "itemName": "ANUA Niacinamide 10% + TXA 4% Serum 30ml",
            "packageQuantity": pack, "websiteDisplayGroup": display_group,
            "browseClassification": {"displayName": "Cura della pelle"},
        }],
        "identifiers": [{"marketplaceId": "IT", "identifiers": [
            {"identifierType": "EAN", "identifier": EAN},
        ]}],
        "productTypes": [{"marketplaceId": "IT", "productType": "SKIN_SERUM"}],
        "salesRanks": [{"marketplaceId": "IT", "displayGroupRanks": [{
            "websiteDisplayGroup": display_group,
            "title": "Bellezza" if display_group == "beauty_display_on_website" else "Salute e cura della persona",
            "rank": bsr,
        }]}],
        "attributes": {
            "item_package_quantity": [{"value": pack}],
            "number_of_items": [{"value": pack}],
            "item_volume": [{"value": volume, "unit": volume_unit}],
            "package_level": [{"value": "unit"}],
            "model_number": [{"value": "ANA006"}],
            "part_number": [{"value": EAN}],
        },
        "relationships": [{"marketplaceId": "IT", "relationships": []}],
        "images": [], "classifications": [], "dimensions": [],
    }


def normalized_candidate():
    candidates, diagnostics = normalize_umma_candidates(
        real_5820_three_mode_rows(), now=NOW
    )
    assert diagnostics["umma_scenarios"] == 3
    candidate = candidates[0]
    candidate["brand"] = "ANUA"
    candidate["package_quantity"] = 1
    candidate["volume_value"] = 30
    candidate["volume_unit"] = "ml"
    return candidate


def supplier_stats(candidate):
    return {
        "initial": 1, "valid_gtin": 1, "qogita_products": 0,
        "qogita_scenarios": 0, "umma_products": 1,
        "umma_scenarios": len(candidate["scenarios"]),
    }


class MultiListingCatalogTests(unittest.TestCase):
    def test_two_equivalent_asins_are_both_compatible(self):
        candidate = normalized_candidate()
        mapping = correlate_catalog_items(
            [EAN], [catalog_item("B0CVFPT7FC"), catalog_item("B0DL92QM79")],
            [candidate],
        )[EAN]
        self.assertEqual(mapping["status"], "resolved")
        self.assertEqual(len(mapping["listings"]), 2)
        self.assertEqual(
            {row["compatibility_status"] for row in mapping["listings"]},
            {"compatible"},
        )

    def test_supplier_single_excludes_multipack(self):
        candidate = normalized_candidate()
        mapping = correlate_catalog_items(
            [EAN], [catalog_item("B000SINGLE", pack=1), catalog_item("B000PACK02", pack=2)],
            [candidate],
        )[EAN]
        by_asin = {row["asin"]: row for row in mapping["listings"]}
        self.assertEqual(by_asin["B000SINGLE"]["compatibility_status"], "compatible")
        self.assertEqual(by_asin["B000PACK02"]["compatibility_status"], "incompatible")
        self.assertIn(
            "package_quantity_mismatch",
            by_asin["B000PACK02"]["compatibility_reason"],
        )

    def test_health_and_beauty_is_not_beauty(self):
        item = catalog_item(
            "B0CVFPT7FC", bsr=16016,
            display_group="health_and_beauty_display_on_website",
        )
        self.assertEqual(beauty_rank(item), (None, "beauty_rank_unverified"))


class ListingMeasureCompatibilityTests(unittest.TestCase):
    def listing(self, supplier_value, supplier_unit, amazon_value, amazon_unit,
                *, pack=1):
        candidate = normalized_candidate()
        candidate["volume_value"] = Decimal(str(supplier_value))
        candidate["volume_unit"] = supplier_unit
        mapping = correlate_catalog_items(
            [EAN], [catalog_item(
                "B000MEASURE", pack=pack, volume=amazon_value,
                volume_unit=amazon_unit,
            )], [candidate],
        )[EAN]
        return mapping["listings"][0]

    def test_same_and_rounded_metric_volumes_match(self):
        for actual in ("150", "150.82", "151"):
            with self.subTest(actual=actual):
                listing = self.listing("150", "ml", actual, "milliliters")
                self.assertEqual(listing["compatibility_status"], "compatible")
                self.assertIn("volume_match", listing["compatibility_reason"])

    def test_materially_different_volumes_remain_hard_conflicts(self):
        for expected, actual in ((150, 152), (150, 200), (30, 50), (100, 120), (30, 60)):
            with self.subTest(expected=expected, actual=actual):
                listing = self.listing(expected, "ml", actual, "ml")
                self.assertEqual(listing["compatibility_status"], "incompatible")
                self.assertIn("volume_mismatch", listing["compatibility_reason"])

    def test_small_format_rounding_matches(self):
        listing = self.listing("30", "ml", "30.1", "ml")
        self.assertEqual(listing["compatibility_status"], "compatible")

    def test_litres_and_fluid_ounces_are_normalized_to_ml(self):
        litre = self.listing("1000", "ml", "1", "liters")
        centilitres = self.listing("150", "ml", "15", "cl")
        fluid_ounces = self.listing("150", "ml", "5.07", "fluid ounces")
        self.assertEqual(litre["compatibility_status"], "compatible")
        self.assertEqual(centilitres["compatibility_status"], "compatible")
        self.assertEqual(fluid_ounces["compatibility_status"], "compatible")
        diagnostic = fluid_ounces["diagnostics"]["measurement_comparison"]
        self.assertEqual(diagnostic["supplier_normalized_unit"], "ml")
        self.assertEqual(diagnostic["amazon_normalized_unit"], "ml")
        self.assertLess(diagnostic["absolute_delta"], Decimal("1"))

    def test_kilograms_are_normalized_to_grams(self):
        listing = self.listing("1000", "g", "1", "kg")
        self.assertEqual(listing["compatibility_status"], "compatible")
        diagnostic = listing["diagnostics"]["measurement_comparison"]
        self.assertEqual(diagnostic["supplier_normalized_unit"], "g")
        self.assertEqual(diagnostic["amazon_normalized_unit"], "g")
        self.assertEqual(diagnostic["absolute_delta"], Decimal("0"))

    def test_weight_and_volume_are_not_equated(self):
        listing = self.listing("150", "g", "150", "ml")
        self.assertEqual(listing["compatibility_status"], "compatible")
        self.assertNotIn("volume_match", listing["compatibility_reason"])
        diagnostic = listing["diagnostics"]["measurement_comparison"]
        self.assertEqual(diagnostic["status"], "different_dimensions")
        self.assertEqual(diagnostic["supplier_dimension"], "weight")
        self.assertEqual(diagnostic["amazon_dimension"], "volume")

    def test_real_abw_150_vs_150_82_is_compatible_with_diagnostics(self):
        listing = self.listing("150", "ml", "150.82", "milliliters")
        self.assertEqual(listing["compatibility_status"], "compatible")
        self.assertNotIn("volume_mismatch", listing["compatibility_reason"])
        diagnostic = listing["diagnostics"]["measurement_comparison"]
        self.assertEqual(diagnostic["absolute_delta"], Decimal("0.82"))
        self.assertLess(diagnostic["relative_delta"], Decimal("0.01"))
        self.assertEqual(diagnostic["absolute_tolerance"], Decimal("1"))
        self.assertEqual(diagnostic["relative_tolerance"], Decimal("0.01"))

    def test_volume_tolerance_does_not_override_multipack_conflict(self):
        listing = self.listing("30", "ml", "30.1", "ml", pack=2)
        self.assertEqual(listing["compatibility_status"], "incompatible")
        self.assertIn("package_quantity_mismatch", listing["compatibility_reason"])
        self.assertIn("number_of_items_mismatch", listing["compatibility_reason"])
        self.assertNotIn("volume_mismatch", listing["compatibility_reason"])

    def test_historical_checkpoint_listing_is_not_reinterpreted(self):
        state = {
            "discovery_schema_version": "supplier_multi_listing_v1",
            "candidates": [{
                "product_key": "product", "scenarios": [],
                "amazon_listings": [{
                    "listing_id": "listing", "asin": "B07QPB2R8N",
                    "compatibility_status": "incompatible",
                    "compatibility_reason": ["volume_mismatch"],
                }],
                "opportunity_combinations": [],
            }],
            "results": [],
            "amazon_listings": [],
            "amazon_observations": [],
        }
        normalized = normalize_discovery_state(state)
        listing = normalized["candidates"][0]["amazon_listings"][0]
        self.assertEqual(listing["compatibility_status"], "incompatible")
        self.assertEqual(listing["compatibility_reason"], ["volume_mismatch"])


class MultiListingPipelineTests(unittest.TestCase):
    def run_pipeline(self, directory, *, catalog_rows=None, pricing_override=None,
                     fees_override=None, filters=None, minimum_prices=None,
                     candidate_override=None):
        candidate = candidate_override or normalized_candidate()
        calls = {"catalog": [], "pricing": [], "fees": []}
        items = catalog_rows or [
            catalog_item("B0CVFPT7FC", bsr=5000),
            catalog_item("B0DL92QM79", bsr=8000),
        ]

        def catalog(gtins, _job_id, products):
            calls["catalog"].append(list(gtins))
            return correlate_catalog_items(gtins, items, products)

        def pricing(asins, _job_id):
            calls["pricing"].append(list(asins))
            values = {
                "B0CVFPT7FC": (30, 2, 4),
                "B0DL92QM79": (40, 1, 3),
            }
            if pricing_override:
                values.update(pricing_override)
            minimums = {
                "B0CVFPT7FC": (29, 28),
                "B0DL92QM79": (39, None),
            }
            if minimum_prices:
                minimums.update(minimum_prices)
            return {
                asin: {
                    "status": "success", "Venditori FBA": values[asin][1],
                    "Venditori totali": values[asin][2],
                    "Seller count source": "summary_number_of_offers",
                    "reference_price": values[asin][0], "price_source": "buy_box",
                    "Prezzo minimo FBA Amount": minimums[asin][0],
                    "Prezzo minimo FBM Amount": minimums[asin][1],
                }
                for asin in asins
            }

        def fees(requests_, _token):
            calls["fees"].append([row["asin"] for row in requests_])
            if fees_override:
                return fees_override(requests_)
            return [
                fee_result(row["asin"], row["identifier"], referral=(4.5 if row["asin"] == "B0CVFPT7FC" else 6))
                for row in requests_
            ]

        state = run_discovery(
            filters or default_filters(), checkpoint_store=DiscoveryCheckpointStore(directory),
            catalog_batch=catalog, pricing_batch=pricing, fees_batch=fees,
            token_provider=type("Tokens", (), {"get": lambda self: "token", "invalidate": lambda self: None})(),
            qogita_loader=lambda: [qogita_row(observed_at="2026-08-24T10:00:00Z")],
            qogita_normalizer=lambda _rows, minimum_stock: (
                [candidate], supplier_stats(candidate)
            ),
            qogita_refresher=lambda _: (_ for _ in ()).throw(AssertionError("fresh")),
            now_provider=lambda: datetime(2026, 8, 24, 12, 0, tzinfo=timezone.utc),
            sleep_func=lambda _: None, pricing_batch_interval=0, fee_batch_interval=0,
        )
        return state, calls

    def test_bsr_zero_minimum_is_valid_and_only_positive_ranks_pass(self):
        filters = default_filters()
        filters.update({"bsr_min": 0, "bsr_max": 20000})
        self.assertEqual(validate_filters(filters)["bsr_min"], 0)
        for bsr, expected in ((653, True), (1, True), (20000, True), (20001, False), (0, False)):
            with self.subTest(bsr=bsr), tempfile.TemporaryDirectory() as directory:
                state, calls = self.run_pipeline(
                    directory,
                    catalog_rows=[catalog_item("B0DL92QM79", bsr=bsr)],
                    filters=filters,
                )
                self.assertEqual(bool(calls["pricing"]), expected)
        missing = catalog_item("B0DL92QM79", bsr=653)
        missing["salesRanks"] = []
        with tempfile.TemporaryDirectory() as directory:
            _, calls = self.run_pipeline(directory, catalog_rows=[missing], filters=filters)
        self.assertEqual(calls["pricing"], [])

    def test_minimum_fba_fbm_prices_survive_checkpoint_round_trip(self):
        with tempfile.TemporaryDirectory() as directory:
            state, _ = self.run_pipeline(directory)
            persisted = DiscoveryCheckpointStore(directory).load(state["job_id"])
        listings = {
            row["asin"]: row for row in persisted["candidates"][0]["amazon_listings"]
        }
        observations = {row["asin"]: row for row in persisted["amazon_observations"]}
        self.assertEqual(listings["B0CVFPT7FC"]["min_fba_price"], 29)
        self.assertEqual(listings["B0CVFPT7FC"]["min_fbm_price"], 28)
        self.assertEqual(observations["B0CVFPT7FC"]["min_fba_price"], Decimal("29"))
        self.assertEqual(observations["B0CVFPT7FC"]["min_fbm_price"], Decimal("28"))
        self.assertEqual(listings["B0DL92QM79"]["min_fba_price"], 39)
        self.assertIsNone(listings["B0DL92QM79"]["min_fbm_price"])
        self.assertIsNone(observations["B0DL92QM79"]["min_fbm_price"])

    def test_checkpoint_restores_all_economic_numeric_contracts(self):
        with tempfile.TemporaryDirectory() as directory:
            state, _ = self.run_pipeline(directory)
            persisted = DiscoveryCheckpointStore(directory).load(state["job_id"])
        product = persisted["candidates"][0]
        scenario = product["scenarios"][0]
        combination = product["opportunity_combinations"][0]
        observation = persisted["amazon_observations"][0]
        for value in (
            scenario["cost_gross_unit_eur"], scenario["margin_percent"],
            combination["cost_gross_unit_eur"], combination["price_reference"],
            combination["profit"], combination["margin_percent"],
            observation["reference_price"], observation["fba_fee_net"],
            observation["fba_fee_gross"], observation["referral_fee"],
            observation["referral_rate"],
        ):
            self.assertIsInstance(value, Decimal)
        self.assertTrue(all(
            isinstance(value, Decimal)
            for value in combination["target_prices"].values()
        ))
        self.assertTrue(all(
            isinstance(value, Decimal)
            for value in combination["economics"]["target_prices"].values()
        ))

    def test_malformed_legacy_economic_value_is_diagnostic_not_invented(self):
        state = {
            "discovery_schema_version": "supplier_multi_listing_v1",
            "funnel": {"qogita_products": 1, "qogita_scenarios": 1},
            "candidates": [],
            "results": [{
                "product_key": "product", "gtin": EAN,
                "scenario_roles": {"scenario_raccomandato": "scenario"},
                "scenarios": [{
                    "scenario_id": "scenario", "product_key": "product",
                    "canonical_ean": EAN, "supplier": "umma",
                    "scenario_type": "umma_u_quick", "account_mov": "700",
                    "cost_net_unit_eur": "8", "cost_gross_unit_eur": "9.76",
                }],
                "combination_roles": {"recommended_combination": "combination"},
                "opportunity_combinations": [{
                    "combination_id": "combination", "scenario_id": "scenario",
                    "asin": "B0C1BN3QWG", "cost_gross_unit_eur": "9.76",
                    "price_reference": "18.35", "profit": "1.95",
                    "margin_percent": "not-a-number", "score": "68",
                }],
            }],
            "amazon_observations": [],
        }
        normalized = normalize_discovery_state(state)
        combination = normalized["results"][0]["opportunity_combinations"][0]
        self.assertIsNone(combination["margin_percent"])
        self.assertIn(
            "OpportunityCombination.margin_percent:invalid_numeric_value",
            combination["numeric_normalization_errors"],
        )
        self.assertEqual(normalized["checkpoint_compatibility"], "legacy_incompatible")

    def test_legacy_multilisting_checkpoint_defaults_missing_minimum_prices(self):
        state = {
            "discovery_schema_version": "supplier_multi_listing_v1",
            "candidates": [{
                "product_key": "product", "scenarios": [],
                "amazon_listings": [{"listing_id": "listing", "asin": "B000000001"}],
                "opportunity_combinations": [],
            }],
            "results": [], "amazon_listings": [{"listing_id": "listing"}],
            "amazon_observations": [{
                "observation_id": "observation", "reference_price": "20",
                "fee_estimate": {"referral_fee": "3"},
            }],
        }
        normalized = normalize_discovery_state(state)
        self.assertIsNone(normalized["candidates"][0]["amazon_listings"][0]["min_fba_price"])
        self.assertIsNone(normalized["candidates"][0]["amazon_listings"][0]["min_fbm_price"])
        self.assertIsNone(normalized["amazon_observations"][0]["min_fba_price"])
        self.assertIsNone(normalized["amazon_observations"][0]["min_fbm_price"])

    def test_three_scenarios_two_asins_make_six_local_combinations(self):
        with tempfile.TemporaryDirectory() as directory:
            state, calls = self.run_pipeline(directory)
            persisted = DiscoveryCheckpointStore(directory).load(state["job_id"])
        product = state["results"][0]
        self.assertEqual(len(product["scenarios"]), 3)
        self.assertEqual(len(product["amazon_listings"]), 2)
        self.assertEqual(len(state["amazon_observations"]), 2)
        self.assertEqual(len(product["opportunity_combinations"]), 6)
        first_scenario = product["scenarios"][0]["scenario_id"]
        pair = [
            row for row in product["opportunity_combinations"]
            if row["scenario_id"] == first_scenario
        ]
        self.assertEqual(len(pair), 2)
        self.assertNotEqual(pair[0]["margin_percent"], pair[1]["margin_percent"])
        self.assertNotEqual(pair[0]["score"], pair[1]["score"])
        self.assertEqual(calls["catalog"], [[EAN]])
        self.assertEqual(set(calls["pricing"][0]), {"B0CVFPT7FC", "B0DL92QM79"})
        self.assertEqual(len(calls["pricing"]), 1)
        self.assertEqual(set(calls["fees"][0]), {"B0CVFPT7FC", "B0DL92QM79"})
        self.assertEqual(len(calls["fees"]), 1)
        self.assertEqual(len(persisted["candidates"][0]["opportunity_combinations"]), 6)
        self.assertIsNotNone(recommended_combination(product))
        self.assertIn(
            product["combination_roles"]["best_listing"],
            {"B0CVFPT7FC", "B0DL92QM79"},
        )
        self.assertIn(
            product["combination_roles"]["best_purchase_scenario"],
            {row["scenario_id"] for row in product["scenarios"]},
        )
        self.assertEqual(state["funnel"]["combinations_evaluated"], 6)

    def test_listing_competition_failure_does_not_block_other_listing(self):
        with tempfile.TemporaryDirectory() as directory:
            state, calls = self.run_pipeline(
                directory, pricing_override={"B0CVFPT7FC": (30, 8, 15)}
            )
        product = state["results"][0]
        by_asin = {row["asin"]: row for row in product["amazon_listings"]}
        self.assertEqual(by_asin["B0CVFPT7FC"]["evaluation_status"], "competition_filtered")
        self.assertEqual(by_asin["B0DL92QM79"]["competition_status"], "passed")
        self.assertEqual(calls["fees"], [["B0DL92QM79"]])
        self.assertEqual(len(product["opportunity_combinations"]), 3)

    def test_out_of_range_listing_does_not_block_other_listing(self):
        rows = [
            catalog_item("B0CVFPT7FC", bsr=90000),
            catalog_item("B0DL92QM79", bsr=8000),
        ]
        with tempfile.TemporaryDirectory() as directory:
            state, calls = self.run_pipeline(directory, catalog_rows=rows)
        self.assertEqual(calls["pricing"], [["B0DL92QM79"]])
        self.assertEqual(len(state["results"][0]["opportunity_combinations"]), 3)

    def test_fee_unavailable_listing_can_be_retried_without_repeating_success(self):
        def partial(requests_):
            return [
                (
                    fee_error_result(row["asin"], row["identifier"])
                    if row["asin"] == "B0DL92QM79"
                    else fee_result(row["asin"], row["identifier"])
                )
                for row in requests_
            ]

        with tempfile.TemporaryDirectory() as directory:
            store = DiscoveryCheckpointStore(directory)
            state, calls = self.run_pipeline(directory, fees_override=partial)
            self.assertEqual(state["status"], "completed")
            unavailable = next(
                row for row in state["amazon_observations"]
                if row["asin"] == "B0DL92QM79"
            )
            self.assertEqual(unavailable["fee_status"], "unavailable")
            self.assertEqual(len(state["results"]), 1)
            self.assertEqual(
                {row["asin"] for row in state["results"][0]["opportunity_combinations"]},
                {"B0CVFPT7FC"},
            )
            self.assertEqual(
                calls["fees"],
                [["B0CVFPT7FC", "B0DL92QM79"], ["B0DL92QM79"], ["B0DL92QM79"]],
            )
            resumed_calls = []

            def resumed_fees(requests_, _token):
                resumed_calls.append([row["asin"] for row in requests_])
                return [
                    fee_result(row["asin"], row["identifier"], referral=6)
                    for row in requests_
                ]

            resumed = run_discovery(
                state["filters"], checkpoint_store=store,
                catalog_batch=lambda *_: (_ for _ in ()).throw(AssertionError("catalog repeated")),
                pricing_batch=lambda *_: (_ for _ in ()).throw(AssertionError("pricing repeated")),
                fees_batch=resumed_fees,
                token_provider=type("Tokens", (), {"get": lambda self: "token", "invalidate": lambda self: None})(),
                qogita_loader=lambda: (_ for _ in ()).throw(AssertionError("supplier repeated")),
                job_id=state["job_id"], sleep_func=lambda _: None,
                pricing_batch_interval=0, fee_batch_interval=0,
            )
        self.assertEqual(resumed_calls, [["B0DL92QM79"]])
        self.assertEqual(resumed["status"], "completed")
        self.assertEqual(len(resumed["results"][0]["opportunity_combinations"]), 6)
        resumed_observation = next(
            row for row in resumed["amazon_observations"]
            if row["asin"] == "B0DL92QM79"
        )
        self.assertEqual(resumed_observation["min_fba_price"], Decimal("39"))
        self.assertIsNone(resumed_observation["min_fbm_price"])

    def test_multilisting_excel_has_combinations_and_two_data_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            state, _ = self.run_pipeline(directory)
            path = Path(directory) / "multi.xlsx"
            write_discovery_excel(state["results"], path)
            workbook = load_workbook(path, data_only=False)
        opportunity = workbook["Opportunità"]
        combinations = workbook["Scenari"]
        data = workbook["Dati"]
        self.assertEqual(opportunity.max_row, 2)
        self.assertEqual(combinations.max_row, 7)
        self.assertEqual(data.max_row, 3)
        self.assertEqual(data.sheet_state, "hidden")
        self.assertFalse(combinations["I2"].protection.locked)
        self.assertIn("MATCH($AB2,'Dati'!$A$2:$A$3,0)", combinations["N2"].value)
        headers = [cell.value for cell in data[1]]
        self.assertEqual(headers[-2:], ["Prezzo minimo FBA", "Prezzo minimo FBM"])
        self.assertEqual(data.cell(2, headers.index("Prezzo minimo FBA") + 1).value, 29)
        self.assertEqual(data.cell(2, headers.index("Prezzo minimo FBM") + 1).value, 28)

    def test_ui_renders_one_product_two_pages_and_local_combinations(self):
        with tempfile.TemporaryDirectory() as directory:
            state, calls = self.run_pipeline(directory)
            state = DiscoveryCheckpointStore(directory).load(state["job_id"])
        app = AppTest.from_file("app_glowup.py", default_timeout=20).run()
        app.session_state["ui_state"] = "discovery_result"
        app.session_state["discovery_status"] = "completed"
        app.session_state["discovery_result"] = {
            "state": state, "output_bytes": b"workbook",
        }
        app.run()
        self.assertFalse(app.exception)
        self.assertTrue(any("2 pagine Amazon" in row.value for row in app.markdown))
        toggle = next(row for row in app.button if row.label == "Confronta scenari")
        toggle.click()
        app.run()
        self.assertFalse(app.exception)
        self.assertEqual(len(app.dataframe), 2)
        listing_table = app.dataframe[0].value
        self.assertIn("Min FBA", listing_table.columns)
        self.assertIn("Min FBM", listing_table.columns)
        self.assertEqual(listing_table.loc[0, "Min FBA"], 29)
        self.assertEqual(listing_table.loc[0, "Min FBM"], 28)
        self.assertEqual(calls["catalog"], [[EAN]])
        self.assertEqual(len(calls["pricing"]), 1)
        self.assertEqual(len(calls["fees"]), 1)

    def test_pilot_equal_economics_recommends_u_quick_after_reload(self):
        candidate = normalized_candidate()
        candidate["scenarios"] = [
            row for row in candidate["scenarios"]
            if row["scenario_type"] in {"umma_u_quick", "umma_standard"}
        ]
        with tempfile.TemporaryDirectory() as directory:
            state, _ = self.run_pipeline(
                directory,
                candidate_override=candidate,
                catalog_rows=[catalog_item("B0C1BN3QWG", bsr=993)],
                pricing_override={"B0C1BN3QWG": (18.35, 5, 7)},
                minimum_prices={"B0C1BN3QWG": (18.35, 27.99)},
                fees_override=lambda requests_: [
                    fee_result(row["asin"], row["identifier"], referral=2.7525)
                    for row in requests_
                ],
                filters={
                    **default_filters(), "bsr_min": 0,
                    "max_fba_sellers": 10, "max_total_sellers": 15,
                },
            )
            persisted = DiscoveryCheckpointStore(directory).load(state["job_id"])
        product = persisted["candidates"][0]
        scenarios = {row["scenario_id"]: row for row in product["scenarios"]}
        recommended = recommended_combination(product)
        selected = scenarios[recommended["scenario_id"]]
        self.assertEqual(selected["scenario_type"], "umma_u_quick")
        self.assertEqual(product["best_purchase_scenario"], selected["scenario_id"])
        self.assertEqual(
            product["scenario_roles"]["scenario_base"], selected["scenario_id"]
        )
        self.assertIsNone(
            product["combination_roles"]["minimum_profitable_combination"]
        )

    def test_umma_operational_tie_break_order_and_supplier_boundaries(self):
        def scenario(identifier, supplier="umma", minimum=5, unit=5, stock=10, mov=700):
            return {
                "scenario_id": identifier, "supplier": supplier,
                "minimum_product_quantity": minimum, "selling_unit": unit,
                "stock": stock, "account_mov": mov,
            }

        def combination(identifier, scenario_id, supplier="umma"):
            return {
                "combination_id": identifier, "scenario_id": scenario_id,
                "supplier": supplier, "asin": "B0C1BN3QWG", "score": 68,
                "margin_percent": Decimal("10.61"),
                "profit": Decimal("1.95"),
                "cost_gross_unit_eur": Decimal("8.0297"),
            }

        scenarios = {
            "standard": scenario("standard", minimum=30, unit=30, stock=None),
            "quick": scenario("quick", minimum=5, unit=5, stock=166),
        }
        rows = [combination("a", "standard"), combination("b", "quick")]
        self.assertEqual(
            _select_recommended_combination(rows, scenarios)["scenario_id"], "quick"
        )

        scenarios = {
            "wide": scenario("wide", minimum=5, unit=10, stock=10),
            "compact": scenario("compact", minimum=5, unit=5, stock=None),
        }
        rows = [combination("a", "wide"), combination("b", "compact")]
        self.assertEqual(
            _select_recommended_combination(rows, scenarios)["scenario_id"], "compact"
        )

        scenarios = {
            "unknown": scenario("unknown", minimum=5, unit=5, stock=None),
            "known": scenario("known", minimum=5, unit=5, stock=1),
        }
        rows = [combination("a", "unknown"), combination("b", "known")]
        self.assertEqual(
            _select_recommended_combination(rows, scenarios)["scenario_id"], "known"
        )

        scenarios = {
            "z": scenario("z", minimum=5, unit=5, stock=1),
            "a": scenario("a", minimum=5, unit=5, stock=1),
        }
        rows = [combination("z", "z"), combination("a", "a")]
        self.assertEqual(
            _select_recommended_combination(rows, scenarios)["scenario_id"], "a"
        )

        cross_scenarios = {
            "qogita": scenario("qogita", supplier="qogita", minimum=None, mov=10000),
            "umma": scenario("umma", supplier="umma", minimum=1, mov=700),
        }
        cross_rows = [
            combination("q", "qogita", supplier="qogita"),
            combination("u", "umma", supplier="umma"),
        ]
        self.assertEqual(
            _select_recommended_combination(cross_rows, cross_scenarios)["supplier"],
            "qogita",
        )

        qogita_scenarios = {
            "a": scenario("a", supplier="qogita", mov=10000),
            "z": scenario("z", supplier="qogita", mov=500),
        }
        qogita_rows = [
            combination("a", "a", supplier="qogita"),
            combination("z", "z", supplier="qogita"),
        ]
        self.assertEqual(
            _select_recommended_combination(qogita_rows, qogita_scenarios)["scenario_id"],
            "a",
        )


if __name__ == "__main__":
    unittest.main()
