import json
import sqlite3
import tempfile
import unittest
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from streamlit.testing.v1 import AppTest

from discovery import (
    DISCOVERY_SCHEMA_VERSION,
    DiscoveryCheckpointStore,
    _build_amazon_observations,
    _evaluate_product_scenarios,
)
from discovery_excel import write_discovery_excel
from purchase_scenarios import (
    assign_scenario_roles,
    merge_product_candidates,
    recommended_scenario,
)
from qogita_discovery import normalize_qogita_candidates
from umma_discovery import (
    _sqlite_rows,
    normalize_umma_barcode,
    normalize_umma_candidates,
    snapshot_is_stale,
    valid_ean13,
)


NOW = datetime(2026, 8, 24, 12, 0, tzinfo=timezone.utc)


def umma_row(
    *, ean="8809532221738", mode="u_quick", price="6.78",
    stock=120, minimum=5, selling_unit=5, maximum=120,
    availability="in_stock", observed_at="2026-08-24T04:10:43Z",
    run_id="run-current", fx_rate="0.854773912300196598",
    fx_stale=False, fx_date="2026-08-21", latest_status="success",
    latest_attempt_at="2026-08-24T04:10:43Z",
):
    return {
        "run_id": run_id, "seller_sku": "HARU04", "gtin": ean,
        "supplier_product_id": "10643", "mapper_sale_product_id": "18734",
        "product_option_id": "18932", "supplier_sku": "BL01WAC06800008000",
        "product_name": "BLACK RICE 10 HYALURONIC CREAM 50ML",
        "sales_mode": mode, "observed_at": observed_at,
        "original_unit_price": price, "original_currency": "USD",
        "price_basis": "net_ex_vat",
        "price_basis_source": "glowup_business_rule",
        "fx_usd_to_eur_rate": fx_rate, "fx_reference_rate": "1.1699",
        "fx_rate_date": fx_date, "fx_source": "ECB_DAILY_REFERENCE_RATES",
        "fx_stale": fx_stale, "net_unit_price_eur": None,
        "vat_rate_percent": "22", "vat_amount_eur": None,
        "gross_unit_price_eur": None, "available_quantity": stock,
        "availability_status": availability,
        "minimum_product_quantity": minimum, "selling_unit": selling_unit,
        "maximum_quantity": maximum,
        "lead_time": "entro 48 ore" if mode == "u_quick" else (
            "Europe Direct" if mode == "europe_direct"
            else "entro 15 giorni lavorativi"
        ),
        "minimum_order_value": "700", "minimum_order_currency": "USD",
        "pricing_scope": "authenticated_buyer",
        "source": "umma_authenticated_buyer_api_v3",
        "latest_attempt_status": latest_status,
        "latest_attempt_at": latest_attempt_at,
    }


def standard_row(**overrides):
    values = dict(
        mode="standard", stock=None, minimum=20, selling_unit=20,
        maximum=10000, availability="available_to_order",
    )
    values.update(overrides)
    return umma_row(**values)


def europe_row(**overrides):
    values = dict(
        mode="europe_direct", stock=90, minimum=30, selling_unit=30,
        maximum=90, availability="in_stock",
    )
    values.update(overrides)
    return umma_row(**values)


def real_5820_three_mode_rows():
    u_quick = umma_row(
        ean="8809640735820", price="10.69", stock=1531,
        minimum=5, selling_unit=5, maximum=1531,
    )
    standard = standard_row(
        ean="8809640735820", price="10.69",
        minimum=27, selling_unit=27,
    )
    for row in (u_quick, standard):
        row.update({
            "supplier_product_id": "19013",
            "mapper_sale_product_id": "27809",
            "product_option_id": "28007",
            "product_name": "NIACINAMIDE 10%+TXA 4% SERUM 30ML (EU)",
        })
    europe = europe_row(
        ean="8809640735820ED", price="10.95", stock=1620,
        minimum=54, selling_unit=54, maximum=1620,
    )
    europe.update({
        "supplier_product_id": "15264",
        "mapper_sale_product_id": "23789",
        "product_option_id": "23987",
        "product_name": "NIACINAMIDE 10%+TXA 4% SERUM 30ML (ED)",
    })
    return [u_quick, standard, europe]


def amazon_observation():
    return {
        "observation_id": "amazon-1", "marketplace": "IT",
        "canonical_ean": "8809532220748", "asin": "B000000001",
        "amazon_brand": "Haruharu", "amazon_title": "Black Bamboo Mist",
        "bsr_beauty": 5000, "reference_price": Decimal("23"),
        "price_source": "buy_box", "fba_sellers": 2, "total_sellers": 4,
        "seller_count_source": "summary_number_of_offers",
        "fee_status": "valid",
        "fee_estimate": {
            "fba_fee_net": Decimal("4"), "fba_fee_gross": Decimal("4.88"),
            "referral_fee": Decimal("3.45"), "referral_rate": Decimal("0.15"),
            "source": "FBAFees",
        },
    }


class UmmaAdapterTests(unittest.TestCase):
    def test_read_only_cache_query_returns_latest_successful_modes(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "manager.db"
            connection = sqlite3.connect(path)
            connection.execute("""CREATE TABLE umma_purchase_price_runs (
                run_id TEXT, seller_sku TEXT, gtin TEXT, status TEXT,
                started_at TEXT, observed_at TEXT
            )""")
            connection.execute("""CREATE TABLE umma_purchase_price_snapshots (
                run_id TEXT, seller_sku TEXT, gtin TEXT,
                supplier_product_id TEXT, mapper_sale_product_id TEXT,
                product_option_id TEXT, supplier_sku TEXT, product_name TEXT,
                sales_mode TEXT, observed_at TEXT, original_unit_price TEXT,
                original_currency TEXT, price_basis TEXT, price_basis_source TEXT,
                fx_usd_to_eur_rate TEXT, fx_reference_rate TEXT, fx_rate_date TEXT,
                fx_source TEXT, fx_stale INTEGER, net_unit_price_eur TEXT,
                vat_rate_percent TEXT, vat_amount_eur TEXT,
                gross_unit_price_eur TEXT, available_quantity INTEGER,
                availability_status TEXT, minimum_product_quantity INTEGER,
                selling_unit INTEGER, maximum_quantity INTEGER, lead_time TEXT,
                minimum_order_value TEXT, minimum_order_currency TEXT,
                pricing_scope TEXT, source TEXT
            )""")
            row = umma_row()
            connection.execute(
                "INSERT INTO umma_purchase_price_runs VALUES (?,?,?,?,?,?)",
                (row["run_id"], row["seller_sku"], row["gtin"], "success",
                 row["observed_at"], row["observed_at"]),
            )
            snapshot_columns = [
                "run_id", "seller_sku", "gtin", "supplier_product_id",
                "mapper_sale_product_id", "product_option_id", "supplier_sku",
                "product_name", "sales_mode", "observed_at",
                "original_unit_price", "original_currency", "price_basis",
                "price_basis_source", "fx_usd_to_eur_rate", "fx_reference_rate",
                "fx_rate_date", "fx_source", "fx_stale", "net_unit_price_eur",
                "vat_rate_percent", "vat_amount_eur", "gross_unit_price_eur",
                "available_quantity", "availability_status",
                "minimum_product_quantity", "selling_unit", "maximum_quantity",
                "lead_time", "minimum_order_value", "minimum_order_currency",
                "pricing_scope", "source",
            ]
            connection.execute(
                f"INSERT INTO umma_purchase_price_snapshots "
                f"({','.join(snapshot_columns)}) VALUES "
                f"({','.join('?' for _ in snapshot_columns)})",
                [row.get(column) for column in snapshot_columns],
            )
            connection.commit()
            connection.close()
            loaded = _sqlite_rows(path)
        self.assertEqual(len(loaded), 1)
        self.assertEqual(loaded[0]["gtin"], "8809532221738")
        self.assertEqual(loaded[0]["sales_mode"], "u_quick")

    def test_uquick_valid_and_zero_stock_unavailable(self):
        candidates, diagnostics = normalize_umma_candidates([umma_row()], now=NOW)
        scenario = candidates[0]["scenarios"][0]
        self.assertEqual(scenario["scenario_type"], "umma_u_quick")
        self.assertEqual(scenario["minimum_product_quantity"], 5)
        self.assertEqual(scenario["selling_unit"], 5)
        self.assertEqual(scenario["stock"], 120)
        self.assertEqual(scenario["lead_time"], "entro 48 ore")
        unavailable, rejected = normalize_umma_candidates(
            [umma_row(stock=0)], now=NOW
        )
        self.assertEqual(unavailable, [])
        self.assertEqual(rejected["umma_unavailable_scenarios"], 1)

    def test_europe_direct_valid_and_unavailable(self):
        candidates, _ = normalize_umma_candidates([europe_row()], now=NOW)
        scenario = candidates[0]["scenarios"][0]
        self.assertEqual(scenario["scenario_type"], "europe_direct")
        self.assertEqual(scenario["warehouse"], "europe")
        self.assertEqual(scenario["minimum_product_quantity"], 30)
        missing, diagnostics = normalize_umma_candidates(
            [europe_row(stock=None)], now=NOW
        )
        self.assertEqual(missing, [])
        self.assertEqual(diagnostics["umma_unavailable_scenarios"], 1)

    def test_europe_direct_ed_barcode_is_strictly_normalized(self):
        canonical, identifier_type, raw, error = normalize_umma_barcode(
            "8809640735820ED", "europe_direct"
        )
        self.assertIsNone(error)
        self.assertEqual(canonical, "8809640735820")
        self.assertEqual(identifier_type, "EAN")
        self.assertEqual(raw, "8809640735820ED")
        self.assertTrue(valid_ean13(canonical))

        candidate, diagnostics = normalize_umma_candidates(
            [real_5820_three_mode_rows()[-1]], now=NOW
        )
        scenario = candidate[0]["scenarios"][0]
        self.assertEqual(diagnostics["invalid_identifier"], 0)
        self.assertEqual(candidate[0]["canonical_ean"], "8809640735820")
        self.assertEqual(candidate[0]["identifier_type"], "EAN")
        self.assertEqual(scenario["canonical_ean"], "8809640735820")
        self.assertEqual(scenario["supplier_barcode_raw"], "8809640735820ED")
        self.assertEqual(
            scenario["source_metadata"]["supplier_barcode_raw"],
            "8809640735820ED",
        )

    def test_europe_direct_invalid_check_digit_is_rejected(self):
        row = real_5820_three_mode_rows()[-1]
        row["gtin"] = "8809640735821ED"
        candidates, diagnostics = normalize_umma_candidates([row], now=NOW)
        self.assertEqual(candidates, [])
        self.assertEqual(diagnostics["invalid_identifier"], 1)
        self.assertEqual(
            diagnostics["rejected_scenarios"][0]["reason"],
            "invalid_ean13_check_digit",
        )
        self.assertEqual(
            diagnostics["rejected_scenarios"][0]["supplier_barcode_raw"],
            "8809640735821ED",
        )

    def test_only_exact_uppercase_ed_suffix_is_supported(self):
        for barcode in (
            "8809640735820EU", "8809640735820ed", "8809640735820EDD",
            "809640735820ED",
        ):
            with self.subTest(barcode=barcode):
                canonical, _, raw, error = normalize_umma_barcode(
                    barcode, "europe_direct"
                )
                self.assertIsNone(canonical)
                self.assertEqual(raw, barcode)
                self.assertEqual(error, "invalid_supplier_barcode_format")
        canonical, _, _, error = normalize_umma_barcode(
            "8809640735820ED", "u_quick"
        )
        self.assertIsNone(canonical)
        self.assertEqual(error, "invalid_supplier_barcode_format")

    def test_all_observed_europe_direct_barcodes_map_to_valid_ean13(self):
        expected = {
            "8809640735820ED": "8809640735820",
            "8809640735288ED": "8809640735288",
            "8809640735790ED": "8809640735790",
            "8809640736025ED": "8809640736025",
            "8809640735653ED": "8809640735653",
        }
        for raw, canonical_expected in expected.items():
            with self.subTest(raw=raw):
                canonical, identifier_type, preserved, error = (
                    normalize_umma_barcode(raw, "europe_direct")
                )
                self.assertIsNone(error)
                self.assertEqual(canonical, canonical_expected)
                self.assertEqual(identifier_type, "EAN")
                self.assertEqual(preserved, raw)
                self.assertTrue(valid_ean13(canonical))

    def test_three_live_modes_merge_under_one_canonical_ean(self):
        candidates, diagnostics = normalize_umma_candidates(
            real_5820_three_mode_rows(), now=NOW
        )
        self.assertEqual(len(candidates), 1)
        product = candidates[0]
        self.assertEqual(product["canonical_ean"], "8809640735820")
        self.assertEqual(product["gtin"], "8809640735820")
        self.assertEqual(diagnostics["umma_scenarios"], 3)
        self.assertEqual(len(product["scenarios"]), 3)
        self.assertEqual(
            {row["shipping_mode"] for row in product["scenarios"]},
            {"standard", "u_quick", "europe_direct"},
        )
        self.assertEqual(len({row["scenario_id"] for row in product["scenarios"]}), 3)
        europe = next(
            row for row in product["scenarios"]
            if row["shipping_mode"] == "europe_direct"
        )
        self.assertEqual(europe["supplier_product_id"], "15264")
        self.assertEqual(europe["variant_id"], "23987")
        self.assertEqual(europe["stock"], 1620)
        self.assertEqual(europe["selling_unit"], 54)

    def test_standard_valid_with_unknown_numeric_stock(self):
        candidates, _ = normalize_umma_candidates([standard_row()], now=NOW)
        scenario = candidates[0]["scenarios"][0]
        self.assertEqual(scenario["scenario_type"], "umma_standard")
        self.assertIsNone(scenario["stock"])
        self.assertEqual(scenario["availability_status"], "available_to_order")
        self.assertEqual(scenario["minimum_product_quantity"], 20)
        self.assertEqual(scenario["lead_time"], "entro 15 giorni lavorativi")

    def test_fx_conversion_vat_and_account_mov_are_distinct_from_moq(self):
        candidate = normalize_umma_candidates([umma_row()], now=NOW)[0][0]
        scenario = candidate["scenarios"][0]
        expected_net = Decimal("6.78") * Decimal("0.854773912300196598")
        self.assertEqual(scenario["cost_net_unit_eur"], expected_net)
        self.assertEqual(scenario["vat_amount_unit"], expected_net * Decimal("0.22"))
        self.assertEqual(
            scenario["cost_gross_unit_eur"], expected_net * Decimal("1.22")
        )
        self.assertEqual(scenario["account_mov"], Decimal("700"))
        self.assertEqual(scenario["account_mov_currency"], "USD")
        self.assertEqual(scenario["minimum_product_quantity"], 5)
        self.assertEqual(
            scenario["account_mov_eur"],
            Decimal("700") * Decimal("0.854773912300196598"),
        )

    def test_missing_or_stale_fx_suppresses_economic_scenario(self):
        for row, expected in (
            (umma_row(fx_rate=None), "fx_missing"),
            (umma_row(fx_rate="bad"), "fx_invalid"),
            (umma_row(fx_stale=True), "fx_stale"),
        ):
            with self.subTest(expected=expected):
                candidates, diagnostics = normalize_umma_candidates([row], now=NOW)
                self.assertEqual(candidates, [])
                self.assertEqual(diagnostics["umma_invalid_fx_scenarios"], 1)
                self.assertEqual(diagnostics["rejected_scenarios"][0]["reason"], expected)

    def test_snapshot_freshness_manager_boundary(self):
        self.assertFalse(snapshot_is_stale(NOW - timedelta(hours=47, minutes=59), now=NOW))
        self.assertFalse(snapshot_is_stale(NOW - timedelta(hours=48), now=NOW))
        self.assertTrue(snapshot_is_stale(NOW - timedelta(hours=48, seconds=1), now=NOW))
        candidates, diagnostics = normalize_umma_candidates([
            umma_row(observed_at="2026-08-22T11:59:59Z")
        ], now=NOW)
        self.assertEqual(candidates, [])
        self.assertEqual(diagnostics["umma_stale_scenarios"], 1)

    def test_later_failed_attempt_makes_snapshot_stale(self):
        candidates, diagnostics = normalize_umma_candidates([
            umma_row(
                latest_status="failed",
                latest_attempt_at="2026-08-24T05:00:00Z",
            )
        ], now=NOW)
        self.assertEqual(candidates, [])
        self.assertEqual(diagnostics["umma_stale_scenarios"], 1)

    def test_real_1738_standard_and_uquick_same_price_remain_distinct(self):
        candidates, diagnostics = normalize_umma_candidates(
            [umma_row(), standard_row()], now=NOW
        )
        scenarios = candidates[0]["scenarios"]
        self.assertEqual(diagnostics["umma_products"], 1)
        self.assertEqual(diagnostics["umma_scenarios"], 2)
        self.assertEqual(len(scenarios), 2)
        self.assertEqual(
            {row["source_net_unit_price"] for row in scenarios}, {Decimal("6.78")}
        )
        self.assertEqual(len({row["scenario_id"] for row in scenarios}), 2)
        self.assertEqual(
            {row["minimum_product_quantity"] for row in scenarios}, {5, 20}
        )
        self.assertEqual({row["stock"] for row in scenarios}, {120, None})

    def test_scenario_id_ignores_snapshot_price_stock_and_fx(self):
        first = normalize_umma_candidates([umma_row()], now=NOW)[0][0]["scenarios"][0]
        changed = umma_row(
            price="7.10", stock=99, run_id="new-run",
            observed_at="2026-08-24T10:00:00Z", fx_rate="0.86",
        )
        second = normalize_umma_candidates([changed], now=NOW)[0][0]["scenarios"][0]
        self.assertEqual(first["scenario_id"], second["scenario_id"])

    def test_checkpoint_json_round_trip_preserves_umma_fields(self):
        candidate = normalize_umma_candidates(
            real_5820_three_mode_rows(), now=NOW
        )[0][0]
        with tempfile.TemporaryDirectory() as directory:
            store = DiscoveryCheckpointStore(directory)
            state = store.create({"supplier": "umma"})
            state["candidates"] = [candidate]
            store.save(state)
            loaded = store.load(state["job_id"])
        standard = next(
            row for row in loaded["candidates"][0]["scenarios"]
            if row["shipping_mode"] == "standard"
        )
        self.assertIsNone(standard["stock"])
        self.assertEqual(standard["minimum_product_quantity"], 27)
        self.assertEqual(standard["account_mov_currency"], "USD")
        self.assertEqual(standard["fx_source"], "ECB_DAILY_REFERENCE_RATES")
        self.assertEqual(standard["freshness_status"], "fresh")
        europe = next(
            row for row in loaded["candidates"][0]["scenarios"]
            if row["shipping_mode"] == "europe_direct"
        )
        self.assertEqual(europe["canonical_ean"], "8809640735820")
        self.assertEqual(europe["supplier_barcode_raw"], "8809640735820ED")


class UmmaCrossSupplierTests(unittest.TestCase):
    @staticmethod
    def qogita_rows():
        return [{
            "run_id": "qogita-run", "gtin": "8809532220748",
            "variant_fid": "variant", "offer_qid": "offer",
            "product_name": "Black Bamboo Mist", "brand": "Haruharu",
            "category_name": "Beauty", "image_url": "", "inventory": 5000,
            "selling_unit": 60, "product_url": "",
            "observed_at": "2026-08-24T04:00:00Z", "seller_alias": "3VKO8Q",
            "tier_mov": mov, "currency": "EUR", "tier_price": price,
            "is_active": False,
        } for mov, price in (
            (500, "6.94"), (1500, "6.83"), (5000, "6.80"),
            (10000, "6.73"), (15000, "6.73"),
        )]

    def merged_product(self):
        qogita = normalize_qogita_candidates(self.qogita_rows(), now=NOW)[0]
        umma = normalize_umma_candidates([
            standard_row(
                ean="8809532220748", price="8.02",
                run_id="umma-standard-0748",
            )
        ], now=NOW)[0]
        return merge_product_candidates(qogita, umma)[0]

    def test_0748_is_one_product_with_five_qogita_and_one_umma_scenario(self):
        product = self.merged_product()
        self.assertEqual(product["canonical_ean"], "8809532220748")
        self.assertEqual(product["suppliers"], ["qogita", "umma"])
        self.assertEqual(len(product["scenarios"]), 6)
        self.assertEqual(
            [row["supplier"] for row in product["scenarios"]].count("qogita"), 5
        )
        self.assertEqual(
            [row["supplier"] for row in product["scenarios"]].count("umma"), 1
        )

    def test_amazon_observation_is_deduplicated_before_scenario_fanout(self):
        product = self.merged_product()
        product.update({
            "asin": "B000000001", "reference_price": Decimal("23"),
            "bsr_beauty": 5000, "price_source": "buy_box",
            "fba_sellers": 2, "total_sellers": 4,
            "seller_count_source": "summary_number_of_offers",
        })
        observations = _build_amazon_observations([product])
        self.assertEqual(len({product["canonical_ean"]}), 1)
        self.assertEqual(len(observations), 1)
        self.assertEqual(len(product["scenarios"]), 6)

    def test_three_umma_modes_share_one_amazon_observation(self):
        product = normalize_umma_candidates(
            real_5820_three_mode_rows(), now=NOW
        )[0][0]
        product.update({
            "asin": "B000000582", "reference_price": Decimal("23"),
            "bsr_beauty": 5000, "price_source": "buy_box",
            "fba_sellers": 2, "total_sellers": 4,
            "seller_count_source": "summary_number_of_offers",
        })
        observations = _build_amazon_observations([product])
        self.assertEqual(len(observations), 1)
        self.assertEqual(observations[0]["canonical_ean"], "8809640735820")
        self.assertEqual(len(product["scenarios"]), 3)
        self.assertEqual(
            {row["amazon_observation_id"] for row in [product]},
            {observations[0]["observation_id"]},
        )

    def test_economics_score_and_supplier_aware_roles_fan_out_offline(self):
        product = self.merged_product()
        observation = amazon_observation()
        passed = _evaluate_product_scenarios(product, observation, 15)
        self.assertTrue(passed)
        self.assertEqual(len(product["scenarios"]), 6)
        self.assertTrue(all("economics" in row and "score" in row for row in product["scenarios"]))
        roles = product["scenario_roles"]
        self.assertIsNone(roles["scenario_base"])
        self.assertEqual(set(roles["scenario_base_by_supplier"]), {"qogita", "umma"})
        self.assertIsNotNone(roles["scenario_raccomandato"])

    def test_discovery_excel_represents_umma_supplier_requirement_and_nullable_stock(self):
        product = self.merged_product()
        observation = amazon_observation()
        _evaluate_product_scenarios(product, observation, 0)
        umma = next(row for row in product["scenarios"] if row["supplier"] == "umma")
        product["scenario_roles"]["scenario_raccomandato"] = umma["scenario_id"]
        product["amazon_observation"] = observation
        product["amazon_offers_url"] = "https://www.amazon.it/gp/offer-listing/B000000001"
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "umma.xlsx"
            write_discovery_excel([product], path)
            workbook = load_workbook(path, data_only=False)
            opportunity = workbook["Opportunità"]
            scenarios = workbook["Scenari"]
            self.assertEqual(opportunity["D2"].value, "UMMA")
            self.assertIn("Min. 20 pz", opportunity["F2"].value)
            umma_row_number = next(
                row for row in range(2, scenarios.max_row + 1)
                if scenarios[f"E{row}"].value == "UMMA"
            )
            self.assertIn("MOV USD 700", scenarios[f"H{umma_row_number}"].value)
            self.assertIsNone(scenarios[f"W{umma_row_number}"].value)
            self.assertIn("15 giorni", scenarios[f"X{umma_row_number}"].value)

    def test_discovery_excel_lists_europe_direct_as_distinct_scenario(self):
        product = normalize_umma_candidates(
            real_5820_three_mode_rows(), now=NOW
        )[0][0]
        observation = amazon_observation()
        observation["canonical_ean"] = "8809640735820"
        _evaluate_product_scenarios(product, observation, 0)
        product["amazon_observation"] = observation
        product["amazon_offers_url"] = (
            "https://www.amazon.it/gp/offer-listing/B000000582"
        )
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "umma-europe-direct.xlsx"
            write_discovery_excel([product], path)
            workbook = load_workbook(path, data_only=False)
            scenarios = workbook["Scenari"]
            labels = [
                scenarios[f"G{row}"].value
                for row in range(2, scenarios.max_row + 1)
            ]
            self.assertEqual(scenarios.max_row - 1, 3)
            self.assertIn("Europe Direct", labels)
            europe_row_number = labels.index("Europe Direct") + 2
            self.assertEqual(scenarios[f"A{europe_row_number}"].value, "8809640735820")
            self.assertEqual(scenarios[f"E{europe_row_number}"].value, "UMMA")
            self.assertIn("Min. 54 pz", scenarios[f"H{europe_row_number}"].value)


class UmmaSupplierNeutralUITests(unittest.TestCase):
    def test_renderer_and_scenario_detail_accept_umma_without_external_calls(self):
        candidate = normalize_umma_candidates(
            real_5820_three_mode_rows(), now=NOW
        )[0][0]
        observation = amazon_observation()
        observation["canonical_ean"] = "8809640735820"
        _evaluate_product_scenarios(candidate, observation, 0)
        candidate["amazon_observation"] = observation
        candidate["amazon_offers_url"] = "https://www.amazon.it/gp/offer-listing/B000000001"
        state = {
            "discovery_schema_version": DISCOVERY_SCHEMA_VERSION,
            "job_id": "umma-ui", "status": "completed", "phase": "completed",
            "results": [candidate], "candidates": [candidate],
            "amazon_observations": [observation],
            "qogita_refresh_status": "cache_fresh", "qogita_snapshot_after": {},
            "funnel": {
                "qogita_products": 0, "qogita_scenarios": 0,
                "amazon_found": 1, "beauty_valid": 1, "bsr_passed": 1,
                "competition_passed": 1, "fee_valid": 1,
                "final_opportunities": 1, "scenarios_evaluated": 3,
                "scenarios_margin_passed": 3,
                "scenarios_margin_below_threshold": 0,
            },
        }
        app = AppTest.from_file("app_glowup.py", default_timeout=20).run()
        app.session_state["ui_state"] = "discovery_result"
        app.session_state["discovery_status"] = "completed"
        app.session_state["discovery_result"] = {
            "state": json.loads(json.dumps(state, default=str)),
            "output_bytes": b"workbook",
        }
        app.run()
        self.assertFalse(app.exception)
        self.assertTrue(any("UMMA" in element.value for element in app.markdown))
        toggle = next(button for button in app.button if button.label == "Confronta scenari")
        toggle.click().run()
        self.assertFalse(app.exception)
        self.assertEqual(len(app.dataframe), 1)
        self.assertEqual(len(app.dataframe[0].value), 3)
        self.assertIn(
            "Europe Direct", app.dataframe[0].value["Scenario"].tolist()
        )


if __name__ == "__main__":
    unittest.main()
