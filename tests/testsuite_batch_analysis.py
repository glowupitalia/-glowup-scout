import tempfile
import unittest
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from batch_analysis import (
    RESULT_COLUMNS,
    analyze_products,
    bsr_points,
    calculate_economics,
    fba_seller_points,
    margin_points,
    opportunity_from_score,
    opportunity_score,
    parse_product_fee_result,
    select_reference_price,
    summarize_results,
    total_seller_points,
    write_results_excel,
)
from product_fees import build_product_fee_requests, search_product_fees_batch


def fee_result(
    asin,
    identifier=None,
    *,
    price=20,
    fba_final=4,
    pick_and_pack_final=4,
    referral_final=3,
    tax=None,
):
    included = []
    if pick_and_pack_final is not None:
        included.append({
            "FeeType": "FBAPickAndPack",
            "FeeAmount": {"Amount": pick_and_pack_final, "CurrencyCode": "EUR"},
            "FeePromotion": {"Amount": 0, "CurrencyCode": "EUR"},
            "FinalFee": {"Amount": pick_and_pack_final, "CurrencyCode": "EUR"},
        })
    fba = {
        "FeeType": "FBAFees",
        "IncludedFeeDetailList": included,
    }
    if fba_final is not None:
        fba.update({
            "FeeAmount": {"Amount": fba_final, "CurrencyCode": "EUR"},
            "FeePromotion": {"Amount": 0, "CurrencyCode": "EUR"},
            "FinalFee": {"Amount": fba_final, "CurrencyCode": "EUR"},
        })
    if tax is not None:
        fba["TaxAmount"] = {"Amount": tax, "CurrencyCode": "EUR"}

    fee_details = [fba]
    if referral_final is not None:
        fee_details.insert(0, {
            "FeeType": "ReferralFee",
            "FeeAmount": {"Amount": referral_final, "CurrencyCode": "EUR"},
            "FeePromotion": {"Amount": 0, "CurrencyCode": "EUR"},
            "FinalFee": {"Amount": referral_final, "CurrencyCode": "EUR"},
        })
    return {
        "FeesEstimateResult": {
            "Status": "Success",
            "FeesEstimateIdentifier": {
                "IdType": "ASIN",
                "IdValue": asin,
                "SellerInputIdentifier": identifier or f"fee|{asin}",
                "PriceToEstimateFees": {
                    "ListingPrice": {"Amount": price, "CurrencyCode": "EUR"},
                },
            },
            "FeesEstimate": {
                "FeeDetailList": fee_details,
                "TotalFeesEstimate": {
                    "Amount": (fba_final or pick_and_pack_final or 0)
                    + (referral_final or 0),
                    "CurrencyCode": "EUR",
                },
            },
        },
    }


class ScoreTests(unittest.TestCase):
    def test_all_bsr_bands(self):
        cases = [
            (None, 0),
            ("invalid", 0),
            (1000, 50),
            (1001, 45),
            (5000, 45),
            (5001, 40),
            (10000, 40),
            (10001, 30),
            (25000, 30),
            (25001, 15),
            (50000, 15),
            (50001, 5),
        ]
        for value, expected in cases:
            with self.subTest(value=value):
                self.assertEqual(bsr_points(value), expected)

    def test_all_fba_seller_bands_and_zero_offer_special_case(self):
        cases = [
            (0, 0, 0),
            (0, 1, 20),
            (1, 1, 18),
            (2, 2, 18),
            (3, 3, 14),
            (4, 4, 14),
            (5, 5, 10),
            (6, 6, 10),
            (7, 7, 5),
            (10, 10, 5),
            (11, 11, 0),
        ]
        for fba, total, expected in cases:
            with self.subTest(fba=fba, total=total):
                self.assertEqual(fba_seller_points(fba, total), expected)

    def test_all_total_seller_bands(self):
        cases = [
            (None, 0),
            (0, 0),
            (1, 10),
            (3, 10),
            (4, 7),
            (6, 7),
            (7, 4),
            (10, 4),
            (11, 0),
        ]
        for value, expected in cases:
            with self.subTest(value=value):
                self.assertEqual(total_seller_points(value), expected)

    def test_zero_offers_zeroes_entire_competition_component(self):
        score, _ = opportunity_score(1000, 0, 0, 25)
        self.assertEqual(score, 70)

    def test_all_margin_bands(self):
        cases = [
            (None, 0),
            (9.99, 0),
            (10, 4),
            (14.99, 4),
            (15, 14),
            (19.99, 14),
            (20, 18),
            (24.99, 18),
            (25, 20),
            (50, 20),
        ]
        for value, expected in cases:
            with self.subTest(value=value):
                self.assertEqual(margin_points(value), expected)

    def test_all_opportunity_bands(self):
        cases = [
            (100, "🟢 Eccellente"),
            (85, "🟢 Eccellente"),
            (84, "🟢 Ottima"),
            (70, "🟢 Ottima"),
            (69, "🟡 Interessante"),
            (55, "🟡 Interessante"),
            (54, "🟠 Da valutare"),
            (40, "🟠 Da valutare"),
            (39, "🔴 Debole"),
        ]
        for score, expected in cases:
            with self.subTest(score=score):
                self.assertEqual(opportunity_from_score(score), expected)


class ProductFeesAndEconomicsTests(unittest.TestCase):
    def test_batch_payload_uses_asin_buy_box_eur_and_fba(self):
        candidates = [{
            "asin": "ASIN1",
            "price": 19.9,
            "identifier": "scout|1|ASIN1",
        }]
        body = build_product_fee_requests(candidates, "IT-MARKETPLACE")
        self.assertEqual(len(body), 1)
        self.assertEqual(body[0]["IdType"], "ASIN")
        self.assertEqual(body[0]["IdValue"], "ASIN1")
        request = body[0]["FeesEstimateRequest"]
        self.assertEqual(request["MarketplaceId"], "IT-MARKETPLACE")
        self.assertTrue(request["IsAmazonFulfilled"])
        self.assertEqual(
            request["PriceToEstimateFees"]["ListingPrice"],
            {"Amount": 19.9, "CurrencyCode": "EUR"},
        )

    def test_fbm_reference_still_requests_amazon_fulfilled_fee(self):
        price, source = select_reference_price({
            "Buy Box Amount": None,
            "Prezzo minimo FBA Amount": None,
            "Prezzo minimo FBM Amount": 18.75,
        })
        body = build_product_fee_requests([{
            "asin": "ASIN1",
            "price": float(price),
            "identifier": "scout|1|ASIN1",
        }], "IT-MARKETPLACE")
        self.assertEqual(source, "min_fbm")
        self.assertEqual(
            body[0]["FeesEstimateRequest"]["PriceToEstimateFees"]
            ["ListingPrice"]["Amount"],
            18.75,
        )
        self.assertTrue(body[0]["FeesEstimateRequest"]["IsAmazonFulfilled"])

    def test_batch_payload_rejects_more_than_twenty_products(self):
        candidates = [
            {"asin": str(index), "price": 20, "identifier": str(index)}
            for index in range(21)
        ]
        with self.assertRaises(ValueError):
            build_product_fee_requests(candidates, "IT-MARKETPLACE")

    def test_http_adapter_does_not_expose_token_in_payload(self):
        calls = []

        class Response:
            def raise_for_status(self):
                return None

            def json(self):
                return []

        def post(url, **kwargs):
            calls.append((url, kwargs))
            return Response()

        search_product_fees_batch(
            [{"asin": "ASIN1", "price": 20, "identifier": "fee|1"}],
            "secret-token",
            marketplace_id="IT-MARKETPLACE",
            request_post=post,
        )
        self.assertEqual(len(calls), 1)
        self.assertNotIn("secret-token", str(calls[0][1]["json"]))
        self.assertEqual(calls[0][1]["timeout"], 30)

    def test_fba_fees_final_is_primary_and_not_summed_with_pick_and_pack(self):
        parsed = parse_product_fee_result(
            fee_result(
                "ASIN1",
                fba_final=4.25,
                pick_and_pack_final=4.10,
                referral_final=3,
            ),
            reference_price=20,
        )
        self.assertEqual(parsed["source"], "FBAFees")
        self.assertEqual(parsed["fba_fee_net"], Decimal("4.25"))
        self.assertNotEqual(parsed["fba_fee_net"], Decimal("8.35"))

    def test_pick_and_pack_is_fallback_when_fba_final_is_missing(self):
        parsed = parse_product_fee_result(
            fee_result(
                "ASIN1",
                fba_final=None,
                pick_and_pack_final=4.10,
            ),
            reference_price=20,
        )
        self.assertEqual(parsed["source"], "FBAPickAndPack")
        self.assertEqual(parsed["fba_fee_net"], Decimal("4.10"))

    def test_manager_vat_is_applied_once_when_amazon_tax_is_absent(self):
        parsed = parse_product_fee_result(
            fee_result("ASIN1", fba_final=4),
            reference_price=20,
        )
        self.assertEqual(parsed["fba_fee_net"], Decimal("4"))
        self.assertEqual(parsed["fba_tax"], Decimal("0.88"))
        self.assertEqual(parsed["fba_fee_gross"], Decimal("4.88"))
        self.assertEqual(parsed["tax_handling"], "manager_vat_22")

    def test_explicit_tax_in_final_is_not_applied_twice(self):
        entry = fee_result("ASIN1", fba_final=4.88, tax=0.88)
        fba = entry["FeesEstimateResult"]["FeesEstimate"]["FeeDetailList"][1]
        fba["FeeAmount"]["Amount"] = 4
        parsed = parse_product_fee_result(entry, reference_price=20)
        self.assertEqual(parsed["fba_fee_net"], Decimal("4"))
        self.assertEqual(parsed["fba_fee_gross"], Decimal("4.88"))
        self.assertEqual(
            parsed["tax_handling"], "amazon_tax_included_in_final"
        )

    def test_referral_fee_is_separate_and_drives_current_margin(self):
        parsed = parse_product_fee_result(
            fee_result("ASIN1", price=20, fba_final=4, referral_final=3),
            reference_price=20,
        )
        economics = calculate_economics(20, 10, parsed)
        self.assertEqual(parsed["referral_fee"], Decimal("3"))
        self.assertEqual(
            economics["referral_source"],
            "amazon_referral_fee",
        )
        self.assertEqual(economics["referral_fee"], Decimal("3.00"))
        self.assertEqual(economics["fba_fee_gross"], Decimal("4.88"))
        self.assertEqual(economics["profit"], Decimal("2.12"))
        self.assertEqual(economics["margin_percent"], Decimal("10.60"))

    def test_amazon_referral_at_fifteen_percent_is_not_replaced(self):
        parsed = parse_product_fee_result(
            fee_result("ASIN1", price=20, fba_final=4, referral_final=3),
            reference_price=20,
        )
        economics = calculate_economics(20, 10, parsed)
        self.assertEqual(economics["referral_rate"], Decimal("0.15"))
        self.assertEqual(economics["referral_fee"], Decimal("3.00"))
        self.assertEqual(
            economics["referral_source"],
            "amazon_referral_fee",
        )

    def assert_fallback_19_percent(self, referral_final):
        parsed = parse_product_fee_result(
            fee_result(
                "ASIN1",
                price=20,
                fba_final=4,
                referral_final=referral_final,
            ),
            reference_price=20,
        )
        economics = calculate_economics(20, 10, parsed)
        self.assertEqual(economics["status"], "ready")
        self.assertEqual(
            economics["referral_source"],
            "fallback_19_percent",
        )
        self.assertEqual(economics["referral_rate"], Decimal("0.19"))
        self.assertEqual(economics["referral_fee"], Decimal("3.80"))
        self.assertEqual(economics["fba_fee_gross"], Decimal("4.88"))
        self.assertEqual(economics["profit"], Decimal("1.32"))
        self.assertEqual(economics["margin_percent"], Decimal("6.60"))
        self.assertEqual(economics["target_prices"][15], Decimal("22.55"))
        self.assertEqual(economics["target_prices"][20], Decimal("24.39"))
        self.assertEqual(economics["target_prices"][25], Decimal("26.57"))

    def test_missing_referral_uses_19_percent_without_extra_vat(self):
        self.assert_fallback_19_percent(None)

    def test_zero_referral_uses_19_percent_without_extra_vat(self):
        self.assert_fallback_19_percent(0)

    def test_target_prices_solve_margin_with_variable_referral_fee(self):
        parsed = parse_product_fee_result(
            fee_result("ASIN1", price=20, fba_final=4, referral_final=3),
            reference_price=20,
        )
        economics = calculate_economics(20, 10, parsed)
        self.assertEqual(economics["target_prices"][15], Decimal("21.26"))
        self.assertEqual(economics["target_prices"][20], Decimal("22.89"))
        self.assertEqual(economics["target_prices"][25], Decimal("24.80"))


class BatchAnalysisTests(unittest.TestCase):
    def run_single_fee_case(self, fee_entries):
        return analyze_products(
            pd.DataFrame({"EAN": ["111"], "Costo": [10]}),
            "Costo",
            "test-token-must-not-be-logged",
            lambda _ean, _token: {
                "ASIN": "ASIN111", "Titolo": "Prodotto",
                "Brand": "Brand", "BSR Beauty": 1000,
            },
            lambda _asin, _token: {
                "Buy Box Amount": 20,
                "Venditori totali": 2,
                "Venditori FBA": 1,
            },
            lambda function, *args: function(*args),
            search_fees_batch=lambda _candidates, _token: fee_entries,
            throttle_seconds=0,
            fee_batch_interval_seconds=0,
            source_file="diagnostic.xlsx",
        )

    def run_reference_price_case(self, pricing, fee_entries=None):
        captured_candidates = []

        def fees(candidates, _token):
            captured_candidates.extend(candidates)
            if fee_entries is not None:
                return fee_entries
            return [
                fee_result(
                    candidate["asin"],
                    identifier=candidate["identifier"],
                    price=candidate["price"],
                )
                for candidate in candidates
            ]

        result = analyze_products(
            pd.DataFrame({"EAN": ["111"], "Costo": [10]}),
            "Costo",
            "token",
            lambda _ean, _token: {
                "ASIN": "ASIN111", "Titolo": "Prodotto",
                "Brand": "Brand", "BSR Beauty": 1000,
            },
            lambda _asin, _token: {
                "Venditori totali": 2,
                "Venditori FBA": 1,
                **pricing,
            },
            lambda function, *args: function(*args),
            search_fees_batch=fees,
            throttle_seconds=0,
            fee_batch_interval_seconds=0,
        )
        return result.iloc[0], captured_candidates

    def test_buy_box_has_priority_as_reference_price(self):
        row, candidates = self.run_reference_price_case({
            "Buy Box Amount": 20,
            "Prezzo minimo FBA Amount": 18,
            "Prezzo minimo FBM Amount": 17,
        })
        self.assertEqual(row["Prezzo riferimento"], 20)
        self.assertEqual(row["_Price source"], "buy_box")
        self.assertEqual(candidates[0]["price"], 20)

    def test_lowest_fba_is_used_without_buy_box(self):
        price, source = select_reference_price({
            "Prezzo minimo FBA Amount": 14.25,
            "Prezzo minimo FBM Amount": 12,
        })
        self.assertEqual(price, Decimal("14.25"))
        self.assertEqual(source, "min_fba")

    def test_lowest_of_multiple_fba_landed_prices_is_propagated(self):
        row, candidates = self.run_reference_price_case({
            "Buy Box Amount": None,
            "Prezzo minimo FBA Amount": min(18.5, 16.25, 17.0),
            "Prezzo minimo FBM Amount": 15,
        })
        self.assertEqual(row["Prezzo riferimento"], 16.25)
        self.assertEqual(row["_Price source"], "min_fba")
        self.assertEqual(candidates[0]["price"], 16.25)

    def test_fbm_is_used_only_when_buy_box_and_fba_are_missing(self):
        row, candidates = self.run_reference_price_case({
            "Buy Box Amount": None,
            "Prezzo minimo FBA Amount": None,
            "Prezzo minimo FBM Amount": 13.75,
        })
        self.assertEqual(row["Prezzo riferimento"], 13.75)
        self.assertEqual(row["_Price source"], "min_fbm")
        self.assertEqual(candidates[0]["price"], 13.75)

    def test_fba_has_priority_even_when_fbm_is_cheaper(self):
        price, source = select_reference_price({
            "Prezzo minimo FBA Amount": 17,
            "Prezzo minimo FBM Amount": 12,
        })
        self.assertEqual(price, Decimal("17"))
        self.assertEqual(source, "min_fba")

    def test_missing_or_non_positive_prices_skip_fees_and_economics(self):
        row, candidates = self.run_reference_price_case({
            "Buy Box Amount": 0,
            "Prezzo minimo FBA Amount": -4,
            "Prezzo minimo FBM Amount": 0,
        })
        self.assertEqual(row["_Price source"], "missing_price")
        self.assertEqual(row["FBA Fee Status"], "missing_price")
        self.assertEqual(row["Economics Status"], "missing_buy_box")
        self.assertEqual(candidates, [])
        self.assertTrue(pd.isna(row["Margine attuale %"]))

    def test_margin_is_calculated_on_reference_price(self):
        row, candidates = self.run_reference_price_case({
            "Prezzo minimo FBA Amount": 20,
        })
        self.assertEqual(candidates[0]["price"], 20)
        self.assertEqual(row["Economics Status"], "ready")
        self.assertEqual(row["Margine attuale %"], 10.60)

    def test_logs_fee_result_missing(self):
        with self.assertLogs("batch_analysis", level="INFO") as captured:
            result = self.run_single_fee_case([])
        messages = "\n".join(captured.output)
        self.assertEqual(result.iloc[0]["FBA Fee Status"], "fee_result_missing")
        self.assertIn("candidates_sent=1 results_received=0", messages)
        self.assertIn("correlated=False correlation_method=none", messages)
        self.assertIn("fba_status=fee_result_missing", messages)
        self.assertIn("economics_status=missing_fba_fee", messages)

    def test_logs_amazon_fee_invalid_and_response_structure(self):
        invalid = fee_result("ASIN111")
        invalid["FeesEstimateResult"]["FeesEstimate"]["FeeDetailList"] = []
        with self.assertLogs("batch_analysis", level="INFO") as captured:
            result = self.run_single_fee_case([invalid])
        messages = "\n".join(captured.output)
        self.assertEqual(result.iloc[0]["FBA Fee Status"], "amazon_fee_invalid")
        self.assertIn("PRODUCT FEES PARSE REJECTED", messages)
        self.assertIn("FBAFees and FBAPickAndPack missing", messages)
        self.assertIn("has_FeeDetailList=True", messages)
        self.assertIn("has_FBAFees=False", messages)
        self.assertIn("fba_status=amazon_fee_invalid", messages)

    def test_logs_non_success_status_without_credentials(self):
        entry = fee_result("ASIN111")
        response = entry["FeesEstimateResult"]
        response["Status"] = "ClientError"
        response["Error"] = {
            "Code": "InvalidParameterValue",
            "Message": (
                "Rejected Authorization: Bearer secret-access-token "
                "client_secret=hunter2 refresh_token=refresh-secret"
            ),
        }
        with self.assertLogs("batch_analysis", level="INFO") as captured:
            result = self.run_single_fee_case([entry])
        messages = "\n".join(captured.output)
        self.assertEqual(result.iloc[0]["FBA Fee Status"], "amazon_fee_invalid")
        self.assertIn("status=ClientError", messages)
        self.assertIn("error_code=InvalidParameterValue", messages)
        self.assertNotIn("secret-access-token", messages)
        self.assertNotIn("hunter2", messages)
        self.assertNotIn("refresh-secret", messages)
        self.assertNotIn("test-token-must-not-be-logged", messages)

    def test_logs_response_without_seller_identifier_correlated_by_asin(self):
        entry = fee_result("ASIN111")
        del entry["FeesEstimateResult"]["FeesEstimateIdentifier"][
            "SellerInputIdentifier"
        ]
        with self.assertLogs("batch_analysis", level="INFO") as captured:
            result = self.run_single_fee_case([entry])
        messages = "\n".join(captured.output)
        self.assertEqual(result.iloc[0]["FBA Fee Status"], "FBAFees")
        self.assertIn("correlated=True correlation_method=ASIN", messages)
        self.assertIn("fba_status=FBAFees", messages)
        self.assertIn("economics_status=ready", messages)

    def test_analysis_keeps_results_when_progress_ui_fails(self):
        df_input = pd.DataFrame({
            "EAN": ["111", "222", "333"],
            "Costo": [10, 20, 30],
        })
        catalogs = {
            "111": {
                "ASIN": "ASIN111",
                "Titolo": "Prodotto 1",
                "Brand": "Brand",
                "BSR Beauty": 1000,
            },
            "222": None,
            "333": {
                "ASIN": "ASIN333",
                "Titolo": "Prodotto 3",
                "Brand": "Brand",
                "BSR Beauty": 5000,
            },
        }
        pricing = {
            "ASIN111": {
                "Buy Box": "20 EUR",
                "Venditori totali": 2,
                "Venditori FBA": 1,
            },
            "ASIN333": {
                "Buy Box": "",
                "Venditori totali": 0,
                "Venditori FBA": 0,
            },
        }
        progress_calls = []

        def failing_progress(value):
            progress_calls.append(value)
            raise RuntimeError("frontend unavailable")

        def search_fees(candidates, _token):
            return [
                fee_result(
                    candidate["asin"],
                    identifier=candidate["identifier"],
                    price=candidate["price"],
                )
                for candidate in candidates
            ]

        result = analyze_products(
            df_input=df_input,
            costo_col="Costo",
            token="token",
            search_catalog=lambda ean, _: catalogs[ean],
            search_pricing=lambda asin, _: pricing[asin],
            search_fees_batch=search_fees,
            safe_call=lambda function, *args: function(*args),
            progress_callback=failing_progress,
            throttle_seconds=0,
            fee_batch_interval_seconds=0,
            source_file="input.xlsx",
        )

        self.assertEqual(len(result), 3)
        self.assertEqual(len(progress_calls), 1)
        self.assertEqual(
            result.loc[result["EAN"] == "111", "Stato"].item(),
            "TROVATO CON OFFERTE",
        )
        self.assertEqual(
            result.loc[result["EAN"] == "222", "Stato"].item(),
            "NON TROVATO SU AMAZON",
        )
        self.assertEqual(
            result.loc[result["EAN"] == "333", "Stato"].item(),
            "TROVATO SENZA OFFERTE",
        )
        self.assertEqual(
            result.loc[result["EAN"] == "111", "Costo"].item(),
            10,
        )
        self.assertEqual(
            summarize_results(result),
            {"total": 3, "eligible": 1, "not_eligible": 2},
        )

    def test_product_fees_are_batched_at_twenty_and_rate_spaced(self):
        count = 45
        df_input = pd.DataFrame({
            "EAN": [str(index) for index in range(count)],
            "Costo": [10] * count,
        })
        batch_sizes = []
        sleeps = []

        def catalog(ean, _token):
            return {
                "ASIN": f"ASIN{ean}",
                "Titolo": ean,
                "Brand": "Brand",
                "BSR Beauty": 1000,
            }

        def pricing(_asin, _token):
            return {
                "Buy Box Amount": 20,
                "Venditori totali": 2,
                "Venditori FBA": 1,
            }

        def fees(candidates, _token):
            batch_sizes.append(len(candidates))
            return [
                fee_result(
                    candidate["asin"],
                    identifier=candidate["identifier"],
                    price=candidate["price"],
                )
                for candidate in candidates
            ]

        result = analyze_products(
            df_input,
            "Costo",
            "token",
            catalog,
            pricing,
            lambda function, *args: function(*args),
            search_fees_batch=fees,
            throttle_seconds=0,
            fee_batch_interval_seconds=2,
            sleep_func=sleeps.append,
        )
        self.assertEqual(batch_sizes, [20, 20, 5])
        self.assertEqual(sleeps, [2, 2])
        self.assertEqual(len(result), count)
        self.assertTrue(result["Margine attuale %"].notna().all())

    def test_fee_batch_failure_isolated_from_other_batches(self):
        df_input = pd.DataFrame({
            "EAN": [str(index) for index in range(21)],
            "Costo": [10] * 21,
        })
        calls = 0

        def fees(candidates, _token):
            nonlocal calls
            calls += 1
            if calls == 1:
                raise RuntimeError("rate limited")
            return [
                fee_result(
                    candidate["asin"],
                    identifier=candidate["identifier"],
                )
                for candidate in candidates
            ]

        result = analyze_products(
            df_input,
            "Costo",
            "token",
            lambda ean, _: {
                "ASIN": f"ASIN{ean}", "Titolo": ean,
                "Brand": "B", "BSR Beauty": 1000,
            },
            lambda asin, _: {
                "Buy Box Amount": 20,
                "Venditori totali": 2,
                "Venditori FBA": 1,
            },
            lambda function, *args: function(*args),
            search_fees_batch=fees,
            throttle_seconds=0,
            fee_batch_interval_seconds=0,
        )
        self.assertEqual(calls, 2)
        self.assertEqual(
            int((result["Economics Status"] == "ready").sum()),
            1,
        )
        self.assertEqual(
            int((result["FBA Fee Status"] == "fee_batch_error").sum()),
            20,
        )


class ExcelExportTests(unittest.TestCase):
    def synthetic_results(self, rows=1):
        return pd.DataFrame([
            {
                "EAN": str(index),
                "Brand": "Brand",
                "Titolo": f"Prodotto {index}",
                "Costo": 10,
                "ASIN": f"ASIN{index}",
                "Decisione": "Compra",
                "BSR Beauty": 1000,
                "Prezzo riferimento": 20,
                "Venditori FBA": 1,
                "Venditori totali": 2,
                "Margine attuale %": 10.60,
                "Prezzo 15%": 21.26,
                "Prezzo 20%": 22.89,
                "Prezzo 25%": 24.80,
                "Score": 82,
                "Opportunità": "🟢 Ottima",
                "Link Offerte Amazon": (
                    f"https://www.amazon.it/gp/offer-listing/ASIN{index}"
                ),
                "_Input order": index,
                "_Price source": "buy_box",
                "Economics Status": "ready",
                "_Economics": {
                    "status": "ready",
                    "fba_fee_net": Decimal("4.00"),
                    "fba_fee_gross": Decimal("4.88"),
                    "referral_fee": Decimal("3.00"),
                    "referral_rate": Decimal("0.15"),
                    "referral_source": "amazon_referral_fee",
                },
            }
            for index in range(rows)
        ])

    def test_excel_has_exact_columns_and_only_offers_hyperlink(self):
        results = self.synthetic_results()
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "glowup_scout_output.xlsx"
            generated_path = write_results_excel(results, output_path)
            self.assertEqual(generated_path, str(output_path))

            workbook = load_workbook(output_path)
            worksheet = workbook["Risultati"]
            data_sheet = workbook["Dati"]
            headers = [worksheet.cell(1, column).value for column in range(1, 16)]
            self.assertEqual(headers, RESULT_COLUMNS)
            self.assertNotIn("ASIN", headers)
            self.assertNotIn("Decisione", headers)
            self.assertIn("Prezzo riferimento", headers)
            self.assertNotIn("Buy Box", headers)
            self.assertEqual(len(headers), 15)
            self.assertEqual(worksheet["P1"].value, "Row ID")
            self.assertTrue(worksheet.column_dimensions["P"].hidden)
            link_column = headers.index("Link Offerte Amazon") + 1
            link_cell = worksheet.cell(row=2, column=link_column)
            self.assertEqual(
                link_cell.hyperlink.target,
                results.iloc[0]["Link Offerte Amazon"],
            )
            self.assertEqual(worksheet.freeze_panes, "A2")
            self.assertEqual(data_sheet.sheet_state, "hidden")
            self.assertEqual(workbook.sheetnames, ["Risultati", "Dati"])

    def test_interactive_formulas_technical_data_and_protection(self):
        results = self.synthetic_results()
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "interactive.xlsx"
            write_results_excel(results, output_path)
            workbook = load_workbook(output_path, data_only=False)
            results_sheet = workbook["Risultati"]
            data_sheet = workbook["Dati"]

            self.assertTrue(results_sheet.protection.sheet)
            self.assertFalse(results_sheet["D2"].protection.locked)
            self.assertTrue(results_sheet["I2"].protection.locked)
            self.assertTrue(results_sheet["D1"].protection.locked)
            self.assertTrue(results_sheet["P2"].protection.locked)
            self.assertEqual(results_sheet["P2"].value, 0)
            self.assertTrue(results_sheet.column_dimensions["P"].hidden)
            self.assertEqual(results_sheet["D2"].fill.fgColor.rgb, "00EAF2FE")
            self.assertIn("Modifica questi valori", results_sheet["D1"].comment.text)

            self.assertTrue(results_sheet["I2"].value.startswith("=IFERROR"))
            self.assertIn("D2", results_sheet["I2"].value)
            self.assertIn("INDEX('Dati'!$D$2:$D$2", results_sheet["I2"].value)
            self.assertIn("INDEX('Dati'!$F$2:$F$2", results_sheet["I2"].value)
            self.assertIn("MATCH($P2,'Dati'!$A$2:$A$2,0)", results_sheet["I2"].value)
            for cell, target in (("J2", "0.15"), ("K2", "0.2"), ("L2", "0.25")):
                self.assertTrue(results_sheet[cell].value.startswith("=IFERROR"))
                self.assertIn("D2", results_sheet[cell].value)
                self.assertIn("INDEX('Dati'!$D$2:$D$2", results_sheet[cell].value)
                self.assertIn("INDEX('Dati'!$F$2:$F$2", results_sheet[cell].value)
                self.assertIn("MATCH($P2,'Dati'!$A$2:$A$2,0)", results_sheet[cell].value)
                self.assertIn(target, results_sheet[cell].value)
            self.assertTrue(results_sheet["M2"].value.startswith("=IF("))
            self.assertIn("I2<10%", results_sheet["M2"].value)
            self.assertIn("🟢 Eccellente", results_sheet["N2"].value)

            self.assertEqual(data_sheet["B2"].value, "ASIN0")
            self.assertEqual(data_sheet["C2"].value, 4)
            self.assertEqual(data_sheet["D2"].value, 4.88)
            self.assertEqual(data_sheet["E2"].value, 3)
            self.assertEqual(data_sheet["F2"].value, 0.15)
            self.assertEqual(data_sheet["G2"].value, "amazon_referral_fee")
            self.assertEqual(data_sheet["H2"].value, "buy_box")
            self.assertEqual(data_sheet["I2"].value, "ready")

            self.assertEqual(results_sheet["I2"].number_format, "0.00%")
            self.assertEqual(results_sheet.auto_filter.ref, "A1:P2")
            self.assertEqual(results_sheet.print_area, "'Risultati'!$A$1:$O$2")
            self.assertEqual(results_sheet.freeze_panes, "A2")
            self.assertEqual(
                results_sheet["O2"].hyperlink.target,
                results.iloc[0]["Link Offerte Amazon"],
            )
            self.assertEqual(workbook.calculation.calcMode, "auto")
            self.assertTrue(workbook.calculation.fullCalcOnLoad)
            self.assertTrue(workbook.calculation.forceFullCalc)
            self.assertTrue(workbook.calculation.calcOnSave)

    def test_fallback_referral_rate_is_preserved_without_extra_vat(self):
        results = self.synthetic_results()
        economics = dict(results.at[0, "_Economics"])
        economics.update({
            "referral_fee": Decimal("3.80"),
            "referral_rate": Decimal("0.19"),
            "referral_source": "fallback_19_percent",
        })
        results.at[0, "_Economics"] = economics
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "fallback.xlsx"
            write_results_excel(results, output_path)
            workbook = load_workbook(output_path, data_only=False)
            data_sheet = workbook["Dati"]
            self.assertEqual(data_sheet["D2"].value, 4.88)
            self.assertEqual(data_sheet["E2"].value, 3.8)
            self.assertEqual(data_sheet["F2"].value, 0.19)
            self.assertEqual(data_sheet["G2"].value, "fallback_19_percent")
            self.assertNotIn("1.22", workbook["Risultati"]["I2"].value)

    def test_rows_without_economics_use_guarded_blank_formulas(self):
        results = self.synthetic_results()
        results.at[0, "_Economics"] = None
        results.at[0, "Economics Status"] = "missing_fba_fee"
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "missing.xlsx"
            write_results_excel(results, output_path)
            workbook = load_workbook(output_path, data_only=False)
            results_sheet = workbook["Risultati"]
            data_sheet = workbook["Dati"]
            self.assertIsNone(data_sheet["D2"].value)
            self.assertIsNone(data_sheet["F2"].value)
            for cell in ("I2", "J2", "K2", "L2"):
                self.assertIn("IFERROR", results_sheet[cell].value)
                self.assertIn('""', results_sheet[cell].value)
            self.assertTrue(results_sheet["M2"].value.startswith("=IF("))

    def test_formula_math_parity_with_python_business_functions(self):
        fee_estimate = {
            "fba_fee_net": Decimal("4.00"),
            "fba_fee_gross": Decimal("4.88"),
            "referral_fee": Decimal("3.00"),
            "referral_rate": Decimal("0.15"),
        }
        economics = calculate_economics(20, 10, fee_estimate)

        excel_margin_fraction = (
            (Decimal("20") - Decimal("10")
             - Decimal("20") * Decimal("0.15") - Decimal("4.88"))
            / Decimal("20")
        ).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
        self.assertEqual(
            excel_margin_fraction * Decimal("100"),
            economics["margin_percent"],
        )
        for target in (15, 20, 25):
            excel_target = (
                (Decimal("10") + Decimal("4.88"))
                / (
                    Decimal("1") - Decimal("0.15")
                    - Decimal(target) / Decimal("100")
                )
            ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            self.assertEqual(excel_target, economics["target_prices"][target])

        python_score, python_opportunity = opportunity_score(
            1000, 1, 2, economics["margin_percent"]
        )
        excel_score = 50 + 18 + 10 + 4
        self.assertEqual(excel_score, python_score)
        self.assertEqual(python_opportunity, "🟢 Ottima")

    def test_row_id_mapping_survives_structural_sort_scenarios(self):
        results = self.synthetic_results(rows=5)
        brands = ["Zeta", "Alfa", "Gamma", "Beta", "Delta"]
        costs = [8, 11, 6, 14, 9]
        bsr_values = [5000, 500, 40000, 9000, 2000]
        scores = [70, 95, 42, 60, 82]
        for index in range(5):
            results.at[index, "Brand"] = brands[index]
            results.at[index, "Costo"] = costs[index]
            results.at[index, "BSR Beauty"] = bsr_values[index]
            results.at[index, "Score"] = scores[index]
            economics = dict(results.at[index, "_Economics"])
            economics.update({
                "fba_fee_gross": Decimal("4.50") + Decimal(index) / 10,
                "referral_rate": Decimal("0.11") + Decimal(index) / 100,
            })
            results.at[index, "_Economics"] = economics

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "sort-safe.xlsx"
            write_results_excel(results, output_path)
            workbook = load_workbook(output_path, data_only=False)
            sheet = workbook["Risultati"]
            data_sheet = workbook["Dati"]

            technical = {
                data_sheet.cell(row, 1).value: (
                    data_sheet.cell(row, 4).value,
                    data_sheet.cell(row, 6).value,
                )
                for row in range(2, data_sheet.max_row + 1)
            }
            scores_by_row_id = dict(
                zip(results["_Input order"], results["Score"])
            )
            visible_rows = [
                {
                    "brand": sheet.cell(row, 2).value,
                    "cost": sheet.cell(row, 4).value,
                    "bsr": sheet.cell(row, 5).value,
                    "row_id": sheet.cell(row, 16).value,
                    "score": scores_by_row_id[sheet.cell(row, 16).value],
                }
                for row in range(2, sheet.max_row + 1)
            ]

            scenarios = (
                ("before_sort", lambda item: visible_rows.index(item), False),
                ("score_asc", lambda item: item["score"], False),
                ("score_desc", lambda item: item["score"], True),
                ("bsr", lambda item: item["bsr"], False),
                ("brand", lambda item: item["brand"], False),
            )
            for name, key, reverse in scenarios:
                sorted_rows = sorted(visible_rows, key=key, reverse=reverse)
                with self.subTest(name=name):
                    for product in sorted_rows:
                        fba_gross, referral_rate = technical[product["row_id"]]
                        original = results.loc[
                            results["_Input order"] == product["row_id"]
                        ].iloc[0]["_Economics"]
                        self.assertEqual(
                            Decimal(str(fba_gross)), original["fba_fee_gross"]
                        )
                        self.assertEqual(
                            Decimal(str(referral_rate)), original["referral_rate"]
                        )

            sorted_rows = sorted(visible_rows, key=lambda item: item["score"])
            edited = dict(sorted_rows[0])
            edited["cost"] = Decimal("5.25")
            fba_gross, referral_rate = technical[edited["row_id"]]
            margin = (
                Decimal("20") - edited["cost"]
                - Decimal("20") * Decimal(str(referral_rate))
                - Decimal(str(fba_gross))
            ) / Decimal("20")
            self.assertIsInstance(margin, Decimal)
            self.assertIn("MATCH($P2", sheet["I2"].value)
            self.assertEqual(sheet.auto_filter.ref, "A1:P6")
            self.assertTrue(sheet.column_dimensions["P"].hidden)

    def test_synthetic_6700_row_export_without_dataframe_rendering(self):
        results = self.synthetic_results(rows=6700)
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "large.xlsx"
            write_results_excel(results, output_path)
            workbook = load_workbook(output_path, read_only=True)
            worksheet = workbook["Risultati"]
            self.assertEqual(worksheet.max_row, 6701)
            self.assertEqual(worksheet.max_column, 16)
            visible_headers = [
                worksheet.cell(1, column).value for column in range(1, 16)
            ]
            self.assertEqual(visible_headers, RESULT_COLUMNS)

        app_source = (Path(__file__).parents[1] / "app_glowup.py").read_text()
        self.assertNotIn("st.dataframe(df_results", app_source)


if __name__ == "__main__":
    unittest.main()
