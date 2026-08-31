import sqlite3
import tempfile
import tracemalloc
import unittest
import zipfile
from pathlib import Path
from unittest.mock import Mock, patch

from openpyxl import load_workbook

from discovery_finalize_worker import (
    export_offline,
    export_operational_offline,
    finalization_state,
    finalize,
)
from discovery_excel import (
    EXCEL_MAX_HYPERLINKS_PER_WORKSHEET,
    validate_excel_compatibility,
)
from discovery_incremental import (
    DiscoveryIncrementalStore,
    IncrementalCandidateCollection,
    LightweightCheckpointStore,
)
from discovery_jobs import DiscoveryJobRegistry
from discovery_resources import ResourcePause, ResourceSnapshot


def completed_fixture(job_id="job"):
    observation_id = "obs-1"
    scenario = {
        "scenario_id": "scenario-1", "supplier": "umma",
        "scenario_label": "U-Quick", "cost_gross_unit_eur": "8.06",
        "currency": "EUR", "minimum_product_quantity": 1,
    }
    combination = {
        "combination_id": "combination-1", "scenario_id": "scenario-1",
        "amazon_observation_id": observation_id, "asin": "B012345678",
        "cost_gross_unit_eur": "8.06", "price_reference": "22.90",
        "profit": "5.61", "margin_percent": "24.50", "score": 81,
        "evaluation_status": "margin_passed",
    }
    product = {
        "canonical_ean": "8809562191179", "gtin": "8809562191179",
        "product_key": "product-1", "identifier_type": "EAN",
        "brand": "Arencia", "title": "Cleanser", "catalog_status": "resolved",
        "amazon_offers_url": "https://www.amazon.it/dp/B012345678",
        "is_final_result": True,
        "scenario_roles": {"scenario_raccomandato": "scenario-1"},
        "combination_roles": {"recommended_combination": "combination-1"},
        "recommended_combination": combination,
        "scenarios": [scenario],
        "amazon_listings": [{
            "asin": "B012345678", "amazon_observation_id": observation_id,
            "compatibility_status": "compatible", "beauty_status": "display_group_beauty",
            "bsr_beauty": 5000, "reference_price": "22.90",
            "price_source": "buy_box", "fba_sellers": 1, "total_sellers": 2,
            "pricing_status": "success", "competition_status": "passed",
            "fee_status": "valid", "evaluation_status": "margin_passed",
        }],
        "opportunity_combinations": [combination],
    }
    observation = {
        "observation_id": observation_id, "asin": "B012345678",
        "amazon_title": "Cleanser", "amazon_brand": "Arencia",
        "bsr_beauty": 5000, "reference_price": "22.90",
        "fba_sellers": 1, "total_sellers": 2, "fee_status": "valid",
        "fee_estimate": {
            "fba_fee_net": "4.00", "fba_fee_gross": "4.88",
            "referral_fee": "3.44", "referral_rate": "0.15",
        },
        "diagnostics": {"product_keys": ["product-1"]},
    }
    state = {
        "job_id": job_id, "status": "completed", "phase": "completed",
        "selected_count": 1, "sampled_identifier_count": 1,
        "progress_current": 1, "progress_total": 1,
        "catalog_completed_count": 1, "catalog_pending_count": 0,
        "selected_suppliers": ["umma"],
        "filters": {
            "bsr_min": 1, "bsr_max": 18000, "max_fba_sellers": 15,
            "max_total_sellers": 15, "minimum_margin": 20,
        },
        "completed_at": "2026-08-28T10:50:26Z", "final_products": 1,
        "fee_target_count": 1, "fee_valid_count": 1,
        "fee_unavailable_count": 0, "fee_pending_count": 0,
    }
    return state, product, observation


class FinalizationTests(unittest.TestCase):
    def setUp(self):
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.store = DiscoveryIncrementalStore(self.root / "incremental.sqlite3")
        self.checkpoints = LightweightCheckpointStore(self.root / "checkpoints")
        self.registry = DiscoveryJobRegistry(self.root / "jobs.sqlite3")
        self.state, product, observation = completed_fixture()
        self.store.create_job(self.state, [product])
        self.store.upsert_observations("job", [observation])
        self.store.update_candidates(
            "job", [product], phase="completed", replace_scenarios=True,
        )
        self.store.set_phase("job", "completed", status="completed")
        self.checkpoints.save(self.state)
        self.registry.register_checkpoint(self.state)

    def tearDown(self):
        self.temporary.cleanup()

    def test_connection_context_really_closes_sqlite_handle(self):
        with self.store._connect() as connection:
            self.assertEqual(connection.execute("SELECT 1").fetchone()[0], 1)
        with self.assertRaises(sqlite3.ProgrammingError):
            connection.execute("SELECT 1")

    def test_notification_summary_is_bounded_and_authoritative(self):
        summary = self.store.notification_summary("job")
        self.assertEqual(summary["selected_count"], 1)
        self.assertEqual(summary["final_opportunity_count"], 1)
        self.assertEqual(summary["combination_count"], 1)
        self.assertEqual(summary["fee_target_count"], 1)
        self.assertEqual(summary["fee_valid_count"], 1)
        self.assertEqual(summary["fee_unavailable_count"], 0)
        self.assertEqual(summary["bsr_passed_count"], 1)
        self.assertEqual(summary["best_opportunity"]["supplier"], "umma")
        self.assertEqual(summary["best_opportunity"]["score"], 81)

    def test_finalization_state_includes_persisted_notification_projection(self):
        state = finalization_state("job", self.store, self.checkpoints)
        self.assertNotIn("results", state)
        self.assertEqual(state["final_opportunity_count"], 1)
        self.assertEqual(state["best_opportunity"]["asin"], "B012345678")

    def test_export_iterator_hydrates_combinations_and_closes_every_batch(self):
        opened = []
        original = self.store._new_connection

        def tracked():
            connection = original()
            opened.append(connection)
            return connection

        self.store._new_connection = tracked
        products = list(self.store.iter_export_candidates("job", batch_size=1))
        self.assertEqual(len(products[0]["opportunity_combinations"]), 1)
        self.assertEqual(len(products[0]["amazon_observations"]), 1)
        for connection in opened:
            with self.assertRaises(sqlite3.ProgrammingError):
                connection.execute("SELECT 1")

    def test_offline_streaming_export_preserves_workbook_contract(self):
        target = self.root / "offline.xlsx"
        result = export_offline(
            "job", target, store=self.store, checkpoints=self.checkpoints,
        )
        self.assertEqual(result["export_state"]["status"], "completed")
        workbook = load_workbook(target, read_only=False, data_only=False)
        self.assertEqual(workbook.sheetnames, [
            "Opportunità", "Tutti i risultati", "Listing Amazon",
            "Scenari", "Dati", "Parametri run",
        ])
        self.assertEqual(workbook["Opportunità"].max_row, 2)
        self.assertEqual(workbook["Scenari"].max_row, 2)
        self.assertTrue(str(workbook["Opportunità"]["M2"].value).startswith("="))
        self.assertTrue(str(workbook["Opportunità"]["N2"].value).startswith("="))
        self.assertEqual(workbook["Opportunità"]["G2"].data_type, "n")
        self.assertEqual(workbook["Opportunità"]["J2"].data_type, "n")
        self.assertEqual(workbook["Opportunità"]["G2"].number_format, '€ #,##0.00')
        self.assertFalse(workbook["Opportunità"]["G2"].protection.locked)
        self.assertTrue(workbook["Opportunità"]["N2"].protection.locked)
        self.assertFalse(workbook["Opportunità"].protection.autoFilter)
        self.assertFalse(workbook["Opportunità"].protection.sort)
        self.assertFalse(workbook["Opportunità"].protection.formatColumns)
        for sheet_name, coordinates in {
            "Scenari": ("I2", "J2"),
            "Tutti i risultati": ("K2", "M2", "N2", "Q2"),
            "Listing Amazon": ("K2", "L2", "M2"),
            "Dati": ("E2", "H2", "I2", "J2", "K2"),
        }.items():
            for coordinate in coordinates:
                self.assertEqual(
                    workbook[sheet_name][coordinate].data_type, "n",
                    f"{sheet_name}!{coordinate} must remain numeric",
                )
        self.assertEqual(workbook["Dati"].sheet_state, "hidden")
        self.assertEqual(workbook["Opportunità"].freeze_panes, "A2")
        listing_link = workbook["Listing Amazon"]["V2"]
        self.assertEqual(listing_link.data_type, "f")
        self.assertTrue(str(listing_link.value).startswith('=HYPERLINK("https://'))
        self.assertIsNone(listing_link.hyperlink)
        compatibility = validate_excel_compatibility(target)
        self.assertTrue(all(
            count <= EXCEL_MAX_HYPERLINKS_PER_WORKSHEET
            for count in compatibility["hyperlinks_by_part"].values()
        ))
        workbook.close()

    def test_operational_export_is_minimal_and_keeps_editable_economics(self):
        target = self.root / "job.operational.xlsx"
        result = export_operational_offline(
            "job", target, store=self.store, checkpoints=self.checkpoints,
        )
        self.assertEqual(result["operational_export"]["status"], "completed")
        workbook = load_workbook(target, read_only=False, data_only=False)
        self.assertEqual(
            workbook.sheetnames, ["Opportunità", "Dati", "Parametri run"],
        )
        self.assertEqual(workbook["Opportunità"].max_row, 2)
        self.assertEqual(workbook["Dati"].max_row, 2)
        self.assertEqual(workbook["Dati"].sheet_state, "hidden")
        self.assertEqual(workbook["Opportunità"]["G2"].data_type, "n")
        self.assertFalse(workbook["Opportunità"]["G2"].protection.locked)
        for coordinate in ("M2", "N2", "O2", "P2", "Q2", "R2", "S2"):
            cell = workbook["Opportunità"][coordinate]
            self.assertEqual(cell.data_type, "f")
            self.assertTrue(cell.protection.locked)
        self.assertIn("G2", workbook["Opportunità"]["N2"].value)
        self.assertIn("Dati", workbook["Opportunità"]["R2"].value)
        self.assertIsNotNone(workbook["Opportunità"]["V2"].hyperlink)
        self.assertFalse(workbook["Opportunità"].protection.autoFilter)
        self.assertFalse(workbook["Opportunità"].protection.sort)
        workbook.close()
        compatibility = validate_excel_compatibility(target)
        self.assertEqual(sum(compatibility["hyperlinks_by_part"].values()), 1)

    def test_excel_compatibility_rejects_worksheet_hyperlink_limit(self):
        target = self.root / "too-many-hyperlinks.xlsx"
        relationships = "".join(
            '<Relationship Id="rId{0}" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink" '
            'Target="https://example.test/{0}" TargetMode="External"/>'.format(index)
            for index in range(1, EXCEL_MAX_HYPERLINKS_PER_WORKSHEET + 2)
        )
        with zipfile.ZipFile(target, "w", compression=zipfile.ZIP_DEFLATED) as package:
            package.writestr(
                "xl/worksheets/_rels/sheet1.xml.rels",
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                f"{relationships}</Relationships>",
            )
        with self.assertRaisesRegex(ValueError, "hyperlink limit exceeded"):
            validate_excel_compatibility(target)

    def test_finalizer_is_restart_safe_and_does_not_rewrite_valid_export(self):
        target = self.root / "final.xlsx"
        first = finalize(
            "job", registry=self.registry, store=self.store,
            checkpoints=self.checkpoints, send_notification=False,
            output_path=target,
        )
        first_mtime = target.stat().st_mtime_ns
        second = finalize(
            "job", registry=self.registry, store=self.store,
            checkpoints=self.checkpoints, send_notification=False,
            output_path=target,
        )
        self.assertEqual(first["export_state"]["sha256"], second["export_state"]["sha256"])
        self.assertEqual(first["operational_export"], first["export_state"])
        self.assertEqual(target.stat().st_mtime_ns, first_mtime)
        runtime = self.registry.get("job")
        self.assertEqual(runtime["status"], "completed")
        self.assertIsNone(runtime["worker_pid"])
        self.assertIsNone(runtime["lease_expires_at"])
        self.assertEqual(list(self.root.glob(".*.xlsx")), [])

    def test_resource_pause_is_export_specific_and_resumable(self):
        snapshot = ResourceSnapshot(
            rss_bytes=2_000_000_000, available_memory_bytes=1_000_000_000,
            swap_used_bytes=0, disk_free_bytes=100_000_000_000,
            database_bytes=1, wal_bytes=0, write_bytes=0,
            write_rate_bytes_per_second=0, db_latency_ms=1,
            observed_monotonic=1,
        )
        governor = Mock()
        governor.before_next_batch.side_effect = ResourcePause(
            "rss_hard", snapshot, 1_280_000_000,
        )
        result = finalize(
            "job", registry=self.registry, store=self.store,
            checkpoints=self.checkpoints, governor=governor,
            send_notification=False, output_path=self.root / "paused.xlsx",
        )
        self.assertEqual(result["status"], "export_resource_paused")
        self.assertEqual(self.registry.get("job")["status"], "export_resource_paused")
        self.assertFalse((self.root / "paused.xlsx").exists())

    def test_launch_finalizer_uses_clean_detached_interpreter(self):
        process = Mock(pid=4242)
        with patch("discovery_jobs.subprocess.Popen", return_value=process) as popen:
            pid = self.registry.launch_finalizer("job")
        self.assertEqual(pid, 4242)
        command = popen.call_args.args[0]
        self.assertTrue(command[1].endswith("discovery_finalize_worker.py"))
        self.assertTrue(popen.call_args.kwargs["start_new_session"])

    def test_large_20k_combination_export_remains_streaming(self):
        large_store = DiscoveryIncrementalStore(self.root / "large.sqlite3")
        large_checkpoints = LightweightCheckpointStore(self.root / "large-checkpoints")
        state, template, observation = completed_fixture("large")
        state.update({"selected_count": 200, "sampled_identifier_count": 200,
                      "catalog_completed_count": 200, "final_products": 143})

        def products():
            for index in range(200):
                row = {**template}
                row["canonical_ean"] = row["gtin"] = f"{index:013d}"
                row["product_key"] = f"product-{index}"
                row["is_final_result"] = False
                row["scenarios"] = [{**template["scenarios"][0]}]
                row["amazon_listings"] = [{**template["amazon_listings"][0]}]
                row["opportunity_combinations"] = []
                yield row

        large_store.create_job(state, products())
        large_store.upsert_observations("large", [observation])
        for index, row in enumerate(products()):
            row["opportunity_combinations"] = [{
                **template["opportunity_combinations"][0],
                "combination_id": f"combination-{index}-{item}",
            } for item in range(100)]
            row["is_final_result"] = index < 143
            row["recommended_combination"] = row["opportunity_combinations"][0]
            large_store.update_candidates("large", [row], replace_scenarios=True)
        large_store.set_phase("large", "completed", status="completed")
        large_checkpoints.save(state)
        target = self.root / "large.xlsx"
        tracemalloc.start()
        export_offline("large", target, store=large_store, checkpoints=large_checkpoints)
        _, peak = tracemalloc.get_traced_memory()
        tracemalloc.stop()
        self.assertLess(peak, 250 * 1024 * 1024)
        workbook = load_workbook(target, read_only=True, data_only=False)
        self.assertEqual(sum(1 for _ in workbook["Opportunità"].iter_rows()), 144)
        self.assertEqual(sum(1 for _ in workbook["Scenari"].iter_rows()), 20_001)
        workbook.close()


if __name__ == "__main__":
    unittest.main()
