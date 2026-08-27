import json
import tempfile
import unittest
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from streamlit.testing.v1 import AppTest

from abw_discovery import normalize_abw_candidates
from discovery import (
    DISCOVERY_SCHEMA_VERSION,
    DiscoveryCheckpointStore,
    _evaluate_product_scenarios,
    _select_recommended_combination,
    default_filters,
    discovery_funnel_view,
    run_discovery,
)
from discovery_amazon import correlate_catalog_items
from discovery_excel import write_discovery_excel
from purchase_scenarios import (
    merge_product_candidates,
    normalize_purchase_scenario,
    recommended_scenario,
    scenario_requirement_label,
)
from qogita_discovery import normalize_qogita_candidates
from qudo_discovery import (
    normalize_qudo_candidates,
    snapshot_is_stale,
    valid_qudo_identifier,
)
from tests.testsuite_abw_discovery import real_0748_rows
from tests.testsuite_discovery import fee_result, qogita_row
from tests.testsuite_multi_listing import EAN as MULTI_EAN, catalog_item
from tests.testsuite_umma_discovery import amazon_observation, standard_row
from umma_discovery import normalize_umma_candidates


EAN = "8809532220748"
NOW = datetime(2026, 8, 24, 12, 0, tzinfo=timezone.utc)


def qudo_row(
    *, ean=EAN, run_id="3a2938ae-08aa-46d9-9772-ad77968957c5",
    observed_at="2026-08-24T04:06:08.597815Z", net="8.0000",
    stock=309, availability="in_stock", minimum=1, selling_unit=1,
    account_mov="300.0000", account_currency="EUR",
    product_id="892", offer_id="47062", supplier_sku="QUDO-HARU-15_1",
    latest_attempt_status="success", latest_attempt_at=None,
):
    return {
        "run_id": run_id, "seller_sku": "HARU06", "gtin": ean,
        "supplier": "QUDO", "supplier_product_id": product_id,
        "supplier_offer_id": offer_id, "supplier_sku": supplier_sku,
        "product_name": "Haruharu Wonder - Black Bamboo Mist, 150ml",
        "observed_at": observed_at, "currency": "EUR", "unit_price": net,
        "price_basis": "net_ex_vat", "pricing_scope": "public_catalog",
        "available_quantity": stock, "availability_status": availability,
        "minimum_product_quantity": minimum, "selling_unit": selling_unit,
        "minimum_order_value": account_mov,
        "minimum_order_currency": account_currency,
        "product_url": "https://qudobeauty.com/product/haruharu-wonder-black-bamboo-mist-150ml/",
        "source": "qudo_woocommerce_store_api",
        "latest_attempt_status": latest_attempt_status,
        "latest_attempt_at": latest_attempt_at or observed_at,
    }


def normalized_qudo(*, rows=None, now=NOW):
    candidates, diagnostics = normalize_qudo_candidates(
        rows or [qudo_row()], now=now
    )
    return candidates[0], diagnostics


class QudoAdapterTests(unittest.TestCase):
    def test_same_gtin_different_supplier_products_keep_distinct_scenarios(self):
        first = qudo_row()
        second = qudo_row(
            product_id="893", offer_id="47063", supplier_sku="QUDO-HARU-15_2",
        )
        candidates, diagnostics = normalize_qudo_candidates([first, second], now=NOW)
        self.assertEqual(len(candidates), 2)
        self.assertEqual(diagnostics["qudo_scenarios"], 2)
        self.assertEqual(
            len({row["scenarios"][0]["scenario_id"] for row in candidates}), 2,
        )

    def test_real_qudo_scenario_maps_identity_price_vat_stock_and_mov(self):
        candidate, diagnostics = normalized_qudo()
        self.assertEqual(candidate["canonical_ean"], EAN)
        self.assertEqual(candidate["identifier_type"], "EAN")
        self.assertEqual(len(candidate["scenarios"]), 1)
        scenario = candidate["scenarios"][0]
        self.assertEqual(scenario["supplier"], "qudo")
        self.assertEqual(scenario["scenario_type"], "qudo_standard")
        self.assertEqual(scenario["scenario_label"], "Qudo")
        self.assertEqual(scenario["supplier_product_id"], "892")
        self.assertEqual(scenario["supplier_offer_id"], "47062")
        self.assertEqual(scenario["supplier_sku"], "QUDO-HARU-15_1")
        self.assertEqual(scenario["cost_net_unit_eur"], Decimal("8"))
        self.assertEqual(scenario["vat_rate"], Decimal("0.22"))
        self.assertEqual(scenario["vat_amount_unit"], Decimal("1.76"))
        self.assertEqual(scenario["cost_gross_unit_eur"], Decimal("9.76"))
        self.assertEqual(scenario["stock"], 309)
        self.assertEqual(scenario["minimum_product_quantity"], 1)
        self.assertEqual(scenario["selling_unit"], 1)
        self.assertEqual(scenario["maximum_product_quantity"], 309)
        self.assertEqual(scenario["account_mov"], Decimal("300"))
        self.assertEqual(scenario["account_mov_currency"], "EUR")
        self.assertEqual(scenario["account_mov_eur"], Decimal("300"))
        self.assertNotEqual(scenario["account_mov"], scenario["minimum_product_quantity"])
        self.assertEqual(scenario["availability_status"], "in_stock")
        self.assertEqual(scenario["freshness_status"], "fresh")
        self.assertIn("qudobeauty.com", scenario["product_url"])
        self.assertEqual(diagnostics["qudo_products"], 1)
        self.assertEqual(diagnostics["qudo_scenarios"], 1)

    def test_invalid_ean_price_currency_and_mov_are_rejected(self):
        self.assertTrue(valid_qudo_identifier(EAN))
        self.assertFalse(valid_qudo_identifier(EAN[:-1] + "9"))
        cases = (
            (qudo_row(ean=EAN[:-1] + "9"), "invalid_identifier"),
            (qudo_row(net="0"), "invalid_price_or_currency"),
            (dict(qudo_row(), currency="USD"), "invalid_price_or_currency"),
            (qudo_row(account_mov="0"), "invalid_account_mov"),
            (qudo_row(account_currency="USD"), "invalid_account_mov"),
        )
        for row, reason in cases:
            with self.subTest(reason=reason):
                candidates, diagnostics = normalize_qudo_candidates([row], now=NOW)
                self.assertEqual(candidates, [])
                self.assertEqual(diagnostics["rejected_scenarios"][0]["reason"], reason)

    def test_stock_availability_and_quantities_are_enforced(self):
        cases = (
            (qudo_row(stock=0), "unavailable"),
            (qudo_row(availability="out_of_stock"), "unavailable"),
            (qudo_row(minimum=0), "invalid_quantity_requirement"),
            (qudo_row(selling_unit=0), "invalid_quantity_requirement"),
            (qudo_row(stock=4, minimum=5), "insufficient_stock"),
            (qudo_row(stock=4, selling_unit=5), "insufficient_stock"),
        )
        for row, reason in cases:
            with self.subTest(reason=reason):
                candidates, diagnostics = normalize_qudo_candidates([row], now=NOW)
                self.assertEqual(candidates, [])
                self.assertEqual(diagnostics["rejected_scenarios"][0]["reason"], reason)

    def test_scenario_id_is_stable_across_price_stock_and_snapshot_changes(self):
        first, _ = normalized_qudo()
        second, _ = normalized_qudo(rows=[qudo_row(
            run_id="new-run", observed_at="2026-08-24T05:00:00Z",
            net="7.50", stock=500,
        )])
        self.assertEqual(
            first["scenarios"][0]["scenario_id"],
            second["scenarios"][0]["scenario_id"],
        )
        other, _ = normalized_qudo(rows=[qudo_row(offer_id="other-offer")])
        self.assertNotEqual(
            first["scenarios"][0]["scenario_id"],
            other["scenarios"][0]["scenario_id"],
        )

    def test_ttl_boundary_and_later_failure_match_manager_semantics(self):
        self.assertFalse(snapshot_is_stale(NOW - timedelta(hours=47, minutes=59), now=NOW))
        self.assertFalse(snapshot_is_stale(NOW - timedelta(hours=48), now=NOW))
        self.assertTrue(snapshot_is_stale(NOW - timedelta(hours=48, seconds=1), now=NOW))
        stale, diagnostics = normalize_qudo_candidates([
            qudo_row(observed_at=(NOW - timedelta(hours=48, seconds=1)).isoformat())
        ], now=NOW)
        self.assertEqual(stale, [])
        self.assertEqual(diagnostics["qudo_stale_scenarios"], 1)
        self.assertEqual(diagnostics["rejected_scenarios"][0]["reason"], "stale_snapshot")
        failed, diagnostics = normalize_qudo_candidates([qudo_row(
            observed_at="2026-08-24T04:00:00Z",
            latest_attempt_status="failed", latest_attempt_at="2026-08-24T05:00:00Z",
        )], now=NOW)
        self.assertEqual(failed, [])
        self.assertEqual(diagnostics["qudo_stale_scenarios"], 1)

    def test_checkpoint_round_trip_preserves_numeric_and_qudo_fields(self):
        candidate, _ = normalized_qudo()
        encoded = json.loads(json.dumps(candidate, default=str))
        scenario = normalize_purchase_scenario(encoded["scenarios"][0])
        self.assertIsInstance(scenario["cost_net_unit_eur"], Decimal)
        self.assertIsInstance(scenario["cost_gross_unit_eur"], Decimal)
        self.assertIsInstance(scenario["account_mov"], Decimal)
        self.assertEqual(scenario["stock"], 309)
        self.assertEqual(scenario["supplier_sku"], "QUDO-HARU-15_1")
        self.assertIn("qudobeauty.com", scenario["product_url"])

    def test_requirement_label_keeps_mov_separate_from_quantity(self):
        candidate, _ = normalized_qudo()
        label = scenario_requirement_label(candidate["scenarios"][0])
        self.assertEqual(label, "Min. 1 pz · MOV EUR 300")


class QudoCrossSupplierTests(unittest.TestCase):
    def full_cross_supplier_candidate(self):
        qogita, _ = normalize_qogita_candidates([
            qogita_row(gtin=EAN, mov=mov, price=price, observed_at="2026-08-24T04:00:00Z")
            for mov, price in (
                (500, "6.94"), (1500, "6.83"), (5000, "6.80"),
                (10000, "6.73"), (15000, "6.73"),
            )
        ], now=NOW)
        umma, _ = normalize_umma_candidates([
            standard_row(ean=EAN, observed_at="2026-08-24T04:00:00Z")
        ], now=NOW)
        abw, _ = normalize_abw_candidates(real_0748_rows(), now=NOW)
        qudo, _ = normalize_qudo_candidates([qudo_row()], now=NOW)
        return merge_product_candidates(qogita, umma, abw, qudo)

    def test_real_0748_merges_four_suppliers_into_one_product_and_13_scenarios(self):
        merged = self.full_cross_supplier_candidate()
        self.assertEqual(len(merged), 1)
        self.assertEqual(len(merged[0]["scenarios"]), 13)
        self.assertEqual(
            {row["supplier"] for row in merged[0]["scenarios"]},
            {"qogita", "umma", "abw", "qudo"},
        )

    def test_qudo_economics_score_and_below_threshold_do_not_remove_others(self):
        candidate = self.full_cross_supplier_candidate()[0]
        observation = amazon_observation()
        passed = _evaluate_product_scenarios(candidate, observation, 15)
        qudo = next(row for row in candidate["scenarios"] if row["supplier"] == "qudo")
        self.assertEqual(qudo["economics_status"], "ready")
        self.assertIsInstance(qudo["economics"]["profit"], Decimal)
        self.assertIsInstance(qudo["margin_percent"], float)
        self.assertEqual(set(qudo["economics"]["target_prices"]), {"15", "20", "25"})
        self.assertIsInstance(qudo["score"], int)
        self.assertTrue(qudo["opportunity"])
        self.assertTrue(passed)
        self.assertEqual(len(candidate["scenarios"]), 13)

    def test_cross_supplier_tie_does_not_compare_incompatible_requirements(self):
        scenarios = {
            "qogita": {"scenario_id": "qogita", "supplier": "qogita", "account_mov": 1},
            "qudo": {
                "scenario_id": "qudo", "supplier": "qudo",
                "account_mov": 999999, "minimum_product_quantity": 1,
            },
        }
        rows = [{
            "combination_id": key, "scenario_id": key, "supplier": key,
            "asin": "B000000001", "score": 70,
            "margin_percent": Decimal("20"), "profit": Decimal("3"),
            "cost_gross_unit_eur": Decimal("8"),
        } for key in ("qogita", "qudo")]
        first = _select_recommended_combination(rows, scenarios)
        scenarios["qogita"]["account_mov"] = 999999999
        scenarios["qudo"]["minimum_product_quantity"] = 999999999
        second = _select_recommended_combination(rows, scenarios)
        self.assertEqual(first["scenario_id"], second["scenario_id"])
        self.assertEqual(first["scenario_id"], "qogita")

    def test_funnel_exposes_qudo_without_breaking_existing_suppliers(self):
        view = discovery_funnel_view({"funnel": {
            "qogita_products": 1, "qogita_scenarios": 5,
            "umma_products": 1, "umma_scenarios": 1,
            "abw_products": 1, "abw_scenarios": 6,
            "qudo_products": 1, "qudo_scenarios": 1,
            "qudo_stale_scenarios": 0,
            "supplier_products_total": 1, "supplier_scenarios_total": 13,
        }})
        self.assertEqual(view["suppliers"]["qudo_products"], 1)
        self.assertEqual(view["suppliers"]["qudo_scenarios"], 1)
        self.assertEqual(view["purchase_scenarios"]["supplier_scenarios_total"], 13)


class QudoIntegrationTests(unittest.TestCase):
    def test_qudo_scenario_does_not_multiply_multilisting_amazon_calls(self):
        candidate, diagnostics = normalized_qudo(rows=[qudo_row(ean=MULTI_EAN)])
        candidate.update({
            "brand": "ANUA", "title": "ANUA Niacinamide Serum 30ml",
            "package_quantity": 1, "volume_value": 30, "volume_unit": "ml",
        })
        calls = {"catalog": [], "pricing": [], "fees": []}

        def catalog(gtins, _job_id, products):
            calls["catalog"].append(list(gtins))
            return correlate_catalog_items(
                gtins, [catalog_item("B0CVFPT7FC"), catalog_item("B0DL92QM79")], products
            )

        def pricing(asins, _job_id):
            calls["pricing"].append(list(asins))
            return {asin: {
                "status": "success", "Venditori FBA": 2,
                "Venditori totali": 4,
                "Seller count source": "summary_number_of_offers",
                "reference_price": 30, "price_source": "buy_box",
            } for asin in asins}

        def fees(requests_, _token):
            calls["fees"].append([row["asin"] for row in requests_])
            return [fee_result(row["asin"], row["identifier"]) for row in requests_]

        with tempfile.TemporaryDirectory() as directory:
            state = run_discovery(
                default_filters(), checkpoint_store=DiscoveryCheckpointStore(directory),
                catalog_batch=catalog, pricing_batch=pricing, fees_batch=fees,
                token_provider=type("Tokens", (), {
                    "get": lambda self: "token", "invalidate": lambda self: None,
                })(),
                qogita_loader=lambda: [qogita_row(observed_at="2026-08-24T10:00:00Z")],
                qogita_normalizer=lambda _rows, minimum_stock: (
                    [candidate], {
                        "initial": 1, "valid_gtin": 1,
                        "qogita_products": 0, "qogita_scenarios": 0,
                        **diagnostics,
                    },
                ),
                qogita_refresher=lambda _: (_ for _ in ()).throw(AssertionError("fresh")),
                now_provider=lambda: NOW, sleep_func=lambda _: None,
                catalog_batch_interval=0, pricing_batch_interval=0,
                fee_batch_interval=0,
            )
        self.assertEqual(len(state["results"][0]["opportunity_combinations"]), 2)
        self.assertEqual(calls["catalog"], [[MULTI_EAN]])
        self.assertEqual(len(calls["pricing"]), 1)
        self.assertEqual(set(calls["pricing"][0]), {"B0CVFPT7FC", "B0DL92QM79"})
        self.assertEqual(len(calls["fees"]), 1)
        self.assertEqual(set(calls["fees"][0]), {"B0CVFPT7FC", "B0DL92QM79"})
        self.assertEqual(state["funnel"]["qudo_products"], 1)
        self.assertEqual(state["funnel"]["qudo_scenarios"], 1)

    def test_ui_renders_qudo_offline_without_api_calls(self):
        candidate, diagnostics = normalized_qudo()
        observation = amazon_observation()
        _evaluate_product_scenarios(candidate, observation, 0)
        candidate["amazon_observation"] = observation
        candidate["amazon_offers_url"] = "https://www.amazon.it/gp/offer-listing/B000000001"
        state = {
            "discovery_schema_version": DISCOVERY_SCHEMA_VERSION,
            "job_id": "qudo-ui", "status": "completed", "phase": "completed",
            "results": [candidate], "candidates": [candidate],
            "amazon_observations": [observation],
            "qogita_refresh_status": "cache_fresh", "qogita_snapshot_after": {},
            "funnel": {
                **diagnostics, "amazon_found": 1, "beauty_valid": 1,
                "bsr_passed": 1, "competition_passed": 1, "fee_valid": 1,
                "final_opportunities": 1, "scenarios_evaluated": 1,
                "scenarios_margin_passed": 1,
                "scenarios_margin_below_threshold": 0,
            },
        }
        app = AppTest.from_file("app_glowup.py", default_timeout=20).run()
        app.session_state["ui_state"] = "discovery_result"
        app.session_state["discovery_status"] = "completed"
        app.session_state["discovery_result"] = {
            "state": json.loads(json.dumps(state, default=str)),
            "output_bytes": b"workbook",
        }
        app.run()
        self.assertFalse(app.exception)
        self.assertTrue(any("QUDO" in element.value for element in app.markdown))
        next(button for button in app.button if button.label == "Confronta scenari").click().run()
        self.assertFalse(app.exception)
        table = app.dataframe[0].value
        self.assertEqual(len(table), 1)
        self.assertEqual(table.loc[0, "Fornitore"], "QUDO")
        self.assertEqual(table.loc[0, "Stock"], "309")
        self.assertIn("MOV EUR 300", table.loc[0, "Requisito"])

    def test_excel_represents_qudo_with_editable_dynamic_cost(self):
        candidate, _ = normalized_qudo()
        observation = amazon_observation()
        candidate["amazon_observation"] = observation
        candidate["amazon_observations"] = [observation]
        _evaluate_product_scenarios(candidate, observation, 0)
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "qudo.xlsx"
            write_discovery_excel([candidate], path)
            workbook = load_workbook(path, data_only=False)
        scenarios = workbook["Scenari"]
        headers = {cell.value: cell.column for cell in scenarios[1]}
        self.assertEqual(scenarios.max_row, 2)
        self.assertEqual(scenarios.cell(2, headers["Fornitore"]).value, "QUDO")
        self.assertEqual(scenarios.cell(2, headers["Scenario"]).value, "Qudo")
        self.assertIn("MOV EUR 300", scenarios.cell(2, headers["Requisito"]).value)
        self.assertEqual(scenarios.cell(2, headers["Stock"]).value, 309)
        self.assertFalse(scenarios.cell(2, headers["Costo"]).protection.locked)
        self.assertIn("MATCH($AB2", scenarios.cell(2, headers["Margine attuale %"]).value)
        self.assertTrue(scenarios.protection.sheet)
        self.assertEqual(workbook["Dati"].sheet_state, "hidden")


if __name__ == "__main__":
    unittest.main()
