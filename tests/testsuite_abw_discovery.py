import json
import tempfile
import unittest
from datetime import datetime, timedelta, timezone
from decimal import Decimal, localcontext
from pathlib import Path

from openpyxl import load_workbook
from streamlit.testing.v1 import AppTest

from abw_discovery import (
    normalize_abw_candidates,
    snapshot_is_stale,
    valid_abw_identifier,
)
from discovery import (
    DISCOVERY_SCHEMA_VERSION,
    DiscoveryCheckpointStore,
    _evaluate_product_scenarios,
    _select_recommended_combination,
    default_filters,
    discovery_funnel_view,
    run_discovery,
)
from discovery_amazon import correlate_catalog_items
from discovery_excel import SCENARIO_COLUMNS, write_discovery_excel
from purchase_scenarios import (
    merge_product_candidates,
    normalize_purchase_scenario,
    scenario_requirement_label,
)
from qogita_discovery import normalize_qogita_candidates
from tests.testsuite_discovery import fee_result, qogita_row
from tests.testsuite_umma_discovery import amazon_observation, standard_row
from tests.testsuite_multi_listing import (
    EAN as MULTI_LISTING_EAN,
    catalog_item,
)
from umma_discovery import normalize_umma_candidates


EAN = "8809532220748"
NOW = datetime(2026, 8, 24, 12, 0, tzinfo=timezone.utc)


def abw_row(
    *, ean=EAN, mode="standard", condition="tier_1_99", minimum=1, maximum=99,
    net="7.53", pack_size=None, pack_price=None, warehouse="core",
    availability="available", observed_at="2026-08-24T04:15:49.911718Z",
    run_id="65c866c0-6b40-49e9-bbb2-082b8eb7ac28", stock=None,
    is_sellable=True, out_of_stock=False,
):
    return {
        "run_id": run_id, "seller_sku": "HARU06", "gtin": ean,
        "supplier_product_id": "10035", "option_product_id": "10035",
        "product_name": "BLACK BAMBOO MIST JUMBO", "brand": "haruharu wonder",
        "mode": mode, "condition_key": condition,
        "condition_label": condition, "tier_min_quantity": minimum,
        "tier_max_quantity": maximum, "pack_size": pack_size,
        "pack_price": pack_price, "net_unit_price_eur": net,
        "currency": "EUR", "price_source": "abw_authenticated_reference_currency",
        "price_basis": "net_ex_vat", "vat_rate": "0.22",
        "vat_amount": None, "gross_unit_price": None,
        "available_quantity": stock, "availability_status": availability,
        "stock_text": "In-Stock 1-2 days", "lead_time": "1-2 days",
        "warehouse": warehouse, "discount_label": "", "product_url": "",
        "minimum_order_value": "250", "minimum_order_currency": "USD",
        "observed_at": observed_at, "source": "abw_authenticated_rest",
        "latest_attempt_status": "found", "latest_attempt_at": observed_at,
        "latest_run_status": "success", "is_sellable": is_sellable,
        "out_of_stock": out_of_stock,
    }


def real_0748_rows():
    return [
        abw_row(condition="tier_1_99", minimum=1, maximum=99, net="7.53"),
        abw_row(condition="tier_100_199", minimum=100, maximum=199, net="7.30"),
        abw_row(condition="tier_200_plus", minimum=200, maximum=None, net="7.15"),
        abw_row(mode="bulk_box", condition="box_15", minimum=15, maximum=15,
                net="6.869333", pack_size=15, pack_price="103.04"),
        abw_row(mode="bulk_box", condition="box_60", minimum=60, maximum=60,
                net="6.593333", pack_size=60, pack_price="395.60"),
        abw_row(mode="bulk_box", condition="box_80", minimum=80, maximum=80,
                net="6.3825", pack_size=80, pack_price="510.60"),
    ]


def normalized_abw():
    candidates, diagnostics = normalize_abw_candidates(real_0748_rows(), now=NOW)
    assert diagnostics["abw_scenarios"] == 6
    return candidates[0], diagnostics


class AbwAdapterTests(unittest.TestCase):
    def test_real_product_has_three_standard_and_three_bulk_scenarios(self):
        candidate, diagnostics = normalized_abw()
        standard = [s for s in candidate["scenarios"] if s["scenario_type"] == "abw_standard"]
        bulk = [s for s in candidate["scenarios"] if s["scenario_type"] == "abw_bulk_box"]
        self.assertEqual(len(standard), 3)
        self.assertEqual(len(bulk), 3)
        self.assertEqual(diagnostics["abw_products"], 1)
        self.assertEqual(diagnostics["abw_standard_scenarios"], 3)
        self.assertEqual(diagnostics["abw_bulk_box_scenarios"], 3)
        self.assertEqual({s["bundle_quantity"] for s in bulk}, {15, 60, 80})

    def test_bulk_uses_authoritative_pack_total_and_vat(self):
        candidate, _ = normalized_abw()
        box = next(s for s in candidate["scenarios"] if s["bundle_quantity"] == 15)
        with localcontext() as context:
            context.prec = 40
            expected_net = Decimal("103.04") / Decimal(15)
            expected_vat = expected_net * Decimal("0.22")
            expected_gross = expected_net * Decimal("1.22")
        self.assertEqual(box["source_pack_total_price"], Decimal("103.04"))
        self.assertEqual(box["cost_net_unit_eur"], expected_net)
        self.assertEqual(box["vat_amount_unit"], expected_vat)
        self.assertEqual(box["cost_gross_unit_eur"], expected_gross)

    def test_ranges_and_bundles_never_collapse_on_equal_price(self):
        rows = [
            abw_row(condition="tier_a", minimum=1, maximum=9, net="7"),
            abw_row(condition="tier_b", minimum=10, maximum=19, net="7"),
            abw_row(mode="bulk_box", condition="box_10", minimum=10, maximum=10,
                    net="7", pack_size=10, pack_price="70"),
            abw_row(mode="bulk_box", condition="box_20", minimum=20, maximum=20,
                    net="7", pack_size=20, pack_price="140"),
        ]
        candidates, _ = normalize_abw_candidates(rows, now=NOW)
        self.assertEqual(len(candidates[0]["scenarios"]), 4)
        self.assertEqual(len({s["scenario_id"] for s in candidates[0]["scenarios"]}), 4)

    def test_nullable_stock_and_unavailable_rows(self):
        candidate, _ = normalized_abw()
        self.assertTrue(all(s["stock"] is None for s in candidate["scenarios"]))
        for changes in (
            {"out_of_stock": True}, {"is_sellable": False}, {"stock": 0},
        ):
            row = abw_row(**changes)
            candidates, diagnostics = normalize_abw_candidates([row], now=NOW)
            self.assertEqual(candidates, [])
            self.assertEqual(diagnostics["abw_unavailable_scenarios"], 1)

    def test_warehouse_and_snapshot_do_not_destabilize_identity(self):
        original = abw_row()
        changed = abw_row(
            net="6.99", stock=50, observed_at="2026-08-24T05:00:00Z",
            run_id="new-run",
        )
        first, _ = normalize_abw_candidates([original], now=NOW)
        second, _ = normalize_abw_candidates([changed], now=NOW)
        self.assertEqual(first[0]["scenarios"][0]["scenario_id"], second[0]["scenarios"][0]["scenario_id"])
        other, _ = normalize_abw_candidates([abw_row(warehouse="eu")], now=NOW)
        self.assertNotEqual(first[0]["scenarios"][0]["scenario_id"], other[0]["scenarios"][0]["scenario_id"])

    def test_identifier_and_ttl_boundaries(self):
        self.assertTrue(valid_abw_identifier(EAN))
        self.assertFalse(valid_abw_identifier(EAN[:-1] + "9"))
        self.assertFalse(snapshot_is_stale(NOW - timedelta(hours=47, minutes=59), now=NOW))
        self.assertFalse(snapshot_is_stale(NOW - timedelta(hours=48), now=NOW))
        self.assertTrue(snapshot_is_stale(NOW - timedelta(hours=48, seconds=1), now=NOW))

    def test_account_mov_is_not_product_quantity(self):
        candidate, _ = normalized_abw()
        scenario = candidate["scenarios"][0]
        self.assertEqual(scenario["account_mov"], Decimal("250"))
        self.assertEqual(scenario["account_mov_currency"], "USD")
        self.assertNotEqual(scenario["account_mov"], scenario["minimum_product_quantity"])

    def test_checkpoint_numeric_round_trip_preserves_abw_fields(self):
        candidate, _ = normalized_abw()
        state = {"candidates": [candidate], "results": [], "amazon_observations": []}
        for scenario in state["candidates"][0]["scenarios"]:
            normalize_purchase_scenario(scenario)
        encoded = json.loads(json.dumps(state, default=str))
        for scenario in encoded["candidates"][0]["scenarios"]:
            normalize_purchase_scenario(scenario)
        scenario = encoded["candidates"][0]["scenarios"][3]
        self.assertIsInstance(scenario["source_pack_total_price"], Decimal)
        self.assertIsInstance(scenario["bundle_quantity"], int)
        self.assertIsNone(encoded["candidates"][0]["scenarios"][0]["stock"])

    def test_cross_supplier_aggregation_is_one_product(self):
        abw, _ = normalized_abw()
        qogita_rows = [
            qogita_row(gtin=EAN, mov=mov, price=price, observed_at="2026-08-24T04:00:00Z")
            for mov, price in ((500, "6.94"), (1500, "6.83"), (5000, "6.80"), (10000, "6.73"), (15000, "6.73"))
        ]
        qogita, _ = normalize_qogita_candidates(qogita_rows, now=NOW)
        umma_rows = [standard_row(ean=EAN, observed_at="2026-08-24T04:00:00Z")]
        umma, _ = normalize_umma_candidates(umma_rows, now=NOW)
        merged = merge_product_candidates(qogita, umma, [abw])
        self.assertEqual(len(merged), 1)
        self.assertEqual(len(merged[0]["scenarios"]), 12)
        self.assertEqual({s["supplier"] for s in merged[0]["scenarios"]}, {"qogita", "umma", "abw"})

    def test_fan_out_economics_scores_and_product_pass(self):
        candidate, _ = normalized_abw()
        observation = amazon_observation()
        _evaluate_product_scenarios(candidate, observation, 15)
        self.assertEqual(len(candidate["scenarios"]), 6)
        self.assertEqual(len({s["margin_percent"] for s in candidate["scenarios"]}), 6)
        self.assertTrue(any(s["evaluation_status"] == "margin_passed" for s in candidate["scenarios"]))
        self.assertEqual(candidate["evaluation_status"], "margin_passed")


class AbwIntegrationTests(unittest.TestCase):
    @staticmethod
    def _combination(identifier, scenario_id):
        return {
            "combination_id": identifier, "scenario_id": scenario_id,
            "supplier": "abw", "asin": "B000000001", "score": 70,
            "margin_percent": Decimal("20"), "profit": Decimal("3"),
            "cost_gross_unit_eur": Decimal("8"),
        }

    def test_supplier_aware_tie_breaks_only_comparable_abw_families(self):
        standard = {
            "s1": {"scenario_id": "s1", "supplier": "abw", "scenario_type": "abw_standard", "minimum_product_quantity": 100, "maximum_product_quantity": None},
            "s2": {"scenario_id": "s2", "supplier": "abw", "scenario_type": "abw_standard", "minimum_product_quantity": 1, "maximum_product_quantity": 99},
        }
        rows = [self._combination("c1", "s1"), self._combination("c2", "s2")]
        self.assertEqual(_select_recommended_combination(rows, standard)["scenario_id"], "s2")
        bulk = {
            "b60": {"scenario_id": "b60", "supplier": "abw", "scenario_type": "abw_bulk_box", "bundle_quantity": 60},
            "b15": {"scenario_id": "b15", "supplier": "abw", "scenario_type": "abw_bulk_box", "bundle_quantity": 15},
        }
        rows = [self._combination("c60", "b60"), self._combination("c15", "b15")]
        self.assertEqual(_select_recommended_combination(rows, bulk)["scenario_id"], "b15")
        mixed = {"a": standard["s2"], "z": bulk["b15"]}
        mixed["a"]["scenario_id"] = "a"
        mixed["z"]["scenario_id"] = "z"
        rows = [self._combination("ca", "a"), self._combination("cz", "z")]
        self.assertEqual(_select_recommended_combination(rows, mixed)["scenario_id"], "a")

    def test_ui_labels_and_funnel_are_supplier_neutral(self):
        candidate, diagnostics = normalized_abw()
        labels = {scenario_requirement_label(s) for s in candidate["scenarios"]}
        self.assertIn("1–99 pz", labels)
        self.assertIn("Da 200 pz", labels)
        self.assertIn("Box 15", labels)
        view = discovery_funnel_view({"funnel": diagnostics})
        self.assertEqual(view["suppliers"]["abw_products"], 1)
        self.assertEqual(view["suppliers"]["abw_scenarios"], 6)

    def test_discovery_excel_represents_abw_and_dynamic_cost(self):
        candidate, _ = normalized_abw()
        observation = amazon_observation()
        candidate["amazon_observation"] = observation
        candidate["amazon_observations"] = [observation]
        _evaluate_product_scenarios(candidate, observation, 15)
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "abw.xlsx"
            write_discovery_excel([candidate], path)
            workbook = load_workbook(path, data_only=False)
        scenarios = workbook["Scenari"]
        self.assertEqual(scenarios.max_row, 7)
        self.assertEqual([scenarios.cell(1, c).value for c in range(1, len(SCENARIO_COLUMNS) + 1)], SCENARIO_COLUMNS)
        self.assertEqual({scenarios.cell(r, 5).value for r in range(2, 8)}, {"ABW"})
        self.assertFalse(scenarios["I2"].protection.locked)
        self.assertIn("MATCH($AC2", scenarios["N2"].value)
        self.assertEqual(scenarios["Y2"].value, "core")
        self.assertEqual(workbook["Dati"].sheet_state, "hidden")

    def test_abw_scenarios_do_not_multiply_amazon_calls(self):
        rows = [dict(row, gtin=MULTI_LISTING_EAN) for row in real_0748_rows()]
        candidates, diagnostics = normalize_abw_candidates(rows, now=NOW)
        self.assertEqual(diagnostics["abw_scenarios"], 6)
        candidate = candidates[0]
        candidate.update({
            "brand": "ANUA", "title": "ANUA Niacinamide Serum 30ml",
            "package_quantity": 1, "volume_value": 30, "volume_unit": "ml",
        })
        calls = {"catalog": [], "pricing": [], "fees": []}

        def catalog(gtins, _job_id, products):
            calls["catalog"].append(list(gtins))
            return correlate_catalog_items(
                gtins,
                [catalog_item("B0CVFPT7FC"), catalog_item("B0DL92QM79")],
                products,
            )

        def pricing(asins, _job_id):
            calls["pricing"].append(list(asins))
            return {asin: {
                "status": "success", "Venditori FBA": 2,
                "Venditori totali": 4,
                "Seller count source": "summary_number_of_offers",
                "reference_price": 30, "price_source": "buy_box",
            } for asin in asins}

        def fees(requests_, _token):
            calls["fees"].append([row["asin"] for row in requests_])
            return [fee_result(row["asin"], row["identifier"]) for row in requests_]

        with tempfile.TemporaryDirectory() as directory:
            state = run_discovery(
                default_filters(),
                checkpoint_store=DiscoveryCheckpointStore(directory),
                catalog_batch=catalog, pricing_batch=pricing, fees_batch=fees,
                token_provider=type("Tokens", (), {
                    "get": lambda self: "token",
                    "invalidate": lambda self: None,
                })(),
                qogita_loader=lambda: [qogita_row(observed_at="2026-08-24T10:00:00Z")],
                qogita_normalizer=lambda _rows, minimum_stock: (
                    [candidate], {
                        "initial": 1, "valid_gtin": 1,
                        "qogita_products": 0, "qogita_scenarios": 0,
                        "abw_products": 1, "abw_scenarios": 6,
                        "supplier_products_total": 1,
                        "supplier_scenarios_total": 6,
                    },
                ),
                qogita_refresher=lambda _: (_ for _ in ()).throw(
                    AssertionError("fresh cache must not refresh")
                ),
                now_provider=lambda: NOW, sleep_func=lambda _: None,
                catalog_batch_interval=0, pricing_batch_interval=0,
                fee_batch_interval=0,
            )
        self.assertEqual(len(state["results"][0]["opportunity_combinations"]), 12)
        self.assertEqual(calls["catalog"], [[MULTI_LISTING_EAN]])
        self.assertEqual(len(calls["pricing"]), 1)
        self.assertEqual(set(calls["pricing"][0]), {"B0CVFPT7FC", "B0DL92QM79"})
        self.assertEqual(len(calls["fees"]), 1)
        self.assertEqual(set(calls["fees"][0]), {"B0CVFPT7FC", "B0DL92QM79"})

    def test_streamlit_detail_renders_abw_ranges_and_boxes_offline(self):
        candidate, diagnostics = normalized_abw()
        observation = amazon_observation()
        _evaluate_product_scenarios(candidate, observation, 0)
        candidate["amazon_observation"] = observation
        candidate["amazon_offers_url"] = "https://www.amazon.it/gp/offer-listing/B000000001"
        state = {
            "discovery_schema_version": DISCOVERY_SCHEMA_VERSION,
            "job_id": "abw-ui", "status": "completed", "phase": "completed",
            "results": [candidate], "candidates": [candidate],
            "amazon_observations": [observation],
            "qogita_refresh_status": "cache_fresh", "qogita_snapshot_after": {},
            "funnel": {
                **diagnostics, "amazon_found": 1, "beauty_valid": 1,
                "bsr_passed": 1, "competition_passed": 1, "fee_valid": 1,
                "final_opportunities": 1, "scenarios_evaluated": 6,
                "scenarios_margin_passed": 6,
                "scenarios_margin_below_threshold": 0,
            },
        }
        app = AppTest.from_file("app_glowup.py", default_timeout=20).run()
        app.session_state["ui_state"] = "discovery_result"
        app.session_state["discovery_status"] = "completed"
        app.session_state["discovery_result"] = {
            "state": json.loads(json.dumps(state, default=str)),
            "output_bytes": b"workbook",
        }
        app.run()
        self.assertFalse(app.exception)
        self.assertTrue(any("ABW" in element.value for element in app.markdown))
        next(button for button in app.button if button.label == "Confronta scenari").click().run()
        self.assertFalse(app.exception)
        table = app.dataframe[0].value
        self.assertEqual(len(table), 6)
        self.assertIn("Warehouse", table.columns)
        self.assertIn("Disponibilità", table.columns)
        self.assertIn("Box 15", table["Requisito"].tolist())
        self.assertIn("Da 200 pz", table["Requisito"].tolist())


if __name__ == "__main__":
    unittest.main()
