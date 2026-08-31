"""Interactive supplier-neutral multi-scenario Discovery Excel export."""

from __future__ import annotations

import math
import os
import tempfile
import zipfile
from decimal import Decimal
from xml.etree import ElementTree

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter

from discovery import normalize_discovery_state
from purchase_scenarios import (
    recommended_combination,
    scenario_requirement_label,
)


OPPORTUNITY_COLUMNS = [
    "EAN", "Brand", "Titolo", "Fornitore", "Scenario", "Requisito", "Costo",
    "ASIN raccomandato", "BSR Beauty", "Prezzo riferimento", "Venditori FBA", "Venditori totali",
    "Margine attuale %", "Utile €", "Prezzo 15%", "Prezzo 20%", "Prezzo 25%",
    "Score", "Opportunità", "Numero scenari acquisto", "Numero pagine Amazon",
    "Link Offerte Amazon",
]
SCENARIO_COLUMNS = [
    "EAN", "Brand", "Titolo", "ASIN", "Fornitore", "Seller alias", "Scenario",
    "Requisito", "Costo", "Prezzo riferimento", "BSR Beauty", "Venditori FBA",
    "Venditori totali", "Margine attuale %", "Utile €", "Prezzo 15%", "Prezzo 20%",
    "Prezzo 25%", "Score", "Opportunità", "Ruolo", "Stato", "Stock",
    "Lead time / Snapshot / freshness", "Warehouse", "Disponibilità",
]
TECHNICAL_COLUMNS = [
    "ObservationRowID", "ProductRowID", "ASIN", "BSR Beauty", "Prezzo riferimento",
    "Venditori FBA", "Venditori totali", "FBA fee netta",
    "FBA fee IVA inclusa", "Referral Fee", "Referral rate",
    "Referral source", "Price source", "Seller count source", "Observed at",
    "Prezzo minimo FBA", "Prezzo minimo FBM",
]
ALL_RESULTS_COLUMNS = [
    "EAN", "Brand", "Titolo", "Fornitori disponibili",
    "Numero scenari acquisto", "ASIN", "Titolo Amazon", "Compatibility",
    "Beauty", "BSR Beauty", "Prezzo riferimento", "Origine prezzo",
    "Min FBA", "Min FBM", "Venditori FBA", "Venditori totali",
    "Miglior costo disponibile", "Miglior margine disponibile %",
    "Score migliore disponibile", "Stato", "Motivo esclusione",
    "Soglia / limite rilevante", "Dettaglio",
]
LISTING_COLUMNS = [
    "EAN", "ASIN", "Titolo Amazon", "Brand Amazon",
    "Compatibility status", "Compatibility reason", "Beauty",
    "Display group", "BSR Beauty", "Catalog status", "Buy Box", "Min FBA",
    "Min FBM", "Price source", "Venditori FBA", "Venditori totali",
    "Pricing status", "Competition status", "Fee status", "Fee attempts",
    "Exclusion reason", "Link Amazon",
]
RUN_COLUMNS = ["Parametro", "Valore"]
# Backwards-compatible import name; Discovery now exports Opportunita, not Risultati.
DISCOVERY_COLUMNS = OPPORTUNITY_COLUMNS
EXCEL_MAX_HYPERLINKS_PER_WORKSHEET = 65_530
_RELATIONSHIP_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_HYPERLINK_RELATIONSHIP_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink"
)


def _value(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, float) and not math.isfinite(value):
        return None
    return value


def _hyperlink_formula(url):
    """Return a relationship-free hyperlink accepted by desktop Excel.

    Excel limits external hyperlink relationships to 65,530 per worksheet.
    Large audit sheets can exceed that limit even though the XLSX package is
    otherwise valid. A HYPERLINK formula keeps every URL clickable without
    creating one package relationship per row.
    """
    if not url:
        return None
    text = str(url).replace('"', '""')
    return f'=HYPERLINK("{text}","{text}")'


def validate_excel_compatibility(path):
    """Validate package constraints that openpyxl/ZIP checks do not enforce."""
    result = {"hyperlinks_by_part": {}}
    with zipfile.ZipFile(path) as package:
        package.testzip()
        for name in package.namelist():
            if not (name.startswith("xl/worksheets/_rels/") and name.endswith(".rels")):
                continue
            root = ElementTree.fromstring(package.read(name))
            identifiers = set()
            hyperlink_count = 0
            for relationship in root.findall(f"{{{_RELATIONSHIP_NS}}}Relationship"):
                relationship_id = relationship.get("Id")
                if relationship_id in identifiers:
                    raise ValueError(f"Duplicate relationship ID in {name}: {relationship_id}")
                identifiers.add(relationship_id)
                if relationship.get("Type") == _HYPERLINK_RELATIONSHIP_TYPE:
                    hyperlink_count += 1
            result["hyperlinks_by_part"][name] = hyperlink_count
            if hyperlink_count > EXCEL_MAX_HYPERLINKS_PER_WORKSHEET:
                raise ValueError(
                    f"Excel worksheet hyperlink limit exceeded in {name}: "
                    f"{hyperlink_count} > {EXCEL_MAX_HYPERLINKS_PER_WORKSHEET}"
                )
    return result


def _as_decimal(value):
    try:
        return Decimal(str(value))
    except Exception:
        return None


def _excel_number(value):
    """Return an Excel numeric value without coercing identifiers or formulas."""
    if value in (None, "") or isinstance(value, bool):
        return value
    if isinstance(value, str) and value.startswith("="):
        return value
    number = _as_decimal(value)
    return float(number) if number is not None and number.is_finite() else value


def _opportunity_sort_key(product):
    """Use persisted economics to physically order the bounded final result set."""
    combination = recommended_combination(product) or {}
    scenario = _recommended(product) or {}

    def descending(value):
        number = _as_decimal(value)
        return -number if number is not None and number.is_finite() else Decimal("Infinity")

    def persisted(field):
        value = combination.get(field)
        return value if value is not None else scenario.get(field)

    return (
        descending(persisted("score")),
        descending(persisted("margin_percent")),
        descending(persisted("profit")),
        str(_product_ean(product) or ""),
    )


def _sorted_final_products(final_products):
    # Final opportunities are intentionally bounded (143 in the first full
    # run). Sorting this view must never materialize the complete candidate or
    # combination universe.
    return sorted(final_products, key=_opportunity_sort_key)


def _export_context(payload):
    if isinstance(payload, dict):
        raw_candidates = payload.get("candidates") or []
        raw_final_products = payload.get("results") or []
        # Incremental rows were normalized before being persisted.  Running
        # normalize_discovery_state here would iterate both SQL views and keep
        # every hydrated product in its ``normalized_products`` list before a
        # temporary workbook even exists, recreating the multi-gigabyte graph
        # that incremental persistence was designed to avoid.
        if getattr(raw_candidates, "store", None) is not None:
            return dict(payload), raw_candidates, raw_final_products
        state = normalize_discovery_state(payload)
        candidates = state.get("candidates") or []
        final_products = state.get("results") or []
        # Incremental collections are deliberately re-iterable SQL views.  Do
        # not materialize tens of thousands of products merely to export them.
        if getattr(candidates, "store", None) is not None:
            return state, candidates, final_products
        return state, list(candidates), list(final_products)
    products = list(payload or [])
    return {}, products, products


def _product_ean(product):
    return product.get("canonical_ean") or product.get("gtin")


def _supplier_names(product):
    values = []
    for scenario in product.get("scenarios") or []:
        supplier = str(scenario.get("supplier") or "").strip()
        if supplier and supplier.casefold() not in {row.casefold() for row in values}:
            values.append(supplier)
    return " · ".join(row.upper() for row in values)


def _listing_combinations(product, listing):
    asin = listing.get("asin")
    observation_id = listing.get("amazon_observation_id")
    return [
        row for row in product.get("opportunity_combinations") or []
        if (
            (observation_id and row.get("amazon_observation_id") == observation_id)
            or (not observation_id and asin and row.get("asin") == asin)
        )
    ]


def _best_combination(rows):
    if not rows:
        return None
    return min(rows, key=lambda row: (
        -int(row.get("score") or 0),
        -float(_as_decimal(row.get("margin_percent")) or Decimal("-999999")),
        -float(_as_decimal(row.get("profit")) or Decimal("-999999")),
        float(_as_decimal(row.get("cost_gross_unit_eur")) or Decimal("999999")),
        str(row.get("combination_id") or ""),
    ))


def _best_cost(product):
    values = [
        _as_decimal(row.get("cost_gross_unit_eur"))
        for row in product.get("scenarios") or []
    ]
    values = [row for row in values if row is not None and row > 0]
    return min(values) if values else None


def _status_detail(product, listing, filters, final_product_keys):
    """Return technical status, Italian label, reason, threshold and detail."""
    if listing is None:
        catalog_status = product.get("catalog_status")
        if catalog_status == "not_found":
            return (
                "catalog_not_found", "Non trovato su Amazon",
                "Nessun listing Amazon restituito per l'EAN", "—",
                "catalog_status=not_found",
            )
        if catalog_status == "catalog_incomplete":
            return (
                "catalog_incomplete", "Catalog Amazon incompleto",
                "La paginazione Catalog non è stata completata; risultato non definitivo",
                "Ripetere Catalog Items", "catalog_status=catalog_incomplete",
            )
        return (
            "economics_unavailable", "Economia non disponibile",
            "Nessun listing Amazon valutabile", "—",
            f"catalog_status={catalog_status or 'unknown'}",
        )

    evaluation = listing.get("evaluation_status")
    compatibility = listing.get("compatibility_status")
    if compatibility == "incompatible" or evaluation == "catalog_incompatible":
        reason = listing.get("compatibility_reason") or listing.get("exclusion_reason")
        return (
            "incompatible_listing", "Listing incompatibile",
            "Listing non compatibile con il prodotto supplier", "Compatibilità prodotto",
            _detail_text(reason),
        )
    if evaluation == "beauty_filtered":
        return (
            "not_beauty", "Non Beauty",
            "Display group Amazon non riconosciuto come Beauty", "Beauty richiesta",
            _detail_text(listing.get("beauty_status") or listing.get("display_group")),
        )
    if evaluation == "bsr_filtered":
        bsr = listing.get("bsr_beauty")
        minimum = filters.get("bsr_min")
        maximum = filters.get("bsr_max")
        if bsr is None or _as_decimal(bsr) is None or _as_decimal(bsr) <= 0:
            reason = "BSR Beauty mancante o non valido"
        elif minimum is not None and _as_decimal(bsr) < _as_decimal(minimum):
            reason = f"BSR {_integer_text(bsr)} < minimo {_integer_text(minimum)}"
        else:
            reason = f"BSR {_integer_text(bsr)} > massimo {_integer_text(maximum)}"
        return (
            "bsr_out_of_range", "Fuori range BSR", reason,
            f"BSR {minimum}–{maximum}", _detail_text(listing.get("exclusion_reason")),
        )
    if evaluation == "competition_filtered":
        reasons = listing.get("exclusion_reasons") or str(
            listing.get("exclusion_reason") or ""
        ).split(",")
        descriptions = []
        if "fba_sellers_above_threshold" in reasons:
            descriptions.append(
                f"Venditori FBA {_integer_text(listing.get('fba_sellers'))} > "
                f"limite {_integer_text(filters.get('max_fba_sellers'))}"
            )
        if "total_sellers_above_threshold" in reasons:
            descriptions.append(
                f"Venditori totali {_integer_text(listing.get('total_sellers'))} > "
                f"limite {_integer_text(filters.get('max_total_sellers'))}"
            )
        if "missing_reference_price" in reasons:
            descriptions.append("Prezzo riferimento non disponibile")
        return (
            "competition_filtered", "Escluso per concorrenza",
            "; ".join(descriptions) or "Concorrenza oltre i limiti configurati",
            f"FBA ≤ {filters.get('max_fba_sellers')} · Totali ≤ {filters.get('max_total_sellers')}",
            _detail_text(reasons),
        )
    if listing.get("fee_status") == "fee_pending":
        return (
            "fee_pending", "Fee in attesa", "Product Fees temporaneamente non disponibile",
            "Retry Product Fees", _detail_text(listing.get("fee_error")),
        )
    if listing.get("fee_status") == "unavailable":
        return (
            "fee_unavailable", "Fee non disponibile",
            "Amazon non ha restituito una Fee valida",
            "Escluso dal ranking economico",
            _detail_text(
                listing.get("fee_unavailable_reason") or listing.get("fee_error")
            ),
        )
    if listing.get("fee_status") in {"invalid", "fee_invalid"}:
        return (
            "fee_invalid", "Fee non valida", "Product Fees non valida",
            "Fee Estimate valida richiesta", _detail_text(listing.get("fee_error")),
        )

    combinations = _listing_combinations(product, listing)
    passed = [row for row in combinations if row.get("evaluation_status") == "margin_passed"]
    product_key = product.get("product_key")
    if passed and product_key in final_product_keys:
        return (
            "opportunity", "Opportunità", "Almeno una combinazione supera il margine minimo",
            f"Margine ≥ {float(filters.get('minimum_margin') or 0):.2f}%", "",
        )
    if combinations:
        best = _best_combination(combinations)
        margin = _as_decimal((best or {}).get("margin_percent"))
        threshold = _as_decimal(filters.get("minimum_margin")) or Decimal("0")
        if margin is not None and margin < threshold:
            return (
                "margin_below_threshold", "Sotto soglia margine",
                f"Margine {float(margin):.2f}% < soglia {float(threshold):.2f}%",
                f"Margine ≥ {float(threshold):.2f}%", "",
            )
    return (
        "economics_unavailable", "Economia non disponibile",
        "Il listing non dispone di una valutazione economica completa", "—",
        _detail_text(listing.get("exclusion_reason") or evaluation),
    )


def _detail_text(value):
    if isinstance(value, (list, tuple, set)):
        return " · ".join(str(row) for row in value if row not in (None, ""))
    if isinstance(value, dict):
        return " · ".join(f"{key}={item}" for key, item in sorted(value.items()))
    return str(value or "")


def _integer_text(value):
    number = _as_decimal(value)
    return f"{int(number):,}".replace(",", ".") if number is not None else "—"


def _lookup(row, observation_id_column, data_column, last_row):
    return (
        f"INDEX('Dati'!${data_column}$2:${data_column}${last_row},"
        f"MATCH(${observation_id_column}{row},'Dati'!$A$2:$A${last_row},0))"
    )


def _economic_formulas(row, *, cost_col, margin_col, profit_col, target_cols, score_col,
                       opportunity_col, observation_id_col, last_data_row):
    price = _lookup(row, observation_id_col, "E", last_data_row)
    bsr = _lookup(row, observation_id_col, "D", last_data_row)
    fba_sellers = _lookup(row, observation_id_col, "F", last_data_row)
    total_sellers = _lookup(row, observation_id_col, "G", last_data_row)
    fba_gross = _lookup(row, observation_id_col, "I", last_data_row)
    referral_rate = _lookup(row, observation_id_col, "K", last_data_row)
    profit = (
        f'=IFERROR(IF(AND(ISNUMBER({price}),{price}>0,ISNUMBER({cost_col}{row}),'
        f'ISNUMBER({fba_gross}),ISNUMBER({referral_rate})),'
        f'ROUND({price}-{cost_col}{row}-({price}*{referral_rate})-'
        f'{fba_gross},2),""),"")'
    )
    margin = (
        f'=IFERROR(IF(AND(ISNUMBER({price}),{price}>0,ISNUMBER({profit_col}{row})),'
        f'ROUND({profit_col}{row}/{price},4),""),"")'
    )
    targets = {}
    for target, column in zip((0.15, 0.20, 0.25), target_cols):
        targets[column] = (
            f'=IFERROR(IF(AND(ISNUMBER({cost_col}{row}),ISNUMBER({fba_gross}),'
            f'ISNUMBER({referral_rate}),(1-{referral_rate}-{target})>0),'
            f'ROUND(({cost_col}{row}+{fba_gross})/(1-{referral_rate}-{target}),2),""),"")'
        )
    bsr_points = (
        f"IF(AND(ISNUMBER({bsr}),{bsr}>=0),IF({bsr}<=1000,50,"
        f"IF({bsr}<=5000,45,IF({bsr}<=10000,40,IF({bsr}<=25000,30,"
        f"IF({bsr}<=50000,15,5))))),0)"
    )
    fba_points = (
        f"IF(AND(ISNUMBER({total_sellers}),{total_sellers}>0,ISNUMBER({fba_sellers}),"
        f"{fba_sellers}>=0),IF({fba_sellers}=0,20,IF({fba_sellers}<=2,18,"
        f"IF({fba_sellers}<=4,14,IF({fba_sellers}<=6,10,IF({fba_sellers}<=10,5,0))))),0)"
    )
    total_points = (
        f"IF(AND(ISNUMBER({total_sellers}),{total_sellers}>0),"
        f"IF({total_sellers}<=3,10,IF({total_sellers}<=6,7,"
        f"IF({total_sellers}<=10,4,0))),0)"
    )
    margin_points = (
        f"IF(ISNUMBER({margin_col}{row}),IF({margin_col}{row}<10%,0,"
        f"IF({margin_col}{row}<15%,4,IF({margin_col}{row}<20%,14,"
        f"IF({margin_col}{row}<25%,18,20)))),0)"
    )
    score = f"={bsr_points}+{fba_points}+{total_points}+{margin_points}"
    opportunity = (
        f'=IF({score_col}{row}>=85,"🟢 Eccellente",IF({score_col}{row}>=70,'
        f'"🟢 Ottima",IF({score_col}{row}>=55,"🟡 Interessante",'
        f'IF({score_col}{row}>=40,"🟠 Da valutare","🔴 Debole"))))'
    )
    return margin, profit, targets, score, opportunity


def _recommended(product):
    scenario_id = (product.get("scenario_roles") or {}).get("scenario_raccomandato")
    return next(
        (row for row in product.get("scenarios") or [] if row.get("scenario_id") == scenario_id),
        None,
    )


def _observations(product):
    rows = list(product.get("amazon_observations") or [])
    legacy = product.get("amazon_observation") or {}
    if legacy and not any(
        row.get("observation_id") == legacy.get("observation_id") for row in rows
    ):
        rows.append(legacy)
    return rows


def _combinations(product):
    rows = list(product.get("opportunity_combinations") or [])
    if rows:
        return rows
    observation = product.get("amazon_observation") or {}
    for scenario in product.get("scenarios") or []:
        economics = scenario.get("economics") or {}
        if not (
            scenario.get("economics_status") == "ready"
            or economics.get("status") == "ready"
            or (
                scenario.get("margin_percent") is not None
                and scenario.get("score") is not None
            )
        ):
            continue
        rows.append({
            "combination_id": f"legacy|{scenario.get('scenario_id')}",
            "scenario_id": scenario.get("scenario_id"),
            "asin": observation.get("asin"),
            "amazon_observation_id": observation.get("observation_id") or product.get("product_key"),
            "price_reference": observation.get("reference_price"),
            "margin_percent": scenario.get("margin_percent"),
            "score": scenario.get("score"), "opportunity": scenario.get("opportunity"),
            "evaluation_status": scenario.get("evaluation_status"),
            "economics": economics,
        })
    return rows


def _style_sheet(ws, *, visible_columns, cost_column, hidden_columns, data_ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws[f"{cost_column}1"].comment = Comment(
        "Modifica questi valori per simulare utile, margine, prezzi target, Score e Opportunità.",
        "GlowUp Scout",
    )
    editable_fill = PatternFill("solid", fgColor="EAF2FE")
    for row in range(2, ws.max_row + 1):
        ws[f"{cost_column}{row}"].protection = Protection(locked=False)
        ws[f"{cost_column}{row}"].fill = editable_fill
    for column in hidden_columns:
        ws.column_dimensions[column].hidden = True
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(visible_columns)}{ws.max_row}"
    ws.print_area = f"A1:{get_column_letter(visible_columns)}{ws.max_row}"
    ws.protection.sheet = True
    ws.protection.autoFilter = False
    ws.protection.sort = False
    ws.protection.formatColumns = False
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False
    data_ws.sheet_state = "hidden"


def _style_audit_sheet(ws, widths, *, currency_columns=(), percent_columns=(),
                       integer_columns=(), text_columns=(), hyperlink_column=None,
                       formula_hyperlink=False):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(1, ws.max_row)}"
    ws.sheet_view.showGridLines = False
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row in range(2, ws.max_row + 1):
        for column in text_columns:
            ws.cell(row, column).number_format = "@"
        for column in currency_columns:
            cell = ws.cell(row, column)
            cell.value = _excel_number(cell.value)
            cell.number_format = '€ #,##0.00'
        for column in percent_columns:
            cell = ws.cell(row, column)
            cell.value = _excel_number(cell.value)
            cell.number_format = '0.00%'
        for column in integer_columns:
            cell = ws.cell(row, column)
            cell.value = _excel_number(cell.value)
            cell.number_format = '#,##0'
        if hyperlink_column:
            cell = ws.cell(row, hyperlink_column)
            if cell.value:
                if formula_hyperlink:
                    cell.value = _hyperlink_formula(cell.value)
                else:
                    cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"


def _run_metadata(state, candidates, final_products):
    filters = state.get("filters") or {}
    funnel = state.get("funnel") or {}
    selected = state.get("selected_suppliers") or []
    if not selected:
        selected = sorted({
            str(scenario.get("supplier") or "").upper()
            for product in candidates
            for scenario in product.get("scenarios") or []
            if scenario.get("supplier")
        }) or ["QOGITA"]
    return [
        ("Job ID", state.get("job_id") or "—"),
        ("Schema checkpoint", state.get("schema_version") or "—"),
        ("Stato job", state.get("status") or "—"),
        ("Fase", state.get("phase") or "—"),
        ("Avviato", state.get("started_at") or state.get("created_at") or "—"),
        ("Completato", state.get("completed_at") or "—"),
        ("Fornitori selezionati", " · ".join(str(row).upper() for row in selected)),
        ("BSR minimo", filters.get("bsr_min")),
        ("BSR massimo", filters.get("bsr_max")),
        ("Venditori FBA massimi", filters.get("max_fba_sellers")),
        ("Venditori totali massimi", filters.get("max_total_sellers")),
        ("Margine minimo %", filters.get("minimum_margin")),
        ("EAN universo supplier", state.get("total_supplier_ean_universe")),
        ("Identificatori idonei", state.get("eligible_identifier_count")),
        ("Budget ricerca", state.get("run_budget") or "all"),
        ("Identificatori campionati", state.get("sampled_identifier_count") or len(candidates)),
        ("Prodotti richiesti", state.get("requested_universe_count")),
        ("Cache Amazon riusata", state.get("cache_reuse_count")),
        ("Dati Amazon aggiornati", state.get("refresh_count")),
        ("Nuovi lookup Amazon", state.get("new_lookup_count")),
        ("Policy freshness Amazon", state.get("freshness_policy_version") or "—"),
        ("Strategia campionamento", state.get("sampling_strategy") or "—"),
        ("Scope rotazione", state.get("rotation_scope") or "—"),
        ("Ciclo rotazione", state.get("rotation_cycle_id")),
        ("Scope rotazione inizializzato", state.get("rotation_scope_initialized")),
        ("Storico Amazon globale", state.get("rotation_global_analyzed_count")),
        ("Nuovi identificatori scope", state.get("rotation_new_identifier_count")),
        ("Analizzati prima della run", state.get("rotation_analyzed_before_run")),
        ("Analizzati in questa run", state.get("rotation_analyzed_this_run")),
        ("Rimanenti nel ciclo", state.get("rotation_remaining_after_run")),
        *[
            (
                f"Freshness {str(supplier).upper()}",
                " · ".join(filter(None, (
                    str((state.get("supplier_snapshot_set") or {}).get(supplier, {}).get("freshness") or "—"),
                    str((state.get("supplier_snapshot_set") or {}).get(supplier, {}).get("snapshot_at") or "—"),
                ))),
            )
            for supplier in selected
        ],
        ("Prodotti analizzati", len(candidates)),
        ("Prodotti trovati Amazon", funnel.get("amazon_found")),
        ("Pagine Amazon trovate", funnel.get("amazon_listings_found")),
        (
            "Prodotti esclusi per concorrenza",
            funnel.get("competition_filtered_products"),
        ),
        (
            "Pagine Amazon con concorrenza valida",
            funnel.get("competition_passed_listings"),
        ),
        ("Fee Amazon target", state.get("fee_target_count") or funnel.get("fee_target_count")),
        ("Fee Amazon valide", state.get("fee_valid_count") or funnel.get("fee_valid_count")),
        ("Fee Amazon non disponibili", state.get("fee_unavailable_count") or funnel.get("fee_unavailable_count")),
        ("Fee Amazon non valide", state.get("fee_invalid_count") or funnel.get("fee_invalid_count")),
        ("Copertura Fee parziale", bool(state.get("fee_coverage_partial") or funnel.get("fee_coverage_partial"))),
        ("Opportunità finali", len(final_products)),
    ]


def write_discovery_excel(results, output_file, *, progress=None):
    state, candidates, final_products = _export_context(results)
    if getattr(candidates, "store", None) is not None:
        return _write_incremental_discovery_excel(
            state, candidates, final_products, output_file, progress=progress,
        )
    final_products = _sorted_final_products(final_products)
    output_path = os.path.abspath(output_file)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    descriptor, temporary_path = tempfile.mkstemp(
        prefix=".glowup_discovery_", suffix=".xlsx",
        dir=os.path.dirname(output_path) or ".",
    )
    os.close(descriptor)
    try:
        workbook = Workbook()
        opportunities = workbook.active
        opportunities.title = "Opportunità"
        all_results_ws = workbook.create_sheet("Tutti i risultati")
        listings_ws = workbook.create_sheet("Listing Amazon")
        scenarios_ws = workbook.create_sheet("Scenari")
        data_ws = workbook.create_sheet("Dati")
        run_ws = workbook.create_sheet("Parametri run")
        opportunities.append(OPPORTUNITY_COLUMNS + [
            "ProductRowID", "ScenarioRowID", "ObservationRowID", "CombinationRowID"
        ])
        all_results_ws.append(ALL_RESULTS_COLUMNS)
        listings_ws.append(LISTING_COLUMNS)
        scenarios_ws.append(SCENARIO_COLUMNS + [
            "ProductRowID", "ScenarioRowID", "ObservationRowID", "CombinationRowID"
        ])
        data_ws.append(TECHNICAL_COLUMNS)
        run_ws.append(RUN_COLUMNS)

        observations = {}
        for observation in state.get("amazon_observations") or []:
            observation_id = observation.get("observation_id")
            if observation_id:
                product_keys = (observation.get("diagnostics") or {}).get("product_keys") or []
                observations.setdefault(
                    observation_id, (product_keys[0] if product_keys else "", observation)
                )
        for product in candidates:
            product_id = product.get("product_key") or f"product-{product.get('gtin')}"
            for observation in _observations(product):
                observation_id = observation.get("observation_id") or product_id
                observations.setdefault(observation_id, (product_id, observation))
        final_product_keys = {
            row.get("product_key") for row in final_products if row.get("product_key")
        }

        for product in final_products:
            product_id = product.get("product_key") or f"product-{_product_ean(product)}"
            recommended_pair = recommended_combination(product)
            scenario_by_id = {
                row.get("scenario_id"): row for row in product.get("scenarios") or []
            }
            recommended = scenario_by_id.get(
                (recommended_pair or {}).get("scenario_id")
            ) or _recommended(product)
            if not recommended:
                continue
            if recommended_pair:
                observation = next(
                    (row for row in _observations(product) if row.get("observation_id") == recommended_pair.get("amazon_observation_id")),
                    product.get("amazon_observation") or {},
                )
            else:
                observation = product.get("amazon_observation") or {}
                recommended_pair = {
                    "combination_id": f"legacy|{recommended.get('scenario_id')}",
                    "scenario_id": recommended.get("scenario_id"),
                    "asin": observation.get("asin"),
                    "amazon_observation_id": observation.get("observation_id") or product_id,
                }
            opportunities.append([
                _product_ean(product), observation.get("amazon_brand") or product.get("brand"),
                observation.get("amazon_title") or product.get("title"),
                str(recommended.get("supplier") or "").upper(),
                recommended.get("scenario_label"), scenario_requirement_label(recommended),
                _excel_number(recommended.get("cost_gross_unit_eur")), recommended_pair.get("asin"),
                observation.get("bsr_beauty"),
                _excel_number(observation.get("reference_price")), observation.get("fba_sellers"),
                observation.get("total_sellers"), None, None, None, None, None, None, None,
                len(product.get("scenarios") or []), len(product.get("amazon_listings") or [observation]),
                product.get("amazon_offers_url"), product_id,
                recommended.get("scenario_id"), recommended_pair.get("amazon_observation_id"),
                recommended_pair.get("combination_id"),
            ])

        for product in candidates:
            product_id = product.get("product_key") or f"product-{_product_ean(product)}"
            recommended_pair = recommended_combination(product)
            if not recommended_pair:
                legacy_recommended = _recommended(product) or {}
                legacy_observation = product.get("amazon_observation") or {}
                recommended_pair = {
                    "combination_id": f"legacy|{legacy_recommended.get('scenario_id')}",
                    "scenario_id": legacy_recommended.get("scenario_id"),
                    "asin": legacy_observation.get("asin"),
                    "amazon_observation_id": (
                        legacy_observation.get("observation_id") or product_id
                    ),
                }
            scenario_by_id = {row.get("scenario_id"): row for row in product.get("scenarios") or []}
            observation_by_id = {row.get("observation_id"): row for row in _observations(product)}
            for combination in _combinations(product):
                scenario = scenario_by_id.get(combination.get("scenario_id"), {})
                combination_observation = observation_by_id.get(
                    combination.get("amazon_observation_id"), {}
                )
                if not combination_observation:
                    combination_observation = observations.get(
                        combination.get("amazon_observation_id"), (None, {})
                    )[1]
                scenarios_ws.append([
                    _product_ean(product), combination_observation.get("amazon_brand") or product.get("brand"),
                    combination_observation.get("amazon_title") or product.get("title"),
                    combination.get("asin"),
                    str(scenario.get("supplier") or "").upper(),
                    scenario.get("supplier_alias"), scenario.get("scenario_label"),
                    scenario_requirement_label(scenario), _excel_number(scenario.get("cost_gross_unit_eur")),
                    _excel_number(combination_observation.get("reference_price")),
                    combination_observation.get("bsr_beauty"),
                    combination_observation.get("fba_sellers"),
                    combination_observation.get("total_sellers"),
                    None, None, None, None, None, None, None,
                    "Raccomandata" if combination.get("combination_id") == recommended_pair.get("combination_id") else "",
                    combination.get("evaluation_status"), scenario.get("stock"),
                    f"{scenario.get('lead_time') or '—'} · "
                    f"{scenario.get('snapshot_at') or ''} · "
                    f"{scenario.get('freshness_status') or ''}",
                    scenario.get("warehouse"),
                    scenario.get("availability_text")
                    or scenario.get("availability_status"),
                    product_id, scenario.get("scenario_id"),
                    combination.get("amazon_observation_id"), combination.get("combination_id"),
                ])

            listings = list(product.get("amazon_listings") or [])
            if not listings:
                technical, label, reason, threshold, detail = _status_detail(
                    product, None, state.get("filters") or {}, final_product_keys,
                )
                all_results_ws.append([
                    _product_ean(product), product.get("brand"), product.get("title"),
                    _supplier_names(product), len(product.get("scenarios") or []),
                    None, None, None, None, None, None, None, None, None, None,
                    None, _excel_number(_best_cost(product)), None, None, label, reason,
                    threshold, f"{technical} · {detail}".strip(" ·"),
                ])
            for listing in listings:
                combinations = _listing_combinations(product, listing)
                best = _best_combination(combinations)
                technical, label, reason, threshold, detail = _status_detail(
                    product, listing, state.get("filters") or {}, final_product_keys,
                )
                all_results_ws.append([
                    _product_ean(product), product.get("brand"), product.get("title"),
                    _supplier_names(product), len(product.get("scenarios") or []),
                    listing.get("asin"), listing.get("title"),
                    listing.get("compatibility_status"),
                    "Sì" if listing.get("beauty_status") == "display_group_beauty" else "No",
                    listing.get("bsr_beauty"), _excel_number(listing.get("reference_price")),
                    listing.get("price_source"), _excel_number(listing.get("min_fba_price")),
                    _excel_number(listing.get("min_fbm_price")), listing.get("fba_sellers"),
                    listing.get("total_sellers"), _excel_number(_best_cost(product)),
                    (_value(_as_decimal((best or {}).get("margin_percent")) / Decimal("100"))
                     if (best or {}).get("margin_percent") is not None else None),
                    (best or {}).get("score"), label, reason, threshold,
                    f"{technical} · {detail}".strip(" ·"),
                ])
                link = (
                    f"https://www.amazon.it/gp/offer-listing/{listing.get('asin')}"
                    if listing.get("asin") else None
                )
                buy_box = (
                    listing.get("reference_price")
                    if listing.get("price_source") == "buy_box" else listing.get("buy_box_price")
                )
                listings_ws.append([
                    _product_ean(product), listing.get("asin"), listing.get("title"),
                    listing.get("brand"), listing.get("compatibility_status"),
                    _detail_text(listing.get("compatibility_reason")),
                    "Sì" if listing.get("beauty_status") == "display_group_beauty" else "No",
                    listing.get("display_group"), listing.get("bsr_beauty"),
                    listing.get("catalog_status"), _excel_number(buy_box),
                    _excel_number(listing.get("min_fba_price")), _excel_number(listing.get("min_fbm_price")),
                    listing.get("price_source"), listing.get("fba_sellers"),
                    listing.get("total_sellers"), listing.get("pricing_status"),
                    listing.get("competition_status"), listing.get("fee_status"),
                    listing.get("fee_attempts"),
                    listing.get("exclusion_reason") or reason, link,
                ])

        for observation_id, (product_id, observation) in observations.items():
            fee = observation.get("fee_estimate") or {}
            data_ws.append([
                observation_id, product_id, observation.get("asin"), observation.get("bsr_beauty"),
                _excel_number(observation.get("reference_price")), observation.get("fba_sellers"),
                observation.get("total_sellers"), _excel_number(fee.get("fba_fee_net")),
                _excel_number(fee.get("fba_fee_gross")), _excel_number(fee.get("referral_fee")),
                _excel_number(observation.get("referral_rate") or fee.get("referral_rate")),
                observation.get("referral_source"),
                observation.get("price_source"), observation.get("seller_count_source"),
                observation.get("observed_at"),
                _excel_number(observation.get("min_fba_price")),
                _excel_number(observation.get("min_fbm_price")),
            ])
        for label, value in _run_metadata(state, candidates, final_products):
            run_ws.append([label, _value(value)])
        last_data_row = max(2, data_ws.max_row)

        for row in range(2, opportunities.max_row + 1):
            margin, profit, targets, score, opportunity = _economic_formulas(
                row, cost_col="G", margin_col="M", profit_col="N",
                target_cols=("O", "P", "Q"), score_col="R", opportunity_col="S",
                observation_id_col="Y",
                last_data_row=last_data_row,
            )
            opportunities[f"M{row}"] = margin
            opportunities[f"N{row}"] = profit
            opportunities[f"O{row}"] = targets["O"]
            opportunities[f"P{row}"] = targets["P"]
            opportunities[f"Q{row}"] = targets["Q"]
            opportunities[f"R{row}"] = score
            opportunities[f"S{row}"] = opportunity
            opportunities[f"M{row}"].number_format = "0.00%"
            opportunities[f"R{row}"].number_format = "0"
            for column in ("G", "J", "N", "O", "P", "Q"):
                opportunities[f"{column}{row}"].number_format = '€ #,##0.00'
            link = opportunities[f"V{row}"]
            if link.value:
                link.hyperlink = link.value
                link.style = "Hyperlink"

        for row in range(2, scenarios_ws.max_row + 1):
            margin, profit, targets, score, opportunity = _economic_formulas(
                row, cost_col="I", margin_col="N", profit_col="O",
                target_cols=("P", "Q", "R"), score_col="S", opportunity_col="T",
                observation_id_col="AC",
                last_data_row=last_data_row,
            )
            scenarios_ws[f"N{row}"] = margin
            scenarios_ws[f"O{row}"] = profit
            scenarios_ws[f"P{row}"] = targets["P"]
            scenarios_ws[f"Q{row}"] = targets["Q"]
            scenarios_ws[f"R{row}"] = targets["R"]
            scenarios_ws[f"S{row}"] = score
            scenarios_ws[f"T{row}"] = opportunity
            scenarios_ws[f"N{row}"].number_format = "0.00%"
            scenarios_ws[f"S{row}"].number_format = "0"
            for column in ("I", "J", "O", "P", "Q", "R"):
                scenarios_ws[f"{column}{row}"].number_format = '€ #,##0.00'

        _style_sheet(
            opportunities, visible_columns=len(OPPORTUNITY_COLUMNS), cost_column="G",
            hidden_columns=("W", "X", "Y", "Z"), data_ws=data_ws,
        )
        _style_sheet(
            scenarios_ws, visible_columns=len(SCENARIO_COLUMNS), cost_column="I",
            hidden_columns=("AA", "AB", "AC", "AD"), data_ws=data_ws,
        )
        _style_audit_sheet(
            all_results_ws,
            [16, 20, 42, 28, 14, 14, 44, 18, 10, 12, 16, 16, 12, 12,
             14, 15, 18, 18, 16, 24, 48, 28, 42],
            currency_columns=(11, 13, 14, 17), percent_columns=(18,),
            integer_columns=(5, 10, 15, 16, 19), text_columns=(1, 6),
        )
        _style_audit_sheet(
            listings_ws,
            [16, 14, 46, 20, 20, 42, 10, 32, 12, 18, 12, 12, 12, 16,
             14, 15, 18, 20, 16, 12, 44, 42],
            currency_columns=(11, 12, 13), integer_columns=(9, 15, 16, 20),
            text_columns=(1, 2), hyperlink_column=22, formula_hyperlink=True,
        )
        _style_audit_sheet(run_ws, [28, 54])
        run_ws.auto_filter.ref = None
        run_ws.freeze_panes = "A2"
        opportunity_widths = [16, 20, 45, 12, 22, 18, 12, 14, 12, 18, 14, 15, 17, 12, 12, 12, 12, 9, 20, 16, 16, 42]
        scenario_widths = [16, 20, 45, 14, 12, 16, 22, 18, 12, 18, 12, 14, 15, 17, 12, 12, 12, 12, 9, 20, 16, 20, 10, 32, 14, 20]
        for ws, widths in ((opportunities, opportunity_widths), (scenarios_ws, scenario_widths)):
            for index, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(index)].width = width
            ws.sheet_view.showGridLines = False
            for row in range(2, ws.max_row + 1):
                ws.cell(row, 1).number_format = "@"
                if ws is opportunities:
                    ws.cell(row, 8).number_format = "@"
                else:
                    ws.cell(row, 4).number_format = "@"
        data_ws.freeze_panes = "A2"
        data_ws.auto_filter.ref = f"A1:{get_column_letter(data_ws.max_column)}{max(1, data_ws.max_row)}"
        for row in range(2, data_ws.max_row + 1):
            for column in (5, 8, 9, 10, 16, 17):
                data_ws.cell(row, column).number_format = '€ #,##0.00'
            data_ws.cell(row, 11).number_format = "0.00%"

        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcOnSave = True
        workbook.save(temporary_path)
        validate_excel_compatibility(temporary_path)
        os.replace(temporary_path, output_path)
    except Exception:
        try:
            os.unlink(temporary_path)
        except FileNotFoundError:
            pass
        raise
    return output_path


def write_discovery_operational_excel(results, output_file, *, progress=None):
    """Write the bounded, daily-use workbook for final opportunities only.

    The full technical export remains available through ``write_discovery_excel``.
    This workbook keeps the editable economics model but includes only the
    observations referenced by final opportunities.
    """
    state, candidates, final_products = _export_context(results)
    final_products = _sorted_final_products(final_products)
    output_path = os.path.abspath(output_file)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    descriptor, temporary_path = tempfile.mkstemp(
        prefix=".glowup_discovery_operational_", suffix=".xlsx",
        dir=os.path.dirname(output_path) or ".",
    )
    os.close(descriptor)
    try:
        workbook = Workbook()
        opportunities = workbook.active
        opportunities.title = "Opportunità"
        data_ws = workbook.create_sheet("Dati")
        run_ws = workbook.create_sheet("Parametri run")
        opportunities.append(OPPORTUNITY_COLUMNS + [
            "ProductRowID", "ScenarioRowID", "ObservationRowID", "CombinationRowID",
        ])
        data_ws.append(TECHNICAL_COLUMNS)
        run_ws.append(RUN_COLUMNS)

        opportunity_records = []
        observations = {}
        for product in final_products:
            product_id = product.get("product_key") or f"product-{_product_ean(product)}"
            recommended_pair = recommended_combination(product)
            scenario_by_id = {
                row.get("scenario_id"): row for row in product.get("scenarios") or []
            }
            recommended = scenario_by_id.get(
                (recommended_pair or {}).get("scenario_id")
            ) or _recommended(product)
            if not recommended:
                continue
            if recommended_pair:
                observation = next(
                    (row for row in _observations(product)
                     if row.get("observation_id") == recommended_pair.get("amazon_observation_id")),
                    product.get("amazon_observation") or {},
                )
            else:
                observation = product.get("amazon_observation") or {}
                recommended_pair = {
                    "combination_id": f"legacy|{recommended.get('scenario_id')}",
                    "scenario_id": recommended.get("scenario_id"),
                    "asin": observation.get("asin"),
                    "amazon_observation_id": observation.get("observation_id") or product_id,
                }
            observation_id = (
                recommended_pair.get("amazon_observation_id")
                or observation.get("observation_id") or product_id
            )
            observations.setdefault(observation_id, (product_id, observation))
            opportunity_records.append((
                product, product_id, recommended_pair, recommended,
                observation_id, observation,
            ))

        for observation_id, (product_id, observation) in observations.items():
            fee = observation.get("fee_estimate") or {}
            data_ws.append([
                observation_id, product_id, observation.get("asin"),
                _excel_number(observation.get("bsr_beauty")),
                _excel_number(observation.get("reference_price")),
                _excel_number(observation.get("fba_sellers")),
                _excel_number(observation.get("total_sellers")),
                _excel_number(fee.get("fba_fee_net")),
                _excel_number(fee.get("fba_fee_gross")),
                _excel_number(fee.get("referral_fee")),
                _excel_number(observation.get("referral_rate") or fee.get("referral_rate")),
                observation.get("referral_source"), observation.get("price_source"),
                observation.get("seller_count_source"), observation.get("observed_at"),
                _excel_number(observation.get("min_fba_price")),
                _excel_number(observation.get("min_fbm_price")),
            ])

        last_data_row = max(2, len(observations) + 1)
        for row_number, record in enumerate(opportunity_records, start=2):
            product, product_id, recommended_pair, recommended, observation_id, observation = record
            margin, profit, targets, score, opportunity = _economic_formulas(
                row_number, cost_col="G", margin_col="M", profit_col="N",
                target_cols=("O", "P", "Q"), score_col="R", opportunity_col="S",
                observation_id_col="Y", last_data_row=last_data_row,
            )
            opportunities.append([
                _product_ean(product), observation.get("amazon_brand") or product.get("brand"),
                observation.get("amazon_title") or product.get("title"),
                str(recommended.get("supplier") or "").upper(),
                recommended.get("scenario_label"), scenario_requirement_label(recommended),
                _excel_number(recommended.get("cost_gross_unit_eur")),
                recommended_pair.get("asin"), observation.get("bsr_beauty"),
                _excel_number(observation.get("reference_price")),
                observation.get("fba_sellers"), observation.get("total_sellers"),
                margin, profit, targets["O"], targets["P"], targets["Q"], score, opportunity,
                len(product.get("scenarios") or []),
                len(product.get("amazon_listings") or [observation]),
                product.get("amazon_offers_url"), product_id,
                recommended.get("scenario_id"), observation_id,
                recommended_pair.get("combination_id"),
            ])
            if progress and row_number % 100 == 0:
                progress("export_operational_rows", row_number - 1, len(opportunity_records))

        for label, value in _run_metadata(state, candidates, final_products):
            run_ws.append([label, _value(value)])
        run_ws.append(["Tipo export", "Operational workbook"])

        _style_sheet(
            opportunities, visible_columns=len(OPPORTUNITY_COLUMNS), cost_column="G",
            hidden_columns=("W", "X", "Y", "Z"), data_ws=data_ws,
        )
        _style_audit_sheet(
            data_ws, [28] * len(TECHNICAL_COLUMNS),
            currency_columns=(5, 8, 9, 10, 16, 17), percent_columns=(11,),
            integer_columns=(4, 6, 7), text_columns=(1, 2, 3),
        )
        _style_audit_sheet(run_ws, [28, 54])
        run_ws.auto_filter.ref = None
        run_ws.freeze_panes = "A2"
        for row in range(2, opportunities.max_row + 1):
            opportunities.cell(row, 1).number_format = "@"
            opportunities.cell(row, 8).number_format = "@"
            opportunities.cell(row, 13).number_format = "0.00%"
            opportunities.cell(row, 18).number_format = "#,##0"
            for column in (7, 10, 14, 15, 16, 17):
                opportunities.cell(row, column).number_format = '€ #,##0.00'
            link = opportunities.cell(row, 22)
            if link.value:
                link.hyperlink = str(link.value)
                link.style = "Hyperlink"

        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcOnSave = True
        workbook.save(temporary_path)
        validate_excel_compatibility(temporary_path)
        os.replace(temporary_path, output_path)
    except Exception:
        try:
            os.unlink(temporary_path)
        except FileNotFoundError:
            pass
        raise
    return output_path


def _stream_cell(ws, value, *, header=False, number_format=None, unlocked=False,
                 hyperlink=None):
    cell = WriteOnlyCell(ws, value=_value(value))
    if header:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if number_format:
        cell.number_format = number_format
    if unlocked:
        cell.protection = Protection(locked=False)
        cell.fill = PatternFill("solid", fgColor="EAF2FE")
    if hyperlink:
        cell.hyperlink = str(hyperlink)
        cell.style = "Hyperlink"
    return cell


def _stream_append(ws, values, *, header=False, currency_columns=(),
                   percent_columns=(), integer_columns=(), text_columns=(),
                   unlocked_columns=(), hyperlink_columns=(),
                   formula_hyperlink_columns=()):
    cells = []
    for index, value in enumerate(values, start=1):
        fmt = None
        numeric = False
        if index in currency_columns:
            fmt = '€ #,##0.00'
            numeric = True
        elif index in percent_columns:
            fmt = '0.00%'
            numeric = True
        elif index in integer_columns:
            fmt = '#,##0'
            numeric = True
        elif index in text_columns:
            fmt = '@'
        hyperlink = value if index in hyperlink_columns and value else None
        if index in formula_hyperlink_columns and value:
            value = _hyperlink_formula(value)
        cells.append(_stream_cell(
            ws, _excel_number(value) if numeric and not header else value,
            header=header, number_format=fmt,
            unlocked=index in unlocked_columns, hyperlink=hyperlink,
        ))
        if index in formula_hyperlink_columns and value:
            cells[-1].style = "Hyperlink"
    ws.append(cells)


def _configure_stream_sheet(ws, widths, *, visible_columns=None,
                            hidden_columns=(), protected=False):
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for column in hidden_columns:
        ws.column_dimensions[column].hidden = True
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    if protected:
        ws.protection.sheet = True
        ws.protection.autoFilter = False
        ws.protection.sort = False
        ws.protection.formatColumns = False
        ws.protection.selectLockedCells = False
        ws.protection.selectUnlockedCells = False


def _write_incremental_discovery_excel(
    state, candidates, final_products, output_file, *, progress=None,
):
    """Write a large incremental job without retaining worksheet cells.

    The collections are repeatable, bounded SQLite views.  openpyxl's
    write-only worksheets keep only the current row in memory while retaining
    formulas, filters, hyperlinks, hidden technical columns and run metadata.
    """
    final_products = _sorted_final_products(final_products)
    output_path = os.path.abspath(output_file)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    descriptor, temporary_path = tempfile.mkstemp(
        prefix=".glowup_discovery_", suffix=".xlsx",
        dir=os.path.dirname(output_path) or ".",
    )
    os.close(descriptor)
    try:
        workbook = Workbook(write_only=True)
        opportunities = workbook.create_sheet("Opportunità")
        all_results_ws = workbook.create_sheet("Tutti i risultati")
        listings_ws = workbook.create_sheet("Listing Amazon")
        scenarios_ws = workbook.create_sheet("Scenari")
        data_ws = workbook.create_sheet("Dati")
        run_ws = workbook.create_sheet("Parametri run")

        opportunity_widths = [16, 20, 45, 12, 22, 18, 12, 14, 12, 18, 14, 15, 17, 12, 12, 12, 12, 9, 20, 16, 16, 42]
        scenario_widths = [16, 20, 45, 14, 12, 16, 22, 18, 12, 18, 12, 14, 15, 17, 12, 12, 12, 12, 9, 20, 16, 20, 10, 32, 14, 20]
        _configure_stream_sheet(
            opportunities, opportunity_widths, visible_columns=len(OPPORTUNITY_COLUMNS),
            hidden_columns=("W", "X", "Y", "Z"), protected=True,
        )
        _configure_stream_sheet(
            scenarios_ws, scenario_widths, visible_columns=len(SCENARIO_COLUMNS),
            hidden_columns=("AA", "AB", "AC", "AD"), protected=True,
        )
        _configure_stream_sheet(
            all_results_ws,
            [16, 20, 42, 28, 14, 14, 44, 18, 10, 12, 16, 16, 12, 12,
             14, 15, 18, 18, 16, 24, 48, 28, 42],
        )
        _configure_stream_sheet(
            listings_ws,
            [16, 14, 46, 20, 20, 42, 10, 32, 12, 18, 12, 12, 12, 16,
             14, 15, 18, 20, 16, 12, 44, 42],
        )
        _configure_stream_sheet(data_ws, [28] * len(TECHNICAL_COLUMNS))
        _configure_stream_sheet(run_ws, [28, 54])
        data_ws.sheet_state = "hidden"

        _stream_append(opportunities, OPPORTUNITY_COLUMNS + [
            "ProductRowID", "ScenarioRowID", "ObservationRowID", "CombinationRowID"
        ], header=True)
        _stream_append(all_results_ws, ALL_RESULTS_COLUMNS, header=True)
        _stream_append(listings_ws, LISTING_COLUMNS, header=True)
        _stream_append(scenarios_ws, SCENARIO_COLUMNS + [
            "ProductRowID", "ScenarioRowID", "ObservationRowID", "CombinationRowID"
        ], header=True)
        _stream_append(data_ws, TECHNICAL_COLUMNS, header=True)
        _stream_append(run_ws, RUN_COLUMNS, header=True)

        observations = {}
        for observation in state.get("amazon_observations") or []:
            observation_id = observation.get("observation_id")
            if observation_id:
                keys = (observation.get("diagnostics") or {}).get("product_keys") or []
                observations.setdefault(observation_id, (keys[0] if keys else "", observation))
        final_product_keys = {
            row.get("product_key") for row in final_products if row.get("product_key")
        }
        last_data_row = max(2, len(observations) + 1)

        opportunity_row = 1
        for product in final_products:
            product_id = product.get("product_key") or f"product-{_product_ean(product)}"
            recommended_pair = recommended_combination(product)
            scenario_by_id = {
                row.get("scenario_id"): row for row in product.get("scenarios") or []
            }
            recommended = scenario_by_id.get(
                (recommended_pair or {}).get("scenario_id")
            ) or _recommended(product)
            if not recommended:
                continue
            if recommended_pair:
                observation = next(
                    (row for row in _observations(product)
                     if row.get("observation_id") == recommended_pair.get("amazon_observation_id")),
                    product.get("amazon_observation") or {},
                )
            else:
                observation = product.get("amazon_observation") or {}
                recommended_pair = {
                    "combination_id": f"legacy|{recommended.get('scenario_id')}",
                    "scenario_id": recommended.get("scenario_id"),
                    "asin": observation.get("asin"),
                    "amazon_observation_id": observation.get("observation_id") or product_id,
                }
            opportunity_row += 1
            margin, profit, targets, score, opportunity = _economic_formulas(
                opportunity_row, cost_col="G", margin_col="M", profit_col="N",
                target_cols=("O", "P", "Q"), score_col="R", opportunity_col="S",
                observation_id_col="Y",
                last_data_row=last_data_row,
            )
            link = product.get("amazon_offers_url")
            values = [
                _product_ean(product), observation.get("amazon_brand") or product.get("brand"),
                observation.get("amazon_title") or product.get("title"),
                str(recommended.get("supplier") or "").upper(),
                recommended.get("scenario_label"), scenario_requirement_label(recommended),
                recommended.get("cost_gross_unit_eur"), recommended_pair.get("asin"),
                observation.get("bsr_beauty"), observation.get("reference_price"),
                observation.get("fba_sellers"), observation.get("total_sellers"),
                margin, profit, targets["O"], targets["P"], targets["Q"], score, opportunity,
                len(product.get("scenarios") or []),
                len(product.get("amazon_listings") or [observation]), link, product_id,
                recommended.get("scenario_id"), recommended_pair.get("amazon_observation_id"),
                recommended_pair.get("combination_id"),
            ]
            _stream_append(
                opportunities, values, currency_columns=(7, 10, 14, 15, 16, 17),
                percent_columns=(13,), integer_columns=(18,), text_columns=(1, 8),
                unlocked_columns=(7,), hyperlink_columns=(22,),
            )

        scenario_row = all_results_row = listing_row = 1
        processed = 0
        for product in candidates:
            product_id = product.get("product_key") or f"product-{_product_ean(product)}"
            recommended_pair = recommended_combination(product)
            if not recommended_pair:
                legacy_recommended = _recommended(product) or {}
                legacy_observation = product.get("amazon_observation") or {}
                recommended_pair = {
                    "combination_id": f"legacy|{legacy_recommended.get('scenario_id')}",
                    "scenario_id": legacy_recommended.get("scenario_id"),
                    "asin": legacy_observation.get("asin"),
                    "amazon_observation_id": legacy_observation.get("observation_id") or product_id,
                }
            scenario_by_id = {row.get("scenario_id"): row for row in product.get("scenarios") or []}
            observation_by_id = {row.get("observation_id"): row for row in _observations(product)}
            for combination in _combinations(product):
                scenario = scenario_by_id.get(combination.get("scenario_id"), {})
                combination_observation = observation_by_id.get(
                    combination.get("amazon_observation_id"), {}
                ) or observations.get(combination.get("amazon_observation_id"), (None, {}))[1]
                scenario_row += 1
                margin, profit, targets, score, opportunity = _economic_formulas(
                    scenario_row, cost_col="I", margin_col="N", profit_col="O",
                    target_cols=("P", "Q", "R"), score_col="S", opportunity_col="T",
                    observation_id_col="AC",
                    last_data_row=last_data_row,
                )
                _stream_append(scenarios_ws, [
                    _product_ean(product), combination_observation.get("amazon_brand") or product.get("brand"),
                    combination_observation.get("amazon_title") or product.get("title"),
                    combination.get("asin"), str(scenario.get("supplier") or "").upper(),
                    scenario.get("supplier_alias"), scenario.get("scenario_label"),
                    scenario_requirement_label(scenario), scenario.get("cost_gross_unit_eur"),
                    combination_observation.get("reference_price"), combination_observation.get("bsr_beauty"),
                    combination_observation.get("fba_sellers"), combination_observation.get("total_sellers"),
                    margin, profit, targets["P"], targets["Q"], targets["R"], score, opportunity,
                    "Raccomandata" if combination.get("combination_id") == recommended_pair.get("combination_id") else "",
                    combination.get("evaluation_status"), scenario.get("stock"),
                    f"{scenario.get('lead_time') or '—'} · {scenario.get('snapshot_at') or ''} · {scenario.get('freshness_status') or ''}",
                    scenario.get("warehouse"), scenario.get("availability_text") or scenario.get("availability_status"),
                    product_id, scenario.get("scenario_id"), combination.get("amazon_observation_id"),
                    combination.get("combination_id"),
                ], currency_columns=(9, 10, 15, 16, 17, 18), percent_columns=(14,),
                    integer_columns=(19,), text_columns=(1, 4), unlocked_columns=(9,))

            listings = product.get("amazon_listings") or []
            if not listings:
                technical, label, reason, threshold, detail = _status_detail(
                    product, None, state.get("filters") or {}, final_product_keys,
                )
                all_results_row += 1
                _stream_append(all_results_ws, [
                    _product_ean(product), product.get("brand"), product.get("title"),
                    _supplier_names(product), len(product.get("scenarios") or []),
                    None, None, None, None, None, None, None, None, None, None, None,
                    _best_cost(product), None, None, label, reason, threshold,
                    f"{technical} · {detail}".strip(" ·"),
                ], currency_columns=(11, 13, 14, 17), percent_columns=(18,),
                    integer_columns=(5, 10, 15, 16, 19), text_columns=(1, 6))
            for listing in listings:
                combinations = _listing_combinations(product, listing)
                best = _best_combination(combinations)
                technical, label, reason, threshold, detail = _status_detail(
                    product, listing, state.get("filters") or {}, final_product_keys,
                )
                all_results_row += 1
                _stream_append(all_results_ws, [
                    _product_ean(product), product.get("brand"), product.get("title"),
                    _supplier_names(product), len(product.get("scenarios") or []),
                    listing.get("asin"), listing.get("title"), listing.get("compatibility_status"),
                    "Sì" if listing.get("beauty_status") == "display_group_beauty" else "No",
                    listing.get("bsr_beauty"), listing.get("reference_price"), listing.get("price_source"),
                    listing.get("min_fba_price"), listing.get("min_fbm_price"), listing.get("fba_sellers"),
                    listing.get("total_sellers"), _best_cost(product),
                    (_as_decimal((best or {}).get("margin_percent")) / Decimal("100")
                     if (best or {}).get("margin_percent") is not None else None),
                    (best or {}).get("score"), label, reason, threshold,
                    f"{technical} · {detail}".strip(" ·"),
                ], currency_columns=(11, 13, 14, 17), percent_columns=(18,),
                    integer_columns=(5, 10, 15, 16, 19), text_columns=(1, 6))
                listing_row += 1
                link = f"https://www.amazon.it/gp/offer-listing/{listing.get('asin')}" if listing.get("asin") else None
                buy_box = listing.get("reference_price") if listing.get("price_source") == "buy_box" else listing.get("buy_box_price")
                _stream_append(listings_ws, [
                    _product_ean(product), listing.get("asin"), listing.get("title"), listing.get("brand"),
                    listing.get("compatibility_status"), _detail_text(listing.get("compatibility_reason")),
                    "Sì" if listing.get("beauty_status") == "display_group_beauty" else "No",
                    listing.get("display_group"), listing.get("bsr_beauty"), listing.get("catalog_status"),
                    buy_box, listing.get("min_fba_price"), listing.get("min_fbm_price"),
                    listing.get("price_source"), listing.get("fba_sellers"), listing.get("total_sellers"),
                    listing.get("pricing_status"), listing.get("competition_status"), listing.get("fee_status"),
                    listing.get("fee_attempts"), listing.get("exclusion_reason") or reason, link,
                ], currency_columns=(11, 12, 13), integer_columns=(9, 15, 16, 20),
                    text_columns=(1, 2), formula_hyperlink_columns=(22,))
            processed += 1
            if progress and processed % 250 == 0:
                progress("export_rows", processed, len(candidates))

        for observation_id, (product_id, observation) in observations.items():
            fee = observation.get("fee_estimate") or {}
            _stream_append(data_ws, [
                observation_id, product_id, observation.get("asin"), observation.get("bsr_beauty"),
                observation.get("reference_price"), observation.get("fba_sellers"),
                observation.get("total_sellers"), fee.get("fba_fee_net"), fee.get("fba_fee_gross"),
                fee.get("referral_fee"), observation.get("referral_rate") or fee.get("referral_rate"),
                observation.get("referral_source"), observation.get("price_source"),
                observation.get("seller_count_source"), observation.get("observed_at"),
                observation.get("min_fba_price"), observation.get("min_fbm_price"),
            ], text_columns=(1, 2, 3), currency_columns=(5, 8, 9, 10, 16, 17),
                percent_columns=(11,), integer_columns=(4, 6, 7))
        for label, value in _run_metadata(state, candidates, final_products):
            _stream_append(run_ws, [label, value])

        opportunities.auto_filter.ref = f"A1:{get_column_letter(len(OPPORTUNITY_COLUMNS))}{opportunity_row}"
        scenarios_ws.auto_filter.ref = f"A1:{get_column_letter(len(SCENARIO_COLUMNS))}{scenario_row}"
        all_results_ws.auto_filter.ref = f"A1:{get_column_letter(len(ALL_RESULTS_COLUMNS))}{all_results_row}"
        listings_ws.auto_filter.ref = f"A1:{get_column_letter(len(LISTING_COLUMNS))}{listing_row}"
        data_ws.auto_filter.ref = f"A1:{get_column_letter(len(TECHNICAL_COLUMNS))}{max(2, len(observations)+1)}"
        workbook.calculation.calcMode = "auto"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcOnSave = True
        if progress:
            progress("saving", processed, len(candidates))
        workbook.save(temporary_path)
        validate_excel_compatibility(temporary_path)
        os.replace(temporary_path, output_path)
    except Exception:
        try:
            os.unlink(temporary_path)
        except FileNotFoundError:
            pass
        raise
    return output_path
